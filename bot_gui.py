import json
import logging
import threading
import asyncio
import os
import re
import sys
import time
import math
import webbrowser
import subprocess
try:
    import winreg
except ImportError:  # non-Windows
    winreg = None
import tkinter as tk
from tkinter import ttk, messagebox
from collections import deque

BASE_DIR = os.path.abspath(os.path.dirname(__file__))

# ── [PATCH-11] 버전 관리 ──
APP_VERSION = "1.1.0"
APP_NAME = "Binance Auto Trading Bot"
GITHUB_REPO = "autobottrading2026-byte/Binance-auto-bot"
# ── 레퍼럴 코드 난독화 + 무결성 체크 ──
import base64 as _b64, hashlib as _hl
_REF_ENC = "QVVUT zIwMjY="  # base64 인코딩된 레퍼럴 코드 (공백 삽입 난독화)
_REF_HASH = "9a61a4dc110e084e"  # 원본 해시 앞 16자리
def _decode_ref() -> str:
    raw = _b64.b64decode(_REF_ENC.replace(" ", "")).decode()
    h = _hl.sha256(raw.encode()).hexdigest()[:16]
    if h != _REF_HASH:
        return ""
    return raw
REFERRAL_CODE = _decode_ref()

# ── Lemon Squeezy 결제 링크 (프리미엄 구독) ──
LEMONSQUEEZY_MONTHLY_URL = ""  # TODO: Lemon Squeezy Payment Link 생성 후 입력
LEMONSQUEEZY_YEARLY_URL  = ""  # TODO: Lemon Squeezy Payment Link 생성 후 입력
PREMIUM_PRICE_MONTHLY = "$9.99"
PREMIUM_PRICE_YEARLY  = "$99/yr"

BOT_ROOT = os.path.join(BASE_DIR, "binance_futures_bot1_1")

# binance_futures_bot1_1 패키지를 찾을 수 있도록 BASE_DIR을 sys.path 최우선에 추가
if BASE_DIR not in sys.path:
    sys.path.insert(0, BASE_DIR)
LOG_DIR = os.path.join(BOT_ROOT, "logs")

# A1: consent version — bump this when the risk notice text changes to force re-acknowledgement
CONSENT_VERSION = "v1.1-d8de2b46"

TITLE = "바이낸스 자동 매매 UI v1.1"
CONFIG_PATH = "gui_config.json"
STATE_PATH = "gui_state.json"
LOG_PATH = os.path.join(LOG_DIR, "bot.log")
NOTIFICATION_PATH = os.path.join(LOG_DIR, "notifications.log")
TRADE_LOG_PATH = os.path.join(LOG_DIR, "trade_history.jsonl")
AUTO_TUNE_STATE_PATH = os.path.join(LOG_DIR, "auto_tuner_state.json")
MAX_LOG_BYTES = 2_000_000
LOG_KEEP_BYTES = 1_500_000



# ------------------------------------------------------------------
# Locked settings: these are enforced on load and on every save.
# Rationale: remove end-user access to auto-tuning controls.
LOCKED_GUI_SETTINGS = {
    "auto_tune_enabled": True,
    # auto_tune_mode는 사용자가 프리셋으로 선택하므로 잠금 제외
}
class BotGUI:
    def __init__(self):
        self.settings_data = self._load_json(CONFIG_PATH, default={})
        defaults = {
            "position_base_pct": 0.055,
            "position_pct": 0.05,
            "top_n": 20,
            "volatility_min": 0.001,
            "momentum_min_long": 0.001,
            "momentum_min_short": -0.001,
            "auto_tune_enabled": True,
            "leverage_min": 5,
            "leverage_max": 25,
            "watch_limit": 10,
            "manual_symbols": ["BTCUSDT", "ETHUSDT"],
            "show_manual_panel": False,
            "auto_boost_position_pct": False,
            "spike_guard_enabled": True,
            "spike_guard_pct": 5.0,   # 8초 내 5% 역방향 급변동 시 즉시 시장가 청산
            "spike_guard_window": 8,
            "spike_guard_interval": 2,
            "global_spike_cooldown_min": 5,
            "spark_reentry_candles": 3,
            "session_loss_limit_pct": 5.0,
            "session_loss_window_minutes": 1440,
            "kill_switch_cooldown_min": 120,
            "auto_tune_cooldown_min": 10,
            "auto_tune_mode": "balanced",
            "risk_acknowledged": False,
            "risk_ack1": False,
            "expert_mode_enabled": False,
            "risk_ack2": False,
            "neural_scorer_enabled": False,
            "neural_license_key": "",
            "consent_version": "",
            "binance_referral_code": REFERRAL_CODE,
            "maker_fee_pct": 0.0002,
            "taker_fee_pct": 0.0005,
            "limit_exit_offset_bps": 2.0,
            "max_loss_per_position": 18.0,
            "max_open_symbols": 10,          # [추가] 동시 최대 오픈 심볼 수 기본값
            "enable_profit_exit_layer": True,
            "enable_partial_take_profit": True,
            "enable_atr_trailing_stop": True,
            "enable_progress_stop": True,
            "partial_tp_levels": [
                {"r": 1.0, "close_frac": 0.30},
                {"r": 1.5, "close_frac": 0.30},
                {"r": 2.0, "close_frac": 1.00},
            ],
            "trail_atr_period": 22,
            "trail_atr_mult": 1.5,
            "trail_activate_pnl_pct": 0.10,
            "trail_use_highest_since_entry": True,
            "trail_recalc_interval_sec": 5,
            "progress_stop_lookback_sec": 1800,
            "progress_stop_no_new_high_sec": 300,
            "progress_stop_drawdown_from_mfe": 0.07,
            "progress_stop_min_pnl_pct": 0.05,
            "progress_stop_action": "partial_or_full",
        }
        for key, value in defaults.items():
            self.settings_data.setdefault(key, value)

        # Enforce locked settings (auto-tune controls are not user-editable).
        for _k, _v in LOCKED_GUI_SETTINGS.items():
            self.settings_data[_k] = _v

        self.language = self.settings_data.get("ui_language", "ko")
        if self.language not in ("ko", "en"):
            self.language = "ko"
        self.translations = {
            "sidebar_env": {"ko": "환경", "en": "Environment"},
            "bot_status": {"ko": "봇 상태", "en": "Bot status"},
            "recent_events": {"ko": "최근 이벤트", "en": "Recent events"},
            "ui_init_done": {"ko": "[INFO] UI 초기화 완료", "en": "[INFO] UI initialized"},

            "settings_title": {"ko": "설정", "en": "Settings"},
            "settings_tab_env": {"ko": "환경설정", "en": "Environment"},
            "settings_tab_trade": {"ko": "거래설정", "en": "Trading"},
            "settings_tab_display": {"ko": "화면설정", "en": "Display"},
            "settings_tab_dev": {"ko": "필수 동의", "en": "Agreement"},
            "settings_tab_report": {"ko": "리포트", "en": "Report"},
            "api_settings_title": {"ko": "API 설정", "en": "API setup"},
            "default_env_title": {"ko": "기본 실행 환경", "en": "Default environment"},
            "language_title": {"ko": "언어 / Language", "en": "Language"},
            "monitoring_panel": {"ko": "모니터링", "en": "Monitoring"},
            "card_opt_unrealized":  {"ko": "미실현 손익 합계", "en": "Unrealized PnL"},
            "card_opt_filter":      {"ko": "필터 통과율",       "en": "Filter Pass Rate"},
            "card_opt_top_symbol":  {"ko": "최다 거래 심볼",    "en": "Top Symbol"},
            "card_opt_pnl15":       {"ko": "15분 손익",          "en": "15m PnL"},
            "card_opt_pnl60":       {"ko": "60분 손익",          "en": "60m PnL"},
            "card_opt_pnl12h":      {"ko": "12시간 손익",        "en": "12h PnL"},
            "card_opt_pnl24h":      {"ko": "24시간 손익",        "en": "24h PnL"},
            "card_opt_rr":          {"ko": "손익비 R:R",         "en": "R:R Ratio"},
            "card_opt_exp":         {"ko": "거래당 기댓값",      "en": "Expectancy"},
            "card_opt_consec":      {"ko": "최대 연속 손실",     "en": "Max Consec. Loss"},
            "display_card_defaults": {"ko": "기본 카드 패널 설정", "en": "Default Card Settings"},
            "display_card_defaults_desc": {"ko": "메인 화면 상단 4개 카드의 기본 표시 항목을 설정합니다.", "en": "Set the default display item for each of the 4 top cards."},
            "card_label_0": {"ko": "카드 1", "en": "Card 1"},
            "card_label_1": {"ko": "카드 2", "en": "Card 2"},
            "card_label_2": {"ko": "카드 3", "en": "Card 3"},
            "card_label_3": {"ko": "카드 4", "en": "Card 4"},
            "log_tab": {"ko": "로그", "en": "Log"},
            "positions_tab": {"ko": "포지션 목록", "en": "Positions"},
            "exits_tab": {"ko": "익절 · Trailing", "en": "Exits · Trailing"},
            "close_selected": {"ko": "선택 청산", "en": "Close selected"},
            "profit_intro": {
                "ko": "수익 보호 스택을 단계별로 구성합니다. 전체 Profit Layer → 부분 익절 → ATR 트레일 → Progress Stop 순으로 평가되어, 위에서 트리거되면 아래 로직은 건너뜁니다.",
                "en": "Profit exits run in stages: master layer → partial take profit → ATR trailing → progress stop. When an upper layer triggers, lower layers are skipped.",
            },
            "profit_layer_hint": {
                "ko": "→ 모든 하위 로직을 사용하려면 ON 상태를 유지하세요",
                "en": "→ Keep this ON to enable every downstream exit",
            },
            "partial_toggle_hint": {
                "ko": "→ ROI%가 설정 레벨을 넘길 때마다 잔여 수량을 부분 청산합니다",
                "en": "→ When ROI reaches each level, close a fraction of the remaining position",
            },
            "trail_toggle_hint": {
                "ko": "→ 최고가/최저가 대비 ATR x 배수로 추적 스톱을 갱신합니다",
                "en": "→ Pulls the stop closer using ATR multiples of the recent high/low",
            },
            "progress_toggle_hint": {
                "ko": "→ 일정 시간 신고점이 없거나 MFE 대비 되밀림이 커지면 부분/전체 청산을 유도합니다",
                "en": "→ Forces partial/full exit when no new high forms or when drawdown from MFE grows",
            },
            "partial_levels_title": {
                "ko": "Partial Levels (예: 30:30,60:30,80:100)",
                "en": "Partial Levels (e.g. 30:30,60:30,80:100)",
            },
            "partial_levels_desc": {
                "ko": "앞 숫자=ROI%, 뒤 숫자=해당 시점 청산 비율입니다. 쉼표로 여러 단계를 나열하세요.",
                "en": "First value = ROI %, second = close fraction at that level. Separate multiple steps with commas.",
            },
            "atr_desc": {
                "ko": "최근 최고/최저 대비 ATR 배수로 스톱을 끌어올립니다.",
                "en": "Tightens the stop using ATR multiples of the recent high/low.",
            },
            "progress_desc": {
                "ko": "신고점이 끊기거나 MFE 대비 되밀림이 커질 때 조기 청산을 유도합니다.",
                "en": "Triggers an early exit when no new highs print or when drawdown from MFE grows.",
            },
            "manual_panel_title": {"ko": "지정 코인 수동 매매", "en": "Manual trading (selected coins)"},
            "manual_size_label": {"ko": "크기 %", "en": "Size %"},
            "manual_lev_label":  {"ko": "배율 ×", "en": "Lev ×"},
            "manual_hide": {"ko": "숨기기", "en": "Hide"},
            "manual_show": {"ko": "보이기", "en": "Show"},
            "manual_no_symbols": {"ko": "수동 매매 심볼이 없습니다", "en": "No manual-trade symbols selected"},
            "manual_card_suffix": {"ko": "수동 매매", "en": "Manual trading"},

            # ── 진입 필터 설정 탭 번역 ────────────────────────────────────────────
            "entry_filter_section":       {"ko": "진입 필터 설정",                       "en": "Entry Filter Settings"},
            "direction_filter_title":     {"ko": "📌 방향 검증 필터",                    "en": "📌 Direction Filters"},
            "mtf_confirm_label":          {"ko": "MTF EMA 하드게이트 (진입 방향과 1m·5m EMA 불일치 시 차단)", "en": "MTF EMA Hard Gate (block when 1m·5m EMA conflicts with direction)"},
            "mtf_confirm_hint":           {"ko": "enable_mtf_ema_confirm: 짧은 EMA 추세가 진입 방향과 반대이면 강제 차단", "en": "enable_mtf_ema_confirm: blocks entry when short-term EMA slope opposes trade direction"},
            "short_ema_conflict_label":   {"ko": "단기 EMA 충돌 필터 (1m·5m 모두 반대 방향이면 차단)", "en": "Short EMA Conflict Filter (block when both 1m·5m EMAs oppose direction)"},
            "short_ema_conflict_hint":    {"ko": "short_ema_conflict_filter: 24h 방향과 단기 EMA 방향이 모두 반대일 때 진입 차단", "en": "short_ema_conflict_filter: blocks entry when 24h trend and short EMAs both oppose"},
            "chop_ema_dir_label":         {"ko": "CHOP 레짐 단기EMA 방향 우선 (24h 방향 대신 단기EMA 방향 따름)", "en": "CHOP Regime: Follow Short EMA (override 24h direction in choppy market)"},
            "chop_ema_dir_hint":          {"ko": "chop_use_short_ema_direction: CHOP 구간에서 24h 기준 방향 대신 1m·5m EMA 방향으로 진입", "en": "chop_use_short_ema_direction: uses 1m·5m EMA direction instead of 24h trend in CHOP regime"},
            "signal_quality_title":       {"ko": "📊 신호 품질 필터",                   "en": "📊 Signal Quality Filters"},
            "composite_signal_label":     {"ko": "복합 신호 스코어링 (모멘텀·거래량·MTF 가중 합산)", "en": "Composite Signal Scoring (momentum · volume · MTF weighted sum)"},
            "composite_min_label":        {"ko": "최소 복합 스코어:",                    "en": "Min composite score:"},
            "rsi_filter_label":           {"ko": "RSI 과열 필터 (고RSI LONG / 저RSI SHORT 차단)", "en": "RSI Extreme Filter (block LONG at high RSI / SHORT at low RSI)"},
            "rsi_ob_label":               {"ko": "과매수 차단:",                          "en": "Overbought block:"},
            "rsi_os_label":               {"ko": "과매도 차단:",                          "en": "Oversold block:"},
            "sizing_title":               {"ko": "💰 포지션 사이징",                     "en": "💰 Position Sizing"},
            "atr_risk_label":             {"ko": "ATR 리스크 사이징 (포지션당 최대 손실 한도로 수량 캡)", "en": "ATR Risk Sizing (cap quantity by max-loss-per-position)"},
            "entry_risk_label":           {"ko": "포지션당 리스크:",                      "en": "Risk per position:"},
            "kelly_label":                {"ko": "Kelly 사이징 (승률 기반 포지션 비율 자동 조정)", "en": "Kelly Sizing (win-rate based dynamic position fraction)"},
            "execution_title":            {"ko": "⚙️ 실행 설정",                         "en": "⚙️ Execution Settings"},
            "breakeven_label":            {"ko": "Breakeven Stop (첫 TP 후 손익분기 스탑 자동 이동)", "en": "Breakeven Stop (auto-move stop to breakeven after first TP)"},
            "funding_filter_label":       {"ko": "펀딩레이트 편향 필터 (과도한 쏠림 시 신호 강도 감소)", "en": "Funding Rate Bias Filter (reduce signal strength on extreme funding skew)"},
            "maker_first_label":          {"ko": "Maker-First 진입 (지정가 먼저 시도 → 실패 시 시장가)", "en": "Maker-First Entry (try limit order first → market fallback)"},
            "diversify_label":            {"ko": "워치리스트 다양성 (변동성 분산 선택)", "en": "Diversify Watchlist (spread across varied volatility symbols)"},
            "min_hold_label":             {"ko": "최소 보유 시간:",                       "en": "Min hold time:"},
            "time_stop_label":            {"ko": "최대 보유 시간:",                       "en": "Max hold time:"},
            "seconds_suffix":             {"ko": "초",                                    "en": "sec"},
            "manual_long_limit": {"ko": "롱 지정가 진입", "en": "Long limit entry"},
            "manual_long_market": {"ko": "롱 시장가 진입", "en": "Long market entry"},
            "manual_short_limit": {"ko": "숏 지정가 진입", "en": "Short limit entry"},
            "manual_short_market": {"ko": "숏 시장가 진입", "en": "Short market entry"},
            "manual_close_limit": {"ko": "지정가 청산", "en": "Close (limit)"},
            "manual_close_market": {"ko": "시장가 청산", "en": "Close (market)"},
            "env_guidance": {"ko": "이 버전부터 API 키/시크릿은 프로그램 안에서 입력하지 않습니다.\nWindows 환경 변수에 아래 키를 등록한 뒤 GUI를 다시 시작해 주세요. (제어판 → 시스템 → 고급 시스템 설정 → 환경 변수)\n필수 변수\n  • TESTNET_API_KEY / TESTNET_API_SECRET (테스트넷)\n  • BINANCE_API_KEY / BINANCE_API_SECRET (라이브)\n각 변수값은 복사/붙여넣기 후 ‘새로 만들기’ 또는 ‘편집’으로 저장하고, 변경 시 GUI를 재실행해야 적용됩니다.", "en": "From this version, API key/secret are not entered inside the app.\nSet them as Windows environment variables and restart the GUI. (Control Panel → System → Advanced system settings → Environment Variables)\nRequired variables\n  • TESTNET_API_KEY / TESTNET_API_SECRET (Testnet)\n  • BINANCE_API_KEY / BINANCE_API_SECRET (Live)\nAfter pasting, click New/Edit to save. Restart the GUI to apply changes."},
            "env_label_testnet_key": {"ko": "테스트넷 API Key", "en": "Testnet API key"},
            "env_label_testnet_secret": {"ko": "테스트넷 API Secret", "en": "Testnet API secret"},
            "env_label_live_key": {"ko": "라이브 API Key", "en": "Live API key"},
            "env_label_live_secret": {"ko": "라이브 API Secret", "en": "Live API secret"},
            "env_status_set": {"ko": "설정됨", "en": "Set"},
            "env_status_unset": {"ko": "미설정", "en": "Not set"},
            "env_notify_popup": {"ko": "주문/체결 완료 시 알림 팝업", "en": "Popup on order/fill completion"},
            "env_auto_start": {"ko": "Windows 시작 시 자동 실행", "en": "Run on Windows startup"},
            "save": {"ko": "저장", "en": "Save"},
            "defaults": {"ko": "기본값", "en": "Defaults"},
            "env_saved_msg": {"ko": "환경 설정이 저장되었습니다. (API 키는 환경 변수에서 불러옵니다)", "en": "Environment settings saved. (API keys are loaded from environment variables)"},
            "restart_prompt": {"ko": "환경 설정이 변경되었습니다. 엔진을 다시 시작할까요?", "en": "Settings changed. Restart the engine now?"},
            "api_env_check_title": {"ko": "API 환경 변수 확인", "en": "API environment variables"},
            "api_env_missing_msg": {"ko": "아직 설정되지 않은 환경 변수가 있습니다.", "en": "Some environment variables are missing."},
            "close": {"ko": "닫기", "en": "Close"},
            "watchlist_title": {"ko": "거래 대상 심볼 상태", "en": "Tracked symbols status"},
            "watchlist_header": {"ko": "최근 모니터링 심볼", "en": "Recently monitored symbols"},
            "watchlist_none": {"ko": "로그에서 추출한 심볼이 없습니다. 엔진이 실행 중인지 확인하세요.", "en": "No symbols were extracted from logs. Check if the engine is running."},
            "filter_reject_title": {"ko": "최근 필터/체결 거부 사유", "en": "Recent filter / fill-reject reasons"},
            "risk_ack_title": {"ko": "위험 고지 동의 필요", "en": "Risk acknowledgement required"},
            "risk_ack_msg": {"ko": "지정 코인 수동 매매를 사용하려면 '필수 동의' 탭에서 동의 체크박스를 활성화하세요.", "en": "To use manual trading, enable the acknowledgement checkbox in the Agreement tab."},
            "price_missing_title": {"ko": "가격 정보 없음", "en": "Price unavailable"},
            "alert": {"ko": "알림", "en": "Notice"},
            "app_title": {"ko": "바이낸스 자동 매매 UI v1.1", "en": "Binance Auto Trading UI v1.1"},
            "quick_links": {"ko": "빠른 링크", "en": "Quick links"},
            "link_binance": {"ko": "바이낸스", "en": "Binance"},
            "link_testnet": {"ko": "테스트넷", "en": "Testnet"},
            "help_btn": {"ko": "도움말", "en": "Help"},
            "settings_btn":       {"ko": "설정", "en": "Settings"},
            "expert_mode_label":  {"ko": "Expert 모드 (Aggressive 프리셋 활성화)", "en": "Expert Mode (Enable Aggressive Preset)"},
            "premium_tab_label":         {"ko": "🧠 프리미엄", "en": "🧠 Premium"},
            "premium_section_title":     {"ko": "AI 학습 스코어러", "en": "AI Learning Scorer"},
            "premium_hero_title":        {"ko": "🧠 AI 학습 스코어러", "en": "🧠 AI Learning Scorer"},
            "premium_badge":             {"ko": "PREMIUM", "en": "PREMIUM"},
            "premium_hero_desc": {
                "ko": "거래 결과를 누적 학습하여 다음 진입의 승률을 예측합니다.\n매 청산 후 자동 학습하며, 재시작해도 학습이 유지됩니다.\n학습이 쌓일수록 좋은 신호는 더 강하게, 나쁜 신호는 걸러냅니다.",
                "en": "Learns from each trade outcome to predict win probability.\nAuto-learns after every close. Learning persists across restarts.\nStrong signals are amplified; weak ones are filtered.",
            },
            "premium_feature_how_title": {"ko": "작동 방식", "en": "How It Works"},
            "premium_feature_h1":        {"ko": "📥  진입 시 12개 피처 수집", "en": "📥  12 features captured at entry"},
            "premium_feature_d1":        {"ko": "모멘텀 강도 · 변동성 · 거래량 서지 · MTF EMA 기울기 · 스프레드 · 펀딩레이트 · 레짐 · 진입 시각 등", "en": "Momentum · Volatility · Volume surge · MTF EMA slope · Spread · Funding rate · Regime · Entry time"},
            "premium_feature_h2":        {"ko": "🎓  청산 후 자동 학습", "en": "🎓  Auto-learns after close"},
            "premium_feature_d2":        {"ko": "수익(✅) / 손실(❌) 결과를 라벨로 신경망 가중치 업데이트.\nExperience Replay로 과거 거래도 반복 학습(망각 방지).", "en": "Win/Loss outcome updates neural network weights.\nExperience Replay re-trains on past trades (prevents forgetting)."},
            "premium_feature_h3":        {"ko": "🎯  다음 진입에 반영", "en": "🎯  Applied to next entry"},
            "premium_feature_d3":        {"ko": "v3: 20개 피처 + 듀얼헤드(승률+ROI예측)로 진입 보정 0.0x~2.0x.\n승률 25% 미만은 진입 차단. 적응형 학습률 자동 조절.", "en": "v3: 20 features + dual-head (win prob + ROI prediction), strength 0.0x–2.0x.\nBlocks entry when win prob < 25%. Adaptive learning rate."},
            "premium_toggle_label":      {"ko": "AI 스코어러 활성화", "en": "Enable AI Scorer"},
            "premium_toggle_on":         {"ko": "ON — 진입 strength 보정 중", "en": "ON — adjusting entry strength"},
            "premium_toggle_off":        {"ko": "OFF — 예측 미적용 (데이터 수집·학습만 진행)", "en": "OFF — prediction not applied (data collection & learning active)"},
            "premium_dashboard_title":   {"ko": "학습 현황 대시보드", "en": "Learning Dashboard"},
            "premium_stat_trades":       {"ko": "학습 거래", "en": "Trained"},
            "premium_stat_wins":         {"ko": "승리", "en": "Wins"},
            "premium_stat_losses":       {"ko": "패배", "en": "Losses"},
            "premium_stat_wr":           {"ko": "승률", "en": "Win Rate"},
            "premium_stat_acc":          {"ko": "예측 정확도", "en": "Accuracy"},
            "premium_stat_roi":          {"ko": "평균 ROI", "en": "Avg ROI"},
            "premium_stat_replay":       {"ko": "Replay 버퍼", "en": "Replay Buffer"},
            "premium_stat_status":       {"ko": "모델 상태", "en": "Model Status"},
            "premium_stat_active":       {"ko": "✅ 예측 중", "en": "✅ Predicting"},
            "premium_stat_warmup":       {"ko": "⏳ 냉각 ({n}/50)", "en": "⏳ Warmup ({n}/50)"},
            "premium_stat_paused":       {"ko": "⚠️ 정확도 미달", "en": "⚠️ Low accuracy"},
            "premium_stat_inactive":     {"ko": "🔘 비활성", "en": "🔘 Inactive"},
            "premium_feat_title":        {"ko": "피처 가중치 (학습된 중요도)", "en": "Feature Weights (Learned Importance)"},
            "premium_feat_names": {
                "ko": ["상대 모멘텀", "변동성", "거래량 서지", "1m EMA 기울기", "5m EMA 기울기",
                       "MTF 정렬도", "스프레드", "펀딩 방향", "레짐", "진입 시각(sin)", "진입 시각(cos)", "방향 일치"],
                "en": ["Rel Momentum", "Volatility", "Vol Surge", "1m EMA Slope", "5m EMA Slope",
                       "MTF Align", "Spread", "Funding", "Regime", "Time(sin)", "Time(cos)", "Dir Match"],
            },
            "premium_reset_btn":         {"ko": "모델 초기화", "en": "Reset Model"},
            "premium_refresh_btn":       {"ko": "새로고침", "en": "Refresh"},
            "premium_future_title":      {"ko": "🌐 서버 연동 (출시 예정)", "en": "🌐 Server Sync (Coming Soon)"},
            "premium_future_desc": {
                "ko": "향후 업데이트에서 각 클라이언트의 학습 데이터를 중앙 서버에 익명으로 수집,\n수만 건의 거래를 기반으로 더 강력한 딥러닝 모델을 배포할 예정입니다.\n현재는 로컬 학습만 지원합니다.",
                "en": "In a future update, anonymized learning data from all clients will be aggregated on a central server.\nA more powerful deep learning model trained on thousands of trades will be deployed.\nCurrently, only local learning is supported.",
            },
            "premium_license_label":     {"ko": "라이선스 키", "en": "License Key"},
            "premium_activate_btn":      {"ko": "활성화", "en": "Activate"},
            "premium_key_ok":            {"ko": "✅ 키 유효 — 활성화됨 ({msg})", "en": "✅ Key valid — activated ({msg})"},
            "premium_key_fail":          {"ko": "❌ {msg}", "en": "❌ {msg}"},
            "bulk_close_collapse":{"ko": "일괄 청산 접기",   "en": "Collapse Bulk Close"},
            "bulk_close_expand":  {"ko": "일괄 청산 펼치기", "en": "Expand Bulk Close"},
            "report_btn": {"ko": "리포트", "en": "Report"},
            "sidebar_manual_section": {"ko": "수동 매매", "en": "Quick close"},
            "sidebar_toggle_hide": {"ko": "숨기기", "en": "Hide"},
            "sidebar_toggle_show": {"ko": "보이기", "en": "Show"},
            "sidebar_close_profit_limit": {"ko": "수익중 · 지정가 청산", "en": "Profit · close (limit)"},
            "sidebar_close_profit_market": {"ko": "수익중 · 시장가 청산", "en": "Profit · close (market)"},
            "sidebar_close_loss_limit": {"ko": "손실중 · 지정가 청산", "en": "Loss · close (limit)"},
            "sidebar_close_loss_market": {"ko": "손실중 · 시장가 청산", "en": "Loss · close (market)"},
            "stat_unrealized_total": {"ko": "미실현 손익 합계", "en": "Unrealized PnL"},
            "stat_filter_pass_rate": {"ko": "필터 통과율", "en": "Filter Pass Rate"},
            "stat_top_symbol": {"ko": "최다 거래 심볼", "en": "Top Symbol"},
            "stat_pnl_24h": {"ko": "24시간 손익", "en": "24h PnL"},
            "stat_win_rate": {"ko": "승률", "en": "Win rate"},
            "stat_trade_count": {"ko": "거래 수", "en": "Trades"},
            "stat_notional": {"ko": "진입 총액 (USDT / 레버리지)", "en": "Notional (USDT / leverage)"},
            "stat_balance": {"ko": "계좌 잔고", "en": "Wallet balance"},
            "trade_settings_title": {"ko": "거래 설정", "en": "Trading settings"},
            "auto_tune_locked": {"ko": "Auto-tune 켜짐; 수동 설정이 잠겨 있습니다.", "en": "Auto-tune is ON; manual controls are locked."},
            "auto_tune_unlocked": {"ko": "수동 설정이 잠금 해제되었습니다.", "en": "Manual controls are unlocked."},
            "auto_tune_restart_prompt": {"ko": "자동 튜닝 토글 변경으로 엔진을 다시 시작할까요?", "en": "Auto-tune setting changed. Restart the engine now?"},
            "auto_tune_toggle_label": {"ko": "Auto-tune 활성화", "en": "Enable auto-tune"},
            "auto_tune_helper_on": {"ko": "Auto-tune이 켜져 있습니다. 최근 시장 데이터를 기반으로 파라미터를 자동 조정합니다.", "en": "Auto-tune is enabled and adjusts parameters using recent market data."},
            "auto_tune_helper_off": {"ko": "파라미터가 수동 값으로 고정됩니다. 필요 시 수정 후 저장하세요.", "en": "Parameters are fixed to manual values. Adjust and save if needed."},
            "auto_boost_label": {"ko": "포지션 비율 자동 보정 (최소 증거금 부족 시 Position % 자동 증가)", "en": "Auto-boost position % (increase entry size when margin is insufficient)"},
            "mode_aggressive": {"ko": "공격", "en": "Aggressive"},
            "mode_balanced": {"ko": "적정", "en": "Balanced"},
            "mode_conservative": {"ko": "보수", "en": "Conservative"},
            "display_title": {"ko": "화면 / 수동 매매 설정", "en": "Display / manual trading"},
            "display_manual_symbols": {"ko": "수동 매매 심볼", "en": "Manual-trade symbols"},
            "dev_title": {"ko": "필수 동의", "en": "Agreement"},
            "risk_ack_complete":        {"ko": "위험 고지 동의 완료", "en": "Risk acknowledgement complete"},
            "risk_ack_revoked":         {"ko": "위험 고지 동의가 해제되었습니다", "en": "Risk acknowledgement revoked"},
            "consent_version_changed":  {"ko": "고지문이 업데이트되었습니다. 필수 동의 탭에서 다시 동의해주세요.", "en": "The risk notice has been updated. Please re-acknowledge in the Agreement tab."},
            "dev_ack_required_note":    {"ko": "※ 두 항목 모두 체크해야 자동/수동 매매가 활성화됩니다.", "en": "※ Both items must be checked to enable auto/manual trading."},
            # A1: split consent items
            "dev_ack1_label": {"ko": "① 비투자자문·손실 가능·사용자 책임", "en": "① No investment advice · Loss possible · User responsibility"},
            "dev_ack1_text":  {"ko": "위험 고지문을 읽었으며, 이 소프트웨어가 투자 조언을 제공하지 않고 원금 전액 손실이 가능하며, 모든 결과에 대한 책임이 사용자에게 있음을 동의합니다.", "en": "I have read the risk notice and agree that this software provides no investment advice, full loss of principal is possible, and I accept full responsibility for all outcomes."},
            "dev_ack2_label": {"ko": "② 거래소 자동화 도구 리스크", "en": "② Exchange automation tool risks"},
            "dev_ack2_text":  {"ko": "Binance 자동화 도구 사용에 따른 API 차단·계정 제재·재무/법적 리스크를 이해하며, 거래소 이용약관 및 API 규정 준수 의무가 사용자에게 있음에 동의합니다.", "en": "I understand the risks of using automated tools on Binance (API bans, account restrictions, financial and legal risks) and agree that compliance with exchange terms and API policies is my responsibility."},
            "dev_ack_checkbox": {"ko": "위험 고지 내용을 읽었으며, 본 소프트웨어 사용으로 인한 모든 결과에 대해 책임집니다.", "en": "I have read the risk notice and accept full responsibility for all outcomes."},
            "profit_exit_layer_toggle": {"ko": "Profit Exit Layer 활성화", "en": "Enable Profit Exit Layer"},
            "symbol_label": {"ko": "심볼", "en": "Symbol"},
            "display_show_manual_panel": {"ko": "지정 코인 수동 매매 패널 표시", "en": "Show manual trading panel (selected coins)"},
            "display_hide_panel_hint": {"ko": "패널을 숨기면 메인 화면에서 포지션 창이 바로 위 카드 밑으로 이동합니다.", "en": "Hiding the panel moves the Positions view directly under the top cards."},
            "auto_tune_reset": {"ko": "Auto-Tune 초기화", "en": "Reset Auto-Tune"},
            "auto_tune_desc": {"ko": "실시간 튜너가 포지션 필터·모멘텀 값을 조정합니다.", "en": "The live tuner adjusts position filters and momentum parameters."},
            "auto_tune_update": {"ko": "업데이트: -", "en": "Updated: -"},
            "manual_tune_section": {"ko": "수동 튜닝", "en": "Manual tuning"},
            "manual_edit_hint": {"ko": "Auto-tune이 OFF일 때만 아래 값을 수정할 수 있습니다.", "en": "You can edit the values below only when auto-tune is OFF."},
            "recalc_sec": {"ko": "재계산(s)", "en": "Recalc (s)"},
            "risk_display_settings": {"ko": "리스크 & 표시 설정", "en": "Risk & display settings"},
            "symbol_filter_settings": {"ko": "심볼/필터 설정", "en": "Symbol / filter settings"},
            "dev_notice_1": {"ko": "※ 동의하지 않으면 자동/수동 매매 기능이 비활성화됩니다.", "en": "※ Auto and manual trading are disabled until you acknowledge the notice above."},
            "dev_notice_2": {"ko": "위 고지 내용을 읽고 이해했으며, 사용으로 인한 모든 결과에 대해 본인이 책임집니다.", "en": "I have read and understood the notice above, and accept full responsibility for all outcomes."},
            "ok": {"ko": "확인", "en": "OK"},
            "profit_exit_trailing_title": {"ko": "익절 / 트레일링", "en": "Profit Exit / Trailing"},
            "partial_tp_label": {"ko": "부분 익절", "en": "Partial Take Profit"},
            "atr_trailing_label": {"ko": "ATR 샹들리에 트레일링", "en": "ATR Chandelier Trailing"},
            "progress_stop_label": {"ko": "프로그레스 스톱", "en": "Progress Stop"},
            "atr_trail_title": {"ko": "ATR 트레일", "en": "ATR Trail"},
            "atr_period": {"ko": "기간", "en": "Period"},
            "atr_multiplier": {"ko": "배수", "en": "Multiplier"},
            "atr_activate_pct": {"ko": "활성화 %", "en": "Activate %"},
            "progress_lookback": {"ko": "룩백 (초)", "en": "Lookback s"},
            "progress_no_new_high": {"ko": "신고점 없음 (초)", "en": "No-new-high s"},
            "progress_drawdown": {"ko": "되밀림 %", "en": "Drawdown%"},
            "progress_min_pnl": {"ko": "최소 PnL %", "en": "Min PnL%"},
            "field_top_n": {"ko": "상위 심볼 수", "en": "Top N symbols"},
            "field_vol_min": {"ko": "최소 변동성", "en": "Min volatility"},
            "field_mom_long": {"ko": "롱 모멘텀 최소", "en": "Min LONG momentum"},
            "field_mom_short": {"ko": "숏 모멘텀 최소", "en": "Min SHORT momentum"},
            "field_watch_limit": {"ko": "최대 모니터링 심볼 수", "en": "Max watchlist size"},
            "field_max_open": {"ko": "최대 동시 진입 심볼 수", "en": "Max open symbols"},
            "field_cooldown": {"ko": "자동 튜닝 쿨다운 (분)", "en": "Auto-tune cooldown (min)"},
            "field_stop_loss_pnl": {"ko": "스톱로스 규칙 (PnL %)", "en": "Stop-loss rule (PnL %)"},
            "spike_guard_label": {"ko": "스파크 방어 (급락 감지)", "en": "Spike guard (crash detection)"},
            "spike_guard_hint": {"ko": "급락 시 포지션을 강제로 정리합니다. 자동 튜닝이 잠겨 있을 때는 변경할 수 없습니다.", "en": "Forces position close on crash. Cannot be changed while auto-tune is locked."},
            "status_mom_long": {"ko": "롱 모멘텀 최소", "en": "Min LONG momentum"},
            "status_mom_short": {"ko": "숏 모멘텀 최소", "en": "Min SHORT momentum"},
            "status_vol_min": {"ko": "최소 변동성", "en": "Min volatility"},
            "status_watch_limit": {"ko": "워치리스트 크기", "en": "Watchlist size"},
            "status_max_symbols": {"ko": "최대 동시 진입 심볼", "en": "Max open symbols"},
            "status_regime": {"ko": "시장 레짐", "en": "Market regime"},
            "status_hits": {"ko": "히스테리시스 횟수", "en": "Hysteresis hits"},
            "status_cooldown": {"ko": "쿨다운 상태", "en": "Cooldown state"},
            "status_shadow": {"ko": "섀도우 튜너", "en": "Shadow tuner"},
            "status_pos_pct": {"ko": "포지션 %", "en": "Position %"},
            "status_lev_range": {"ko": "레버리지 범위", "en": "Leverage range"},
            "status_stop_loss": {"ko": "스톱로스 규칙", "en": "Stop-loss rule"},
            "status_quality": {"ko": "퀄리티 점수", "en": "Quality score"},
            "status_noise": {"ko": "노이즈 지수", "en": "Noise index"},
            "status_pass_entry": {"ko": "통과/진입 비율", "en": "Pass/entry ratio"},
            "status_fill_rate": {"ko": "체결률", "en": "Fill rate"},
            "status_pnl_30m": {"ko": "최근 30분 손익", "en": "Recent 30m PnL"},
            "status_active_snap": {"ko": "활성 스냅샷", "en": "Active snapshot"},
            "status_staging": {"ko": "스테이징", "en": "Staging"},
            "status_proposed": {"ko": "제안/롤백", "en": "Proposed/rollback"},
            "status_last_apply": {"ko": "마지막 적용/롤백", "en": "Last apply/rollback"},
            "display_max_symbols_hint": {"ko": "최대 2개까지 선택할 수 있습니다.", "en": "You can select up to 2 symbols."},
            "auto_tune_reset_tooltip": {"ko": "Auto-Tune 학습 데이터를 초기화합니다. 엔진 재시작 후 기본값으로 재적용됩니다.", "en": "Resets Auto-Tune learned state. Default values will be re-applied after engine restart."},

            # ── 다크 다이얼로그 / 알림창 공통 ──────────────────────────
            "dlg_ok": {"ko": "확인", "en": "OK"},
            "dlg_yes": {"ko": "예", "en": "Yes"},
            "dlg_no": {"ko": "아니오", "en": "No"},
            "dlg_close": {"ko": "닫기", "en": "Close"},
            # ── 엔진 시작/종료 알림 ──────────────────────────────────────
            "bot_start_title": {"ko": "봇 시작", "en": "Bot starting"},
            "bot_start_msg": {"ko": "자동 매매 엔진을 실행합니다.", "en": "Starting the trading engine."},
            "bot_stop_title": {"ko": "봇 종료", "en": "Bot stopped"},
            "bot_stop_already": {"ko": "엔진이 이미 종료된 상태입니다.", "en": "The engine is already stopped."},
            "bot_stop_msg": {"ko": "자동 매매 엔진이 종료되었습니다.", "en": "The trading engine has been stopped."},
            "test_order_title": {"ko": "테스트 주문", "en": "Test order"},
            "close_sel_select_first": {"ko": "먼저 포지션 목록에서 심볼을 선택해주세요.", "en": "Select a symbol from the positions list first."},
            "auto_tune_reset_fail": {"ko": "상태 파일 삭제에 실패했습니다.", "en": "Failed to delete the state file."},
            "auto_tune_reset_done": {"ko": "학습 상태가 초기화되었습니다. 엔진을 다시 시작하면 기본값으로 재적용됩니다.", "en": "Learned state has been reset. Default values will be re-applied after engine restart."},
            "auto_tune_reset_restart": {"ko": "Auto-Tune 상태를 리셋했습니다. 엔진을 다시 시작할까요?", "en": "Auto-Tune state has been reset. Restart the engine now?"},
            "engine_stop_fail": {"ko": "엔진 정지 실패", "en": "Engine stop failed"},
            "error_title": {"ko": "오류", "en": "Error"},
            # ── 엔진 시작 / 재시작 ──────────────────────────────────────
            "engine_start_risk_title": {"ko": "위험 고지 동의 필요", "en": "Risk acknowledgement required"},
            "engine_start_risk_msg": {"ko": "자동 매매를 실행하려면 '필수 동의' 탭에서 위험 고지에 동의해주세요.", "en": "To start the engine, acknowledge the risk notice in the Agreement tab."},
            "engine_config_error": {"ko": "엔진 설정을 불러올 수 없습니다.", "en": "Failed to load engine configuration."},
            "engine_restart_title": {"ko": "엔진 재시작 필요", "en": "Engine restart required"},
            "engine_restart_log": {"ko": "[INFO] 설정 변경으로 엔진 재시작", "en": "[INFO] Restarting engine after settings change"},
            "engine_restart_pending": {"ko": "[INFO] 설정 변경 후 재시작 보류", "en": "[INFO] Engine restart deferred"},
            "engine_restart_done_title": {"ko": "재시작 완료", "en": "Restart complete"},
            "engine_restart_done_msg": {"ko": "자동 매매 엔진이 재시작되었습니다.", "en": "The trading engine has been restarted."},
            "engine_start_fail_title": {"ko": "엔진 시작 실패", "en": "Engine start failed"},
            "spike_guard_restart": {"ko": "스파크 방어 설정이 변경되었습니다. 엔진을 다시 시작할까요?", "en": "Spike guard setting changed. Restart the engine now?"},
            "trade_settings_restart": {"ko": "거래 설정이 변경되었습니다. 엔진을 다시 시작할까요?", "en": "Trading settings changed. Restart the engine now?"},
            "trade_settings_saved": {"ko": "거래 설정이 저장되었습니다.", "en": "Trading settings saved."},
            "trade_settings_save_title": {"ko": "저장", "en": "Saved"},
            # ── 포지션 청산 버튼 알림 ────────────────────────────────────
            "close_positions_title": {"ko": "포지션 청산", "en": "Close positions"},
            "close_positions_none": {"ko": "열린 포지션이 없습니다", "en": "No open positions"},
            "close_positions_no_profit": {"ko": "현재 수익 중인 포지션이 없습니다", "en": "No profitable positions are open"},
            "close_positions_no_loss": {"ko": "현재 손실 중인 포지션이 없습니다", "en": "No losing positions are open"},
            "close_positions_sent_profit": {"ko": "{n}개 수익 포지션에 청산 명령 전송", "en": "Close order sent for {n} profitable position(s)"},
            "close_positions_sent_loss": {"ko": "{n}개 손실 포지션에 청산 명령 전송", "en": "Close order sent for {n} losing position(s)"},
            "close_all_title": {"ko": "전체 청산", "en": "Close all"},
            "close_all_sent": {"ko": "모든 포지션 청산 명령을 전송했습니다", "en": "Close order sent for all positions"},
            # ── 위험 고지 관련 청산 차단 ─────────────────────────────────
            "close_risk_msg": {"ko": "청산 기능을 사용하려면 '필수 동의' 탭에서 동의 체크박스를 활성화하세요.", "en": "Enable the risk acknowledgement checkbox in the Agreement tab to use close functions."},
            # ── 트레일링 패널 상태 텍스트 ────────────────────────────────
            "status_waiting_plain": {"ko": "대기", "en": "Waiting"},
            "status_done_fraction": {"ko": "완료 ({done}/{total})", "en": "Done ({done}/{total})"},
            "trail_waiting_pct": {"ko": "대기 ≥ {pct:.0f}%", "en": "Wait ≥ {pct:.0f}%"},
            "progress_waiting_stale": {"ko": "대기 (stale {m}m)", "en": "Wait (stale {m}m)"},
            "progress_waiting_pct_stale": {"ko": "대기 ≥ {pct:.0f}% (stale {m}m)", "en": "Wait ≥ {pct:.0f}% (stale {m}m)"},
            "profit_exit_based": {"ko": "(Profit Exit 설정 기반)", "en": "(based on Profit Exit settings)"},
            # ── 거래 확인 다이얼로그 텍스트 ─────────────────────────────
            "trade_confirm_title": {"ko": "거래 확인", "en": "Order confirmation"},
            "trade_confirm_symbol": {"ko": "심볼", "en": "Symbol"},
            "trade_confirm_order": {"ko": "주문", "en": "Order"},
            "trade_confirm_type": {"ko": "유형", "en": "Type"},
            "trade_confirm_close_label": {"ko": "청산", "en": "Close"},
            "trade_confirm_entry_label": {"ko": "진입", "en": "Entry"},
            "trade_confirm_ref_price": {"ko": "기준 가격", "en": "Ref. price"},
            "trade_confirm_qty": {"ko": "수량(추정)", "en": "Est. qty"},
            "trade_confirm_pos_amount": {"ko": "포지션 금액(추정)", "en": "Est. position"},
            "trade_confirm_notional": {"ko": "명목금(추정)", "en": "Est. notional"},
            "trade_confirm_margin": {"ko": "증거금(추정)", "en": "Est. margin"},
            "trade_confirm_question": {"ko": "거래하시겠습니까?", "en": "Confirm this order?"},
            "trade_confirm_yes": {"ko": "거래 실행 (Y)", "en": "Execute (Y)"},
            "trade_confirm_no": {"ko": "취소 (N)", "en": "Cancel (N)"},
            # ── 수동 매매 액션 레이블 ────────────────────────────────────
            "action_long_limit": {"ko": "롱 지정가 진입", "en": "Long limit entry"},
            "action_long_market": {"ko": "롱 시장가 진입", "en": "Long market entry"},
            "action_short_limit": {"ko": "숏 지정가 진입", "en": "Short limit entry"},
            "action_short_market": {"ko": "숏 시장가 진입", "en": "Short market entry"},
            "action_close_limit": {"ko": "지정가 청산", "en": "Close (limit)"},
            "action_close_market": {"ko": "시장가 청산", "en": "Close (market)"},
            # ── 주문 결과 알림 ───────────────────────────────────────────
            "order_sent_title": {"ko": "주문", "en": "Order"},
            "order_sent_entry": {"ko": "{sym} 진입 주문 전송", "en": "{sym} entry order sent"},
            "order_sent_close": {"ko": "{sym} 청산 주문 전송", "en": "{sym} close order sent"},
            "order_fail_prefix": {"ko": "{sym} 수동 주문 실패", "en": "{sym} manual order failed"},
            # ── 수동 매매 설정 저장 ──────────────────────────────────────
            "manual_symbol_title": {"ko": "수동 매매 심볼", "en": "Manual trade symbol"},
            "manual_symbol_none_msg": {"ko": "최소 1개 이상의 심볼을 선택해주세요.", "en": "Please select at least one symbol."},
            "display_settings_saved": {"ko": "수동 매매 설정이 저장되었습니다.", "en": "Display settings saved."},
            # ── 선택 청산 ────────────────────────────────────────────────
            "close_sel_no_sym": {"ko": "선택된 행에서 심볼을 찾을 수 없습니다.", "en": "Could not find a symbol in the selected row."},
            # ── 가격/수량 오류 ───────────────────────────────────────────
            "price_convert_error_title": {"ko": "가격 변환 오류", "en": "Price conversion error"},
            "qty_error_title": {"ko": "수량 오류", "en": "Quantity error"},
            "env_change_title": {"ko": "환경 변경", "en": "Change environment"},
            "env_change_msg": {"ko": "봇이 실행 중입니다. {env} 환경으로 전환하려면 재시작이 필요합니다. 지금 재시작할까요?", "en": "The bot is running. Switching to {env} requires a restart. Restart now?"},
            "field_leverage": {"ko": "레버리지", "en": "Leverage"},
            "lang_change_title": {"ko": "언어 변경됨", "en": "Language changed"},
            "lang_change_msg": {"ko": "UI가 새로고침되었습니다. 일부 레이블이 잘못된 언어로 보이면 앱을 재시작해주세요.", "en": "UI has been refreshed. If any labels still appear in the wrong language, please restart the app."},
            "monitor_tab":              {"ko": "모니터링",              "en": "Monitoring"},
            "monitor_filter_title":     {"ko": "필터 현황",             "en": "Filter Status"},
            "monitor_watch_title":      {"ko": "모니터링 심볼",          "en": "Watched Symbols"},
            "monitor_reject_title":     {"ko": "필터 탈락 / 스킵",       "en": "Rejected / Skipped"},
            "monitor_autotune_title":   {"ko": "Auto-tune 상태",        "en": "Auto-tune State"},
            "monitor_col_symbol":       {"ko": "심볼",                  "en": "Symbol"},
            "monitor_col_status":       {"ko": "상태",                  "en": "Status"},
            "monitor_col_reason":       {"ko": "사유",                  "en": "Reason"},
            "monitor_col_time":         {"ko": "시각",                  "en": "Time"},
            "monitor_no_data":          {"ko": "데이터 없음 — 엔진 실행 후 표시됩니다", "en": "No data — run the engine to populate"},
            "monitor_input":            {"ko": "전체",                  "en": "Input"},
            "monitor_topn":             {"ko": "Top-N",                 "en": "Top-N"},
            "monitor_passed":           {"ko": "통과",                  "en": "Passed"},
            "monitor_vol_fail":         {"ko": "변동성 미달",             "en": "Vol fail"},
            "monitor_mom_fail":         {"ko": "모멘텀 미달",             "en": "Mom fail"},
            "monitor_refresh":          {"ko": "새로고침",               "en": "Refresh"},
            "monitor_autotune_mode":    {"ko": "모드",                  "en": "Mode"},
            "monitor_autotune_vol":     {"ko": "변동성",                 "en": "Volatility"},
            "monitor_autotune_mom":     {"ko": "모멘텀",                 "en": "Momentum"},
            "monitor_autotune_pos":     {"ko": "포지션 크기",             "en": "Position size"},
            "monitor_autotune_score":   {"ko": "스코어",                 "en": "Score"},
            "monitor_col_param":        {"ko": "파라미터",               "en": "Parameter"},
            "monitor_col_value":        {"ko": "현재값",                 "en": "Current"},
            "monitor_col_base":         {"ko": "기본값",                 "en": "Base"},
            "monitor_last_filter":      {"ko": "최근 필터 실행",          "en": "Last filter run"},
            "monitor_engine_off":       {"ko": "엔진이 정지 상태입니다",   "en": "Engine is stopped"},


    
        }

        allowed_modes = {"aggressive", "balanced", "conservative"}
        mode_value = str(self.settings_data.get("auto_tune_mode", "balanced")).lower()
        if mode_value not in allowed_modes:
            mode_value = "balanced"
        self.settings_data["auto_tune_mode"] = mode_value

        legacy_momentum = self.settings_data.get("momentum_min")
        if legacy_momentum is not None:
            try:
                legacy_value = float(legacy_momentum)
                self.settings_data.setdefault("momentum_min_long", legacy_value)
                self.settings_data.setdefault("momentum_min_short", -abs(legacy_value))
            except (TypeError, ValueError):
                pass
        if "momentum_min" in self.settings_data and all(
            key in self.settings_data for key in ("momentum_min_long", "momentum_min_short")
        ):
            self.settings_data.pop("momentum_min", None)
        self.settings_data.pop("atr_min", None)
        self.settings_data.pop("show_command_buttons", None)
        if "theme" in self.settings_data or "font_scale" in self.settings_data:
            self.settings_data.pop("theme", None)
            self.settings_data.pop("font_scale", None)
        self.manual_symbols = list(self.settings_data.get("manual_symbols", ["BTCUSDT", "ETHUSDT"]))
        self.alert_enabled = bool(self.settings_data.get("alert_enabled", True))
        self.manual_panel = None
        self.positions_panel = None
        self.manual_symbol_options = ["BTCUSDT", "ETHUSDT", "BNBUSDT", "SOLUSDT", "XRPUSDT", "ADAUSDT", "DOGEUSDT", "LINKUSDT"]
        self.risk_acknowledged = bool(self.settings_data.get("risk_acknowledged", False))
        # A1: consent version check — reset if notice was updated
        self._check_consent_version_early()
        self._configure_auto_start(bool(self.settings_data.get("auto_start", False)))
        # [PATCH-5 수정] 환경 설정: default_env_testnet을 기준으로 사용 (유저가 GUI에서 선택한 값)
        self.state_data = self._load_json(STATE_PATH, default={})
        _cfg_default = self.settings_data.get("default_env_testnet", True)
        self.env_mode = "TESTNET" if _cfg_default else "LIVE"
        # use_testnet을 default_env_testnet에 동기화 (역방향 불일치 방지)
        _cfg_use = self.settings_data.get("use_testnet", True)
        if _cfg_use != (not not _cfg_default):
            logging.info(f"[ENV] use_testnet={_cfg_use}을 default_env_testnet={_cfg_default}에 동기화")
            self.settings_data["use_testnet"] = _cfg_default
        self.state_data["env_label"] = self.env_mode
        self.session_start_ms = self.state_data.get("session_start_ms")
        self.notification_path = NOTIFICATION_PATH
        os.makedirs(os.path.dirname(self.notification_path), exist_ok=True)
        self.log_trim_targets = {
            "notifications": self.notification_path,
            "gui_log": LOG_PATH,
        }
        for target in self.log_trim_targets.values():
            self._trim_file(target)
        try:
            self.notification_pointer = os.path.getsize(self.notification_path)
        except FileNotFoundError:
            self.notification_pointer = 0

        self.engine_running = False
        self.balance_history = deque(maxlen=960)
        self.last_account_balance = 0.0
        self.loop = asyncio.new_event_loop()
        self.loop_thread = threading.Thread(target=self.loop.run_forever, daemon=True)
        self.loop_thread.start()

        self.active_alerts = []
        self.top_symbols = []
        self.symbol_price_map = {}
        self.position_refresh_job = None
        self.position_countdown_job = None
        self.next_position_refresh_ts = None
        self.stats_refresh_job = None
        self.current_positions = {}
        self.manual_highlight = {}
        self.symbol_filters_cache = {}
        self.log_box = None
        self.stats_labels = {}
        self.position_refresh_label = None
        self.stat_defaults = {}
        self.stat_resets = {"win_rate": None, "trade_count": None}
        self.pnl_reset_ms: int = 0  # 손익 카드 초기화 기준 시각
        # 상단 4개 전환 카드: 현재 모드 저장 (state_data로 영속)
        _default_modes = self.settings_data.get("default_card_modes", {
            "card0": "unrealized_total", "card1": "filter_pass_rate",
            "card2": "top_symbol",       "card3": "pnl_24h",
        })
        _saved_modes = self.state_data.get("card_modes", {})
        self.card_modes: dict = {
            "card0": _saved_modes.get("card0", _default_modes.get("card0", "unrealized_total")),
            "card1": _saved_modes.get("card1", _default_modes.get("card1", "filter_pass_rate")),
            "card2": _saved_modes.get("card2", _default_modes.get("card2", "top_symbol")),
            "card3": _saved_modes.get("card3", _default_modes.get("card3", "pnl_24h")),
        }
        self.card_title_labels: dict = {}   # card_id → title Label
        self.card_frame_refs:   dict = {}   # card_id → card Frame
        self.env_request_token = 0
        self.last_total_margin = 0.0
        self.auto_tune_status_labels = {}
        self.auto_tune_last_update_label = None
        self.auto_tune_refresh_job = None
        self._auto_tune_refresh_started = False
        self.trade_field_vars = {}
        self.stats_labels = {}
        self._auto_launch_checked = False
        self._restart_prompt_open = False
        self._restart_in_progress = False
        # 모니터링 탭 상태
        self._monitor_refresh_job = None
        self.monitor_watch_tree = None
        self.monitor_reject_tree = None
        self.monitor_filter_stat_labels = {}
        self.monitor_autotune_tree = None
        self.monitor_last_update_label = None

        self.root = tk.Tk()
        self.root.withdraw()
        self.root.title(f"{self._t('app_title', TITLE)} v{APP_VERSION}")
        # ── 윈도우 아이콘 설정 ──
        _icon_path = os.path.join(BASE_DIR, "assets", "bot_converted.ico")
        if not os.path.isfile(_icon_path):
            _icon_path = os.path.join(BASE_DIR, "assets", "bot.ico")
        if os.path.isfile(_icon_path):
            try:
                self.root.iconbitmap(_icon_path)
            except Exception:
                # ico가 아닌 PNG 등일 경우 iconphoto fallback
                try:
                    _icon_img = tk.PhotoImage(file=os.path.join(BASE_DIR, "assets", "bot.ico"))
                    self.root.iconphoto(True, _icon_img)
                except Exception:
                    pass
        self.root.geometry("1280x780")
        self.root.configure(bg="#181A20")
        self.root.resizable(False, False)
        self.root.protocol("WM_DELETE_WINDOW", self._on_close)

        # ── 글로벌 스크롤바 다크 테마 스타일 ──
        _gs = ttk.Style()
        try:
            _gs.theme_use("clam")
        except Exception:
            pass
        _gs.configure("Vertical.TScrollbar",
                      background="#353b50", troughcolor="#12141a",
                      arrowcolor="#4a5270", borderwidth=0, relief="flat",
                      width=8)
        _gs.map("Vertical.TScrollbar",
                background=[("active", "#4a5270"), ("pressed", "#5a6488")])

        self.checkbox_images = self._load_checkbox_images()

        self.root.bind("<F1>", lambda event=None: self._open_help_modal())
        self._build_layout()
        self.root.deiconify()
        self._schedule_notification_poll()
        if not self.risk_acknowledged:
            self.root.after(800, lambda: self.open_settings_modal(initial_tab="dev"))
            # 설정창 위에 시작하기 다이얼로그도 띄움
            self.root.after(1200, self._show_referral_onboarding)
        else:
            # 동의 완료 상태 → 레퍼럴 온보딩만
            self.root.after(500, self._show_referral_onboarding)
        self.root.after(1500, self._maybe_auto_launch_engine)
        self._shutdown_alert_suppressed = False
        self._restart_notice_pending = False

    # ------------------------------------------------------------------
    def _load_json(self, path, default):
        try:
            with open(path, "r", encoding="utf-8") as fh:
                return json.load(fh)
        except Exception:
            return default.copy()

    def _save_json(self, path, data):
        # Always enforce locked settings when writing gui_config.json.
        if os.path.basename(path) == os.path.basename(CONFIG_PATH):
            for _k, _v in LOCKED_GUI_SETTINGS.items():
                data[_k] = _v
        with open(path, "w", encoding="utf-8") as fh:
            json.dump(data, fh, ensure_ascii=False, indent=2)

    # ------------------------------------------------------------------
    def _t(self, key, fallback=""):
        table = getattr(self, "translations", {})
        entry = table.get(key)
        if isinstance(entry, dict):
            return entry.get(self.language, entry.get("ko", fallback))
        if isinstance(entry, str):
            return entry
        return fallback


    def _rebuild_ui(self):
        """Rebuild the main UI to apply language changes to already-created widgets."""
        # 모든 위젯 파괴 전 grab/modal 상태 정리 — 파괴된 위젯을 가리키는
        # _active_modal이 남으면 이후 다이얼로그가 grab을 잘못 관리함
        self._active_modal = None
        self.log_box = None  # 파괴 전 참조 해제 (stale 위젯 방지)
        try:
            for child in list(self.root.winfo_children()):
                try:
                    child.grab_release()
                except Exception:
                    pass
                try:
                    child.destroy()
                except Exception:
                    pass
        except Exception:
            pass
        self._build_layout()
        # _build_layout은 버튼/라벨을 항상 초기(STOPPED) 상태로 만듦.
        # 엔진이 실행 중이었다면 UI 상태를 실제 상태와 동기화.
        self._sync_engine_ui_state()

    def _build_layout(self):
        self.root.columnconfigure(1, weight=1)
        self.root.rowconfigure(0, weight=1)

        self.sidebar = tk.Frame(self.root, width=220, bg="#0c1017")
        self.sidebar.grid(row=0, column=0, sticky="nswe")
        self.sidebar.grid_propagate(False)

        self.main_area = tk.Frame(self.root, bg="#181A20")
        self.main_area.grid(row=0, column=1, sticky="nsew", padx=10, pady=10)
        self.main_area.columnconfigure(0, weight=1)
        self.main_area.rowconfigure(0, weight=0)
        self.main_area.rowconfigure(1, weight=0)
        self.main_area.rowconfigure(2, weight=1)

        self._build_sidebar()
        self._build_stats()
        self._build_command_panel()
        self._build_manual_panel()
        self._build_positions_panel()
        self._add_settings_button()
        # [PATCH-12] 레퍼럴 코드 무결성 체크
        if not REFERRAL_CODE:
            self.root.after(200, self._show_integrity_warning)
        # [PATCH-11] 첫 실행 시 레퍼럴 가입 안내
        # → __init__에서 risk_acknowledged 여부에 따라 호출 시점 결정
        # [PATCH-11] 자동 업데이트 체크
        self.root.after(3000, self._check_for_updates)

    def _show_integrity_warning(self):
        """레퍼럴 코드 변조 감지 시 경고."""
        from tkinter import messagebox as _mb
        _is_ko = self.language == "ko"
        _mb.showwarning(
            "무결성 경고" if _is_ko else "Integrity Warning",
            ("이 프로그램은 비공식 수정 버전일 수 있습니다.\n"
             "공식 버전을 다운로드하세요:\n"
             f"https://github.com/{GITHUB_REPO}/releases")
            if _is_ko else
            ("This program may be an unofficial modified version.\n"
             "Please download the official version:\n"
             f"https://github.com/{GITHUB_REPO}/releases")
        )

    def _show_referral_onboarding(self):
        """신규 사용자(API 키 미설정)에게 레퍼럴 링크로 가입 유도."""
        # API 키가 하나라도 설정되어 있으면 스킵 (매 실행마다 체크)
        has_key = bool(os.environ.get("BINANCE_API_KEY") or os.environ.get("TESTNET_API_KEY"))
        if has_key:
            self.settings_data["referral_onboarding_shown"] = True
            self._save_json(CONFIG_PATH, self.settings_data)
            return
        ref_code = self.settings_data.get("binance_referral_code", "")
        if not ref_code:
            return
        # 설정창이 열려 있으면 grab 해제 (레퍼럴 다이얼로그가 위에 뜰 수 있게)
        _modal = getattr(self, "_active_modal", None)
        if _modal and _modal.winfo_exists():
            try:
                _modal.grab_release()
            except Exception:
                pass
        _is_ko = self.language == "ko"
        import webbrowser as _wb

        ref_url = f"https://www.binance.com/register?ref={ref_code}"
        api_url = "https://www.binance.com/en/my/settings/api-management"
        testnet_api_url = "https://testnet.binancefuture.com/en/futures/BTC_USDT"

        _BG = "#181A20"
        _GOLD = "#F0B90B"
        _GOLD_DIM = "#c49a09"
        _GREEN = "#2EBD85"
        _CARD_BG = "#1e2230"

        _parent = self.root
        if _modal and _modal.winfo_exists():
            _parent = _modal

        dialog = tk.Toplevel(_parent)
        dialog.title("바이낸스 시작하기" if _is_ko else "Get Started with Binance")
        self._apply_icon(dialog)
        dialog.configure(bg=_BG)
        dialog.geometry("500x440")
        dialog.resizable(False, False)
        dialog.transient(_parent)
        dialog.grab_set()
        dialog.lift()
        dialog.focus_force()

        # ── 제목 ──
        tk.Label(dialog,
                 text="🚀 시작하기" if _is_ko else "🚀 Getting Started",
                 bg=_BG, fg=_GOLD,
                 font=("Malgun Gothic", 16, "bold")).pack(pady=(28, 2))
        tk.Label(dialog,
                 text="현재 상태에 맞는 항목을 선택해주세요." if _is_ko else "Select the option that matches your situation.",
                 bg=_BG, fg="#6a7080",
                 font=("Malgun Gothic", 9)).pack(pady=(0, 20))

        # ── 레퍼럴 링크 바 (2줄 구조로 짤림 방지) ──
        ref_bar = tk.Frame(dialog, bg=_CARD_BG)
        ref_bar.pack(fill="x", padx=36, pady=(0, 18))
        # 상단: 라벨 + 복사버튼
        ref_top = tk.Frame(ref_bar, bg=_CARD_BG)
        ref_top.pack(fill="x", padx=12, pady=(8, 0))
        tk.Label(ref_top, text="Referral Link",
                 bg=_CARD_BG, fg="#6a7080",
                 font=("Malgun Gothic", 8)).pack(side=tk.LEFT)
        def _copy_ref():
            dialog.clipboard_clear()
            dialog.clipboard_append(ref_url)
            copy_lbl.configure(text=" Copied! ", fg=_GREEN)
            dialog.after(1500, lambda: copy_lbl.configure(text=" Copy ", fg="#6a7080"))
        copy_lbl = tk.Label(ref_top, text=" Copy ", bg="#282d3a", fg="#6a7080",
                            font=("Malgun Gothic", 8), cursor="hand2", padx=6)
        copy_lbl.pack(side=tk.RIGHT)
        copy_lbl.bind("<Button-1>", lambda e: _copy_ref())
        # 하단: URL (전체 너비 사용)
        ref_bottom = tk.Frame(ref_bar, bg=_CARD_BG)
        ref_bottom.pack(fill="x", padx=12, pady=(2, 8))
        tk.Label(ref_bottom, text=ref_url,
                 bg=_CARD_BG, fg=_GOLD,
                 font=("Consolas", 9), anchor="w").pack(side=tk.LEFT, fill="x")

        # ── 3가지 선택 버튼 (바이낸스 골드 톤 통일) ──
        choices_frame = tk.Frame(dialog, bg=_BG)
        choices_frame.pack(fill="x", padx=36, pady=(0, 8))

        _btn_common = {
            "relief": tk.FLAT, "cursor": "hand2",
            "font": ("Malgun Gothic", 10, "bold"),
            "padx": 16, "pady": 10, "anchor": "w", "bd": 0,
        }

        # 1) 바이낸스 계정이 없어요 → 가입
        def _open_register():
            _wb.open(ref_url)

        tk.Button(choices_frame,
                  text=("🔗  계정이 없어요  ─  가입하기" if _is_ko
                        else "🔗  No account  ─  Sign up"),
                  command=_open_register,
                  bg="#2b2311", fg=_GOLD,
                  activebackground="#3a3018", activeforeground="#ffe082",
                  **_btn_common).pack(fill="x", pady=(0, 5))

        # 2) API 키가 없어요 → 발급 안내
        def _show_api_guide():
            self._show_api_key_guide(dialog, _is_ko, api_url, testnet_api_url)

        tk.Button(choices_frame,
                  text=("🔑  API 키가 없어요  ─  발급 안내" if _is_ko
                        else "🔑  No API key  ─  How to create"),
                  command=_show_api_guide,
                  bg="#1f2816", fg=_GREEN,
                  activebackground="#2a381e", activeforeground="#80e8b0",
                  **_btn_common).pack(fill="x", pady=(0, 5))

        # 3) API 키 발급 완료 → 환경설정
        def _go_to_env_settings():
            dialog.destroy()
            if _modal and _modal.winfo_exists():
                try:
                    _modal.grab_set()
                    _modal.lift()
                    _modal.focus_force()
                except Exception:
                    pass
                if hasattr(self, "_settings_show_section"):
                    self.root.after(100, lambda: self._settings_show_section("env"))
            else:
                self.root.after(150, lambda: self.open_settings_modal(initial_tab="env"))

        tk.Button(choices_frame,
                  text=("⚙  API 키 발급 완료  ─  환경변수 설정" if _is_ko
                        else "⚙  Have API keys  ─  Set up now"),
                  command=_go_to_env_settings,
                  bg=_CARD_BG, fg="#c0c6dc",
                  activebackground="#282d3a", activeforeground="#ffffff",
                  **_btn_common).pack(fill="x", pady=(0, 0))

        # ── 하단 닫기 ──
        def _restore_parent_grab():
            """레퍼럴 다이얼로그 닫힌 후 설정창 grab/focus 복원."""
            if _modal and _modal.winfo_exists():
                try:
                    _modal.grab_set()
                    _modal.lift()
                    _modal.focus_force()
                except Exception:
                    pass

        def _close():
            dialog.destroy()
            _restore_parent_grab()

        tk.Label(dialog, text="", bg=_BG).pack(expand=True)  # spacer
        tk.Button(dialog,
                  text="나중에 할게요" if _is_ko else "Later",
                  command=_close,
                  bg=_BG, fg="#4a4f5e",
                  activebackground=_BG, activeforeground="#6a7080", relief="flat",
                  font=("Malgun Gothic", 9), cursor="hand2",
                  padx=10, pady=3).pack(pady=(0, 18))

    def _show_api_key_guide(self, parent_dialog, _is_ko, api_url, testnet_api_url):
        """API 키 발급 방법을 상세 안내하는 다이얼로그."""
        import webbrowser as _wb

        _BG = "#181A20"
        _GOLD = "#F0B90B"
        _GREEN = "#2EBD85"
        _CARD = "#1e2230"
        _TEXT = "#c0c6dc"
        _DIM = "#6a7080"

        guide = tk.Toplevel(parent_dialog)
        guide.title("API 키 발급 안내" if _is_ko else "API Key Guide")
        self._apply_icon(guide)
        guide.configure(bg=_BG)
        guide.geometry("520x540")
        guide.resizable(False, False)
        guide.transient(parent_dialog)
        guide.grab_set()

        # 스크롤 가능한 영역
        canvas = tk.Canvas(guide, bg=_BG, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(guide, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        body = tk.Frame(canvas, bg=_BG)
        canvas.create_window((0, 0), window=body, anchor="nw")

        def _on_frame_configure(e):
            canvas.configure(scrollregion=canvas.bbox("all"))
        body.bind("<Configure>", _on_frame_configure)

        def _guide_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
        canvas.bind_all("<MouseWheel>", _guide_mousewheel, add="+")
        def _guide_cleanup(*_):
            try: canvas.unbind_all("<MouseWheel>")
            except Exception: pass
        canvas.bind("<Destroy>", _guide_cleanup, add="+")

        px = 28

        # ── 제목 ──
        tk.Label(body, text="API 키 발급 방법" if _is_ko else "How to Create API Keys",
                 bg=_BG, fg="#ffffff",
                 font=("Malgun Gothic", 14, "bold")).pack(anchor="w", padx=px, pady=(22, 16))

        # ── 실거래 섹션 ──
        live_card = tk.Frame(body, bg=_CARD)
        live_card.pack(fill="x", padx=px, pady=(0, 12))

        tk.Label(live_card, text="실거래 (Live)" if _is_ko else "Live",
                 bg=_CARD, fg=_GOLD,
                 font=("Malgun Gothic", 11, "bold")).pack(anchor="w", padx=14, pady=(12, 6))

        live_steps = [
            ("1.", "바이낸스 로그인" if _is_ko else "Log in to Binance"),
            ("2.", "프로필 → API 관리 클릭" if _is_ko else "Profile → API Management"),
            ("3.", "API 생성 → 라벨 입력 (예: AutoBot)" if _is_ko else "Create API → Enter label (e.g. AutoBot)"),
            ("4.", "보안 인증 완료 (이메일/2FA)" if _is_ko else "Complete security verification"),
            ("5.", "수정 → Enable Futures 체크" if _is_ko else "Edit → Enable Futures"),
            ("6.", "IP 제한 설정 권장" if _is_ko else "Set IP restriction (recommended)"),
            ("7.", "API Key, Secret Key 복사 후 보관" if _is_ko else "Copy & save API Key and Secret Key"),
        ]
        for num, text in live_steps:
            row = tk.Frame(live_card, bg=_CARD)
            row.pack(fill="x", padx=14, pady=1)
            tk.Label(row, text=num, bg=_CARD, fg=_GOLD, width=3, anchor="e",
                     font=("Consolas", 10, "bold")).pack(side=tk.LEFT, padx=(0, 6))
            tk.Label(row, text=text, bg=_CARD, fg=_TEXT,
                     font=("Malgun Gothic", 9), anchor="w").pack(side=tk.LEFT)

        # 실거래 링크
        live_btn = tk.Frame(live_card, bg="#2b2311")
        live_btn.pack(fill="x", padx=14, pady=(8, 12))
        def _open_live(): _wb.open(api_url)
        tk.Label(live_btn, text="API Management 페이지 열기  →" if _is_ko else "Open API Management  →",
                 bg="#2b2311", fg=_GOLD, cursor="hand2",
                 font=("Malgun Gothic", 9, "bold")).pack(padx=10, pady=6)
        live_btn.bind("<Button-1>", lambda e: _open_live())
        for child in live_btn.winfo_children():
            child.bind("<Button-1>", lambda e: _open_live())

        # ── 테스트넷 섹션 ──
        test_card = tk.Frame(body, bg=_CARD)
        test_card.pack(fill="x", padx=px, pady=(0, 12))

        tk.Label(test_card, text="테스트넷 (Testnet)" if _is_ko else "Testnet",
                 bg=_CARD, fg=_GREEN,
                 font=("Malgun Gothic", 11, "bold")).pack(anchor="w", padx=14, pady=(12, 6))

        testnet_steps = [
            ("1.", "테스트넷 사이트 접속 → 로그인" if _is_ko else "Go to Testnet → Log in"),
            ("2.", "API Key 메뉴 클릭" if _is_ko else "Click API Key menu"),
            ("3.", "Create API 클릭 → 즉시 발급" if _is_ko else "Create API → Instant generation"),
            ("4.", "API Key, Secret Key 복사" if _is_ko else "Copy API Key & Secret Key"),
        ]
        for num, text in testnet_steps:
            row = tk.Frame(test_card, bg=_CARD)
            row.pack(fill="x", padx=14, pady=1)
            tk.Label(row, text=num, bg=_CARD, fg=_GREEN, width=3, anchor="e",
                     font=("Consolas", 10, "bold")).pack(side=tk.LEFT, padx=(0, 6))
            tk.Label(row, text=text, bg=_CARD, fg=_TEXT,
                     font=("Malgun Gothic", 9), anchor="w").pack(side=tk.LEFT)

        # 테스트넷 링크
        test_btn = tk.Frame(test_card, bg="#1f2816")
        test_btn.pack(fill="x", padx=14, pady=(8, 12))
        def _open_testnet(): _wb.open(testnet_api_url)
        tk.Label(test_btn, text="Testnet Futures 페이지 열기  →" if _is_ko else "Open Testnet Futures  →",
                 bg="#1f2816", fg=_GREEN, cursor="hand2",
                 font=("Malgun Gothic", 9, "bold")).pack(padx=10, pady=6)
        test_btn.bind("<Button-1>", lambda e: _open_testnet())
        for child in test_btn.winfo_children():
            child.bind("<Button-1>", lambda e: _open_testnet())

        # ── 주의사항 ──
        warn_card = tk.Frame(body, bg="#2a2215")
        warn_card.pack(fill="x", padx=px, pady=(0, 16))
        tk.Label(warn_card, text="주의사항" if _is_ko else "Important",
                 bg="#2a2215", fg=_GOLD,
                 font=("Malgun Gothic", 10, "bold")).pack(anchor="w", padx=14, pady=(10, 4))
        warn_items = [
            ("Secret Key는 생성 시 한 번만 표시됩니다." if _is_ko else "Secret Key is shown only once."),
            ("출금(Withdraw) 권한은 절대 활성화하지 마세요." if _is_ko else "Never enable Withdraw permission."),
            ("먼저 테스트넷에서 충분히 테스트 후 실거래하세요." if _is_ko else "Test on Testnet first before Live trading."),
        ]
        for item in warn_items:
            row = tk.Frame(warn_card, bg="#2a2215")
            row.pack(fill="x", padx=14, pady=1)
            tk.Label(row, text="•", bg="#2a2215", fg="#c49a09",
                     font=("Consolas", 10)).pack(side=tk.LEFT, padx=(0, 6))
            tk.Label(row, text=item, bg="#2a2215", fg="#c8b870",
                     font=("Malgun Gothic", 9), anchor="w").pack(side=tk.LEFT)
        tk.Label(warn_card, text="", bg="#2a2215").pack(pady=(0, 6))  # bottom padding

        # ── 하단 닫기 ──
        bottom = tk.Frame(body, bg=_BG)
        bottom.pack(fill="x", padx=px, pady=(0, 22))
        tk.Button(bottom,
                  text="확인" if _is_ko else "OK",
                  command=guide.destroy,
                  bg=_CARD, fg="#c0c6dc",
                  activebackground="#282d3a", activeforeground="#ffffff",
                  relief=tk.FLAT, cursor="hand2",
                  font=("Malgun Gothic", 10, "bold"),
                  padx=24, pady=6).pack(anchor="center")

    def _check_for_updates(self):
        """GitHub Releases에서 최신 버전 확인 (백그라운드)."""
        if not GITHUB_REPO:
            return
        def _do_check():
            try:
                import urllib.request, json as _json
                url = f"https://api.github.com/repos/{GITHUB_REPO}/releases/latest"
                req = urllib.request.Request(url, headers={"User-Agent": "BinanceAutoBot"})
                with urllib.request.urlopen(req, timeout=5) as resp:
                    data = _json.loads(resp.read().decode())
                latest = data.get("tag_name", "").lstrip("v")
                if latest and latest != APP_VERSION:
                    dl_url = data.get("html_url", f"https://github.com/{GITHUB_REPO}/releases/latest")
                    self.root.after(0, lambda: self._show_update_banner(latest, dl_url))
            except Exception:
                pass  # 네트워크 오류 시 무시
        threading.Thread(target=_do_check, daemon=True).start()

    def _show_update_banner(self, new_version, download_url):
        """상단에 업데이트 알림 배너 표시."""
        _is_ko = self.language == "ko"
        banner = tk.Frame(self.root, bg="#F0B90B", height=32)
        banner.place(relx=0, rely=0, relwidth=1, height=32)
        tk.Label(banner,
                 text=f"{'새 버전' if _is_ko else 'New version'} v{new_version} {'출시!' if _is_ko else 'available!'} (현재 v{APP_VERSION})" if _is_ko else f"New version v{new_version} available! (current v{APP_VERSION})",
                 bg="#F0B90B", fg="#181A20",
                 font=("Malgun Gothic", 10, "bold")).pack(side=tk.LEFT, padx=16)
        tk.Button(banner,
                  text="다운로드" if _is_ko else "Download",
                  command=lambda: webbrowser.open(download_url),
                  bg="#181A20", fg="#F0B90B",
                  relief="flat", font=("Malgun Gothic", 9, "bold"),
                  cursor="hand2", padx=10).pack(side=tk.RIGHT, padx=8, pady=4)
        tk.Button(banner, text="✕", command=banner.destroy,
                  bg="#F0B90B", fg="#181A20", relief="flat",
                  font=("Malgun Gothic", 10, "bold"), cursor="hand2",
                  padx=4).pack(side=tk.RIGHT)

    def _sync_engine_ui_state(self):
        """_rebuild_ui 후 실제 engine_running 상태에 맞게 버튼/라벨 동기화."""
        if not self.engine_running:
            return
        try:
            self.start_btn.configure(state=tk.DISABLED, bg="#1a5c3a", fg="#6abf95")
            self.stop_btn.configure(state=tk.NORMAL, bg="#c62828")
            self.status_label.configure(
                text=f'● {self._t("bot_status", "봇 상태")}: RUNNING',
                fg="#2EBD85",
            )
        except Exception:
            pass

    def _build_sidebar(self):
        _SB_BG = "#0c1017"

        # ── 타이틀 영역 (골드 하단 라인) ──
        title_area = tk.Frame(self.sidebar, bg=_SB_BG)
        title_area.pack(fill="x")

        tk.Label(
            title_area,
            text=self._t('app_title', TITLE),
            bg=_SB_BG,
            fg="#F0B90B",
            font=("Segoe UI", 12, "bold"),
            wraplength=180,
            justify="left",
        ).pack(anchor="w", padx=16, pady=(14, 4))
        tk.Label(
            title_area,
            text=f"v{APP_VERSION}",
            bg=_SB_BG, fg="#4a5068",
            font=("Consolas", 8),
        ).pack(anchor="w", padx=16, pady=(0, 10))
        tk.Frame(title_area, bg="#F0B90B", height=1).pack(fill="x", padx=16)

        env_frame = tk.Frame(self.sidebar, bg=_SB_BG)
        env_frame.pack(fill="x", padx=16, pady=(14, 18))
        tk.Label(
            env_frame,
            text=self._t("sidebar_env", "환경"),
            bg=_SB_BG,
            fg="#9aa5c6",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x", pady=(0, 6))
        self.env_canvas = tk.Canvas(env_frame, width=180, height=36, bg="#0c1017", highlightthickness=0)
        self.env_canvas.pack()
        self.env_canvas.bind("<Button-1>", lambda _e: self._toggle_env_mode())
        self._render_env_toggle()

        self.status_label = tk.Label(
            self.sidebar,
            text=f'● {self._t("bot_status","봇 상태")}: STOPPED',
            bg="#0c1017",
            fg="#c62828",
            font=("Malgun Gothic", 10, "bold"),
        )
        self.status_label.pack(fill="x", padx=16, pady=(0, 10))

        btn_frame = tk.Frame(self.sidebar, bg="#0c1017")
        btn_frame.pack(fill="x", padx=16)
        self.start_btn = tk.Button(
            btn_frame,
            text="▶  START",
            command=self.start_engine,
            bg="#1a8f63",
            fg="white",
            font=("Malgun Gothic", 11, "bold"),
            relief=tk.FLAT,
            cursor="hand2",
            activebackground="#2EBD85",
            activeforeground="white",
        )
        self.start_btn.pack(fill="x", pady=(0, 8))
        self.stop_btn = tk.Button(
            btn_frame,
            text="■  STOP",
            command=self.stop_engine,
            bg="#b03040",
            fg="white",
            font=("Malgun Gothic", 11, "bold"),
            relief=tk.FLAT,
            state=tk.DISABLED,
            cursor="hand2",
            activebackground="#F6465D",
            activeforeground="white",
        )
        self.stop_btn.pack(fill="x")

        # ── 구분선 + 빠른 링크 ──
        tk.Frame(self.sidebar, bg="#1e2438", height=1).pack(fill="x", padx=16, pady=(14, 0))
        tk.Label(
            self.sidebar,
            text=self._t("quick_links","빠른 링크"),
            bg=_SB_BG, fg="#6a7490",
            font=("Malgun Gothic", 9, "bold"),
            anchor="w",
        ).pack(fill="x", padx=16, pady=(10, 6))
        _link_kwargs = {"font": ("Segoe UI", 9), "relief": tk.FLAT, "height": 1,
                        "cursor": "hand2", "bd": 0}
        tk.Button(
            self.sidebar,
            text="↗  " + self._t("link_binance","바이낸스"),
            command=lambda: webbrowser.open("https://www.binance.com/kr/futures"),
            bg="#141820", fg="#F0B90B",
            activebackground="#1a2030", activeforeground="#F0B90B",
            **_link_kwargs,
        ).pack(fill="x", padx=16, pady=2)
        tk.Button(
            self.sidebar,
            text="↗  " + self._t("link_testnet","테스트넷"),
            command=lambda: webbrowser.open("https://testnet.binancefuture.com"),
            bg="#141820", fg="#8890a8",
            activebackground="#1a2030", activeforeground="#c0c6dc",
            **_link_kwargs,
        ).pack(fill="x", padx=16, pady=2)

        # ── 하단 구분선 ──
        tk.Frame(self.sidebar, bg="#1e2438", height=1).pack(
            side=tk.BOTTOM, fill="x", padx=16, pady=(0, 0))

        # ── 도움말 버튼 ──
        help_btn = tk.Button(
            self.sidebar,
            text="?  " + self._t("help_btn","도움말"),
            command=self._open_help_modal,
            bg=_SB_BG, fg="#606878",
            font=("Malgun Gothic", 9),
            relief=tk.FLAT,
            cursor="hand2",
            activebackground="#141820",
            activeforeground="#c0c6dc",
        )
        help_btn.pack(side=tk.BOTTOM, fill="x", padx=16, pady=(6, 10))

        # ── 설정 버튼 ──
        settings_btn = tk.Button(
            self.sidebar,
            text="⚙  " + self._t("settings_btn","설정"),
            command=self.open_settings_modal,
            bg="#141820", fg="#c0c6dc",
            font=("Malgun Gothic", 10, "bold"),
            relief=tk.FLAT,
            cursor="hand2",
            activebackground="#1a2030",
            activeforeground="#F0B90B",
        )
        settings_btn.pack(side=tk.BOTTOM, fill="x", padx=16, pady=(6, 2))


    def _open_help_modal(self):
        dialog = tk.Toplevel(self.root)
        dialog.title(self._t("help_btn","도움말"))
        self._apply_icon(dialog)
        width, height = (960, 640) if self.language == "en" else (860, 580)
        dialog.geometry(f"{width}x{height}")
        dialog.configure(bg="#0f131c")
        dialog.transient(self.root)
        self._center_modal(dialog, width, height)
        # grab_set() 제거 — AI 어시스턴트 등 다른 Toplevel과 동시 조작 가능하도록
        # 대신 transient + focus_force 로 의사-모달 유지
        dialog.focus_force()

        container = tk.Frame(dialog, bg="#0f131c")
        container.pack(fill=tk.BOTH, expand=True)
        self._active_modal = dialog

        def _clear_modal(_event=None):
            if getattr(self, "_active_modal", None) is dialog:
                self._active_modal = None

        def _on_settings_close():
            _clear_modal()
            dialog.destroy()
            # 설정창 닫힌 뒤 AI 어시스턴트 창이 있으면 활성화
            _ai_win = getattr(self, "_ai_assistant_win", None)
            if _ai_win and _ai_win.winfo_exists():
                try:
                    _ai_win.lift()
                except Exception:
                    pass

        dialog.protocol("WM_DELETE_WINDOW", _on_settings_close)
        dialog.bind("<Destroy>", _clear_modal)

        sidebar = tk.Frame(container, width=220, bg="#080b12")
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        content = tk.Frame(container, bg="#151b24")
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        sections = self._get_help_sections()

        detail_label = tk.Label(content, bg="#151b24", fg="#d8def8", font=("Malgun Gothic", 11), justify="left", wraplength=580, anchor="nw")
        detail_label.pack(fill=tk.BOTH, expand=True, padx=24, pady=20)

        def show_section(sec_id):
            for sid, _, body in sections:
                if sid == sec_id:
                    detail_label.configure(text=body)
                    break
            for sid, btn in buttons.items():
                btn.configure(bg="#1b2334" if sid == sec_id else "#080b12")

        buttons = {}
        for idx, (sec_id, label, _body) in enumerate(sections):
            btn = tk.Button(
                sidebar,
                text=label,
                command=lambda s=sec_id: show_section(s),
                bg="#080b12",
                fg="#f5f7ff",
                relief=tk.FLAT,
                anchor="w",
                padx=12,
                pady=8,
                font=("Malgun Gothic", 9),
                wraplength=196,
                justify="left",
            )
            btn.pack(fill="x", padx=8, pady=(16 if idx == 0 else 4, 0))
            buttons[sec_id] = btn

        # 버튼이 모두 생성된 뒤 overview 표시 (after로 렌더링 보장)
        dialog.after(0, lambda: show_section(sections[0][0] if sections else "overview"))

        tk.Button(
            content,
            text=self._t("close", "Close"),
            command=dialog.destroy,
            bg="#2EBD85",
            fg="#0b1b15",
            font=("Malgun Gothic", 11, "bold"),
            relief=tk.FLAT,
            padx=18,
            pady=8,
        ).pack(pady=(0, 20))

    def _generate_metrics_report(self, window_hours=12, fmt="markdown"):
        script_path = os.path.join(BASE_DIR, "reports", "metrics_report.py")
        if not os.path.exists(script_path):
            return False, ("reports/metrics_report.py not found." if self.language == "en" else "reports/metrics_report.py 파일을 찾을 수 없습니다.")
        cmd = [sys.executable or "python", script_path, "--window", str(window_hours), "--format", fmt]
        try:
            result = subprocess.run(cmd, capture_output=True, text=True, cwd=BASE_DIR, timeout=30)
        except Exception as exc:
            return False, (f"Failed to generate report: {exc}" if self.language == "en" else f"리포트를 생성하지 못했습니다: {exc}")
        if result.returncode != 0:
            stderr = result.stderr.strip()
            stdout = result.stdout.strip()
            message = stderr or stdout or (f"metrics_report.py exit code {result.returncode}" if self.language == "en" else f"metrics_report.py 종료 코드 {result.returncode}")
            return False, message
        output = result.stdout.strip()
        if not output:
            output = ("Report output is empty." if self.language == "en" else "리포트 출력이 비어 있습니다.")
        return True, output

    def _analyze_metrics_report(self, report_text: str) -> str:
        lines = [line.strip() for line in (report_text or "").splitlines() if line.strip()]
        if not lines:
            return ""
        def extract_float(pattern):
            match = re.search(pattern, report_text, re.IGNORECASE)
            if match:
                try:
                    return float(match.group(1))
                except (TypeError, ValueError):
                    return None
            return None
        samples = extract_float(r"Samples:\s*([0-9]+)")
        win_rate = extract_float(r"Win rate[^0-9]*([0-9]+\.?[0-9]*)")
        expectancy = extract_float(r"Expectancy[:=]\s*([-0-9.]+)")
        avg_hold = extract_float(r"Avg hold[^0-9]*([0-9.]+)")
        summary = []
        if samples is not None:
            summary.append(f"{int(samples)} samples" if self.language == "en" else f"샘플 {int(samples)}개")
        if win_rate is not None:
            summary.append(f"Win rate {win_rate:.2f}%" if self.language == "en" else f"평균 승률 {win_rate:.2f}%")
        if expectancy is not None:
            summary.append(f"Expectancy {expectancy:.4f}" if self.language == "en" else f"기대값 {expectancy:.4f}")
        if avg_hold is not None:
            summary.append(f"Avg hold {avg_hold:.1f}s" if self.language == "en" else f"평균 보유 {avg_hold:.1f}s")
        bullet_lines = []
        if summary:
            bullet_lines.append("· " + ", ".join(summary))
        def top_metrics(prefix):
            items = []
            capture = False
            for line in lines:
                if line.lower().startswith(prefix.lower()):
                    capture = True
                    continue
                if capture:
                    if line.startswith("-") or line.startswith("  ") or line.startswith("\t"):
                        clean = line.lstrip("- ")
                        if clean:
                            items.append(clean)
                        if len(items) >= 2:
                            break
                    else:
                        break
            return items
        entry_items = top_metrics("Entry blocks")
        exit_items = top_metrics("Exits")
        if entry_items:
            bullet_lines.append(("· Entry limits: " if self.language == "en" else "· 주요 진입 제한: ") + ", ".join(entry_items))
        if exit_items:
            bullet_lines.append(("· Exit freq: " if self.language == "en" else "· Exit 빈도: ") + ", ".join(exit_items))
        if not bullet_lines:
            return ""
        return "\n".join(bullet_lines)

    def _open_metrics_report_modal(self, window_hours=12):
        success, report_text = self._generate_metrics_report(window_hours)
        analysis_text = self._analyze_metrics_report(report_text) if success else ("Failed to generate report." if self.language == "en" else "리포트 생성 실패로 요약을 제공할 수 없습니다.")
        dialog = tk.Toplevel(self.root)
        dialog.title("12h Report" if self.language == "en" else "12시간 리포트")
        self._apply_icon(dialog)
        width, height = 780, 560
        dialog.geometry(f"{width}x{height}")
        dialog.configure(bg="#111521")
        dialog.transient(self.root)
        dialog.focus_force()
        self._center_modal(dialog, width, height)

        container = tk.Frame(dialog, bg="#111521")
        container.pack(fill=tk.BOTH, expand=True, padx=12, pady=12)
        header = tk.Frame(container, bg="#111521")
        header.pack(fill="x", pady=(0, 8))
        tk.Label(
            header,
            text="metrics_report.py --window 12",
            bg="#111521",
            fg="#f5f7ff",
            font=("Malgun Gothic", 12, "bold"),
            anchor="w",
        ).pack(side=tk.LEFT)
        tk.Button(
            header,
            text=self._t("close", "Close"),
            command=dialog.destroy,
            bg="#2EBD85" if success else "#F6465D",
            fg="#0b1b15" if success else "#ffffff",
            font=("Malgun Gothic", 10, "bold"),
            relief=tk.FLAT,
            padx=14,
            pady=4,
        ).pack(side=tk.RIGHT)

        analysis_frame = tk.Frame(container, bg="#141a2a", highlightbackground="#2a3550", highlightthickness=1)
        analysis_frame.pack(fill="x", pady=(0, 10))
        tk.Label(
            analysis_frame,
            text=("Summary" if self.language == "en" else "요약"),
            bg="#141a2a",
            fg="#f0f4ff",
            font=("Malgun Gothic", 11, "bold"),
            anchor="w",
        ).pack(fill="x", padx=12, pady=(8, 2))
        analysis_text_widget = tk.Text(
            analysis_frame,
            bg="#0f1420",
            fg="#cfe0ff",
            insertbackground="#ffffff",
            font=("Malgun Gothic", 10),
            height=4,
            wrap="word",
        )
        analysis_text_widget.insert("1.0", analysis_text or ("Summary unavailable." if self.language == "en" else "요약 정보를 추출할 수 없습니다."))
        analysis_text_widget.configure(state=tk.DISABLED)
        analysis_text_widget.pack(fill="x", padx=12, pady=(0, 8))

        text_frame = tk.Frame(container, bg="#111521")
        text_frame.pack(fill=tk.BOTH, expand=True)
        scrollbar = tk.Scrollbar(text_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        text_widget = tk.Text(
            text_frame,
            bg="#0b101a",
            fg="#d6defc" if success else "#ffb3b3",
            insertbackground="#ffffff",
            font=("Consolas", 10),
            wrap="word",
        )
        text_widget.insert("1.0", report_text)
        text_widget.configure(state=tk.DISABLED)
        text_widget.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        text_widget.configure(yscrollcommand=scrollbar.set)
        scrollbar.configure(command=text_widget.yview)
    def _get_help_sections(self):
        """Help modal content. 모든 섹션은 3-튜플 (id, label, body) 로 통일."""
        _en = self.language == "en"

        def _t(ko, en):
            return en if _en else ko

        return [
            (
                "overview",
                _t("개요", "Overview"),
                _t(
                    "이 UI는 Binance USDT-M 선물 자동매매 엔진(v1.1.0)을 운영·모니터링하는 콘솔입니다.\n\n"
                    "• START/STOP으로 엔진 프로세스를 제어합니다.\n"
                    "• 프리셋(공격/기본/보수) 버튼으로 자동매매 전략을 전환합니다.\n"
                    "• Auto-tune이 켜져 있으면 시장 상황에 따라 파라미터가 자동 조정됩니다.\n"
                    "• 거래 로직 패널에서 현재 엔진에 적용 중인 유효 파라미터를 확인할 수 있습니다.\n\n"
                    "설정 탭 구성:\n"
                    "  환경설정  — API 환경(테스트넷/실거래), 언어 설정, 환경변수 가이드\n"
                    "  화면설정  — 수동 매매 패널, 세션 상태 초기화\n"
                    "  거래설정  — 포지션 크기, 레버리지, 오토튜닝, 전략 체크박스\n"
                    "  리포트    — 거래 기록 조회, 통계, 메이커/테이커 수수료 현황\n"
                    "  필수 동의  — 리스크 경고 동의 (2항목 모두 체크 필요)\n"
                    "  프리미엄  — Neural Scorer 라이선스 키 입력, 구독 결제\n"
                    "  정보      — 프로그램 정보, 버전, 문의 이메일",
                    "This UI is an operations console for the Binance USDT-M Futures auto-trading engine (v1.1.0).\n\n"
                    "• START/STOP controls the engine process.\n"
                    "• Use preset buttons (Aggressive / Balanced / Conservative) to switch strategy.\n"
                    "• When Auto-tune is ON, parameters adjust automatically based on market conditions.\n"
                    "• The Trading Logic panel shows the effective parameters currently applied by the engine.\n\n"
                    "Settings tabs:\n"
                    "  Env         — API environment (testnet / live), language, env variable guide\n"
                    "  Display     — Manual trading panel, session reset\n"
                    "  Trade       — Position size, leverage, auto-tuning, strategy checkboxes\n"
                    "  Report      — Trade history, stats, maker/taker fee breakdown\n"
                    "  Agreement   — Risk acknowledgment (both checkboxes required)\n"
                    "  Premium     — Neural Scorer license key, subscription\n"
                    "  About       — Program info, version, contact email"
                ),
            ),
            (
                "order_type",
                _t("주문 방식 (메이커/테이커)", "Order Type (Maker/Taker)"),
                _t(
                    "■ 진입 주문\n"
                    "  • 기본: 시장가(테이커) 0.05% 수수료\n"
                    "  • 필수 동의 탭에서 Maker First를 활성화하면 post-only 지정가로 먼저 시도합니다.\n"
                    "    - offset bps 만큼 호가 안쪽에 지정가 제출 → 타임아웃 내 미체결 시 시장가 fallback\n\n"
                    "■ 청산 주문\n"
                    "  • 일반 청산 (손절 · 트레일링 · 부분익절 · Progress Stop · 신호소멸)\n"
                    "    → 지정가(GTC) 우선 시도 (메이커 0.02%) → 2초 내 미체결 시 시장가 fallback\n"
                    "  • 긴급 청산 (스파크 가드 발동 시)\n"
                    "    → 즉시 시장가(테이커) 0.05%\n\n"
                    "■ 화면 표시 기준\n"
                    "  • 미실현 손익 카드: 보수적으로 테이커 0.05% 차감 (아직 청산 전)\n"
                    "  • 실현 손익 카드(15m/24h 등): 실제 체결 방식의 수수료가 반영된 값\n"
                    "  • 리포트 테이블 Fee 컬럼: M(메이커) / T(테이커) + 실제 수수료 금액 표시\n"
                    "  • 구 데이터(업데이트 전 거래): T≈추정값 으로 표시됩니다.",
                    "■ Entry orders\n"
                    "  • Default: Market (taker) 0.05% fee\n"
                    "  • Enable 'Maker First' in Agreement tab to attempt post-only LIMIT first.\n"
                    "    - Submits limit order inside the spread by offset bps → Market fallback if not filled in timeout\n\n"
                    "■ Exit orders\n"
                    "  • Normal exits (stop-loss · trailing · partial TP · progress stop · signal decay)\n"
                    "    → GTC Limit first (maker 0.02%) → Market fallback if not filled within 2 seconds\n"
                    "  • Urgent exits (spike guard triggered)\n"
                    "    → Immediate Market (taker) 0.05%\n\n"
                    "■ Display basis\n"
                    "  • Unrealized PnL card: conservatively deducts taker 0.05% (position not yet closed)\n"
                    "  • Realized PnL cards (15m/24h etc.): reflects actual fill type fee\n"
                    "  • Report table Fee column: M(maker) / T(taker) + actual fee amount\n"
                    "  • Legacy data (trades before this update): shown as T≈estimated"
                ),
            ),
            (
                "presets",
                _t("프리셋 모드", "Preset Mode"),
                _t(
                    "Auto-tune이 켜져 있으면 시장 상황에 따라 파라미터가 자동 조정됩니다.\n\n"
                    "프리셋 버튼:\n"
                    "• 공격(Aggressive): 참여 폭↑ / 리스크·거래빈도↑\n"
                    "• 기본(Balanced): 기본 프로파일\n"
                    "• 보수(Conservative): 필터 강화 / 리스크·거래빈도↓\n\n"
                    "프리셋을 누르면 해당 버튼이 하이라이트되고 gui_config.json에 저장됩니다.\n"
                    "엔진 실행 중이라면 재시작(STOP → START) 후 적용됩니다.",
                    "When Auto-tune is ON, parameters adjust automatically.\n\n"
                    "Preset buttons:\n"
                    "• Aggressive: wider participation / higher risk & frequency\n"
                    "• Balanced: default profile\n"
                    "• Conservative: stricter filters / lower risk & frequency\n\n"
                    "Selecting a preset highlights the button and saves to gui_config.json.\n"
                    "If the engine is running, restart (STOP → START) to apply."
                ),
            ),
            (
                "filters",
                _t("진입 필터", "Entry Filters"),
                _t(
                    "수수료/슬리피지로 기대값이 무너지는 구간을 막기 위한 진입 가드:\n\n"
                    "• 스프레드(bps) 필터: bid/ask 스프레드가 크면 진입 차단\n"
                    "• 레짐(ADX) 필터: 추세 약/횡보 구간 회피\n"
                    "• RSI 필터: 과매수/과매도 구간 진입 제한\n"
                    "• 펀딩 편향 필터: 펀딩비 방향이 포지션에 불리하면 패널티 적용\n"
                    "• 복합 신호 필터: 여러 지표 점수 합산이 임계값 미달 시 차단\n"
                    "• 스파크 가드: 급등락 감지 시 진입 차단 및 기존 포지션 긴급 청산\n\n"
                    "최근 거부 사유는 Watchlist/Skip 패널과 로그에서 확인할 수 있습니다.",
                    "Entry guards to prevent negative-EV trades:\n\n"
                    "• Spread (bps) filter: blocks entry when bid/ask spread is wide\n"
                    "• Regime (ADX) filter: avoids low-trend/choppy markets\n"
                    "• RSI filter: restricts entries in overbought/oversold zones\n"
                    "• Funding bias filter: applies penalty when funding direction is unfavorable\n"
                    "• Composite signal filter: blocks entry when multi-indicator score is below threshold\n"
                    "• Spike guard: blocks new entries and urgently closes existing positions on sudden spike\n\n"
                    "Recent rejection reasons are shown in the Watchlist/Skip panel and logs."
                ),
            ),
            (
                "exits",
                _t("수익 보호 (익절 스택)", "Profit Protection (Exit Stack)"),
                _t(
                    "수익 보호는 상단부터 단계별로 평가됩니다 (상단 트리거 시 하단 스킵):\n\n"
                    "1) Profit Exit Layer — 마스터 스위치\n"
                    "2) 부분 익절(Partial TP) — ROI 단계별 분할 청산\n"
                    "3) ATR 트레일링 스톱 — 이익 구간에서 추적 손절\n"
                    "4) Progress Stop — 신고점 정체 / MFE 되밀림 감지\n\n"
                    "손절(Stop-Loss)은 별도로 항상 동작합니다.\n"
                    "• max_loss_per_position: 진입가 대비 최대 허용 손실(%)\n"
                    "• Breakeven Stop: 설정 이익 구간 진입 후 손절을 진입가로 이동\n\n"
                    "청산 주문 방식은 '주문 방식' 탭을 참조하세요.",
                    "Profit protection is evaluated top-down (upper layers skip lower ones if triggered):\n\n"
                    "1) Profit Exit Layer — master switch\n"
                    "2) Partial Take-Profit — staged partial closes by ROI levels\n"
                    "3) ATR Trailing Stop — trailing stop in profit zone\n"
                    "4) Progress Stop — detects stale new-high / MFE drawdown\n\n"
                    "Stop-Loss runs independently at all times.\n"
                    "• max_loss_per_position: maximum allowed loss (%) from entry price\n"
                    "• Breakeven Stop: moves stop to entry after reaching a set profit level\n\n"
                    "See the 'Order Type' tab for exit order mechanics."
                ),
            ),
            (
                "report",
                _t("리포트 탭", "Report Tab"),
                _t(
                    "거래 기록 조회 및 통계 분석:\n\n"
                    "• 환경 필터: 테스트넷 / 실거래 / 전체 선택\n"
                    "• 손익 필터: 전체 / 수익 / 손실 거래만 표시\n"
                    "• Fee 컬럼: M(메이커) 또는 T(테이커) + 실제 수수료 금액\n"
                    "• 통계 카드: 총 손익, 승률, 손익비(R:R), 기댓값, 최대/최소, 주요 청산 사유\n"
                    "• 주요 청산 사유 카드 하단: 메이커/테이커 거래 수 및 총 수수료 합계\n\n"
                    "기록 삭제: 헤더의 '🗑 기록 삭제' 버튼으로 거래 기록과 로그를 삭제할 수 있습니다.",
                    "Trade history and statistics analysis:\n\n"
                    "• Environment filter: testnet / live / all\n"
                    "• PnL filter: all / wins only / losses only\n"
                    "• Fee column: M(maker) or T(taker) + actual fee amount\n"
                    "• Stat cards: total PnL, win rate, R:R ratio, expectancy, best/worst, top exit reason\n"
                    "• Top exit reason card footer: maker/taker trade count and total fee sum\n\n"
                    "Clear records: use '🗑 Clear Records' in the header to delete trade history and logs."
                ),
            ),
            (
                "logs",
                _t("로그 · 진단", "Logs & Diagnostics"),
                _t(
                    "• logs/trade_history.jsonl — ENTRY/EXIT 이벤트\n"
                    "  포함 필드: symbol, pnl, roi_pct, trigger, fee_type, fee_amount, fee_rate,\n"
                    "             slippage_bps, spread_bps, leverage, order_id\n"
                    "• logs/auto_tuner_state.json — 현재 오토튜닝 모드와 유효 파라미터\n"
                    "• logs/bot.log — 런타임 로그\n"
                    "• logs/notifications.log — 알림 이벤트\n\n"
                    "리포트 생성:\n"
                    "  python reports/metrics_report.py --window 12 --format markdown",
                    "• logs/trade_history.jsonl — ENTRY/EXIT events\n"
                    "  Fields: symbol, pnl, roi_pct, trigger, fee_type, fee_amount, fee_rate,\n"
                    "          slippage_bps, spread_bps, leverage, order_id\n"
                    "• logs/auto_tuner_state.json — current auto-tune mode and active parameters\n"
                    "• logs/bot.log — runtime log\n"
                    "• logs/notifications.log — alert events\n\n"
                    "Generate a report:\n"
                    "  python reports/metrics_report.py --window 12 --format markdown"
                ),
            ),
            (
                "shortcuts",
                _t("단축키", "Shortcuts"),
                _t(
                    "• F1: 도움말 열기\n"
                    "• ESC: 대부분 모달 닫기\n\n"
                    "첫 실행 안내:\n"
                    "• 프로그램을 처음 실행하면 '시작하기' 안내가 표시됩니다.\n"
                    "• API 키가 설정되지 않은 상태에서는 매 실행 시 안내가 반복됩니다.\n"
                    "• API 키를 환경변수에 등록하면 안내가 더 이상 표시되지 않습니다.\n\n"
                    "문의:\n"
                    "• 설정 → 정보 탭에서 개발자 이메일을 확인할 수 있습니다.\n"
                    "• autobot.trading2026@gmail.com",
                    "• F1: Open help\n"
                    "• ESC: Close most dialogs\n\n"
                    "First-time guide:\n"
                    "• A 'Getting Started' dialog appears on first launch.\n"
                    "• It will repeat each launch until API keys are configured.\n"
                    "• Once API keys are set in environment variables, it will stop appearing.\n\n"
                    "Contact:\n"
                    "• Check the developer email in Settings → About tab.\n"
                    "• autobot.trading2026@gmail.com"
                ),
            ),
        ]


    def _render_env_toggle(self):
        canvas = self.env_canvas
        canvas.delete("all")
        width = 180
        height = 36
        pad = 4
        gap = 2
        mid = width / 2
        # 박스 좌표
        lx1, ly1, lx2, ly2 = pad, pad, mid - gap, height - pad
        rx1, ry1, rx2, ry2 = mid + gap, pad, width - pad, height - pad
        # 각 박스 정중앙
        left_cx  = (lx1 + lx2) / 2
        left_cy  = (ly1 + ly2) / 2
        right_cx = (rx1 + rx2) / 2
        right_cy = (ry1 + ry2) / 2

        canvas.create_rectangle(pad, pad, width - pad, height - pad, fill="#1c2130", outline="")
        if self.env_mode == "TESTNET":
            canvas.create_rectangle(lx1, ly1, lx2, ly2, fill="#2EBD85", outline="")
            canvas.create_rectangle(rx1, ry1, rx2, ry2, fill="#2d3243", outline="")
            test_color = "white"
            live_color = "#7a8299"
        else:
            canvas.create_rectangle(lx1, ly1, lx2, ly2, fill="#2d3243", outline="")
            canvas.create_rectangle(rx1, ry1, rx2, ry2, fill="#F0B90B", outline="")
            test_color = "#98a2bf"
            live_color = "#ffffff"
        canvas.create_text(left_cx, left_cy, text="TESTNET", fill=test_color, font=("Segoe UI", 9, "bold"))
        canvas.create_text(right_cx, right_cy, text="LIVE", fill=live_color, font=("Segoe UI", 9, "bold"))
        canvas.create_rectangle(pad, pad, width - pad, height - pad, outline="#2a3145", width=1)

    def _toggle_env_mode(self):
        target = "LIVE" if self.env_mode == "TESTNET" else "TESTNET"
        if self.engine_running:
            confirm = self._ask_yes_no(
                self._t("env_change_title", "Change environment"),
                self._t("env_change_msg", "The bot is running. Switching to {env} requires a restart. Restart now?").format(env=target),
            )
            if not confirm:
                self._render_env_toggle()
                return

            def apply_and_restart():
                self._apply_env_switch(target)
                self.start_engine()

            self._restart_notice_pending = True
            self.stop_engine(on_stopped=apply_and_restart)
        else:
            self._apply_env_switch(target)

    def _rebuild_stats_cards(self):
        """카드 모드 변경 후 stats 영역 위젯을 재빌드."""
        # 기존 stats 위젯 파괴
        for cid, frm in list(self.card_frame_refs.items()):
            try:
                if frm and frm.winfo_exists():
                    frm.destroy()
            except Exception:
                pass
        self.card_frame_refs.clear()
        self.card_title_labels.clear()
        self.stats_labels.clear()
        self.stat_defaults.clear()
        self._build_stats()
        # place()로 띄운 우측 상단 버튼들은 _build_stats 후 재배치해야 가려지지 않음
        self._add_settings_button()

    def _build_stats(self):
        frame = tk.Frame(self.main_area, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        frame.grid(row=0, column=0, sticky="ew", pady=(0, 8))
        for i in range(4):
            frame.columnconfigure(i, weight=1, uniform="stats")

        # ── 드롭다운 전환 가능한 카드 메타데이터 ──────────────────────
        # (key, ko_label, en_label, default_value)
        SWITCHABLE_CARD_META = {
            "unrealized_total": ("미실현 손익 합계", "Unrealized PnL",  "0.00 USDT"),
            "filter_pass_rate": ("필터 통과율",       "Filter Pass Rate","— %"),
            "top_symbol":       ("최다 거래 심볼",    "Top Symbol",      "—"),
            "pnl_15":           ("15분 손익",          "15m PnL",         "+0.00 USDT"),
            "pnl_60":           ("60분 손익",          "60m PnL",         "+0.00 USDT"),
            "pnl_12h":          ("12시간 손익",        "12h PnL",         "+0.00 USDT"),
            "pnl_24h":          ("24시간 손익",        "24h PnL",         "+0.00 USDT"),
            "rr_ratio":         ("손익비 R:R",         "Win/Loss Ratio",  "—"),
            "expectancy":       ("거래당 기댓값",      "Expectancy",      "+0.00 USDT"),
            "max_consec_loss":  ("최대 연속 손실",     "Max Consec. Loss","—"),
        }
        # 드롭다운 표시 순서
        SWITCHABLE_ORDER = [
            "unrealized_total", "filter_pass_rate", "top_symbol",
            "pnl_15", "pnl_60", "pnl_12h", "pnl_24h",
            "rr_ratio", "expectancy", "max_consec_loss",
        ]

        stats_config = [
            ("win_rate",        self._t("stat_win_rate","승률"),                        "0%"),
            ("trade_count",     self._t("stat_trade_count","거래 수"),                  "0"),
            ("notional",        self._t("stat_notional","진입 총액 (USDT / 레버리지)"), "0 / 0 USDT"),
            ("account_balance", self._t("stat_balance","계좌 잔고"),                    "0 USDT"),
        ]
        self.stats_labels.clear()
        self.stat_defaults.clear()
        self.card_title_labels.clear()
        self.card_frame_refs.clear()

        # ── 상단 4개: 전환 가능한 드롭다운 카드 ─────────────────────
        BG = "#181A20"
        for card_idx in range(4):
            card_id   = f"card{card_idx}"
            mode_key  = self.card_modes.get(card_id, list(SWITCHABLE_ORDER)[card_idx])
            meta      = SWITCHABLE_CARD_META.get(mode_key, SWITCHABLE_CARD_META["pnl_24h"])
            title_txt = meta[1] if self.language == "en" else meta[0]
            def_val   = meta[2]

            card = tk.Frame(frame, bg=BG, highlightbackground="#343942", highlightthickness=1)
            card.grid(row=0, column=card_idx, padx=6, pady=6, sticky="nsew")
            self.card_frame_refs[card_id] = card

            header = tk.Frame(card, bg=BG)
            header.pack(fill="x", padx=8, pady=(8, 0))

            title_lbl = tk.Label(header, text=title_txt, bg=BG, fg="#e0e6ff", font=("Segoe UI", 10))
            title_lbl.pack(side=tk.LEFT, anchor="w")
            self.card_title_labels[card_id] = title_lbl

            # ▾ 드롭다운 버튼
            def _make_dropdown(cid=card_id, h=header, meta_dict=SWITCHABLE_CARD_META, order=SWITCHABLE_ORDER):
                def _show_menu():
                    menu = tk.Menu(self.root, tearoff=0,
                                   bg="#23293a", fg="#e0e6ff",
                                   activebackground="#3a4460", activeforeground="#ffffff",
                                   font=("Segoe UI", 10), relief="flat",
                                   bd=1, activeborderwidth=0)
                    for opt_key in order:
                        opt_meta = meta_dict[opt_key]
                        label_str = opt_meta[1] if self.language == "en" else opt_meta[0]
                        def _select(k=opt_key, c=cid):
                            self._switch_card(c, k, meta_dict)
                        menu.add_command(label=label_str, command=_select)
                    try:
                        btn_widget = h.winfo_children()[-1]
                        x = btn_widget.winfo_rootx()
                        y = btn_widget.winfo_rooty() + btn_widget.winfo_height()
                        menu.tk_popup(x, y)
                    except Exception:
                        menu.tk_popup(self.root.winfo_pointerx(), self.root.winfo_pointery())
                    finally:
                        menu.grab_release()
                return _show_menu

            dd_btn = tk.Label(header, text="▾", bg=BG, fg="#5a6480",
                              font=("Segoe UI", 11), cursor="hand2")
            dd_btn.pack(side=tk.LEFT, padx=(4, 0))
            dd_btn.bind("<Button-1>", lambda e, fn=_make_dropdown(): fn())
            dd_btn.bind("<Enter>", lambda e, w=dd_btn: w.config(fg="#a0aacc"))
            dd_btn.bind("<Leave>", lambda e, w=dd_btn: w.config(fg="#5a6480"))

            value_label = tk.Label(card, text=def_val, bg=BG, fg="#ffffff",
                                   font=("Segoe UI", 15, "bold"), cursor="hand2")
            value_label.pack(anchor="w", padx=8, pady=(2, 0))
            self.stats_labels[mode_key] = value_label
            self.stat_defaults[mode_key] = def_val

            # 수수료 서브라인 (unrealized_total 전용)
            sub_label = tk.Label(card, text="", bg=BG, fg="#6e7da0",
                                 font=("Segoe UI", 9), anchor="w")
            sub_label.pack(anchor="w", padx=8, pady=(0, 6))
            if not hasattr(self, "_card_sub_labels"):
                self._card_sub_labels = {}
            self._card_sub_labels[mode_key] = sub_label
            if not hasattr(self, "_card_value_labels"):
                self._card_value_labels = {}
            self._card_value_labels[card_id] = value_label
            # 클릭으로 값 복사
            def _bind_copy(lbl=value_label, c=card):
                def _copy(_e):
                    try:
                        self.root.clipboard_clear()
                        self.root.clipboard_append(lbl.cget("text"))
                        _orig = c.cget("highlightbackground")
                        c.config(highlightbackground="#2EBD85")
                        self.root.after(400, lambda: c.config(highlightbackground="#343942"))
                    except Exception:
                        pass
                lbl.bind("<Button-1>", _copy)
                def _hover_in(_e): lbl.config(fg="#a0d0ff")
                def _hover_out(_e): lbl.config(fg="#ffffff")
                lbl.bind("<Enter>", _hover_in)
                lbl.bind("<Leave>", _hover_out)
            _bind_copy()

        # ── 하단 4개: 고정 카드 ──────────────────────────────────────
        for idx, (key, title, default_value) in enumerate(stats_config):
            row, col = 1, idx
            card = tk.Frame(frame, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
            card.grid(row=row, column=col, padx=6, pady=6, sticky="nsew")
            header = tk.Frame(card, bg="#181A20")
            header.pack(fill="x", padx=10, pady=(8, 0))
            label = tk.Label(header, text=title, bg="#181A20", fg="#e0e6ff", font=("Segoe UI", 10))
            label.pack(side=tk.LEFT, anchor="w")
            value_label = tk.Label(card, text=default_value, bg="#181A20", fg="#ffffff",
                                   font=("Segoe UI", 16, "bold"))
            value_label.pack(anchor="w", padx=8, pady=(0, 8))
            self.stats_labels[key] = value_label
            self.stat_defaults[key] = default_value

        self._schedule_stats_refresh()

    # ── 드롭다운 카드 전환 ──────────────────────────────────────────
    SWITCHABLE_CARD_META_CLS = {
        "unrealized_total": ("미실현 손익 합계", "Unrealized PnL",  "0.00 USDT"),
        "filter_pass_rate": ("필터 통과율",       "Filter Pass Rate","— %"),
        "top_symbol":       ("최다 거래 심볼",    "Top Symbol",      "—"),
        "pnl_15":           ("15분 손익",          "15m PnL",         "+0.00 USDT"),
        "pnl_60":           ("60분 손익",          "60m PnL",         "+0.00 USDT"),
        "pnl_12h":          ("12시간 손익",        "12h PnL",         "+0.00 USDT"),
        "pnl_24h":          ("24시간 손익",        "24h PnL",         "+0.00 USDT"),
        "rr_ratio":         ("손익비 R:R",         "Win/Loss Ratio",  "—"),
        "expectancy":       ("거래당 기댓값",      "Expectancy",      "+0.00 USDT"),
        "max_consec_loss":  ("최대 연속 손실",     "Max Consec. Loss","—"),
    }

    def _switch_card(self, card_id: str, new_mode: str, meta_dict: dict):
        """드롭다운 선택 시 카드 모드 전환."""
        old_mode = self.card_modes.get(card_id)
        if old_mode == new_mode:
            return
        meta = meta_dict.get(new_mode, meta_dict.get("pnl_24h"))
        new_title = meta[1] if self.language == "en" else meta[0]
        def_val   = meta[2]

        # 이전 모드 stats_labels 키 제거 (같은 키가 다른 카드에 없을 때만)
        if old_mode and old_mode not in [self.card_modes.get(f"card{i}") for i in range(4) if f"card{i}" != card_id]:
            self.stats_labels.pop(old_mode, None)
            self.stat_defaults.pop(old_mode, None)

        # 새 모드 등록
        self.card_modes[card_id] = new_mode
        self.state_data["card_modes"] = dict(self.card_modes)

        # 제목 라벨 텍스트 변경
        title_lbl = self.card_title_labels.get(card_id)
        if title_lbl:
            try:
                title_lbl.config(text=new_title)
            except Exception:
                pass

        # 값 라벨 참조 재배정
        val_lbl = self._card_value_labels.get(card_id) if hasattr(self, "_card_value_labels") else None
        if val_lbl:
            self.stats_labels[new_mode] = val_lbl
            self.stat_defaults[new_mode] = def_val
            try:
                val_lbl.config(text=def_val, fg="#ffffff")
            except Exception:
                pass

        # 즉시 새 모드 데이터로 갱신
        self.root.after(100, lambda: self._refresh_switchable_card(new_mode))

    def _refresh_switchable_card(self, mode_key: str):
        """전환 직후 해당 카드 값 즉시 갱신."""
        try:
            now_ms = int(time.time() * 1000)
            if mode_key in ("pnl_15", "pnl_60", "pnl_12h", "pnl_24h"):
                # PnL 카드: _apply_income_history 트리거
                self._trigger_stat_refresh(0)
            elif mode_key == "unrealized_total":
                # 진입+청산 수수료 차감 순손익
                total_ur = self._calc_unrealized_net()
                sign  = "+" if total_ur > 0 else ""
                color = "#2EBD85" if total_ur > 0 else ("#F6465D" if total_ur < 0 else "#ffffff")
                lbl = self.stats_labels.get(mode_key)
                if lbl:
                    lbl.config(text=self._trim_usdt(f"{sign}{total_ur:.2f} USDT"), fg=color)
            elif mode_key == "filter_pass_rate":
                self._update_filter_pass_rate_card()
            elif mode_key == "top_symbol":
                self._update_top_symbol_card([])
            elif mode_key == "rr_ratio":
                self._update_rr_ratio_card()
            elif mode_key == "expectancy":
                self._update_expectancy_card()
            elif mode_key == "max_consec_loss":
                self._update_max_consec_loss_card()
        except Exception:
            pass

    def _reset_stat_metric(self, key):
        self.stat_resets[key] = int(time.time() * 1000)
        default_value = self.stat_defaults.get(key, "0")
        self._set_stat_value(key, default_value)

    def _reset_pnl_cards(self):
        """활성화된 PnL 카드를 현재 시각 기준으로 초기화."""
        self.pnl_reset_ms = int(time.time() * 1000)
        pnl_keys = {"pnl_15", "pnl_60", "pnl_12h", "pnl_24h"}
        active = set(getattr(self, "card_modes", {}).values())
        for k in pnl_keys:
            if k in active:
                self._set_stat_value(k, "+0 USDT")

    @staticmethod
    def _trim_usdt(text: str) -> str:
        """'0.00 USDT' → '0 USDT', '+1234.50 USDT' → '+1234.5 USDT'
        USDT 앞 숫자의 불필요한 trailing zero 제거."""
        import re
        def _trim(m):
            num = m.group(1)
            stripped = num.rstrip("0").rstrip(".")
            return stripped
        return re.sub(r"([+-]?\d+\.\d+)(?=\s*USDT)", _trim, text)

    def _set_stat_value(self, key, text):
        label = self.stats_labels.get(key)
        if label:
            label.config(text=self._trim_usdt(text))

    def _build_command_panel(self):
        pass

    def _build_log_tab(self, parent):
        """로그 탭 내용 구성 — log_box를 이 탭에 배치."""
        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=1)
        # 툴바
        toolbar = tk.Frame(parent, bg="#0a0d14")
        toolbar.grid(row=0, column=0, sticky="ew", padx=4, pady=(4, 0))
        tk.Label(toolbar, text=self._t("recent_events","최근 이벤트"),
                 bg="#0a0d14", fg="#8b93b7",
                 font=("Segoe UI", 9)).pack(side=tk.LEFT, padx=6)
        tk.Button(toolbar, text="🗑 Clear",
                  command=self._clear_log_box,
                  bg="#0a0d14", fg="#8b93b7", relief=tk.FLAT,
                  font=("Segoe UI", 8), padx=8, pady=2,
                  cursor="hand2",
                  activebackground="#1a2030", activeforeground="#c0d4f0"
                  ).pack(side=tk.RIGHT, padx=4)
        self._log_autoscroll = True
        def _toggle_autoscroll():
            self._log_autoscroll = not self._log_autoscroll
            if self.language == "en":
                _label = "↓ Auto-scroll ON" if self._log_autoscroll else "— Auto-scroll OFF"
            else:
                _label = "↓ 자동스크롤 ON" if self._log_autoscroll else "— 자동스크롤 OFF"
            _as_btn.config(text=_label, fg=("#2EBD85" if self._log_autoscroll else "#5a6280"))
        _as_init = "↓ Auto-scroll ON" if self.language == "en" else "↓ 자동스크롤 ON"
        _as_btn = tk.Button(toolbar, text=_as_init,
                  command=_toggle_autoscroll,
                  bg="#0a0d14", fg="#2EBD85", relief=tk.FLAT,
                  font=("Segoe UI", 8), padx=8, pady=2,
                  cursor="hand2",
                  activebackground="#1a2030", activeforeground="#c0d4f0",
                  )
        _as_btn.pack(side=tk.RIGHT, padx=0)
        # 텍스트 + 스크롤바
        txt_frame = tk.Frame(parent, bg="#0a0d14")
        txt_frame.grid(row=1, column=0, sticky="nsew", padx=4, pady=(2, 4))
        txt_frame.columnconfigure(0, weight=1)
        txt_frame.rowconfigure(0, weight=1)
        parent.rowconfigure(1, weight=1)
        sb = tk.Scrollbar(txt_frame, bg="#1a1f2e", troughcolor="#0a0d14")
        sb.grid(row=0, column=1, sticky="ns")
        self.log_box = tk.Text(
            txt_frame,
            bg="#0a0d14", fg="#c0c6dc",
            insertbackground="white",
            font=("Consolas", 9),
            wrap="none",
            yscrollcommand=sb.set,
            state=tk.NORMAL,
        )
        self.log_box.grid(row=0, column=0, sticky="nsew")
        sb.config(command=self.log_box.yview)
        self._append_log(self._t("ui_init_done","[INFO] UI 초기화 완료"))

    def _clear_log_box(self):
        try:
            if self.log_box:
                self.log_box.config(state=tk.NORMAL)
                self.log_box.delete("1.0", tk.END)
        except Exception:
            pass

    def _build_log_panel(self):
        # 기존 하단 로그 패널 — 로그 탭으로 이전되어 사용하지 않음
        pass

    def _append_log(self, text):
        print(text)  # 콘솔 병행 출력
        if not self.log_box:
            return
        try:
            if not self.log_box.winfo_exists():
                self.log_box = None
                return
            from datetime import datetime as _dt
            ts = _dt.now().strftime("%H:%M:%S")
            self.log_box.config(state=tk.NORMAL)
            self.log_box.insert(tk.END, f"[{ts}] {text}\n")
            self.log_box.see(tk.END)
            # 최대 500줄 유지
            lines = int(self.log_box.index(tk.END).split(".")[0])
            if lines > 502:
                self.log_box.delete("1.0", f"{lines - 500}.0")
        except Exception:
            pass

    def _show_dark_dialog(self, dialog_type, title, message, *, parent=None):
        """
        다크 테마 커스텀 다이얼로그 (거래 확인창 스타일 통일).
        dialog_type: "info" | "warning" | "error" | "yesno"
        반환값: "info"/"warning"/"error" → None, "yesno" → bool
        """
        modal = getattr(self, "_active_modal", None)
        target_parent = parent or (modal if modal and modal.winfo_exists() else self.root)

        # 부모 modal에 grab이 걸려 있을 수 있으므로 일단 해제
        if modal and modal.winfo_exists():
            try:
                modal.grab_release()
            except Exception:
                pass

        try:
            dialog = tk.Toplevel(target_parent)
            self._apply_icon(dialog)
            dialog.configure(bg="#0C1017", highlightthickness=1, highlightbackground="#394058")
            dialog.resizable(False, False)
            dialog.transient(target_parent)
            dialog.grab_set()

            # 아이콘 & 색 설정
            _icon_map = {
                "info":    ("ℹ",  "#5EB0FF"),
                "warning": ("⚠",  "#F7C948"),
                "error":   ("✖",  "#F6465D"),
                "yesno":   ("⚠",  "#F7C948"),
            }
            icon_char, icon_color = _icon_map.get(dialog_type, ("ℹ", "#5EB0FF"))

            # 폭은 메시지 길이에 따라 동적 결정
            lines_cnt = max(message.count("\n") + 1, 1)
            max_line_len = max((len(ln) for ln in message.splitlines()), default=10)
            _w = min(max(400, max_line_len * 10), 560)
            _h = min(max(260, lines_cnt * 24 + 190), 520)

            dialog.geometry(f"{_w}x{_h}")
            self._center_modal(dialog, _w, _h)
            dialog.title(title)

            container = tk.Frame(dialog, bg="#0C1017")
            container.pack(fill="both", expand=True, padx=20, pady=16)

            tk.Label(
                container,
                text=icon_char,
                bg="#0C1017",
                fg=icon_color,
                font=("Segoe UI Symbol", 24, "bold"),
            ).pack(anchor="center", pady=(0, 6))

            tk.Label(
                container,
                text=title,
                bg="#0C1017",
                fg="#ffffff",
                font=("Malgun Gothic", 11, "bold"),
            ).pack(anchor="center", pady=(0, 4))

            tk.Label(
                container,
                text=message,
                bg="#0C1017",
                fg="#e0e6ff",
                font=("Malgun Gothic", 10),
                justify="left",
                wraplength=_w - 60,
            ).pack(fill="both", expand=True, pady=(0, 12))

            result = tk.BooleanVar(value=False)

            # 버튼 구분선
            tk.Frame(container, bg="#394058", height=1).pack(fill="x", pady=(8, 0))
            btn_row = tk.Frame(container, bg="#0C1017")
            btn_row.pack(fill="x", pady=(10, 0))

            _btn_base = {
                "font": ("Malgun Gothic", 10, "bold"),
                "relief": tk.FLAT,
                "cursor": "hand2",
                "padx": 20,
                "pady": 8,
            }

            def _close(val=None):
                if val is not None:
                    result.set(val)
                # Windows Tkinter bug: grab_set() 후 destroy()만으로는 grab이
                # 해제되지 않는 경우가 있어 START 버튼 등이 클릭 불가 상태가 됨.
                # 반드시 grab_release() 먼저 호출.
                try:
                    dialog.grab_release()
                except Exception:
                    pass
                dialog.destroy()

            if dialog_type == "yesno":
                yes_text = self._t("dlg_yes", "Yes")
                no_text  = self._t("dlg_no",  "No")
                tk.Button(btn_row, text=yes_text,
                          bg="#2a3a4a", fg="#7ecbf5",
                          activebackground="#34506a", activeforeground="#ffffff",
                          command=lambda: _close(True),
                          width=10,
                          **_btn_base).pack(side=tk.LEFT, padx=(0, 8))
                tk.Button(btn_row, text=no_text,
                          bg="#1e2230", fg="#8892a8",
                          activebackground="#2a3040", activeforeground="#c8cfe8",
                          command=lambda: _close(False),
                          width=10,
                          **_btn_base).pack(side=tk.LEFT)
                dialog.bind("<Return>", lambda _e: _close(True))
                dialog.bind("<Escape>", lambda _e: _close(False))
                dialog.protocol("WM_DELETE_WINDOW", lambda: _close(False))
            else:
                ok_text = self._t("dlg_ok", "OK")
                btn_bg     = "#1e2d3d" if dialog_type == "info" else ("#2e2a1e" if dialog_type == "warning" else "#2e1e1e")
                btn_fg     = "#7ecbf5" if dialog_type == "info" else ("#e8c96a" if dialog_type == "warning" else "#f08080")
                _active_bg = "#2a4060" if dialog_type == "info" else ("#3d3828" if dialog_type == "warning" else "#3d2828")
                tk.Button(btn_row, text=ok_text,
                          bg=btn_bg, fg=btn_fg,
                          activebackground=_active_bg, activeforeground="#ffffff",
                          command=lambda: _close(),
                          width=16,
                          **_btn_base).pack(anchor="center")
                dialog.bind("<Return>", lambda _e: _close())
                dialog.bind("<Escape>", lambda _e: _close())
                dialog.protocol("WM_DELETE_WINDOW", lambda: _close())

            target_parent.wait_window(dialog)

            if dialog_type == "yesno":
                return result.get()
        finally:
            # 루트 윈도우 grab 강제 해제 (Windows grab 잔류 방지)
            try:
                self.root.grab_release()
            except Exception:
                pass
            # 알림창 닫힌 후 원래 modal을 최상위로 복원
            try:
                _restore_modal = modal if (modal and modal.winfo_exists()) else None
                if _restore_modal is None:
                    _restore_modal = getattr(self, "_active_modal", None)
                if _restore_modal and _restore_modal.winfo_exists():
                    _restore_modal.attributes("-topmost", True)
                    _restore_modal.lift()
                    _restore_modal.focus_force()
                    # topmost 해제 (일시적으로만 사용)
                    _restore_modal.after(100, lambda m=_restore_modal: (
                        m.attributes("-topmost", False) if m.winfo_exists() else None
                    ))
            except Exception:
                pass

    # 하위 호환 래퍼 ─ 내부에서 다크 다이얼로그 사용
    def _call_messagebox(self, func, *args, parent=None, **kwargs):
        # 레거시 호출 경로 — 직접 호출은 하지 않음
        return func(*args, **kwargs)

    def _show_info(self, title, message, **kwargs):
        return self._show_dark_dialog("info", title, message, parent=kwargs.get("parent"))

    def _show_warning(self, title, message, **kwargs):
        return self._show_dark_dialog("warning", title, message, parent=kwargs.get("parent"))

    def _show_error(self, title, message, **kwargs):
        return self._show_dark_dialog("error", title, message, parent=kwargs.get("parent"))

    def _show_yesno(self, title, message, **kwargs):
        return self._show_dark_dialog("yesno", title, message, parent=kwargs.get("parent"))

    def _ask_yes_no(self, title, message, **kwargs):
        return self._show_dark_dialog("yesno", title, message, parent=kwargs.get("parent"))

    def _format_percent_display(self, value):
        try:
            val = float(value)
        except (TypeError, ValueError):
            return "0%"
        if abs(val) <= 1.0:
            val *= 100.0
        return f"{val:.0f}%"

    def _normalize_ratio_input(self, value, clamp_zero_one=False):
        try:
            val = float(value)
        except (TypeError, ValueError):
            return 0.0
        if abs(val) > 1.0:
            val = val / 100.0
        if clamp_zero_one:
            val = max(0.0, min(val, 1.0))
        return val

    def _percent_value(self, value):
        try:
            val = float(value)
        except (TypeError, ValueError):
            return 0.0
        if abs(val) <= 1.0:
            return val * 100.0
        return val

    def _format_partial_levels_for_display(self, levels):
        formatted = []
        for level in levels or []:
            frac = level.get("close_frac", 0.0)
            frac_txt = self._format_percent_display(frac)
            if "r" in level or "r_multiple" in level:
                r_mult = level.get("r", level.get("r_multiple", 0.0))
                try:
                    r_val = float(r_mult)
                except (TypeError, ValueError):
                    r_val = 0.0
                formatted.append(f"{r_val:g}R:{frac_txt}")
            else:
                pnl = level.get("pnl_pct", 0.0)
                pnl_txt = self._format_percent_display(pnl).rstrip('%')
                formatted.append(f"{pnl_txt}:{frac_txt}")
        return ", ".join(formatted)

    def _parse_partial_levels_field(self, text_value: str):
        entries = []
        if not text_value:
            return []
        parts = [token.strip() for token in re.split("[,\n]+", text_value) if token.strip()]
        for token in parts:
            if ':' not in token:
                raise ValueError(f"잘못된 형식: {token}")
            pnl_str, frac_str = token.split(':', 1)
            frac = self._normalize_ratio_input(frac_str, clamp_zero_one=True)
            if frac <= 0:
                continue
            pnl_token = (pnl_str or '').strip()
            if pnl_token.lower().endswith('r'):
                r_txt = pnl_token[:-1].strip()
                r_val = self._normalize_ratio_input(r_txt)
                if r_val <= 0:
                    continue
                entries.append({"r": r_val, "close_frac": frac})
            else:
                pnl = self._normalize_ratio_input(pnl_str)
                if pnl <= 0:
                    continue
                entries.append({"pnl_pct": pnl, "close_frac": frac})
        return entries

    def _format_partial_status(self, roi_percent: float) -> str:
        if not self.settings_data.get("enable_partial_take_profit", True):
            return "OFF"
        levels = [self._percent_value(level.get("pnl_pct", 0.0)) for level in self.settings_data.get("partial_tp_levels", [])]
        levels = [value for value in levels if value > 0]
        if not levels:
            return "OFF"
        ready = sum(1 for value in levels if roi_percent >= value)
        total = len(levels)
        if ready >= total:
            return self._t("status_done_fraction", "Done ({done}/{total})").format(done=total, total=total)
        next_target = levels[ready]
        return f"{ready}/{total} → {next_target:.0f}%"

    def _format_trail_status(self, roi_percent: float) -> str:
        if not self.settings_data.get("enable_atr_trailing_stop", True):
            return "OFF"
        trigger = self._percent_value(self.settings_data.get("trail_activate_pnl_pct", 0.0))
        mult = float(self.settings_data.get("trail_atr_mult", 3.0) or 0.0)
        if trigger <= 0 or roi_percent >= trigger:
            return f"ARMED ({mult:.1f}x)"
        return self._t("trail_waiting_pct", "Wait ≥ {pct:.0f}%").format(pct=trigger)

    def _format_progress_status(self, roi_percent: float) -> str:
        if not self.settings_data.get("enable_progress_stop", True):
            return "OFF"
        min_pnl = self._percent_value(self.settings_data.get("progress_stop_min_pnl_pct", 0.0))
        stale = int(self.settings_data.get("progress_stop_no_new_high_sec", 1800) or 0)
        action = str(self.settings_data.get("progress_stop_action", "partial_or_full")).replace('_', '/').upper()
        if roi_percent >= max(0.0, min_pnl):
            return f"READY ({action})"
        if min_pnl <= 0:
            if stale:
                return self._t("progress_waiting_stale", "Wait (stale {m}m)").format(m=stale // 60)
            return self._t("status_waiting_plain", "Waiting")
        return self._t("progress_waiting_pct_stale", "Wait ≥ {pct:.0f}% (stale {m}m)").format(pct=min_pnl, m=stale // 60)

    def _write_manual_notification(self, message, level="MANUAL"):
        try:
            timestamp = time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
            line = f"{timestamp}|{level}|{message}\n"
            with open(self.notification_path, "a", encoding="utf-8") as fh:
                fh.write(line)
        except Exception as exc:
            self._append_log(f"[WARN] 수동 알림 로그 기록 실패: {exc}")

    # ------------------------------------------------------------------
    def open_settings_modal(self, initial_tab="env"):
        dialog = tk.Toplevel(self.root)
        dialog.title(self._t("settings_title","설정"))
        self._apply_icon(dialog)
        width, height = 1100, 780
        dialog.geometry(f"{width}x{height}")
        dialog.configure(bg="#181A20")
        dialog.transient(self.root)
        dialog.focus_force()
        self._center_modal(dialog, width, height)
        # settings_dialog를 _active_modal에 등록해야 내부에서 호출되는
        # _show_dark_dialog가 올바른 parent/grab 체계를 사용함.
        _prev_modal = getattr(self, "_active_modal", None)
        self._active_modal = dialog

        def _on_settings_dialog_close():
            if getattr(self, "_active_modal", None) is dialog:
                self._active_modal = _prev_modal
            try:
                dialog.destroy()
            except Exception:
                pass
            # 설정창 닫힌 뒤 AI 어시스턴트 창이 있으면 활성화
            _ai_win = getattr(self, "_ai_assistant_win", None)
            if _ai_win:
                try:
                    if _ai_win.winfo_exists():
                        _ai_win.lift()
                except Exception:
                    pass

        dialog.protocol("WM_DELETE_WINDOW", _on_settings_dialog_close)

        container = tk.Frame(dialog, bg="#181A20")
        container.pack(fill=tk.BOTH, expand=True)

        _SET_SB_BG = "#0c1017"
        _SET_SB_SEL = "#181e2c"

        sidebar = tk.Frame(container, width=180, bg=_SET_SB_BG)
        sidebar.pack(side=tk.LEFT, fill=tk.Y)
        sidebar.pack_propagate(False)

        # 설정 사이드바 타이틀
        tk.Label(sidebar, text=self._t("settings_title", "설정"),
                 bg=_SET_SB_BG, fg="#F0B90B",
                 font=("Malgun Gothic", 12, "bold")).pack(
            anchor="w", padx=16, pady=(18, 4))
        tk.Frame(sidebar, bg="#F0B90B", height=1).pack(fill="x", padx=16, pady=(0, 8))

        content = tk.Frame(container, bg="#181A20")
        content.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)

        sections = {}
        buttons = {}

        def create_section(key):
            frame = tk.Frame(content, bg="#181A20")
            sections[key] = frame
            return frame

        def show_section(key):
            for sec in sections.values():
                sec.pack_forget()
            sections[key].pack(fill=tk.BOTH, expand=True)
            for name, btn in buttons.items():
                if name == key:
                    btn.configure(bg=_SET_SB_SEL, fg="#F0B90B")
                elif name == "premium":
                    btn.configure(bg=_SET_SB_BG, fg="#F0B90B")
                else:
                    btn.configure(bg=_SET_SB_BG, fg="#8890a8")
            # 탭 전환 시 새 탭에 Visibility 이벤트 발행 → 스크롤 바인딩
            try:
                sections[key].event_generate("<Visibility>")
            except Exception:
                pass
        self._settings_show_section = show_section
        tabs = [
            ("env",     self._t("settings_tab_env",    "환경설정")),
            ("display", self._t("settings_tab_display", "화면설정")),
            ("trade",   self._t("settings_tab_trade",  "거래설정")),
            ("report",  self._t("settings_tab_report", "리포트")),
            ("dev",     self._t("settings_tab_dev",    "필수 동의")),
            ("premium", self._t("premium_tab_label",   "🧠 프리미엄")),
            ("about",   self._t("settings_tab_about",  "정보")),
        ]
        for idx, (key, label) in enumerate(tabs):
            is_premium = (key == "premium")
            btn = tk.Button(
                sidebar, text=f"  {label}",
                command=lambda k=key: show_section(k),
                bg=_SET_SB_BG,
                fg="#F0B90B" if is_premium else "#8890a8",
                activebackground=_SET_SB_SEL,
                activeforeground="#F0B90B",
                relief=tk.FLAT, pady=8, anchor="w",
                font=("Malgun Gothic", 9, "bold") if is_premium else ("Malgun Gothic", 9),
            )
            btn.pack(fill="x", padx=8, pady=(12 if idx == 0 else 3, 0))
            buttons[key] = btn

        env_section     = create_section("env")
        trade_section   = create_section("trade")
        display_section = create_section("display")
        dev_section     = create_section("dev")
        report_section  = create_section("report")
        premium_section = create_section("premium")
        about_section   = create_section("about")
        self._build_env_tab(env_section)
        self._build_display_tab(display_section)
        self._build_trade_tab(trade_section)
        self._build_developer_tab(dev_section)
        self._build_report_tab(report_section)
        try:
            self._build_premium_tab(premium_section)
        except Exception as _prem_err:
            import traceback as _tb; _tb.print_exc()
            try:
                import tkinter as _tk2
                _ef = _tk2.Frame(premium_section, bg="#2a0000")
                _ef.pack(fill="both", expand=True, padx=10, pady=10)
                _tk2.Label(_ef, text=f"⚠ 프리미엄 탭 렌더링 오류:\n{_prem_err}",
                           bg="#2a0000", fg="#FF6B6B", font=("Malgun Gothic", 9),
                           wraplength=550, justify="left", anchor="nw"
                ).pack(fill="x", padx=10, pady=10)
            except Exception:
                pass
        self._build_about_tab(about_section)
        show_section(initial_tab if initial_tab in sections else "env")

    # ─────────────────────────────────────────────────────────────
    # 정보(About) 탭
    # ─────────────────────────────────────────────────────────────
    def _build_about_tab(self, frame):
        _is_ko = (self.language == "ko")
        _BG   = "#181A20"
        _LINE = "#2a2f42"
        _GOLD = "#F0B90B"
        _GREEN = "#2EBD85"
        _FG   = "#c8cee0"
        _DIM  = "#606878"

        frame.columnconfigure(0, weight=1)

        # 스크롤 가능한 구조
        canvas = tk.Canvas(frame, bg=_BG, highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
        body = tk.Frame(canvas, bg=_BG)
        canvas.create_window((0, 0), window=body, anchor="nw")
        def _on_cfg(e): canvas.configure(scrollregion=canvas.bbox("all"))
        body.bind("<Configure>", _on_cfg)
        def _mw(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception: pass
        def _about_bind(e=None):
            canvas.bind_all("<MouseWheel>", _mw)
        def _about_unbind(e=None):
            try: canvas.unbind_all("<MouseWheel>")
            except Exception: pass
        frame.bind("<Visibility>", _about_bind, add="+")
        canvas.bind("<Enter>", _about_bind, add="+")
        canvas.bind("<Leave>", _about_unbind, add="+")
        body.bind("<Enter>", _about_bind, add="+")
        body.bind("<Leave>", _about_unbind, add="+")

        body.columnconfigure(0, weight=1)
        px = 36

        # ── 헬퍼: 구분선 ──
        def _divider(parent, pad_top=16, pad_bot=16):
            tk.Frame(parent, bg=_LINE, height=1).pack(
                fill="x", padx=px, pady=(pad_top, pad_bot))

        # ══════════════════════════════════════════════════════════
        # 타이틀
        # ══════════════════════════════════════════════════════════
        tk.Label(body, text="Binance Auto Trading Bot",
                 bg=_BG, fg="#ffffff",
                 font=("Malgun Gothic", 15, "bold")).pack(
            anchor="w", padx=px, pady=(28, 2))
        tk.Label(body, text=f"Version {APP_VERSION}",
                 bg=_BG, fg=_GOLD,
                 font=("Consolas", 11)).pack(anchor="w", padx=px, pady=(0, 0))

        # ── 구분선 ──
        _divider(body, pad_top=18, pad_bot=14)

        # ══════════════════════════════════════════════════════════
        # 프로그램 정보
        # ══════════════════════════════════════════════════════════
        info_items = [
            ("엔진" if _is_ko else "Engine",
             "Binance USDT-M Futures Auto Trading"),
            ("전략" if _is_ko else "Strategy",
             "복합 시그널 스코어링 + Kelly Criterion" if _is_ko else
             "Composite Signal Scoring + Kelly Criterion"),
            ("환경" if _is_ko else "Environment",
             "테스트넷 / 실거래" if _is_ko else "Testnet / Live"),
            ("언어" if _is_ko else "Language",
             "한국어 / English"),
        ]
        for label, value in info_items:
            row = tk.Frame(body, bg=_BG)
            row.pack(fill="x", padx=px, pady=3)
            tk.Label(row, text=label, bg=_BG, fg=_DIM, width=10, anchor="w",
                     font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
            tk.Label(row, text=value, bg=_BG, fg=_FG, anchor="w",
                     font=("Malgun Gothic", 9)).pack(side=tk.LEFT)

        # ── 구분선 ──
        _divider(body)

        # ══════════════════════════════════════════════════════════
        # 주요 기능
        # ══════════════════════════════════════════════════════════
        tk.Label(body,
                 text="주요 기능" if _is_ko else "Key Features",
                 bg=_BG, fg="#ffffff",
                 font=("Malgun Gothic", 10, "bold")).pack(
            anchor="w", padx=px, pady=(0, 10))

        features = [
            ("6개 지표 복합 시그널 스코어링" if _is_ko else
             "6-indicator Composite Signal Scoring"),
            ("Kelly Criterion + ATR 동적 포지션 사이징" if _is_ko else
             "Kelly Criterion + ATR Dynamic Sizing"),
            ("3단계 부분 익절 + ATR 트레일링 스탑" if _is_ko else
             "3-level Partial TP + ATR Trailing Stop"),
            ("메이커 우선 주문으로 수수료 절감" if _is_ko else
             "Maker-first Orders for Fee Optimization"),
            ("횡보장 자동 감지 및 방어" if _is_ko else
             "Auto Chop Regime Detection & Defense"),
            ("실시간 대시보드 모니터링" if _is_ko else
             "Real-time Dashboard Monitoring"),
        ]
        for feat in features:
            tk.Label(body, text=f"  ·  {feat}", bg=_BG, fg=_FG,
                     font=("Malgun Gothic", 9), anchor="w").pack(
                fill="x", padx=px, pady=2)

        # ── 구분선 ──
        _divider(body)

        # ══════════════════════════════════════════════════════════
        # 문의
        # ══════════════════════════════════════════════════════════
        tk.Label(body,
                 text="문의 및 개선 요청" if _is_ko else "Contact & Feedback",
                 bg=_BG, fg="#ffffff",
                 font=("Malgun Gothic", 10, "bold")).pack(
            anchor="w", padx=px, pady=(0, 6))

        tk.Label(body,
                 text=("버그 리포트, 기능 개선, 라이선스 문의" if _is_ko else
                       "Bug reports, feature requests, license inquiries"),
                 bg=_BG, fg=_DIM, justify="left",
                 font=("Malgun Gothic", 9)).pack(anchor="w", padx=px, pady=(0, 10))

        # 이메일 행 (배경 없이, 텍스트만)
        email_row = tk.Frame(body, bg=_BG)
        email_row.pack(fill="x", padx=px, pady=(0, 4))
        tk.Label(email_row, text="autobot.trading2026@gmail.com",
                 bg=_BG, fg=_GOLD,
                 font=("Consolas", 10)).pack(side=tk.LEFT)

        def _copy_email():
            frame.clipboard_clear()
            frame.clipboard_append("autobot.trading2026@gmail.com")
            copy_lbl.configure(text="Copied!", fg=_GREEN)
            frame.after(1500, lambda: copy_lbl.configure(
                text="복사" if _is_ko else "Copy", fg=_DIM))

        copy_lbl = tk.Label(email_row,
                            text="복사" if _is_ko else "Copy",
                            bg=_BG, fg=_DIM,
                            font=("Malgun Gothic", 8),
                            cursor="hand2")
        copy_lbl.pack(side=tk.LEFT, padx=(12, 0))
        copy_lbl.bind("<Button-1>", lambda e: _copy_email())

        # ── 구분선 ──
        _divider(body)

        # ══════════════════════════════════════════════════════════
        # 감사 + 면책
        # ══════════════════════════════════════════════════════════
        tk.Label(body,
                 text=("이 프로그램을 이용해 주셔서 감사합니다.\n"
                       "더 나은 서비스를 위해 지속적으로 개선하겠습니다." if _is_ko else
                       "Thank you for using this program.\n"
                       "We are committed to continuous improvement."),
                 bg=_BG, fg=_FG, justify="left",
                 font=("Malgun Gothic", 9)).pack(anchor="w", padx=px, pady=(0, 14))

        tk.Label(body,
                 text=("※ 본 소프트웨어는 투자 조언을 제공하지 않으며, "
                       "자동 거래에 따른 모든 손익의 책임은 사용자에게 있습니다. "
                       "테스트넷에서 충분히 검증한 후 실거래에 적용하시기 바랍니다." if _is_ko else
                       "※ This software does not provide investment advice. "
                       "All profits and losses are the user's responsibility. "
                       "Please test on Testnet before Live trading."),
                 bg=_BG, fg=_DIM, justify="left",
                 font=("Malgun Gothic", 8),
                 wraplength=520).pack(anchor="w", padx=px, pady=(0, 20))

        # ── 저작권 ──
        tk.Label(body,
                 text="© 2026 Binance Auto Trading Bot",
                 bg=_BG, fg="#303848",
                 font=("Consolas", 8)).pack(pady=(4, 24))

    # ─────────────────────────────────────────────────────────────
    # 리포트 탭
    # ─────────────────────────────────────────────────────────────
    def _build_report_tab(self, frame):
        """거래 리포트 탭: 요약 통계 + 거래 기록 테이블 + 카드 초기화."""
        BG      = "#181A20"
        BG2     = "#1c2030"
        BG3     = "#141822"
        FG      = "#e0e6ff"
        FG2     = "#8b93b7"
        ACC     = "#2EBD85"
        WARN    = "#F7C948"
        ERR     = "#F6465D"
        TESTNET_COL = "#6ca4f8"   # 테스트넷 행 색상
        LIVE_COL    = "#e0e6ff"   # 실거래 행 색상
        BORDER  = "#2a3050"
        _en     = self.language == "en"

        frame.columnconfigure(0, weight=1)
        frame.rowconfigure(0, weight=0)   # 헤더 (타이틀 + 버튼들)
        frame.rowconfigure(1, weight=0)   # 요약 통계 카드
        frame.rowconfigure(2, weight=1)   # 거래 기록 테이블

        # ── 함수 전방 선언 (버튼 클릭 핸들러용) ────────────────
        _load_stats_ref  = [None]
        _reload_table_ref = [None]

        # ── 초기화 함수들 ────────────────────────────────────────
        def _reset_all_cards():
            now = int(time.time() * 1000)
            self.pnl_reset_ms = now
            pnl_keys = {"pnl_15", "pnl_60", "pnl_12h", "pnl_24h"}
            active = set(getattr(self, "card_modes", {}).values())
            for k in pnl_keys:
                if k in active:
                    self._set_stat_value(k, "+0 USDT")
            self.stat_resets["win_rate"]    = now
            self.stat_resets["trade_count"] = now
            self._set_stat_value("trade_count", "0")
            self._set_stat_value("win_rate",    "0%")
            for k, dflt in {"unrealized_total": "0.00 USDT", "filter_pass_rate": "— %",
                            "top_symbol": "—", "rr_ratio": "—",
                            "expectancy": "+0.00 USDT", "max_consec_loss": "—"}.items():
                self._set_stat_value(k, dflt)
            self._append_log("[REPORT] 카드 패널 전체 초기화" if not _en else "[REPORT] All cards reset")

        def _reset_pnl_only():
            now = int(time.time() * 1000)
            self.pnl_reset_ms = now
            pnl_keys = {"pnl_15", "pnl_60", "pnl_12h", "pnl_24h"}
            active = set(getattr(self, "card_modes", {}).values())
            for k in pnl_keys:
                if k in active:
                    self._set_stat_value(k, "+0 USDT")
            self._append_log("[REPORT] 손익 카드 초기화" if not _en else "[REPORT] PnL cards reset")

        def _reset_win_rate_only():
            now = int(time.time() * 1000)
            self.stat_resets["win_rate"]    = now
            self.stat_resets["trade_count"] = now
            self._set_stat_value("trade_count", "0")
            self._set_stat_value("win_rate",    "0%")
            self._append_log("[REPORT] 승률·거래수 초기화" if not _en else "[REPORT] Win rate / trade count reset")

        def _refresh_report():
            try:
                if _load_stats_ref[0]:
                    _load_stats_ref[0]()
                if _reload_table_ref[0]:
                    _reload_table_ref[0]()
            except Exception:
                pass

        # ── 헤더 행: 타이틀 + 초기화 버튼들 + 새로고침 ──────────
        hdr = tk.Frame(frame, bg=BG)
        hdr.grid(row=0, column=0, sticky="ew", padx=24, pady=(16, 0))

        tk.Label(hdr, text="리포트" if not _en else "Report",
                 bg=BG, fg=FG, font=("Malgun Gothic", 14, "bold")).pack(side=tk.LEFT)

        # 새로고침 버튼 (맨 오른쪽)
        tk.Button(hdr, text="⟳  " + ("새로고침" if not _en else "Refresh"),
                  command=_refresh_report,
                  bg="#23293a", fg=FG, relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=10, pady=4,
                  cursor="hand2").pack(side=tk.RIGHT, padx=(4, 0))

        # 구분선
        tk.Label(hdr, text="|", bg=BG, fg="#3a4060",
                 font=("Segoe UI", 12)).pack(side=tk.RIGHT, padx=6)
        
        # ═══════════════════════════════════════════════════════════
        # 🆕 내보내기 버튼 (CSV & Excel)
        # ═══════════════════════════════════════════════════════════
        def _export_trades():
            """거래 기록을 CSV/Excel로 내보내기"""
            try:
                import tkinter.filedialog as filedialog
                from datetime import datetime
                
                # 거래 기록 로드
                if not os.path.exists(TRADE_LOG_PATH):
                    tk.messagebox.showwarning(
                        "내보내기 실패" if not _en else "Export Failed",
                        "거래 기록이 없습니다." if not _en else "No trade history.",
                        parent=self.root
                    )
                    return
                
                trades = []
                with open(TRADE_LOG_PATH, "r", encoding="utf-8") as f:
                    for line in f:
                        try:
                            trades.append(json.loads(line.strip()))
                        except:
                            continue
                
                if not trades:
                    tk.messagebox.showwarning(
                        "내보내기 실패" if not _en else "Export Failed",
                        "거래 기록이 없습니다." if not _en else "No trade history.",
                        parent=self.root
                    )
                    return
                
                # 파일 형식 선택
                export_type = tk.messagebox.askquestion(
                    "파일 형식" if not _en else "File Format",
                    "Excel(.xlsx) 형식으로 내보낼까요?\n\n'아니오'를 선택하면 CSV 형식으로 저장됩니다." 
                    if not _en else 
                    "Export as Excel (.xlsx)?\n\n'No' will save as CSV format.",
                    parent=self.root
                )
                
                is_excel = (export_type == 'yes')
                ext = ".xlsx" if is_excel else ".csv"
                default_name = f"trade_history_{datetime.now().strftime('%Y%m%d_%H%M%S')}{ext}"
                
                # 저장 경로 선택
                file_path = filedialog.asksaveasfilename(
                    defaultextension=ext,
                    initialfile=default_name,
                    filetypes=[
                        ("Excel files", "*.xlsx") if is_excel else ("CSV files", "*.csv"),
                        ("All files", "*.*")
                    ],
                    parent=self.root
                )
                
                if not file_path:
                    return
                
                # CSV 내보내기
                if not is_excel:
                    import csv
                    with open(file_path, 'w', newline='', encoding='utf-8-sig') as csvfile:
                        fieldnames = ['시각', '심볼', '방향', '수량', '진입가', '청산가', 
                                     '손익(수수료 전)', '수수료', '순손익(수수료 후)', 
                                     'ROI(%)', '레버리지', '청산사유', 
                                     '수수료타입', '환경']
                        writer = csv.DictWriter(csvfile, fieldnames=fieldnames)
                        writer.writeheader()
                        
                        for t in trades:
                            ts = datetime.fromtimestamp(t.get('ts', 0)).strftime('%Y-%m-%d %H:%M:%S')
                            pnl = t.get('pnl', 0)
                            fee = t.get('fee_amount', 0)
                            net_pnl = pnl - fee
                            
                            writer.writerow({
                                '시각': ts,
                                '심볼': t.get('symbol', ''),
                                '방향': t.get('side', ''),
                                '수량': t.get('quantity', 0),
                                '진입가': t.get('entry_price', 0),
                                '청산가': t.get('exit_price', 0),
                                '손익(수수료 전)': f"{pnl:.4f}",
                                '수수료': f"{fee:.4f}",
                                '순손익(수수료 후)': f"{net_pnl:.4f}",
                                'ROI(%)': f"{t.get('roi_pct', 0):.2f}",
                                '레버리지': f"{t.get('leverage', 0):.1f}x",
                                '청산사유': t.get('trigger', ''),
                                '수수료타입': t.get('fee_type', ''),
                                '환경': t.get('env', 'live')
                            })
                    
                    tk.messagebox.showinfo(
                        "완료" if not _en else "Success",
                        f"CSV 파일로 저장되었습니다:\n{file_path}" if not _en else 
                        f"Saved as CSV:\n{file_path}",
                        parent=self.root
                    )
                
                # Excel 내보내기
                else:
                    try:
                        import openpyxl
                        from openpyxl.styles import Font, PatternFill, Alignment
                    except ImportError:
                        tk.messagebox.showerror(
                            "모듈 없음" if not _en else "Missing Module",
                            "openpyxl 모듈이 필요합니다.\n\nCSV 형식을 사용해주세요." if not _en else
                            "openpyxl module is required.\n\nPlease use CSV format.",
                            parent=self.root
                        )
                        return
                    
                    wb = openpyxl.Workbook()
                    ws = wb.active
                    ws.title = "거래기록"
                    
                    # 헤더
                    headers = ['시각', '심볼', '방향', '수량', '진입가', '청산가', 
                              '손익(수수료 전)', '수수료', '순손익(수수료 후)', 
                              'ROI(%)', '레버리지', '청산사유', 
                              '수수료타입', '환경']
                    ws.append(headers)
                    
                    # 헤더 스타일
                    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                    header_font = Font(bold=True, color="FFFFFF")
                    for cell in ws[1]:
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal="center")
                    
                    # 데이터
                    for t in trades:
                        ts = datetime.fromtimestamp(t.get('ts', 0)).strftime('%Y-%m-%d %H:%M:%S')
                        pnl = t.get('pnl', 0)
                        fee = t.get('fee_amount', 0)
                        net_pnl = pnl - fee
                        
                        row = [
                            ts,
                            t.get('symbol', ''),
                            t.get('side', ''),
                            t.get('quantity', 0),
                            t.get('entry_price', 0),
                            t.get('exit_price', 0),
                            pnl,
                            fee,
                            net_pnl,
                            t.get('roi_pct', 0),
                            f"{t.get('leverage', 0):.1f}x",
                            t.get('trigger', ''),
                            t.get('fee_type', ''),
                            t.get('env', 'live')
                        ]
                        ws.append(row)
                        
                        # 순손익(수수료 후)에 따른 색상
                        row_idx = ws.max_row
                        
                        # 손익(수수료 전) 컬럼 (7번)
                        pnl_cell = ws.cell(row=row_idx, column=7)
                        if pnl > 0:
                            pnl_cell.font = Font(color="888888")
                        elif pnl < 0:
                            pnl_cell.font = Font(color="888888")
                        
                        # 순손익(수수료 후) 컬럼 (9번) - 중요!
                        net_pnl_cell = ws.cell(row=row_idx, column=9)
                        if net_pnl > 0:
                            net_pnl_cell.font = Font(color="2EBD85", bold=True)
                        elif net_pnl < 0:
                            net_pnl_cell.font = Font(color="F6465D", bold=True)
                    
                    # 열 너비 자동 조정
                    for column in ws.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        adjusted_width = min(max_length + 2, 50)
                        ws.column_dimensions[column_letter].width = adjusted_width
                    
                    wb.save(file_path)
                    
                    tk.messagebox.showinfo(
                        "완료" if not _en else "Success",
                        f"Excel 파일로 저장되었습니다:\n{file_path}" if not _en else 
                        f"Saved as Excel:\n{file_path}",
                        parent=self.root
                    )
                
                self._append_log(f"[EXPORT] 거래 기록 내보내기: {os.path.basename(file_path)}")
                
            except Exception as e:
                tk.messagebox.showerror(
                    "오류" if not _en else "Error",
                    f"내보내기 실패:\n{str(e)}" if not _en else f"Export failed:\n{str(e)}",
                    parent=self.root
                )
        
        tk.Button(hdr, text="📊  " + ("내보내기" if not _en else "Export"),
                  command=_export_trades,
                  bg="#1a2a3a", fg="#4fb3d4", relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=10, pady=4,
                  cursor="hand2",
                  activebackground="#2a3a4a", activeforeground="#6fc3e4",
                  ).pack(side=tk.RIGHT, padx=(4, 0))
        
        # 구분선
        tk.Label(hdr, text="|", bg=BG, fg="#3a4060",
                 font=("Segoe UI", 12)).pack(side=tk.RIGHT, padx=6)

        # 배포 초기화 버튼
        # 기록 삭제 버튼
        def _do_clear_records():
            parent_w = getattr(self, "_active_modal", self.root)
            answer = tk.messagebox.askyesno(
                "기록 삭제 확인" if not _en else "Clear Records",
                ("⚠️  다음 파일이 완전히 삭제됩니다:\n\n"
                 "  • 거래 기록 (trade_history.jsonl)\n"
                 "  • 알림 로그 (notifications.log)\n"
                 "  • 봇 실행 로그 (bot.log)\n\n"
                 "설정 및 세션 상태는 영향 없습니다.\n계속하시겠습니까?")
                if not _en else
                ("⚠️  The following files will be permanently deleted:\n\n"
                 "  • Trade history (trade_history.jsonl)\n"
                 "  • Notification log\n"
                 "  • Bot run log (bot.log)\n\n"
                 "Settings and session state are not affected.\nContinue?"),
                parent=parent_w,
            )
            if not answer:
                return
            errors = []
            for fp in [LOG_PATH, NOTIFICATION_PATH, TRADE_LOG_PATH]:
                try:
                    if os.path.exists(fp):
                        os.remove(fp)
                except Exception as e:
                    errors.append(f"{os.path.basename(fp)}: {e}")
            if errors:
                tk.messagebox.showwarning(
                    "일부 오류" if not _en else "Partial Error",
                    "\n".join(errors), parent=parent_w)
            else:
                self._append_log("[REPORT] 거래·로그 기록 삭제 완료")
                tk.messagebox.showinfo(
                    "완료" if not _en else "Done",
                    "삭제 완료!" if not _en else "Cleared!", parent=parent_w)
            try:
                if _reload_table_ref[0]:
                    _reload_table_ref[0]()
            except Exception:
                pass

        tk.Button(hdr,
                  text="🗑  " + ("기록 삭제" if not _en else "Clear Records"),
                  command=_do_clear_records,
                  bg="#2a1520", fg="#ff6b6b", relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=10, pady=4, cursor="hand2",
                  activebackground="#3d1f28", activeforeground="#ff9999",
                  ).pack(side=tk.RIGHT, padx=(4, 0))

        tk.Label(hdr, text="|", bg=BG, fg="#3a4060",
                 font=("Segoe UI", 12)).pack(side=tk.RIGHT, padx=6)

        # 카드 초기화 버튼들 (새로고침 왼쪽에 나란히)
        def _mk_reset_btn(parent, label, cmd, tip=None):
            b = tk.Button(parent, text=label, command=cmd,
                          bg="#241a2a", fg="#ff8fa0", relief=tk.FLAT,
                          font=("Segoe UI", 8), padx=8, pady=4,
                          cursor="hand2",
                          activebackground="#3a2030", activeforeground="#ffaabb")
            b.pack(side=tk.RIGHT, padx=(3, 0))
            return b

        _mk_reset_btn(hdr,
                      "🔄 " + ("전체 초기화" if not _en else "Reset All"),
                      _reset_all_cards)
        _mk_reset_btn(hdr,
                      ("손익만" if not _en else "PnL only"),
                      _reset_pnl_only)
        _mk_reset_btn(hdr,
                      ("승률·거래수" if not _en else "Win/Trades"),
                      _reset_win_rate_only)

        tk.Label(hdr, text=("카드 초기화:" if not _en else "Reset:"),
                 bg=BG, fg=FG2, font=("Segoe UI", 8)).pack(side=tk.RIGHT, padx=(0, 4))

        # ── 요약 통계 카드 6개 ────────────────────────────────────
        summary_outer = tk.Frame(frame, bg=BG)
        summary_outer.grid(row=1, column=0, sticky="ew", padx=24, pady=(10, 0))
        for c in range(6):
            summary_outer.columnconfigure(c, weight=1, uniform="rpt")

        stat_refs = {}

        def _make_stat_card(parent, col, key, title):
            card = tk.Frame(parent, bg=BG2,
                            highlightbackground=BORDER, highlightthickness=1)
            card.grid(row=0, column=col, padx=4, pady=2, sticky="nsew")
            tk.Label(card, text=title, bg=BG2, fg=FG2,
                     font=("Segoe UI", 9)).pack(anchor="w", padx=10, pady=(8, 0))
            val_lbl = tk.Label(card, text="—", bg=BG2, fg=FG,
                               font=("Segoe UI", 13, "bold"))
            val_lbl.pack(anchor="w", padx=10, pady=(0, 2))
            sub_lbl = tk.Label(card, text="", bg=BG2, fg=FG2,
                               font=("Segoe UI", 8))
            sub_lbl.pack(anchor="w", padx=10, pady=(0, 6))
            stat_refs[key] = (val_lbl, sub_lbl)

        cards_meta = [
            ("total_pnl",   "총 손익"          if not _en else "Total PnL"),
            ("win_rate",    "승률 / 거래수"    if not _en else "Win Rate / Trades"),
            ("rr",          "손익비 R:R"       if not _en else "Win/Loss R:R"),
            ("expectancy",  "거래당 기댓값"    if not _en else "Expectancy"),
            ("best_worst",  "최대 이익 / 손실" if not _en else "Best / Worst"),
            ("top_trigger", "주요 청산 사유"   if not _en else "Top Exit Reason"),
        ]
        for col, (key, title) in enumerate(cards_meta):
            _make_stat_card(summary_outer, col, key, title)

        # ── 거래 기록 테이블 ──────────────────────────────────────
        table_frame = tk.Frame(frame, bg=BG3,
                               highlightbackground=BORDER, highlightthickness=1)
        table_frame.grid(row=2, column=0, sticky="nsew", padx=24, pady=(8, 16))
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(1, weight=1)

        # 테이블 헤더 바
        th_row = tk.Frame(table_frame, bg="#111622")
        th_row.grid(row=0, column=0, sticky="ew")

        tbl_label = tk.Label(th_row,
                             text="거래 기록" if not _en else "Trade History",
                             bg="#111622", fg=FG,
                             font=("Segoe UI", 10, "bold"))
        tbl_label.pack(side=tk.LEFT, padx=12, pady=7)

        filter_frame = tk.Frame(th_row, bg="#111622")
        filter_frame.pack(side=tk.RIGHT, padx=8, pady=4)

        # 환경 필터 (테스트넷 / 실거래 / 전체) — 오른쪽 끝
        _env_filter_var = tk.StringVar(value="all")
        env_btns = {}
        for ev, etxt in [("all",      "전체환경"  if not _en else "All Env"),
                         ("testnet",  "테스트넷"  if not _en else "Testnet"),
                         ("live",     "실거래"    if not _en else "Live")]:
            def _make_env_cmd(v=ev):
                def _cmd():
                    _env_filter_var.set(v)
                    for k, b2 in env_btns.items():
                        b2.config(bg="#2563eb" if k == v else "#1c2438",
                                  fg="#ffffff"  if k == v else FG2)
                    if _reload_table_ref[0]:
                        _reload_table_ref[0]()
                return _cmd
            col_ev = "#2563eb" if ev == "all" else "#1c2438"
            b = tk.Button(filter_frame, text=etxt, command=_make_env_cmd(),
                          bg=col_ev, fg="#ffffff" if ev == "all" else FG2,
                          relief=tk.FLAT, font=("Segoe UI", 8),
                          padx=7, pady=2, cursor="hand2")
            b.pack(side=tk.LEFT, padx=2)
            env_btns[ev] = b

        # 구분선
        tk.Label(filter_frame, text="│", bg="#111622", fg="#3a4060").pack(side=tk.LEFT, padx=4)

        # 손익 필터 (전체 / 이익 / 손실)
        _pnl_filter_var = tk.StringVar(value="all")
        pnl_btns = {}
        for fval, ftxt, fcol in [("all",  "전체" if not _en else "All",  "#23293a"),
                                  ("win",  "이익" if not _en else "Win",  "#1a2e1a"),
                                  ("loss", "손실" if not _en else "Loss", "#2e1a1a")]:
            def _make_pnl_cmd(v=fval):
                def _cmd():
                    _pnl_filter_var.set(v)
                    for k, b2 in pnl_btns.items():
                        active = k == v
                        b2.config(bg={"all": "#23293a", "win": "#1a3a1a", "loss": "#3a1a1a"}[k] if active
                                  else {"all": "#1c2438", "win": "#1c2438", "loss": "#1c2438"}[k],
                                  fg={"all": FG, "win": ACC, "loss": ERR}[k] if active else FG2)
                    if _reload_table_ref[0]:
                        _reload_table_ref[0]()
                return _cmd
            b = tk.Button(filter_frame, text=ftxt, command=_make_pnl_cmd(),
                          bg="#23293a" if fval == "all" else "#1c2438",
                          fg=FG if fval == "all" else FG2,
                          relief=tk.FLAT, font=("Segoe UI", 8),
                          padx=7, pady=2, cursor="hand2")
            b.pack(side=tk.LEFT, padx=2)
            pnl_btns[fval] = b

        # Treeview 컬럼
        cols = ("env", "time", "symbol", "side", "lev", "qty",
                "entry", "exit", "pnl", "roi", "fee", "trigger", "mode")
        col_cfg = {
            "env":     (64,  "환경"     if not _en else "Env"),
            "time":    (120, "시간"     if not _en else "Time"),
            "symbol":  (88,  "심볼"     if not _en else "Symbol"),
            "side":    (46,  "방향"     if not _en else "Side"),
            "lev":     (40,  "레버"     if not _en else "Lev"),
            "qty":     (72,  "수량"     if not _en else "Qty"),
            "entry":   (90,  "진입가"   if not _en else "Entry"),
            "exit":    (90,  "청산가"   if not _en else "Exit"),
            "pnl":     (90,  "손익(U)"  if not _en else "PnL (U)"),
            "roi":     (68,  "ROI%"),
            "fee":     (80,  "수수료(M/T)" if not _en else "Fee(M/T)"),
            "trigger": (120, "청산사유" if not _en else "Trigger"),
            "mode":    (50,  "모드"     if not _en else "Mode"),
        }

        style = ttk.Style()
        style.configure("Report.Treeview",
                        background=BG3, foreground=FG,
                        rowheight=22, fieldbackground=BG3,
                        borderwidth=0, font=("Segoe UI", 9))
        style.configure("Report.Treeview.Heading",
                        background="#111622", foreground=FG2,
                        relief="flat", font=("Segoe UI", 9, "bold"))
        style.map("Report.Treeview",
                  background=[("selected", "#2a3860")],
                  foreground=[("selected", "#ffffff")])

        tree_wrap = tk.Frame(table_frame, bg=BG3)
        tree_wrap.grid(row=1, column=0, sticky="nsew")
        tree_wrap.columnconfigure(0, weight=1)
        tree_wrap.rowconfigure(0, weight=1)

        vsb = ttk.Scrollbar(tree_wrap, orient="vertical")
        hsb = ttk.Scrollbar(tree_wrap, orient="horizontal")
        tree = ttk.Treeview(tree_wrap, columns=cols, show="headings",
                            style="Report.Treeview",
                            yscrollcommand=vsb.set, xscrollcommand=hsb.set)
        vsb.config(command=tree.yview)
        hsb.config(command=tree.xview)

        for cid, (w, hdr_txt) in col_cfg.items():
            tree.heading(cid, text=hdr_txt,
                         command=lambda c=cid: _sort_by(c))
            tree.column(cid, width=w, minwidth=w, anchor="center")
        tree.column("time",    anchor="w")
        tree.column("trigger", anchor="w")
        tree.column("fee",     anchor="center")
        tree.column("env",     anchor="center")

        tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")

        # 태그: 환경별 + 손익별 조합
        tree.tag_configure("win_live",     foreground=ACC)
        tree.tag_configure("loss_live",    foreground=ERR)
        tree.tag_configure("win_testnet",  foreground="#5db8f0")
        tree.tag_configure("loss_testnet", foreground="#c97aaa")
        tree.tag_configure("neutral",      foreground=FG2)
        tree.tag_configure("even",         background="#181e2c")
        tree.tag_configure("odd",          background=BG3)

        # ── 요약 통계 로드 ─────────────────────────────────────────
        def _load_stats():
            all_trades = _read_all_trades()
            env_f = _env_filter_var.get()
            if env_f != "all":
                all_trades = [t for t in all_trades if t.get("env", "live") == env_f]

            wins   = [t for t in all_trades if float(t.get("pnl", 0)) > 0]
            losses = [t for t in all_trades if float(t.get("pnl", 0)) < 0]
            n_total   = len(all_trades)
            n_wins    = len(wins)
            total_pnl = sum(float(t.get("pnl", 0)) for t in all_trades)
            win_rate  = n_wins / n_total * 100 if n_total else 0
            avg_win   = sum(float(t.get("pnl", 0)) for t in wins)   / n_wins       if wins   else 0
            avg_loss  = sum(float(t.get("pnl", 0)) for t in losses) / len(losses)  if losses else 0
            rr        = abs(avg_win / avg_loss)  if avg_loss else 0
            expectancy = total_pnl / n_total     if n_total  else 0
            best  = max((float(t.get("pnl", 0)) for t in all_trades), default=0)
            worst = min((float(t.get("pnl", 0)) for t in all_trades), default=0)

            from collections import Counter as _Cnt
            trig_counts = _Cnt(t.get("trigger", "—") for t in all_trades)
            top_trig, top_trig_n = trig_counts.most_common(1)[0] if trig_counts else ("—", 0)

            def _set(key, val, sub="", color=FG):
                if key in stat_refs:
                    v_l, s_l = stat_refs[key]
                    try:
                        v_l.config(text=val, fg=color)
                        s_l.config(text=sub)
                    except Exception:
                        pass

            sign = "+" if total_pnl >= 0 else ""
            env_label = ("" if env_f == "all"
                         else (" [테스트넷]" if env_f == "testnet" else " [실거래]"))
            _set("total_pnl",
                 f"{sign}{total_pnl:.4f} U",
                 f"{n_total}건 합산{env_label}" if not _en else f"{n_total} trades{env_label}",
                 ACC if total_pnl >= 0 else ERR)
            _set("win_rate",
                 f"{win_rate:.1f}%",
                 f"{n_wins}승 / {n_total - n_wins}패" if not _en else f"{n_wins}W / {n_total - n_wins}L",
                 ACC if win_rate >= 50 else WARN)
            _set("rr",
                 f"{rr:.2f}x" if rr else "—",
                 f"+{avg_win:.2f} / {avg_loss:.2f} U",
                 ACC if rr >= 1 else WARN)
            exp_sign = "+" if expectancy >= 0 else ""
            _set("expectancy",
                 f"{exp_sign}{expectancy:.4f} U",
                 "거래당 평균" if not _en else "per trade avg",
                 ACC if expectancy >= 0 else ERR)
            _set("best_worst",
                 f"+{best:.2f} / {worst:.2f}",
                 "이익 / 손실" if not _en else "best / worst",
                 WARN)
            _set("top_trigger",
                 top_trig,
                 f"{top_trig_n}건" if not _en else f"{top_trig_n} trades",
                 FG2)

            # 메이커/테이커 비율 + 총 수수료 → top_trigger 카드 서브라벨에 표시
            _n_maker   = sum(1 for t in all_trades if t.get("fee_type") == "maker")
            _n_taker   = sum(1 for t in all_trades if t.get("fee_type") == "taker")
            _fee_total = sum(float(t.get("fee_amount", 0) or 0) for t in all_trades)
            if n_total > 0 and (_n_maker + _n_taker) > 0:
                _maker_pct = _n_maker / n_total * 100
                _fee_sub = (f"M {_n_maker}건({_maker_pct:.0f}%) / T {_n_taker}건 | 수수료합 {_fee_total:.3f}U"
                            if not _en else
                            f"M {_n_maker}({_maker_pct:.0f}%) / T {_n_taker} | fee Σ {_fee_total:.3f}U")
                _set("top_trigger",
                     top_trig,
                     _fee_sub,
                     FG2)

        # ── 공통 데이터 읽기 ──────────────────────────────────────
        def _read_all_trades():
            trades = []
            try:
                if os.path.exists(TRADE_LOG_PATH):
                    with open(TRADE_LOG_PATH, "r", encoding="utf-8") as fh:
                        for ln in fh:
                            ln = ln.strip()
                            if ln:
                                try:
                                    trades.append(json.loads(ln))
                                except Exception:
                                    pass
            except Exception:
                pass
            return trades

        # ── 테이블 로드 ────────────────────────────────────────────
        _sort_state = {"col": "time", "reverse": True}

        def _fmt_fee(t: dict) -> str:
            """fee_type + fee_amount 표시. 구 데이터는 taker 추정."""
            ft  = t.get("fee_type")       # "maker" | "taker" | None
            fa  = t.get("fee_amount")     # float | None
            if fa is not None:
                label = "M" if ft == "maker" else "T"
                return f"{label} {float(fa):.4f}"
            # 구 데이터: fee 필드 없음 → taker 추정
            try:
                taker = float(t.get("taker_fee_pct") or 0.0005)
                ep    = float(t.get("entry_price", 0))
                xp    = float(t.get("exit_price", 0))
                qty   = float(t.get("quantity", 0))
                est   = xp * qty * taker
                return f"T≈{est:.4f}"
            except Exception:
                return "—"

        def _sort_by(col):
            rev = not _sort_state["reverse"] if _sort_state["col"] == col else False
            _sort_state.update(col=col, reverse=rev)
            if _reload_table_ref[0]:
                _reload_table_ref[0]()

        def _reload_table():
            for row in tree.get_children():
                tree.delete(row)
            all_trades = _read_all_trades()

            # 환경 필터
            env_f = _env_filter_var.get()
            if env_f != "all":
                all_trades = [t for t in all_trades if t.get("env", "live") == env_f]

            # 손익 필터
            pnl_f = _pnl_filter_var.get()
            if pnl_f == "win":
                all_trades = [t for t in all_trades if float(t.get("pnl", 0)) > 0]
            elif pnl_f == "loss":
                all_trades = [t for t in all_trades if float(t.get("pnl", 0)) < 0]

            # 정렬
            col  = _sort_state["col"]
            rev  = _sort_state["reverse"]
            smap = {
                "env":     lambda t: t.get("env", "live"),
                "time":    lambda t: float(t.get("ts", 0)),
                "symbol":  lambda t: t.get("symbol", ""),
                "side":    lambda t: t.get("side", ""),
                "lev":     lambda t: float(t.get("leverage", 0)),
                "qty":     lambda t: float(t.get("quantity", 0)),
                "entry":   lambda t: float(t.get("entry_price", 0)),
                "exit":    lambda t: float(t.get("exit_price", 0)),
                "pnl":     lambda t: float(t.get("pnl", 0)),
                "roi":     lambda t: float(t.get("roi_pct") or 0),
                "fee":     lambda t: t.get("fee_amount", 0.0),
                "trigger": lambda t: t.get("trigger", ""),
                "mode":    lambda t: t.get("mode", ""),
            }
            try:
                all_trades.sort(key=smap.get(col, smap["time"]), reverse=rev)
            except Exception:
                pass

            from datetime import datetime as _dt
            for idx, t in enumerate(all_trades):
                ts  = float(t.get("ts", 0))
                env = t.get("env", "live")
                try:
                    t_str = _dt.fromtimestamp(ts).strftime("%m-%d %H:%M:%S")
                except Exception:
                    t_str = "—"

                env_label = ("TN" if env == "testnet" else "LIVE")
                pnl  = float(t.get("pnl", 0))
                roi  = t.get("roi_pct", None)
                sign = "+" if pnl >= 0 else ""

                # 태그 결정: 환경 + 손익 조합
                if pnl > 0:
                    pnl_tag = "win_testnet"  if env == "testnet" else "win_live"
                elif pnl < 0:
                    pnl_tag = "loss_testnet" if env == "testnet" else "loss_live"
                else:
                    pnl_tag = "neutral"
                even_tag = "even" if idx % 2 == 0 else "odd"

                tree.insert("", tk.END, tags=(pnl_tag, even_tag), values=(
                    env_label,
                    t_str,
                    t.get("symbol", "—"),
                    t.get("side", "—"),
                    f'{t.get("leverage", "—")}x',
                    f'{float(t.get("quantity", 0)):.4f}',
                    f'{float(t.get("entry_price", 0)):.4f}',
                    f'{float(t.get("exit_price", 0)):.4f}',
                    f'{sign}{pnl:.4f}',
                    f'{roi:+.2f}%' if roi is not None else "—",
                    _fmt_fee(t),
                    t.get("trigger", "—"),
                    t.get("mode", "—"),
                ))

            try:
                tbl_label.config(
                    text=("거래 기록" if not _en else "Trade History")
                    + f"  ({len(all_trades)}건)")
            except Exception:
                pass

            # 통계도 같이 갱신
            try:
                _load_stats()
            except Exception:
                pass

        # 참조 등록 후 초기 로드
        _load_stats_ref[0]   = _load_stats
        _reload_table_ref[0] = _reload_table
        _load_stats()
        _reload_table()


    def _build_env_tab(self, frame):
        frame.columnconfigure(0, weight=1)

        # ── 스크롤 가능한 컨테이너 구성 ──
        btn_container = tk.Frame(frame, bg="#181A20")
        btn_container.pack(side=tk.BOTTOM, fill="x", padx=40, pady=(8, 12))
        tk.Frame(btn_container, bg="#343942", height=1).pack(fill="x")

        canvas = tk.Canvas(frame, bg="#181A20", highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        body = tk.Frame(canvas, bg="#181A20")
        canvas.create_window((0, 0), window=body, anchor="nw")
        def _on_env_configure(event):
            canvas.configure(scrollregion=canvas.bbox("all"))
        body.bind("<Configure>", _on_env_configure)
        def _on_env_canvas_configure(event):
            canvas.itemconfig(canvas.find_withtag("all")[0], width=event.width)
        canvas.bind("<Configure>", _on_env_canvas_configure)
        def _env_mousewheel(event):
            try:
                if canvas.winfo_exists():
                    canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")
            except Exception:
                pass
        def _env_bind_mw(e=None):
            canvas.bind_all("<MouseWheel>", _env_mousewheel)
        def _env_unbind_mw(e=None):
            try: canvas.unbind_all("<MouseWheel>")
            except Exception: pass
        frame.bind("<Visibility>", _env_bind_mw, add="+")
        canvas.bind("<Enter>", _env_bind_mw, add="+")
        canvas.bind("<Leave>", _env_unbind_mw, add="+")
        body.bind("<Enter>", _env_bind_mw, add="+")
        body.bind("<Leave>", _env_unbind_mw, add="+")

        # 이하 body에 pack (기존 frame → body로 변경)
        _is_ko = self.language == "ko"
        entry_pad = {"padx": 40, "pady": (0, 10)}

        # ── 환경변수 설정 가이드 (최상단) ──
        tk.Label(body,
                 text="환경변수 설정 방법" if _is_ko else "How to Set Environment Variables",
                 bg="#181A20", fg="white",
                 font=("Malgun Gothic", 13, "bold"), anchor="w"
                 ).pack(fill="x", padx=40, pady=(30, 4))
        tk.Label(body,
                 text=("이 버전부터 API 키/시크릿은 프로그램 안에서 입력하지 않습니다.\n"
                       "Windows 환경 변수에 아래 키를 등록한 뒤 GUI를 다시 시작해 주세요."
                       if _is_ko else
                       "From this version, API keys are not entered inside the program.\n"
                       "Register the keys below as Windows environment variables, then restart the GUI."),
                 bg="#181A20", fg="#c0c6dc",
                 font=("Malgun Gothic", 10), justify="left", wraplength=720, anchor="w"
                 ).pack(fill="x", padx=40, pady=(0, 12))

        # 단계별 가이드
        _guide_frame = tk.Frame(body, bg="#1c1f2b", highlightbackground="#343942", highlightthickness=1)
        _guide_frame.pack(fill="x", padx=40, pady=(0, 8))

        if _is_ko:
            _steps = [
                ("①", "Windows 검색창에 '환경 변수' 입력 → '시스템 환경 변수 편집' 클릭"),
                ("②", "'환경 변수(N)...' 버튼 클릭"),
                ("③", "'사용자 변수' 영역에서 '새로 만들기(N)...' 클릭"),
                ("④", "아래 변수명과 값을 하나씩 추가 (총 4개)"),
                ("⑤", "모두 입력 후 '확인' 버튼 클릭 (모든 창 닫기)"),
                ("⑥", "이 프로그램을 완전히 종료한 후 다시 실행"),
            ]
        else:
            _steps = [
                ("①", "Search 'Environment Variables' in Windows → 'Edit system environment variables'"),
                ("②", "Click 'Environment Variables...' button"),
                ("③", "Under 'User variables', click 'New...'"),
                ("④", "Add the following variables one by one (4 total)"),
                ("⑤", "Click 'OK' on all dialogs"),
                ("⑥", "Close and restart this program"),
            ]

        for _num, _desc in _steps:
            _step_row = tk.Frame(_guide_frame, bg="#1c1f2b")
            _step_row.pack(fill="x", padx=16, pady=3)
            tk.Label(_step_row, text=_num, bg="#1c1f2b", fg="#F0B90B",
                     font=("Malgun Gothic", 10, "bold"), width=3, anchor="e"
                     ).pack(side=tk.LEFT, padx=(0, 8))
            tk.Label(_step_row, text=_desc, bg="#1c1f2b", fg="#f5f7ff",
                     font=("Malgun Gothic", 10), anchor="w"
                     ).pack(side=tk.LEFT, fill="x", expand=True)
        tk.Frame(_guide_frame, bg="#1c1f2b", height=6).pack()

        # 변수명 테이블
        _var_table_frame = tk.Frame(body, bg="#1c1f2b", highlightbackground="#343942", highlightthickness=1)
        _var_table_frame.pack(fill="x", padx=40, pady=(0, 8))

        _tbl_hdr = tk.Frame(_var_table_frame, bg="#0D1117")
        _tbl_hdr.pack(fill="x", padx=1, pady=(1, 0))
        for _col_text, _col_w in [("환경" if _is_ko else "Env", 8),
                                   ("변수명" if _is_ko else "Variable Name", 28),
                                   ("값" if _is_ko else "Value", 30)]:
            tk.Label(_tbl_hdr, text=_col_text, bg="#0D1117", fg="#F0B90B",
                     font=("Malgun Gothic", 9, "bold"), width=_col_w, anchor="w"
                     ).pack(side=tk.LEFT, padx=4, pady=4)

        _var_rows = [
            ("TESTNET", "TESTNET_API_KEY",
             "테스트넷 API Key" if _is_ko else "Testnet API Key"),
            ("TESTNET", "TESTNET_API_SECRET",
             "테스트넷 Secret Key" if _is_ko else "Testnet Secret Key"),
            ("LIVE", "BINANCE_API_KEY",
             "실거래 API Key" if _is_ko else "Live API Key"),
            ("LIVE", "BINANCE_API_SECRET",
             "실거래 Secret Key" if _is_ko else "Live Secret Key"),
        ]
        for _idx, (_env_label, _var_name, _var_desc) in enumerate(_var_rows):
            _bg = "#1c1f2b" if _idx % 2 == 0 else "#151922"
            _env_color = "#2EBD85" if _env_label == "TESTNET" else "#F0B90B"
            _tbl_row = tk.Frame(_var_table_frame, bg=_bg)
            _tbl_row.pack(fill="x", padx=1)
            tk.Label(_tbl_row, text=_env_label, bg=_bg, fg=_env_color,
                     font=("Malgun Gothic", 9, "bold"), width=8, anchor="w"
                     ).pack(side=tk.LEFT, padx=4, pady=3)
            tk.Label(_tbl_row, text=_var_name, bg=_bg, fg="#58a6ff",
                     font=("Consolas", 10), width=28, anchor="w"
                     ).pack(side=tk.LEFT, padx=4, pady=3)
            tk.Label(_tbl_row, text=_var_desc, bg=_bg, fg="#c0c6dc",
                     font=("Malgun Gothic", 9), width=30, anchor="w"
                     ).pack(side=tk.LEFT, padx=4, pady=3)

        tk.Label(_var_table_frame,
                 text=("※ 각 변수값은 복사/붙여넣기 후 '새로 만들기' 또는 '편집'으로 저장하고,\n"
                       "   변경 시 GUI를 재실행해야 적용됩니다."
                       if _is_ko else
                       "※ Copy/paste each value into 'New' or 'Edit', then save.\n"
                       "   Restart the GUI after changes to apply."),
                 bg="#1c1f2b", fg="#888e9e",
                 font=("Malgun Gothic", 9), anchor="w", justify="left"
                 ).pack(fill="x", padx=16, pady=(4, 10))

        # ── API 키 상태 확인 ──
        tk.Label(body, text=self._t("api_settings_title","API 설정 상태"), bg="#181A20", fg="white",
                 font=("Malgun Gothic", 13, "bold"), anchor="w").pack(fill="x", padx=40, pady=(20, 8))

        status_frame = tk.Frame(body, bg="#1c1f2b", highlightbackground="#343942", highlightthickness=1)
        status_frame.pack(fill="x", padx=40, pady=(0, 20))
        env_rows = [
            ("TESTNET_API_KEY", self._t("env_label_testnet_key","테스트넷 API Key")),
            ("TESTNET_API_SECRET", self._t("env_label_testnet_secret","테스트넷 API Secret")),
            ("BINANCE_API_KEY", self._t("env_label_live_key","라이브 API Key")),
            ("BINANCE_API_SECRET", self._t("env_label_live_secret","라이브 API Secret")),
        ]
        for env_name, label_text in env_rows:
            row = tk.Frame(status_frame, bg="#1c1f2b")
            row.pack(fill="x", padx=16, pady=6)
            tk.Label(row, text=label_text, bg="#1c1f2b", fg="#f5f7ff", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(side=tk.LEFT)
            value = os.environ.get(env_name)
            if value:
                preview = value[-4:]
                status_text = f"{self._t('env_status_set','설정됨')} (…{preview})"
                status_color = "#2EBD85"
            else:
                status_text = self._t("env_status_unset","미설정")
                status_color = "#F6465D"
            tk.Label(row, text=status_text, bg="#1c1f2b", fg=status_color, font=("Malgun Gothic", 10, "bold")).pack(side=tk.RIGHT)

        # ── [PATCH-11] 바이낸스 레퍼럴 코드 설정 ──
        tk.Label(body, text=self._t("referral_title", "레퍼럴 코드 / Referral Code"),
                 bg="#181A20", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"),
                 anchor="w").pack(fill="x", padx=40, pady=(20, 6))
        ref_frame = tk.Frame(body, bg="#1c1f2b", highlightbackground="#343942", highlightthickness=1)
        ref_frame.pack(fill="x", padx=40, pady=(0, 4))
        ref_inner = tk.Frame(ref_frame, bg="#1c1f2b")
        ref_inner.pack(fill="x", padx=16, pady=10)
        _is_ko = self.language == "ko"
        tk.Label(ref_inner,
                 text="레퍼럴 코드" if _is_ko else "Referral Code",
                 bg="#1c1f2b", fg="#f5f7ff",
                 font=("Malgun Gothic", 10, "bold"), anchor="w"
                 ).pack(side=tk.LEFT)
        # [PATCH-12] 레퍼럴 코드 고정 표시 (읽기 전용)
        self.settings_data["binance_referral_code"] = REFERRAL_CODE
        _ref_var = tk.StringVar(value=REFERRAL_CODE)
        _ref_entry = tk.Entry(ref_inner, textvariable=_ref_var, width=20,
                              bg="#0D1117", fg="#f5f7ff",
                              insertbackground="#f5f7ff",
                              relief="flat", font=("Courier New", 10),
                              highlightbackground="#343942", highlightthickness=1,
                              state="readonly", readonlybackground="#0D1117")
        _ref_entry.pack(side=tk.LEFT, padx=(12, 8))
        _ref_msg = tk.Label(ref_inner, text="", bg="#1c1f2b", fg="#888e9e",
                            font=("Malgun Gothic", 9))
        _ref_msg.pack(side=tk.LEFT, padx=(4, 0))

        _ref_msg.configure(
            text=f"🔒 {'고정됨' if _is_ko else 'Fixed'}: {REFERRAL_CODE}",
                fg="#888e9e")

        # [PATCH-12] 레퍼럴 혜택 안내
        _benefit_text = (
            "💰 이 코드로 가입 시 선물 거래 수수료 5% 할인 혜택이 적용됩니다"
            if _is_ko else
            "💰 Sign up with this code to get 5% off on futures trading fees"
        )
        tk.Label(ref_frame, text=_benefit_text,
                 bg="#1c1f2b", fg="#F0B90B",
                 font=("Malgun Gothic", 9), anchor="w"
                 ).pack(fill="x", padx=16, pady=(0, 10))

        tk.Label(body, text=self._t("default_env_title","기본 실행 환경"), bg="#181A20", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=40, pady=(20, 6))
        current_env = self.env_mode if self.env_mode in {"TESTNET", "LIVE"} else ("TESTNET" if self.settings_data.get("default_env_testnet", True) else "LIVE")
        env_choice = tk.StringVar(value=current_env)
        env_frame = tk.Frame(body, bg="#181A20")
        env_frame.pack(fill="x", padx=40)
        def set_env(value):
            # [PATCH-5] LIVE 전환 시 확인 다이얼로그
            if value == "LIVE" and env_choice.get() != "LIVE":
                from tkinter import messagebox as _mb
                _en = self.language == "en"
                confirmed = _mb.askyesno(
                    "Confirm LIVE Switch" if _en else "실거래 전환 확인",
                    ("Switching from Testnet to LIVE.\n"
                     "Real funds will be used.\n\nContinue?"
                     if _en else
                     "테스트넷에서 실거래(LIVE)로 전환합니다.\n"
                     "실제 자금이 사용됩니다.\n\n계속하시겠습니까?"),
                    icon="warning"
                )
                if not confirmed:
                    return
            env_choice.set(value)

        style = ttk.Style()
        style.configure("Env.TRadiobutton", background="#181A20", foreground="#f5f7ff", font=("Malgun Gothic", 11, "bold"))
        style.map("Env.TRadiobutton", background=[("selected", "#181A20")], foreground=[("selected", "#2EBD85")])
        style.configure("EnvLive.TRadiobutton", background="#181A20", foreground="#f5f7ff", font=("Malgun Gothic", 11, "bold"))
        style.map("EnvLive.TRadiobutton", background=[("selected", "#181A20")], foreground=[("selected", "#F0B90B")])
        style.configure("EnvSave.TButton", background="#1e2d3d", foreground="#7ecbf5", font=("Malgun Gothic", 10, "bold"), padding=(20, 8))
        style.map("EnvSave.TButton", background=[("active", "#2a4060")], foreground=[("active", "#ffffff")])
        style.configure("EnvDefault.TButton", background="#1e2230", foreground="#8892a8", font=("Malgun Gothic", 10), padding=(20, 8))
        style.map("EnvDefault.TButton", background=[("active", "#2a3040")], foreground=[("active", "#c8cfe8")])
        testnet_btn = ttk.Radiobutton(env_frame, text="TESTNET", value="TESTNET", variable=env_choice, command=lambda: set_env("TESTNET"), style="Env.TRadiobutton")
        live_btn = ttk.Radiobutton(env_frame, text="LIVE", value="LIVE", variable=env_choice, command=lambda: set_env("LIVE"), style="EnvLive.TRadiobutton")
        testnet_btn.pack(side=tk.LEFT, padx=(0, 20))
        live_btn.pack(side=tk.LEFT)

        tk.Label(body, text=self._t("language_title","언어 / Language"), bg="#181A20", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=40, pady=(20, 6))
        language_var = tk.StringVar(value=self.language)
        lang_frame = tk.Frame(body, bg="#181A20")
        lang_frame.pack(fill="x", padx=40)
        ttk.Radiobutton(lang_frame, text="한국어 (Korean)", value="ko", variable=language_var, style="Env.TRadiobutton").pack(side=tk.LEFT, padx=(0, 20))
        ttk.Radiobutton(lang_frame, text="English", value="en", variable=language_var, style="EnvLive.TRadiobutton").pack(side=tk.LEFT)

        notify_var = tk.BooleanVar(value=bool(self.settings_data.get("alert_enabled", True)))
        auto_start_var = tk.BooleanVar(value=bool(self.settings_data.get("auto_start", False)))

        def build_custom_checkbox(parent, text, variable, pady=(12, 0)):
            row = tk.Frame(parent, bg="#181A20")
            row.pack(fill="x", padx=40, pady=pady)
            indicator = tk.Label(row, bg="#181A20")
            indicator.pack(side=tk.LEFT, padx=(0, 8))
            lbl = tk.Label(row, text=text, bg="#181A20", fg="#f5f7ff", font=("Malgun Gothic", 10, "bold"))
            lbl.pack(side=tk.LEFT)

            def update_indicator(*_):
                if variable.get() and self.checkbox_images.get("on"):
                    indicator.configure(image=self.checkbox_images["on"], text="")
                    indicator.image = self.checkbox_images["on"]
                elif not variable.get() and self.checkbox_images.get("off"):
                    indicator.configure(image=self.checkbox_images["off"], text="")
                    indicator.image = self.checkbox_images["off"]
                else:
                    indicator.configure(text="●" if variable.get() else "○", fg="#2EBD85" if variable.get() else "#f5f7ff")

            def toggle(_event=None):
                variable.set(not variable.get())
                update_indicator()

            for widget in (row, indicator, lbl):
                widget.bind("<Button-1>", toggle)
            update_indicator()
            return row

        build_custom_checkbox(body, self._t("env_notify_popup","주문/체결 완료 시 알림 팝업"), notify_var, pady=(12, 0))
        build_custom_checkbox(body, self._t("env_auto_start","Windows 시작 시 자동 실행"), auto_start_var, pady=(6, 0))

        # 하단 여백
        tk.Frame(body, bg="#181A20", height=20).pack(fill="x")

        def save_env():
            settings_dialog = frame.winfo_toplevel()
            selected_env = env_choice.get()
            prev_env = self.env_mode
            prev_language = self.language
            self.settings_data.update(
                {
                    "default_env_testnet": selected_env == "TESTNET",
                    "alert_enabled": notify_var.get(),
                    "auto_start": auto_start_var.get(),
                }
            )
            self.alert_enabled = notify_var.get()
            self._configure_auto_start(auto_start_var.get())
            self.env_mode = selected_env
            selected_language = language_var.get() or "ko"
            if selected_language not in ("ko", "en"):
                selected_language = "ko"
            self.language = selected_language
            self.settings_data["ui_language"] = self.language
            # 실행 중인 엔진 config에도 즉시 반영 (재시작 없이 로그 언어 전환)
            try:
                from binance_futures_bot1_1 import main as _eng_main
                if _eng_main.current_engine is not None:
                    _eng_main.current_engine.config.ui_language = self.language
            except Exception:
                pass
            if selected_env != prev_env:
                self.env_request_token += 1
                self._clear_env_specific_views()
                self._append_log(f"[INFO] Switching engine environment {prev_env} → {selected_env}")
            self.state_data["env_label"] = self.env_mode
            self._save_json(STATE_PATH, self.state_data)
            self._render_env_toggle()
            self._save_json(CONFIG_PATH, self.settings_data)
            if self.language == prev_language:
                lang_note = ""
            else:
                lang_note_msg = (
                    "Language changes apply more broadly after reopening or restarting."
                    if self.language == "en"
                    else "언어 설정은 새 창을 열거나 재시작하면 더 넓게 적용됩니다."
                )
                lang_note = "\n" + lang_note_msg

            self._show_info(self._t("save","저장"), self._t("env_saved_msg","환경 설정이 저장되었습니다. (API 키는 환경 변수에서 불러옵니다)") + lang_note)
            if selected_env != prev_env:
                self._prompt_restart_after_setting_change(self._t("restart_prompt","환경 설정이 변경되었습니다. 엔진을 다시 시작할까요?"))

            if self.language != prev_language:
                # _active_modal 정리 후 destroy
                if getattr(self, "_active_modal", None) is settings_dialog:
                    self._active_modal = None
                try:
                    pass  # grab_set 제거됨 — grab_release 불필요
                except Exception:
                    pass
                try:
                    settings_dialog.destroy()
                except Exception:
                    pass
                # 언어 변경: UI 즉시 재빌드 (두 번째 팝업 제거 - 첫 저장 알림에 이미 안내 포함)
                self.root.after(50, self._rebuild_ui)
            if not self.engine_running:
                missing_env = []
                for env_var in ("TESTNET_API_KEY", "TESTNET_API_SECRET", "BINANCE_API_KEY", "BINANCE_API_SECRET"):
                    if not os.environ.get(env_var):
                        missing_env.append(env_var)
                if missing_env:
                    _warn_parent = settings_dialog if settings_dialog.winfo_exists() else None
                    self._show_warning(
                        self._t("api_env_check_title","API 환경 변수 확인"),
                        self._t("api_env_missing_msg","아직 설정되지 않은 환경 변수가 있습니다.") + "\n" + "\n".join(missing_env),
                        parent=_warn_parent,
                    )

        # btn_container는 상단에서 이미 생성 (side=BOTTOM, 스크롤 영역 밖)
        def restore_env_defaults():
            env_choice.set("TESTNET")
            language_var.set("ko")

        btn_row = tk.Frame(btn_container, bg="#181A20")
        btn_row.pack(anchor="e", padx=20, pady=8)
        ttk.Button(btn_row, text=self._t("defaults","DEFAULT"), command=restore_env_defaults, style="EnvDefault.TButton").pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(btn_row, text=self._t("save","SAVE"), command=save_env, style="EnvSave.TButton").pack(side=tk.LEFT)

    def _build_trade_tab(self, frame):
        frame.columnconfigure(0, weight=1)
        _is_ko = (self.language == "ko")

        container = tk.Frame(frame, bg="#181A20")
        container.pack(fill=tk.BOTH, expand=True)

        # Button bar pinned at the bottom BEFORE canvas so it always stays visible
        trade_btn_container = tk.Frame(container, bg="#181A20")
        trade_btn_container.pack(side=tk.BOTTOM, fill="x")
        tk.Frame(trade_btn_container, bg="#343942", height=1).pack(fill="x")
        trade_btn_row = tk.Frame(trade_btn_container, bg="#181A20")
        trade_btn_row.pack(anchor="e", padx=20, pady=12)

        canvas = tk.Canvas(container, bg="#181A20", highlightthickness=0, bd=0)
        scrollbar = ttk.Scrollbar(container, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=scrollbar.set)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        body = tk.Frame(canvas, bg="#181A20")
        canvas.create_window((0, 0), window=body, anchor="nw")

        def _update_scroll_region(_event=None):
            canvas.configure(scrollregion=canvas.bbox("all"))

        body.bind("<Configure>", _update_scroll_region)

        def _on_mousewheel(event):
            canvas.yview_scroll(int(-1 * (event.delta / 120)), "units")

        def _bind_mousewheel(_event=None):
            canvas.bind_all("<MouseWheel>", _on_mousewheel)

        def _unbind_mousewheel(_event=None):
            canvas.unbind_all("<MouseWheel>")

        body.bind("<Enter>", _bind_mousewheel)
        body.bind("<Leave>", _unbind_mousewheel)

        title_row = tk.Frame(body, bg="#181A20")
        title_row.pack(fill="x", padx=40, pady=(30, 12))
        tk.Label(title_row, text=self._t("trade_settings_title","거래 설정"), bg="#181A20", fg="white", font=("Malgun Gothic", 13, "bold"), anchor="w").pack(side=tk.LEFT)
        _reset_btn = ttk.Button(
            title_row,
            text=self._t("auto_tune_reset", "Reset Auto-Tune"),
            command=self._reset_auto_tune_state,
            style="EnvSave.TButton",
        )
        _reset_btn.pack(side=tk.RIGHT)
        _ks_reset_btn = ttk.Button(
            title_row,
            text="Kill Switch 초기화" if self.language == "ko" else "Reset Kill Switch",
            command=self._reset_kill_switch,
            style="EnvSave.TButton",
        )
        _ks_reset_btn.pack(side=tk.RIGHT, padx=(0, 8))
        _tooltip_text = self._t("auto_tune_reset_tooltip", "Auto-Tune 학습 데이터를 초기화합니다. 엔진 재시작 후 기본값으로 재적용됩니다.")
        _tooltip = tk.Label(
            body,
            text=_tooltip_text,
            bg="#2b3350",
            fg="#d0d8f8",
            font=("Malgun Gothic", 9),
            padx=8,
            pady=4,
            relief="flat",
            wraplength=340,
            justify="left",
        )
        def _show_tooltip(event):
            x = event.widget.winfo_rootx() - body.winfo_rootx()
            y = event.widget.winfo_rooty() - body.winfo_rooty() + event.widget.winfo_height() + 4
            _tooltip.place(x=x, y=y)
            _tooltip.lift()
        def _hide_tooltip(_event=None):
            _tooltip.place_forget()
        _reset_btn.bind("<Enter>", _show_tooltip)
        _reset_btn.bind("<Leave>", _hide_tooltip)

        auto_tune_var = tk.BooleanVar(value=bool(self.settings_data.get("auto_tune_enabled", True)))
        auto_tune_mode_var = tk.StringVar(value=self.settings_data.get("auto_tune_mode", "balanced"))
        spike_guard_var = tk.BooleanVar(value=bool(self.settings_data.get("spike_guard_enabled", True)))
        self._trade_controls_locked = auto_tune_var.get()
        trade_lock_widgets = []
        lock_msg_label = None
        self.auto_tune_status_labels = {}
        self.auto_tune_last_update_label = None

        custom_toggle_widgets = []

        def register_custom_toggle(widget):
            custom_toggle_widgets.append(widget)
            return widget

        def register_control(widget):
            trade_lock_widgets.append(widget)
            return widget

        def set_trade_controls_locked(lock: bool):
            self._trade_controls_locked = lock
            if lock:
                self.settings_data["auto_boost_position_pct"] = bool(self.settings_data.get("auto_boost_position_pct", False))
            for widget in trade_lock_widgets:
                if isinstance(widget, tk.Entry):
                    widget.configure(state=tk.DISABLED if lock else tk.NORMAL)
                elif isinstance(widget, ttk.Checkbutton):
                    if lock:
                        widget.state(["disabled"])
                    else:
                        widget.state(["!disabled"])
                elif isinstance(widget, ttk.Button):
                    if lock:
                        widget.state(["disabled"])
                    else:
                        widget.state(["!disabled"])
            for toggle in custom_toggle_widgets:
                indicator, label = toggle
                color = "#555c76" if lock else "#f5f7ff"
                label.configure(fg=color)
                indicator.configure(cursor="arrow" if lock else "hand2")
            if lock_msg_label is not None:
                if lock:
                    lock_msg_label.configure(text=self._t("auto_tune_locked","Auto-tune is ON; manual controls are locked."), fg="#F0B90B")
                else:
                    lock_msg_label.configure(text=self._t("auto_tune_unlocked","Manual controls are unlocked."), fg="#2EBD85")

        def on_auto_tune_toggle():
            lock = auto_tune_var.get()
            self.settings_data["auto_tune_enabled"] = lock
            self._save_json(CONFIG_PATH, self.settings_data)
            set_trade_controls_locked(lock)
            update_auto_tune_visual()
            if lock:
                self._refresh_auto_tune_state()
            self._prompt_restart_after_setting_change(self._t("auto_tune_restart_prompt","자동 튜닝 토글 변경으로 엔진을 다시 시작할까요?"))

        auto_tune_panel = tk.Frame(body, bg="#1f2128", highlightbackground="#2b313d", highlightthickness=1)
        auto_tune_panel.pack(fill="x", padx=40, pady=(0, 18))

        panel_header = tk.Frame(auto_tune_panel, bg="#1f2128")
        panel_header.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(panel_header, text="🤖 AUTO-TUNE", bg="#1f2128", fg="#c0c6dc", font=("Malgun Gothic", 11, "bold")).pack(side=tk.LEFT)

        mode_switch = tk.Frame(panel_header, bg="#131722", highlightbackground="#2b313d", highlightthickness=1)
        mode_switch.pack(side=tk.LEFT, padx=(10, 0))
        for col_idx in range(3):
            mode_switch.grid_columnconfigure(col_idx, weight=1)
        mode_labels = {}
        mode_options = [(self._t("mode_aggressive","Aggressive"), "aggressive"), (self._t("mode_balanced","Balanced"), "balanced"), (self._t("mode_conservative","Conservative"), "conservative")]

        def update_mode_slider():
            current = auto_tune_mode_var.get()
            for value, widget in mode_labels.items():
                if value == current:
                    widget.configure(bg="#2EBD85", fg="#0f1a14", relief="sunken", bd=1)
                else:
                    widget.configure(bg="#131722", fg="#8f96b7", relief="flat", bd=1)


        def change_mode(value, *, user=True):
            if value not in {opt[1] for opt in mode_options}:
                return
            if auto_tune_mode_var.get() == value and user:
                return
            # D: Aggressive — expert mode guard + 2차 확인
            if value == "aggressive" and user:
                if not bool(self.settings_data.get("expert_mode_enabled", False)):
                    tip = self._t("aggressive_locked_tip",
                                  (self._t("aggressive_expert_warn",
                                   "Aggressive 모드는 Expert 모드 활성화 시 사용 가능합니다 (필수 동의 탭).")
                                  if self.language == "ko" else
                                  "Aggressive mode requires Expert Mode to be enabled (Agreement tab)."))
                    self._show_warning(self._t("aggressive_confirm_title", "Aggressive 모드"), tip)
                    return
                msg = self._t(
                    "aggressive_confirm_msg",
                    "Aggressive 모드는 leverage_max 및 position_pct를 상향하고 필터를 완화합니다.\n"
                    "변동성이 높은 시장에서 원금 대비 대폭 손실이 발생할 수 있습니다.\n\n"
                    "계속하시겠습니까?"
                )
                title = self._t("aggressive_confirm_title", "Aggressive 모드 — 고위험 확인")
                if not self._show_yesno(title, msg):
                    return
            auto_tune_mode_var.set(value)
            self.settings_data["auto_tune_mode"] = value
            self._save_json(CONFIG_PATH, self.settings_data)
            update_mode_slider()
            try:
                _update_mode_desc()
            except Exception:
                pass
            update_auto_tune_visual()
            if user and self.engine_running:
                self._append_log(f"[INFO] Auto-tune mode changed to {value}; engine restart required")

        for col, (label_text, value) in enumerate(mode_options):
            seg = tk.Label(
                mode_switch,
                text=label_text,
                bg="#131722",
                fg="#8f96b7",
                padx=10,
                pady=3,
                font=("Malgun Gothic", 9, "bold"),
                relief="flat",
                bd=1,
                cursor="hand2",
            )
            seg.grid(row=0, column=col, sticky="nsew")
            seg.bind("<Button-1>", lambda _e, v=value: change_mode(v))
            mode_labels[value] = seg

        status_badge = tk.Label(panel_header, text="", bg="#2EBD85", fg="#0f1a14", font=("Malgun Gothic", 9, "bold"), padx=10, pady=2)
        status_badge.pack(side=tk.RIGHT)
        update_mode_slider()

        # ── 모드별 설명 라벨 ─────────────────────────────────────────
        _mode_desc_map = {
            "aggressive": ("높은 레버리지 · 넓은 진입 조건 · 고수익 고위험" if _is_ko
                           else "High leverage · Loose entry · High risk/reward"),
            "balanced":   ("중간 레버리지 · 표준 필터 · 균형 잡힌 전략" if _is_ko
                           else "Moderate leverage · Standard filters · Balanced strategy"),
            "conservative": ("낮은 레버리지 · 엄격한 필터 · 안정적 수익" if _is_ko
                             else "Low leverage · Strict filters · Stable returns"),
        }
        mode_desc_label = tk.Label(
            auto_tune_panel, text=_mode_desc_map.get(auto_tune_mode_var.get(), ""),
            bg="#1f2128", fg="#8893b8", font=("Malgun Gothic", 9),
            anchor="w", wraplength=500, justify="left")
        mode_desc_label.pack(fill="x", padx=18, pady=(0, 4))

        def _update_mode_desc(*_):
            mode_desc_label.configure(text=_mode_desc_map.get(auto_tune_mode_var.get(), ""))

        toggle_row = tk.Frame(auto_tune_panel, bg="#1f2128")
        toggle_row.pack(fill="x", padx=18, pady=(0, 4))
        toggle_indicator = tk.Label(toggle_row, bg="#1f2128")
        toggle_indicator.pack(side=tk.LEFT, padx=(0, 10))
        toggle_label = tk.Label(toggle_row, text=self._t("auto_tune_toggle_label","Enable auto-tune"), bg="#1f2128", fg="#f5f7ff", font=("Malgun Gothic", 11, "bold"))
        toggle_label.pack(side=tk.LEFT)

        helper_label = tk.Label(
            auto_tune_panel,
            text=self._t("auto_tune_desc","The live tuner adjusts position filters and momentum parameters."),
            bg="#1f2128",
            fg="#9aa5c6",
            font=("Malgun Gothic", 9),
            anchor="w",
            justify="left",
        )
        helper_label.pack(fill="x", padx=18, pady=(0, 2))

        # ── 오토튜너 기능 설명 카드 ──────────────────────────────────────
        _at_info_frame = tk.Frame(auto_tune_panel, bg="#171b26")
        _at_info_frame.pack(fill="x", padx=18, pady=(0, 8))

        _is_ko = (self.language == "ko")
        _at_features = [
            ("📊", "모멘텀 & 변동성 자동 조절" if _is_ko else "Auto-adjust momentum & volatility",
             "시장 상태에 따라 진입 모멘텀/변동성 임계값을 실시간 조정합니다." if _is_ko
             else "Dynamically tunes entry momentum/volatility thresholds based on market conditions."),
            ("⚖️", "포지션 크기 & 레버리지 범위" if _is_ko else "Position size & leverage range",
             "승률과 변동성에 따라 포지션 비중과 레버리지 min/max를 자동 조정합니다." if _is_ko
             else "Adjusts position % and leverage min/max based on win rate and volatility."),
            ("🛡️", "손절폭 & 쿨다운" if _is_ko else "Stop-loss & cooldown",
             "연속 손실 시 손절 한도를 조이고, 쿨다운 주기로 과매매를 방지합니다." if _is_ko
             else "Tightens stop-loss after consecutive losses and enforces cooldown to prevent overtrading."),
            ("🔄", "Shadow → Apply 검증" if _is_ko else "Shadow → Apply validation",
             "새 파라미터를 Shadow 테스트 후 성과가 좋을 때만 실제 적용합니다." if _is_ko
             else "Shadow-tests new parameters and only applies them when they prove profitable."),
        ]
        for _icon, _title, _desc in _at_features:
            _feat_row = tk.Frame(_at_info_frame, bg="#171b26")
            _feat_row.pack(fill="x", padx=8, pady=2)
            tk.Label(_feat_row, text=_icon, bg="#171b26", fg="#F0B90B",
                     font=("Malgun Gothic", 10)).pack(side=tk.LEFT, padx=(0, 6))
            _feat_text = tk.Frame(_feat_row, bg="#171b26")
            _feat_text.pack(side=tk.LEFT, fill="x", expand=True)
            tk.Label(_feat_text, text=_title, bg="#171b26", fg="#d0d8f8",
                     font=("Malgun Gothic", 9, "bold"), anchor="w").pack(fill="x")
            tk.Label(_feat_text, text=_desc, bg="#171b26", fg="#6c7490",
                     font=("Malgun Gothic", 8), anchor="w", wraplength=480,
                     justify="left").pack(fill="x")

        lock_msg_label = tk.Label(auto_tune_panel, text="", bg="#1f2128", fg="#2EBD85", font=("Malgun Gothic", 10, "bold"))
        lock_msg_label.pack(fill="x", padx=18, pady=(4, 8))

        def build_image_toggle(parent, text, variable, on_change=None):
            row = tk.Frame(parent, bg="#151b24")
            row.pack(fill="x", padx=18, pady=(0, 6))
            indicator = tk.Label(row, bg="#151b24")
            indicator.pack(side=tk.LEFT, padx=(0, 8))
            lbl = tk.Label(row, text=text, bg="#151b24", fg="#f5f7ff", font=("Malgun Gothic", 10, "bold"))
            lbl.pack(side=tk.LEFT)

            def update_indicator(*_):
                if variable.get() and self.checkbox_images.get("on"):
                    indicator.configure(image=self.checkbox_images["on"], text="")
                    indicator.image = self.checkbox_images["on"]
                elif not variable.get() and self.checkbox_images.get("off"):
                    indicator.configure(image=self.checkbox_images["off"], text="")
                    indicator.image = self.checkbox_images["off"]
                else:
                    indicator.configure(text="●" if variable.get() else "○", fg="#2EBD85" if variable.get() else "#f5f7ff")
                    indicator.image = None

            def toggle(_event=None):
                if getattr(self, "_trade_controls_locked", False):
                    return
                variable.set(not variable.get())
                update_indicator()
                if on_change:
                    on_change(variable.get())

            for widget in (row, indicator, lbl):
                widget.bind("<Button-1>", toggle)
            update_indicator()
            register_custom_toggle((indicator, lbl))
            return row

        # ── 진입 차단 경고 배너 (momentum 높아 거래 안 될 때) ───────────────
        self._trade_block_banner = tk.Label(
            auto_tune_panel,
            text="",
            bg="#3a1a00",
            fg="#ffaa44",
            font=("Malgun Gothic", 9, "bold"),
            anchor="w",
            padx=14,
            pady=6,
            wraplength=520,
            justify="left",
        )
        # 배너는 필요할 때만 pack
        self._trade_block_banner_visible = False

        status_panel = tk.Frame(auto_tune_panel, bg="#151b24", highlightbackground="#2b313d", highlightthickness=1)
        status_panel.pack(fill="x", padx=18, pady=(0, 12))
        status_panel.columnconfigure(0, weight=1)
        status_panel.columnconfigure(1, weight=1)

        def add_status_metric(row, column, title, key, columnspan=1):
            cell = tk.Frame(status_panel, bg="#151b24")
            cell.grid(row=row, column=column, columnspan=columnspan, sticky="nsew", padx=10, pady=8)
            tk.Label(cell, text=title, bg="#151b24", fg="#8893b8", font=("Malgun Gothic", 9, "bold")).pack(anchor="w")
            value_label = tk.Label(cell, text="--", bg="#151b24", fg="#f5f7ff", font=("Malgun Gothic", 11, "bold"))
            value_label.pack(anchor="w", pady=(2, 0))
            self.auto_tune_status_labels[key] = value_label

        # ── 행 0: 핵심 진입 파라미터 (차단 원인 최다) ──────────────────────
        add_status_metric(0, 0, "모멘텀 롱" if self.language=="ko" else "Mom Long", "momentum_long")
        add_status_metric(0, 1, "모멘텀 숏" if self.language=="ko" else "Mom Short", "momentum_short")
        # ── 행 1: 변동성 + 레짐 ──────────────────────────────────────────
        add_status_metric(1, 0, "변동성 최소" if self.language=="ko" else "Volatility Min", "volatility_min")
        add_status_metric(1, 1, "시장 레짐" if self.language=="ko" else "Regime", "regime")
        # ── 행 2: 포지션 크기 + 레버리지 ────────────────────────────────
        add_status_metric(2, 0, "포지션 크기" if self.language=="ko" else "Position %", "position_pct")
        add_status_metric(2, 1, "레버리지" if self.language=="ko" else "Leverage", "leverage_range")
        # ── 행 3: 손절 + 쿨다운 ─────────────────────────────────────────
        add_status_metric(3, 0, "손절 한도" if self.language=="ko" else "Stop Loss", "stop_loss")
        add_status_metric(3, 1, "튜닝 쿨다운" if self.language=="ko" else "Cooldown", "cooldown")
        # ── 행 4: 신뢰도 + 통과율 ───────────────────────────────────────
        add_status_metric(4, 0, "신뢰도 | 노이즈" if self.language=="ko" else "Conf | Noise", "confidence")
        add_status_metric(4, 1, "통과 | 진입률" if self.language=="ko" else "Pass | Entry%", "pass_entry")
        # ── 행 5: 30분 PnL ───────────────────────────────────────────────
        add_status_metric(5, 0, "최근 30분 손익" if self.language=="ko" else "30m Net PnL", "pnl_30m", columnspan=2)
        # ── 행 6: 활성 스냅샷 ────────────────────────────────────────────
        add_status_metric(6, 0, "적용중 파라미터" if self.language=="ko" else "Active Params", "active_snapshot", columnspan=2)

        self.auto_tune_last_update_label = tk.Label(
            auto_tune_panel,
            text=self._t("auto_tune_update","Updated: -"),
            bg="#1f2128",
            fg="#6c738a",
            font=("Malgun Gothic", 9, "italic"),
            anchor="w",
        )
        self.auto_tune_last_update_label.pack(fill="x", padx=18, pady=(0, 8))



        def update_auto_tune_visual():
            lock = auto_tune_var.get()
            if lock and self.checkbox_images.get("on"):
                toggle_indicator.configure(image=self.checkbox_images["on"], text="")
                toggle_indicator.image = self.checkbox_images["on"]
            elif (not lock) and self.checkbox_images.get("off"):
                toggle_indicator.configure(image=self.checkbox_images["off"], text="")
                toggle_indicator.image = self.checkbox_images["off"]
            else:
                toggle_indicator.configure(text="●" if lock else "○", fg="#2EBD85" if lock else "#f5f7ff")
                toggle_indicator.image = None

            mode_text_map = {"aggressive": "AGG", "balanced": "MID", "conservative": "DEF"}
            mode_display = mode_text_map.get(auto_tune_mode_var.get(), "MID")
            status_badge.configure(
                text=f"{mode_display} · {'ON' if lock else 'OFF'}",
                bg="#2EBD85" if lock else "#3f4659",
                fg="#0f1a14" if lock else "#c0c6dc",
            )
            helper_label.configure(
                text=(self._t("auto_tune_helper_on","Auto-tune is enabled and adjusts parameters using recent market data.") if lock else self._t("auto_tune_helper_off","Parameters are fixed to manual values. Adjust and save if needed.")),
                fg="#2EBD85" if lock else "#9aa5c6",
            )

        def toggle_auto_tune(_event=None):
            auto_tune_var.set(not auto_tune_var.get())
            on_auto_tune_toggle()

        for widget in (toggle_row, toggle_indicator, toggle_label, status_badge):
            widget.bind("<Button-1>", toggle_auto_tune)

        update_auto_tune_visual()
        if not self._auto_tune_refresh_started:
            self._auto_tune_refresh_started = True
            self._schedule_auto_tune_state_refresh()
        else:
            self._refresh_auto_tune_state()

        boost_var = tk.BooleanVar(value=bool(self.settings_data.get("auto_boost_position_pct", False)))

        def toggle_boost(_event=None):
            boost_var.set(not boost_var.get())
            self._set_auto_boost_position_pct(boost_var.get())
            update_boost_indicator()

        boost_toggle = tk.Frame(auto_tune_panel, bg="#151b24", highlightbackground="#2b313d", highlightthickness=1)
        boost_toggle.pack(fill="x", padx=18, pady=(0, 12))
        boost_indicator = tk.Label(boost_toggle, bg="#151b24")
        boost_indicator.pack(side=tk.LEFT, padx=(10, 6), pady=6)
        boost_text = tk.Label(
            boost_toggle,
            text=("포지션 % 자동 부스트 (증거금 부족 시 진입 크기 자동 증가)" if _is_ko
                  else "Auto-boost position % (increase entry size when margin is insufficient)"),
            bg="#151b24",
            fg="#f5f7ff",
            font=("Malgun Gothic", 10, "bold"),
        )
        boost_text.pack(side=tk.LEFT, pady=6)

        def update_boost_indicator():
            if boost_var.get() and self.checkbox_images.get("on"):
                boost_indicator.configure(image=self.checkbox_images["on"], text="")
                boost_indicator.image = self.checkbox_images["on"]
            elif (not boost_var.get()) and self.checkbox_images.get("off"):
                boost_indicator.configure(image=self.checkbox_images["off"], text="")
                boost_indicator.image = self.checkbox_images["off"]
            else:
                boost_indicator.configure(text="●" if boost_var.get() else "○", fg="#2EBD85" if boost_var.get() else "#f5f7ff")
                boost_indicator.image = None
            self.settings_data["auto_boost_position_pct"] = bool(boost_var.get())
            self._sync_position_pct_status()

        for widget in (boost_toggle, boost_indicator, boost_text):
            widget.bind("<Button-1>", toggle_boost)
        update_boost_indicator()

        manual_panel = tk.Frame(body, bg="#121620", highlightbackground="#2b313d", highlightthickness=1)
        manual_panel.pack(fill="x", padx=40, pady=(0, 24))
        manual_header = tk.Frame(manual_panel, bg="#121620")
        manual_header.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(manual_header, text="⚙️ " + ("수동 파라미터 (직접 조정)" if _is_ko else "Manual Parameters"), bg="#121620", fg="white", font=("Malgun Gothic", 11, "bold")).pack(side=tk.LEFT)
        tk.Label(
            manual_header,
            text="  " + ("Auto-Tune OFF 시 직접 편집 가능" if _is_ko else "Editable when Auto-Tune is OFF"),
            bg="#121620",
            fg="#7d86a8",
            font=("Malgun Gothic", 9),
        ).pack(side=tk.LEFT, padx=(4, 0))
        manual_body = tk.Frame(manual_panel, bg="#121620")
        manual_body.pack(fill="x", padx=18, pady=(0, 16))

        def _make_dark_entry(parent, variable, row, column, width=10, padx=(0, 0), pady=(4, 0), sticky="w", columnspan=1):
            wrap = tk.Frame(parent, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
            wrap.grid(row=row, column=column, padx=padx, pady=pady, sticky=sticky, columnspan=columnspan)
            entry = tk.Entry(
                wrap,
                textvariable=variable,
                width=width,
                bg="#0b111a",
                fg="#f5f7ff",
                relief=tk.FLAT,
                insertbackground="#ffffff",
                borderwidth=0,
                disabledbackground="#0b111a",
                disabledforeground="#4a5270",
            )
            entry.pack(fill="x", padx=6, pady=4)
            register_control(entry)
            return entry

        profit_panel = tk.Frame(body, bg="#10131d", highlightbackground="#2b313d", highlightthickness=1)
        profit_panel.pack(fill="x", padx=40, pady=(0, 24))
        tk.Label(profit_panel, text="💹 " + ("수익 실현 & 추적 손절" if _is_ko else "Profit Exit & Trailing Stop"),
                 bg="#10131d", fg="white", font=("Malgun Gothic", 11, "bold"), anchor="w").pack(fill="x", padx=18, pady=(18, 4))

        # ── 수익 전략 요약 카드 ──────────────────────────────────────
        _profit_info = tk.Frame(profit_panel, bg="#151924")
        _profit_info.pack(fill="x", padx=18, pady=(0, 8))
        _profit_features = [
            ("🎯", "분할 익절 (Partial TP)" if _is_ko else "Partial Take-Profit",
             "R배수 단계별로 포지션을 나눠서 익절합니다 (기본: 1.2R → 30%, 2.0R → 40%, 3.0R → 100%)." if _is_ko
             else "Closes position in stages at R-multiple levels (default: 1.2R→30%, 2.0R→40%, 3.0R→100%)."),
            ("📈", "ATR 추적 손절" if _is_ko else "ATR Trailing Stop",
             "수익 구간 진입 후 ATR 기반 트레일링 스탑으로 이익을 보호합니다." if _is_ko
             else "Protects profit with ATR-based trailing stop once in profit zone."),
            ("⏳", "진행 정체 스탑" if _is_ko else "Progress Stop",
             "일정 시간 신고점 없이 하락 시 자동 청산하여 수익 반납을 방지합니다." if _is_ko
             else "Auto-exits if price stalls with no new high, preventing profit give-back."),
        ]
        for _icon, _title, _desc in _profit_features:
            _pf_row = tk.Frame(_profit_info, bg="#151924")
            _pf_row.pack(fill="x", padx=8, pady=2)
            tk.Label(_pf_row, text=_icon, bg="#151924", fg="#F0B90B",
                     font=("Malgun Gothic", 10)).pack(side=tk.LEFT, padx=(0, 6))
            _pf_text = tk.Frame(_pf_row, bg="#151924")
            _pf_text.pack(side=tk.LEFT, fill="x", expand=True)
            tk.Label(_pf_text, text=_title, bg="#151924", fg="#d0d8f8",
                     font=("Malgun Gothic", 9, "bold"), anchor="w").pack(fill="x")
            tk.Label(_pf_text, text=_desc, bg="#151924", fg="#6c7490",
                     font=("Malgun Gothic", 8), anchor="w", wraplength=480,
                     justify="left").pack(fill="x")
        pe_var = tk.BooleanVar(value=bool(self.settings_data.get("enable_profit_exit_layer", True)))
        partial_tp_var = tk.BooleanVar(value=bool(self.settings_data.get("enable_partial_take_profit", True)))
        trail_tp_var = tk.BooleanVar(value=bool(self.settings_data.get("enable_atr_trailing_stop", True)))
        progress_var = tk.BooleanVar(value=bool(self.settings_data.get("enable_progress_stop", True)))
        build_image_toggle(profit_panel, self._t("profit_exit_layer_toggle","Enable Profit Exit Layer"), pe_var)
        tk.Label(profit_panel, text=self._t("profit_layer_hint"), bg="#10131d", fg="#7f8ab3", font=("Malgun Gothic", 9), anchor="w", wraplength=560, justify="left").pack(fill="x", padx=36, pady=(0, 6))
        build_image_toggle(profit_panel, self._t("partial_tp_label"), partial_tp_var)
        tk.Label(profit_panel, text=self._t("partial_toggle_hint"), bg="#10131d", fg="#7f8ab3", font=("Malgun Gothic", 9), anchor="w", wraplength=560, justify="left").pack(fill="x", padx=36, pady=(0, 6))
        build_image_toggle(profit_panel, self._t("atr_trailing_label"), trail_tp_var)
        tk.Label(profit_panel, text=self._t("trail_toggle_hint"), bg="#10131d", fg="#7f8ab3", font=("Malgun Gothic", 9), anchor="w", wraplength=560, justify="left").pack(fill="x", padx=36, pady=(0, 6))
        build_image_toggle(profit_panel, self._t("progress_stop_label"), progress_var)
        tk.Label(profit_panel, text=self._t("progress_toggle_hint"), bg="#10131d", fg="#7f8ab3", font=("Malgun Gothic", 9), anchor="w", wraplength=560, justify="left").pack(fill="x", padx=36, pady=(0, 10))

        partial_frame = tk.Frame(profit_panel, bg="#10131d")
        partial_frame.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(partial_frame, text=self._t("partial_levels_title", "Partial Levels"), bg="#10131d", fg="#c0c6dc", font=("Malgun Gothic", 9), anchor="w", wraplength=600).pack(fill="x")
        tk.Label(partial_frame, text=self._t("partial_levels_desc"), bg="#10131d", fg="#6f7da4", font=("Malgun Gothic", 9), anchor="w", wraplength=600).pack(fill="x", pady=(0, 4))
        partial_levels_var = tk.StringVar(value=self._format_partial_levels_for_display(self.settings_data.get("partial_tp_levels", [])))
        partial_entry_wrap = tk.Frame(partial_frame, bg="#0c101a", highlightbackground="#27334a", highlightthickness=1)
        partial_entry_wrap.pack(fill="x")
        partial_entry = tk.Entry(partial_entry_wrap, textvariable=partial_levels_var, bg="#0c101a", fg="#f5f7ff", insertbackground="#f5f7ff", relief=tk.FLAT, borderwidth=0, disabledbackground="#0c101a", disabledforeground="#4a5270")
        partial_entry.pack(fill="x", padx=8, pady=6)
        register_control(partial_entry)

        atr_frame = tk.Frame(profit_panel, bg="#10131d")
        atr_frame.pack(fill="x", padx=18, pady=(12, 6))
        tk.Label(atr_frame, text=self._t("atr_trail_title"), bg="#10131d", fg="#c8d1f2", font=("Malgun Gothic", 10, "bold"), anchor="w").grid(row=0, column=0, columnspan=5, sticky="w")
        tk.Label(atr_frame, text=self._t("atr_desc"), bg="#10131d", fg="#6e7da0", font=("Malgun Gothic", 9), anchor="w", wraplength=600, justify="left").grid(row=1, column=0, columnspan=5, sticky="w", pady=(4, 4))
        tk.Label(atr_frame, text=self._t("atr_period"), bg="#10131d", fg="#8e96b8").grid(row=2, column=0, padx=(0, 8), sticky="w")
        tk.Label(atr_frame, text=self._t("atr_multiplier"), bg="#10131d", fg="#8e96b8").grid(row=2, column=1, padx=(0, 8), sticky="w")
        tk.Label(atr_frame, text=self._t("atr_activate_pct"), bg="#10131d", fg="#8e96b8").grid(row=2, column=2, padx=(0, 8), sticky="w")
        tk.Label(atr_frame, text=self._t("recalc_sec","Recalc (s)"), bg="#10131d", fg="#8e96b8").grid(row=2, column=3, padx=(0, 8), sticky="w")
        trail_period_var = tk.IntVar(value=int(self.settings_data.get("trail_atr_period", 22)))
        trail_mult_var = tk.DoubleVar(value=float(self.settings_data.get("trail_atr_mult", 3.0)))
        trail_activate_var = tk.DoubleVar(value=float(self.settings_data.get("trail_activate_pnl_pct", 0.20) * 100 if abs(self.settings_data.get("trail_activate_pnl_pct", 0.20)) <= 1 else self.settings_data.get("trail_activate_pnl_pct", 0.20)))
        trail_interval_var = tk.IntVar(value=int(self.settings_data.get("trail_recalc_interval_sec", 5)))
        _make_dark_entry(atr_frame, trail_period_var, row=3, column=0, width=10)
        _make_dark_entry(atr_frame, trail_mult_var, row=3, column=1, width=10)
        _make_dark_entry(atr_frame, trail_activate_var, row=3, column=2, width=10)
        _make_dark_entry(atr_frame, trail_interval_var, row=3, column=3, width=10)

        prog_frame = tk.Frame(profit_panel, bg="#10131d")
        prog_frame.pack(fill="x", padx=18, pady=(12, 16))
        tk.Label(prog_frame, text=self._t("progress_stop_label"), bg="#10131d", fg="#c8d1f2", font=("Malgun Gothic", 10, "bold"), anchor="w").grid(row=0, column=0, columnspan=5, sticky="w")
        tk.Label(prog_frame, text=self._t("progress_desc"), bg="#10131d", fg="#6e7da0", font=("Malgun Gothic", 9), anchor="w", wraplength=600, justify="left").grid(row=1, column=0, columnspan=5, sticky="w", pady=(4, 4))
        tk.Label(prog_frame, text=self._t("progress_lookback"), bg="#10131d", fg="#8e96b8").grid(row=2, column=0, sticky="w")
        tk.Label(prog_frame, text=self._t("progress_no_new_high"), bg="#10131d", fg="#8e96b8").grid(row=2, column=1, sticky="w")
        tk.Label(prog_frame, text=self._t("progress_drawdown"), bg="#10131d", fg="#8e96b8").grid(row=2, column=2, sticky="w")
        tk.Label(prog_frame, text=self._t("progress_min_pnl"), bg="#10131d", fg="#8e96b8").grid(row=2, column=3, sticky="w")
        prog_look_var = tk.IntVar(value=int(self.settings_data.get("progress_stop_lookback_sec", 1800)))
        prog_nohigh_var = tk.IntVar(value=int(self.settings_data.get("progress_stop_no_new_high_sec", 1800)))
        prog_drawdown_var = tk.DoubleVar(value=float(self.settings_data.get("progress_stop_drawdown_from_mfe", 0.15) * 100))
        prog_minpnl_var = tk.DoubleVar(value=float(self.settings_data.get("progress_stop_min_pnl_pct", 0.25) * 100))
        prog_action_var = tk.StringVar(value=str(self.settings_data.get("progress_stop_action", "partial_or_full")))
        _make_dark_entry(prog_frame, prog_look_var, row=3, column=0, width=10)
        _make_dark_entry(prog_frame, prog_nohigh_var, row=3, column=1, width=10, padx=(8, 0))
        _make_dark_entry(prog_frame, prog_drawdown_var, row=3, column=2, width=10, padx=(8, 0))
        _make_dark_entry(prog_frame, prog_minpnl_var, row=3, column=3, width=10, padx=(8, 0))
        _make_dark_entry(prog_frame, prog_action_var, row=4, column=0, width=16, columnspan=2, pady=(8, 0))

        base_var = tk.DoubleVar(value=float(self.settings_data.get("position_base_pct", 0.055)))
        base_pct_text = tk.StringVar()
        base_hint_text = tk.StringVar()
        base_hint_label = None

        def _update_base_pct_label(*_):
            _lbl = "진입 크기" if self.language == "ko" else "Entry size"
            _acct = "계좌의" if self.language == "ko" else "account"
            base_pct_text.set(f"{_lbl} ({_acct} {base_var.get() * 100:.1f}%)")
            self._sync_position_pct_status(base_var.get())

        base_var.trace_add("write", _update_base_pct_label)
        _update_base_pct_label()

        def build_round_slider(parent, variable, min_val=0.01, max_val=0.99, length=400, state_callback=None, disabled_check=None):
            canvas_slider = tk.Canvas(parent, width=length, height=40, bg="#181A20", highlightthickness=0, bd=0)
            track_y = 22
            margin = 14
            slider_width = max(1, length - margin * 2)
            canvas_slider.create_line(margin, track_y, length - margin, track_y, fill="#3f4659", width=4, capstyle=tk.ROUND)
            knob_radius = 8
            knob = canvas_slider.create_oval(0, 0, knob_radius * 2, knob_radius * 2, fill="#2EBD85", outline="#2EBD85", width=2)
            trace_handle = {"id": None}

            def cleanup_trace():
                if trace_handle["id"] is not None:
                    try:
                        variable.trace_remove("write", trace_handle["id"])
                    except tk.TclError:
                        pass
                    trace_handle["id"] = None

            def value_to_x(val):
                ratio = (val - min_val) / (max_val - min_val)
                return margin + ratio * slider_width

            def clamp_value(val):
                try:
                    val = float(val)
                except (TypeError, ValueError, tk.TclError):
                    val = min_val
                return max(min_val, min(max_val, val))

            def update_knob(*_):
                if not canvas_slider.winfo_exists():
                    cleanup_trace()
                    return
                val = clamp_value(variable.get())
                x = value_to_x(val)
                canvas_slider.coords(knob, x - knob_radius, track_y - knob_radius, x + knob_radius, track_y + knob_radius)
                knob_color = "#2EBD85"
                if state_callback:
                    knob_color = state_callback(val) or knob_color
                canvas_slider.itemconfigure(knob, outline=knob_color, fill=knob_color)

            def update_from_event(event):
                if disabled_check and disabled_check():
                    return
                ratio = (event.x - margin) / slider_width
                ratio = max(0.0, min(1.0, ratio))
                value = min_val + ratio * (max_val - min_val)
                value = round(value / 0.01) * 0.01
                variable.set(round(value, 4))

            def on_destroy(_event):
                cleanup_trace()

            canvas_slider.bind("<Button-1>", update_from_event)
            canvas_slider.bind("<B1-Motion>", update_from_event)
            canvas_slider.bind("<Destroy>", on_destroy)
            trace_handle["id"] = variable.trace_add("write", lambda *_: update_knob())
            update_knob()
            return canvas_slider

        def update_guidance(val):
            if val < 0.03:
                message = ("3% 미만: 포지션이 작아 수익 실현이 어려울 수 있습니다." if _is_ko
                           else "Below 3%: position size may be too small to realize gains meaningfully.")
                knob_color = "#F7C948"
                label_color = "#F7C948"
            elif val > 0.06:
                message = ("6% 초과: 손실이 확대될 위험이 있습니다. 레버리지와 손절 설정을 확인하세요." if _is_ko
                           else "Above 6%: risk of amplified losses. Re-check leverage and stop-loss settings.")
                knob_color = "#FF5F5F"
                label_color = "#FF5F5F"
            else:
                message = ("권장 3~6%: 횡보장은 ~3%, 확신이 높을 때만 ~5%. 손절 설정 필수." if _is_ko
                           else "Recommended 3-6%: stay near 3% in choppy markets; use ~5% only with strong conviction.")
                knob_color = "#2EBD85"
                label_color = "#c0c6dc"
            base_hint_text.set(message)
            if base_hint_label is not None:
                base_hint_label.configure(fg=label_color)
            return knob_color

        tk.Label(
            manual_body,
            textvariable=base_pct_text,
            bg="#121620",
            fg="#c0c6dc",
            font=("Malgun Gothic", 10, "bold"),
            anchor="w",
        ).pack(fill="x")
        build_round_slider(
            manual_body,
            base_var,
            state_callback=update_guidance,
            disabled_check=lambda: getattr(self, "_trade_controls_locked", False),
        ).pack(fill="x", pady=6)
        base_hint_label = tk.Label(manual_body, textvariable=base_hint_text, bg="#121620", fg="#c0c6dc", font=("Malgun Gothic", 10), anchor="w")
        base_hint_label.pack(fill="x", pady=(0, 12))
        update_guidance(base_var.get())

        lev_limit = 125 if self.env_mode == "TESTNET" else 150
        lev_min = tk.IntVar(value=min(lev_limit, max(1, int(self.settings_data.get("leverage_min", 5)))))
        lev_max = tk.IntVar(value=min(lev_limit, max(lev_min.get(), int(self.settings_data.get("leverage_max", 25)))))
        lev_label_text = tk.StringVar()
        lev_hint_text = tk.StringVar()

        def update_lev_label(*_):
            _unit = "배" if self.language == "ko" else "x"
            lev_label_text.set(f"Leverage min/max (max {lev_limit}{_unit}): {lev_min.get()}x ~ {lev_max.get()}x")

        def update_lev_hint(*_):
            if _is_ko:
                lev_hint_text.set(
                    f"동적 레버리지: 신호 강도 × AI 확률 × 변동성으로 min~max 범위 내 자동 결정 "
                    f"(환경: {'테스트넷 최대 125x' if self.env_mode == 'TESTNET' else '실거래 최대 150x'})")
            else:
                lev_hint_text.set(
                    f"Dynamic leverage: auto-calculated from signal strength × AI probability × volatility "
                    f"within min~max range ({'TESTNET cap 125x' if self.env_mode == 'TESTNET' else 'LIVE up to 150x'})")

        update_lev_label()
        update_lev_hint()

        def build_leverage_slider(parent, min_var, max_var, min_val=1, max_val=150, length=400, disabled_check=None):
            canvas_slider = tk.Canvas(parent, width=length, height=52, bg="#181A20", highlightthickness=0, bd=0)
            track_y = 30
            margin = 16
            slider_width = max(1, length - margin * 2)
            canvas_slider.create_line(margin, track_y, length - margin, track_y, fill="#3f4659", width=4, capstyle=tk.ROUND)
            knob_radius = 8
            knob_min = canvas_slider.create_oval(0, 0, knob_radius * 2, knob_radius * 2, fill="#2EBD85", outline="#2EBD85", width=2)
            knob_max = canvas_slider.create_oval(0, 0, knob_radius * 2, knob_radius * 2, fill="#2EBD85", outline="#2EBD85", width=2)
            active = {"target": None}
            trace_handles = {"min": None, "max": None}

            def cleanup_traces():
                for key, var in (("min", min_var), ("max", max_var)):
                    handle = trace_handles[key]
                    if handle is None:
                        continue
                    try:
                        var.trace_remove("write", handle)
                    except tk.TclError:
                        pass
                    trace_handles[key] = None

            def value_to_x(val):
                ratio = (val - min_val) / (max_val - min_val)
                return margin + ratio * slider_width

            def clamp_values():
                min_value = max(min_val, min(max_val, min_var.get()))
                max_value = max(min_value, min(max_val, max_var.get()))
                if min_value != min_var.get():
                    min_var.set(int(min_value))
                if max_value != max_var.get():
                    max_var.set(int(max_value))
                return int(min_value), int(max_value)

            def update_knobs(*_):
                if not canvas_slider.winfo_exists():
                    cleanup_traces()
                    return
                min_value, max_value = clamp_values()
                x_min = value_to_x(min_value)
                x_max = value_to_x(max_value)
                canvas_slider.coords(knob_min, x_min - knob_radius, track_y - knob_radius, x_min + knob_radius, track_y + knob_radius)
                canvas_slider.coords(knob_max, x_max - knob_radius, track_y - knob_radius, x_max + knob_radius, track_y + knob_radius)

            def set_value_from_event(event):
                if disabled_check and disabled_check():
                    return
                ratio = (event.x - margin) / slider_width
                ratio = max(0.0, min(1.0, ratio))
                value = int(round(min_val + ratio * (max_val - min_val)))
                if active["target"] == "max":
                    max_var.set(max(min_var.get(), value))
                else:
                    min_var.set(min(value, max_var.get()))
                update_knobs()
                update_lev_label()

            def pick_target(event):
                x_min = value_to_x(min_var.get())
                x_max = value_to_x(max_var.get())
                active["target"] = "min" if abs(event.x - x_min) <= abs(event.x - x_max) else "max"

            def on_click(event):
                pick_target(event)
                set_value_from_event(event)

            def on_drag(event):
                set_value_from_event(event)

            def on_destroy(_event):
                cleanup_traces()

            canvas_slider.bind("<Button-1>", on_click)
            canvas_slider.bind("<B1-Motion>", on_drag)
            canvas_slider.bind("<Destroy>", on_destroy)

            def on_min_change(*_):
                update_knobs()
                update_lev_label()

            def on_max_change(*_):
                update_knobs()
                update_lev_label()

            trace_handles["min"] = min_var.trace_add("write", on_min_change)
            trace_handles["max"] = max_var.trace_add("write", on_max_change)
            update_knobs()
            return canvas_slider

        tk.Label(manual_body, textvariable=lev_label_text, bg="#121620", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", pady=(6, 6))
        build_leverage_slider(
            manual_body,
            lev_min,
            lev_max,
            min_val=1,
            max_val=lev_limit,
            disabled_check=lambda: getattr(self, "_trade_controls_locked", False),
        ).pack(fill="x", pady=(0, 6))
        tk.Label(manual_body, textvariable=lev_hint_text, bg="#121620", fg="#8e96b8", font=("Malgun Gothic", 10), anchor="w").pack(fill="x", pady=(0, 12))

        panel_bg = "#11141c"
        panel_border = "#2b313d"
        field_bg = "#0c111b"

        def build_field(panel, label_text, variable, width=12, suffix=None, bottom_pad=0):
            row = tk.Frame(panel, bg=panel_bg)
            row.pack(fill="x", padx=18, pady=(6, bottom_pad))
            tk.Label(row, text=label_text, bg=panel_bg, fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(anchor="w")
            wrapper = tk.Frame(row, bg=field_bg, highlightbackground="#2d3342", highlightcolor="#2d3342", highlightthickness=1)
            wrapper.pack(fill="x", pady=(4, 0))
            trace_id = None
            entry = tk.Entry(
                wrapper,
                textvariable=variable,
                width=width,
                bg=field_bg,
                fg="#f5f7ff",
                insertbackground="#f5f7ff",
                relief=tk.FLAT,
                borderwidth=0,
                highlightthickness=0,
                disabledbackground=field_bg,
                disabledforeground="#5f6682",
            )
            entry.pack(side=tk.LEFT, fill="x", expand=True, padx=8, pady=6)
            if suffix:
                tk.Label(wrapper, text=suffix, bg=field_bg, fg="#6b7394", font=("Malgun Gothic", 10, "bold")).pack(side=tk.RIGHT, padx=(0, 10))
            register_control(entry)

            def cleanup_trace(event=None):
                nonlocal trace_id
                if trace_id is not None:
                    try:
                        variable.trace_remove("write", trace_id)
                    except Exception:
                        pass
                    trace_id = None

            entry.bind("<Destroy>", cleanup_trace)
            return entry

        trade_style = ttk.Style()
        trade_style.configure(
            "Trade.TCheckbutton",
            background="#181A20",
            foreground="#f5f7ff",
            font=("Malgun Gothic", 10, "bold"),
        )
        trade_style.map(
            "Trade.TCheckbutton",
            foreground=[("disabled", "#5f6682")],
            background=[("active", "#1f2128"), ("selected", "#181A20")],
        )

        _sf_row = tk.Frame(manual_body, bg="#121620")
        _sf_row.pack(fill="x", pady=(12, 4))
        tk.Label(_sf_row, text="📊 " + ("심볼 & 임계값 설정" if _is_ko else "Symbol & Threshold Settings"),
                 bg="#121620", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(side=tk.LEFT)
        tk.Label(_sf_row, text="  — " + ("거래대금 상위 심볼 수, 감시 한도, 동시 진입 수" if _is_ko else "top symbols, watch limit, max concurrent"),
                 bg="#121620", fg="#555f7a", font=("Malgun Gothic", 9)).pack(side=tk.LEFT, pady=2)
        filter_panel = tk.Frame(manual_body, bg=panel_bg, highlightbackground=panel_border, highlightthickness=1)
        filter_panel.pack(fill="x", pady=(0, 16))

        top_n_var = tk.IntVar(value=int(self.settings_data.get("top_n", 20)))
        watch_limit_ui_var = tk.IntVar(value=int(self.settings_data.get("watch_limit", 20)))
        max_open_ui_var = tk.IntVar(value=int(self.settings_data.get("max_open_symbols", 5)))
        vol_var = tk.DoubleVar(value=float(self.settings_data.get("volatility_min", 0.002)))
        mom_long_var = tk.DoubleVar(value=float(self.settings_data.get("momentum_min_long", 0.002)))
        mom_short_var = tk.DoubleVar(value=float(self.settings_data.get("momentum_min_short", -0.002)))

        # ── 심볼 감시/진입 수 ────────────────────────────────────────────
        _sym_grid = tk.Frame(filter_panel, bg=panel_bg)
        _sym_grid.pack(fill="x", padx=18, pady=(8, 0))
        _sym_grid.columnconfigure(0, weight=1)
        _sym_grid.columnconfigure(1, weight=1)
        _sym_grid.columnconfigure(2, weight=1)

        def _sym_field(parent, lbl, variable, col, hint=""):
            cell = tk.Frame(parent, bg=panel_bg)
            cell.grid(row=0, column=col, sticky="nsew", padx=(0, 12))
            _lrow = tk.Frame(cell, bg=panel_bg)
            _lrow.pack(fill="x")
            tk.Label(_lrow, text=lbl, bg=panel_bg, fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(side=tk.LEFT)
            if hint:
                tk.Label(_lrow, text=hint, bg=panel_bg, fg="#555f7a", font=("Malgun Gothic", 8)).pack(side=tk.LEFT, padx=(6,0))
            wrap = tk.Frame(cell, bg="#0c111b", highlightbackground="#2d3342", highlightthickness=1)
            wrap.pack(fill="x", pady=(4, 0))
            e = tk.Entry(wrap, textvariable=variable, width=8, bg="#0c111b", fg="#f5f7ff",
                         insertbackground="#f5f7ff", relief=tk.FLAT, borderwidth=0,
                         disabledbackground="#0c111b", disabledforeground="#5f6682")
            e.pack(side=tk.LEFT, fill="x", expand=True, padx=8, pady=6)
            register_control(e)
            return e

        _top_n_hint = "거래대금 상위" if self.language=="ko" else "top by volume"
        _wl_hint = "감시 심볼 수" if self.language=="ko" else "symbols monitored"
        _mo_hint = "동시 진입 한도" if self.language=="ko" else "max concurrent"
        _sym_field(_sym_grid, "Top-N", top_n_var, 0, _top_n_hint)
        _sym_field(_sym_grid, "워치리스트" if self.language=="ko" else "Watch Limit", watch_limit_ui_var, 1, _wl_hint)
        _sym_field(_sym_grid, "동시 포지션" if self.language=="ko" else "Max Open", max_open_ui_var, 2, _mo_hint)

        # ── 진입 임계값 ──────────────────────────────────────────────────
        tk.Frame(filter_panel, bg="#2b313d", height=1).pack(fill="x", padx=18, pady=(10, 0))
        _thresh_lbl = ("📉 진입 임계값 (Auto-Tune ON 시 실시간 자동 조정)" if _is_ko
                       else "📉 Entry Thresholds (auto-adjusted in real-time when Auto-Tune ON)")
        tk.Label(filter_panel, text=_thresh_lbl, bg=panel_bg, fg="#6c7490", font=("Malgun Gothic", 8), anchor="w").pack(fill="x", padx=18, pady=(4,0))
        build_field(filter_panel, self._t("field_vol_min"), vol_var)
        build_field(filter_panel, self._t("field_mom_long"), mom_long_var)
        build_field(filter_panel, self._t("field_mom_short"), mom_short_var, bottom_pad=10)

        tk.Label(manual_body, text="🛡️ " + ("리스크 & 종료 설정" if _is_ko else "Risk & Exit Settings"),
                 bg="#121620", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", pady=(0, 2))
        tk.Label(manual_body, text=("손절 한도, 쿨다운, 최소 증거금 등 리스크 관련 설정" if _is_ko
                                    else "Stop-loss limits, cooldown period, minimum margin settings"),
                 bg="#121620", fg="#6c7490", font=("Malgun Gothic", 8), anchor="w").pack(fill="x", pady=(0, 6))
        limits_panel = tk.Frame(manual_body, bg=panel_bg, highlightbackground=panel_border, highlightthickness=1)
        limits_panel.pack(fill="x", pady=(0, 16))

        # watch_limit_var / max_open_var → watch_limit_ui_var / max_open_ui_var 로 통합 (위에서 선언)
        watch_limit_var = watch_limit_ui_var   # 내부 호환
        max_open_var = max_open_ui_var         # 내부 호환
        auto_tune_cooldown_var = tk.IntVar(value=int(self.settings_data.get("auto_tune_cooldown_min", 10)))
        max_loss_var = tk.DoubleVar(value=float(self.settings_data.get("max_loss_per_position", 18.0)))

        # watch_limit / max_open_symbols: auto-tune 조정 대상 제외됨 → 필드 숨김
        build_field(limits_panel, self._t("field_cooldown"), auto_tune_cooldown_var, width=10)
        build_field(limits_panel, self._t("field_stop_loss_pnl"), max_loss_var, width=10, suffix="%", bottom_pad=10)
        min_margin_var = tk.DoubleVar(value=float(self.settings_data.get("min_margin_usdt", 1.0)))
        build_field(limits_panel, self._t("field_min_margin_usdt", "최소 증거금 (USDT)"), min_margin_var, width=10, suffix=" USDT", bottom_pad=10)

        def handle_spike_guard_toggle(value):
            self.settings_data["spike_guard_enabled"] = value
            self._save_json(CONFIG_PATH, self.settings_data)
            self._prompt_restart_after_setting_change(self._t("spike_guard_restart", "Spike guard setting changed. Restart the engine now?"))

        build_image_toggle(limits_panel,
                           ("급등락 방어 (Spike Guard)" if _is_ko else "Spike Guard"), spike_guard_var, on_change=handle_spike_guard_toggle)
        tk.Label(
            limits_panel,
            text=("급격한 가격 변동(스파이크) 감지 시 진입을 일시 차단하여 슬리피지 손실을 방지합니다." if _is_ko
                  else "Temporarily blocks entry when sudden price spikes are detected to prevent slippage losses."),
            bg="#11141c",
            fg="#7f86a6",
            font=("Malgun Gothic", 9),
            anchor="w",
            wraplength=540,
        ).pack(fill="x", padx=18, pady=(0, 6))

        # ─── 진입 필터 설정 패널 ────────────────────────────────────────────────
        tk.Label(manual_body, text="🔍 " + ("진입 필터 설정" if _is_ko else "Entry Filter Settings"),
                 bg="#121620", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", pady=(12, 2))
        tk.Label(manual_body,
                 text=("진입 신호를 걸러내는 필터들을 ON/OFF합니다. 필터가 많을수록 보수적, 적을수록 공격적." if _is_ko
                       else "Toggle filters that gate entry signals. More filters = conservative, fewer = aggressive."),
                 bg="#121620", fg="#6c7490", font=("Malgun Gothic", 8), anchor="w", wraplength=540).pack(fill="x", pady=(0, 6))
        filter_adv_panel = tk.Frame(manual_body, bg=panel_bg, highlightbackground=panel_border, highlightthickness=1)
        filter_adv_panel.pack(fill="x", pady=(0, 16))

        # BooleanVars
        rsi_filter_var = tk.BooleanVar(value=bool(self.settings_data.get("rsi_filter_enabled", False)))
        composite_var = tk.BooleanVar(value=bool(self.settings_data.get("composite_signal_enabled", True)))
        breakeven_var = tk.BooleanVar(value=bool(self.settings_data.get("breakeven_stop_enabled", True)))
        kelly_var = tk.BooleanVar(value=bool(self.settings_data.get("kelly_sizing_enabled", True)))
        atr_risk_var = tk.BooleanVar(value=bool(self.settings_data.get("atr_risk_sizing_enabled", True)))
        funding_filter_var = tk.BooleanVar(value=bool(self.settings_data.get("funding_filter_enabled", True)))
        maker_first_var = tk.BooleanVar(value=bool(self.settings_data.get("maker_first_enabled", True)))
        diversify_var = tk.BooleanVar(value=bool(self.settings_data.get("diversify_watchlist", False)))
        mtf_confirm_var = tk.BooleanVar(value=bool(self.settings_data.get("enable_mtf_ema_confirm", True)))
        short_ema_conflict_var = tk.BooleanVar(value=bool(self.settings_data.get("short_ema_conflict_filter", True)))
        chop_ema_dir_var = tk.BooleanVar(value=bool(self.settings_data.get("chop_use_short_ema_direction", True)))

        # 숫자 변수
        rsi_ob_var = tk.DoubleVar(value=float(self.settings_data.get("rsi_overbought", 75.0)))
        rsi_os_var = tk.DoubleVar(value=float(self.settings_data.get("rsi_oversold", 25.0)))
        composite_score_var = tk.DoubleVar(value=float(self.settings_data.get("composite_min_score", 0.9)))
        entry_risk_var = tk.DoubleVar(value=float(self.settings_data.get("entry_risk_pct", 0.01) * 100))
        min_hold_var = tk.IntVar(value=int(self.settings_data.get("min_hold_seconds", 45)))
        time_stop_var = tk.IntVar(value=int(self.settings_data.get("time_stop_seconds", 1800)))

        # ── 방향 필터 ──────────────────────────────────────────────────────
        tk.Label(filter_adv_panel, text="📌 " + ("방향 검증 필터" if _is_ko else "Direction Filters"),
                 bg=panel_bg, fg="#9bb5e8", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=18, pady=(10, 4))
        build_image_toggle(filter_adv_panel,
                           ("MTF EMA 하드게이트" if _is_ko else "MTF EMA Hard Gate"), mtf_confirm_var)
        tk.Label(filter_adv_panel,
                 text=("1m·5m EMA 추세가 진입 방향과 반대이면 강제 차단합니다." if _is_ko
                       else "Blocks entry when 1m/5m EMA trend opposes entry direction."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,4))
        build_image_toggle(filter_adv_panel,
                           ("단기 EMA 충돌 필터" if _is_ko else "Short EMA Conflict Filter"), short_ema_conflict_var)
        tk.Label(filter_adv_panel,
                 text=("24h 방향과 단기 EMA(1m·5m) 방향이 모두 반대일 때 진입을 차단합니다." if _is_ko
                       else "Blocks entry when both 1m & 5m EMA direction conflict with 24h direction."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,4))
        build_image_toggle(filter_adv_panel,
                           ("CHOP 레짐: 단기 EMA 우선" if _is_ko else "CHOP Regime: Short EMA Priority"), chop_ema_dir_var)
        tk.Label(filter_adv_panel,
                 text=("횡보(CHOP) 구간에서 24h 기준 대신 1m·5m EMA 방향으로 진입합니다." if _is_ko
                       else "In CHOP regime, uses 1m/5m EMA direction instead of 24h direction for entry."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,8))

        tk.Frame(filter_adv_panel, bg="#2b313d", height=1).pack(fill="x", padx=18, pady=(0, 8))

        # ── 신호 품질 필터 ──────────────────────────────────────────────────────
        tk.Label(filter_adv_panel, text="📊 " + ("신호 품질 필터" if _is_ko else "Signal Quality Filters"),
                 bg=panel_bg, fg="#9bb5e8", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=18, pady=(0, 4))
        build_image_toggle(filter_adv_panel,
                           ("복합 신호 스코어링" if _is_ko else "Composite Signal Scoring"), composite_var)
        tk.Label(filter_adv_panel,
                 text=("모멘텀·거래량·MTF 가중 합산으로 신호 품질을 점수화합니다." if _is_ko
                       else "Scores signal quality by weighted sum of momentum, volume, and MTF."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        _cs_row = tk.Frame(filter_adv_panel, bg=panel_bg)
        _cs_row.pack(fill="x", padx=36, pady=(0,4))
        tk.Label(_cs_row, text=("최소 스코어:" if _is_ko else "Min score:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _cs_entry_wrap = tk.Frame(_cs_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _cs_entry_wrap.pack(side=tk.LEFT, padx=(6,0))
        tk.Entry(_cs_entry_wrap, textvariable=composite_score_var, width=6, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)

        build_image_toggle(filter_adv_panel,
                           ("RSI 과열 필터" if _is_ko else "RSI Extremes Filter"), rsi_filter_var)
        tk.Label(filter_adv_panel,
                 text=("RSI가 과매수 구간이면 LONG 차단, 과매도 구간이면 SHORT 차단합니다." if _is_ko
                       else "Blocks LONG at overbought RSI and SHORT at oversold RSI."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        _rsi_row = tk.Frame(filter_adv_panel, bg=panel_bg)
        _rsi_row.pack(fill="x", padx=36, pady=(0,8))
        tk.Label(_rsi_row, text=self._t("rsi_ob_label","과매수 차단:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _ob_w = tk.Frame(_rsi_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _ob_w.pack(side=tk.LEFT, padx=(6,12))
        tk.Entry(_ob_w, textvariable=rsi_ob_var, width=5, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)
        tk.Label(_rsi_row, text=self._t("rsi_os_label","과매도 차단:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _os_w = tk.Frame(_rsi_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _os_w.pack(side=tk.LEFT, padx=(6,0))
        tk.Entry(_os_w, textvariable=rsi_os_var, width=5, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)

        tk.Frame(filter_adv_panel, bg="#2b313d", height=1).pack(fill="x", padx=18, pady=(0, 8))

        # ── 포지션 사이징 ────────────────────────────────────────────────────────
        tk.Label(filter_adv_panel, text="💰 " + ("포지션 사이징" if _is_ko else "Position Sizing"),
                 bg=panel_bg, fg="#9bb5e8", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=18, pady=(0, 4))
        build_image_toggle(filter_adv_panel,
                           ("ATR 리스크 사이징" if _is_ko else "ATR Risk Sizing"), atr_risk_var)
        tk.Label(filter_adv_panel,
                 text=("ATR 기반으로 포지션당 최대 손실 한도를 계산해 수량을 자동 제한합니다." if _is_ko
                       else "Caps position size based on ATR to limit max loss per trade."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        _ar_row = tk.Frame(filter_adv_panel, bg=panel_bg)
        _ar_row.pack(fill="x", padx=36, pady=(0,4))
        tk.Label(_ar_row, text=self._t("entry_risk_label","포지션당 리스크:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _ar_w = tk.Frame(_ar_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _ar_w.pack(side=tk.LEFT, padx=(6,0))
        tk.Entry(_ar_w, textvariable=entry_risk_var, width=5, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)
        tk.Label(_ar_row, text="%", bg=panel_bg, fg="#6b7394", font=("Malgun Gothic", 9)).pack(side=tk.LEFT, padx=(2,0))

        build_image_toggle(filter_adv_panel,
                           ("Kelly 사이징" if _is_ko else "Kelly Sizing"), kelly_var)
        tk.Label(filter_adv_panel,
                 text=("과거 승률과 손익비를 기반으로 최적 포지션 비율을 자동 산출합니다." if _is_ko
                       else "Calculates optimal position fraction based on historical win rate and payoff ratio."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,4))

        tk.Frame(filter_adv_panel, bg="#2b313d", height=1).pack(fill="x", padx=18, pady=(0, 8))

        # ── 실행 설정 ────────────────────────────────────────────────────────────
        tk.Label(filter_adv_panel, text="⚙️ " + ("실행 설정" if _is_ko else "Execution Settings"),
                 bg=panel_bg, fg="#9bb5e8", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=18, pady=(0, 4))
        build_image_toggle(filter_adv_panel,
                           ("Breakeven Stop" if _is_ko else "Breakeven Stop"), breakeven_var)
        tk.Label(filter_adv_panel,
                 text=("첫 번째 TP 달성 후 손절가를 진입가(손익분기)로 자동 이동합니다." if _is_ko
                       else "Auto-moves stop-loss to breakeven after first take-profit hit."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        build_image_toggle(filter_adv_panel,
                           ("펀딩레이트 편향 필터" if _is_ko else "Funding Rate Bias Filter"), funding_filter_var)
        tk.Label(filter_adv_panel,
                 text=("펀딩레이트가 과도하게 한 방향으로 쏠릴 때 진입 신호 강도를 낮춥니다." if _is_ko
                       else "Reduces signal strength when funding rate shows extreme directional bias."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        build_image_toggle(filter_adv_panel,
                           ("Maker-First 진입" if _is_ko else "Maker-First Entry"), maker_first_var)
        tk.Label(filter_adv_panel,
                 text=("지정가(메이커) 주문을 먼저 시도하고, 체결 실패 시 시장가로 전환합니다. 수수료 절약." if _is_ko
                       else "Tries limit (maker) order first, falls back to market if unfilled. Saves on fees."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,2))
        build_image_toggle(filter_adv_panel,
                           ("워치리스트 다양성" if _is_ko else "Diversify Watchlist"), diversify_var)
        tk.Label(filter_adv_panel,
                 text=("변동성이 유사한 종목 편중을 피하고, 다양한 변동성의 심볼을 선택합니다." if _is_ko
                       else "Avoids clustering similar-volatility symbols; selects diversified watchlist."),
                 bg=panel_bg, fg="#60698a", font=("Malgun Gothic", 9), anchor="w", wraplength=540).pack(fill="x", padx=36, pady=(0,4))

        _time_row = tk.Frame(filter_adv_panel, bg=panel_bg)
        _time_row.pack(fill="x", padx=18, pady=(6, 10))
        tk.Label(_time_row, text=self._t("min_hold_label","최소 보유 시간:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _mh_w = tk.Frame(_time_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _mh_w.pack(side=tk.LEFT, padx=(6,0))
        tk.Entry(_mh_w, textvariable=min_hold_var, width=5, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)
        tk.Label(_time_row, text=self._t("seconds_suffix","초"), bg=panel_bg, fg="#6b7394", font=("Malgun Gothic", 9)).pack(side=tk.LEFT, padx=(2,16))
        tk.Label(_time_row, text=self._t("time_stop_label","최대 보유 시간:"), bg=panel_bg, fg="#8e96b8", font=("Malgun Gothic", 9)).pack(side=tk.LEFT)
        _ts_w = tk.Frame(_time_row, bg="#0b111a", highlightbackground="#2d3548", highlightthickness=1)
        _ts_w.pack(side=tk.LEFT, padx=(6,0))
        tk.Entry(_ts_w, textvariable=time_stop_var, width=5, bg="#0b111a", fg="#e8edff", insertbackground="white", relief="flat", font=("Malgun Gothic", 10)).pack(padx=4, pady=2)
        tk.Label(_time_row, text=self._t("seconds_suffix","초"), bg=panel_bg, fg="#6b7394", font=("Malgun Gothic", 9)).pack(side=tk.LEFT, padx=(2,0))

        self.trade_field_vars = {
            "position_pct": base_var,
            "leverage_min": lev_min,
            "leverage_max": lev_max,
            "top_n": top_n_var,
            "watch_limit": watch_limit_ui_var,
            "max_open_symbols": max_open_ui_var,
            "volatility_min": vol_var,
            "momentum_min_long": mom_long_var,
            "momentum_min_short": mom_short_var,
            "auto_tune_cooldown_min": auto_tune_cooldown_var,
            "max_loss_per_position": max_loss_var,
            "min_margin_usdt": min_margin_var,
        }

        def save_trade():
            mom_long = float(mom_long_var.get())
            try:
                mom_short_value = float(mom_short_var.get())
            except (TypeError, ValueError):
                mom_short_value = -0.002
            mom_short = -abs(mom_short_value)
            stop_loss_pct = max(0.1, float(max_loss_var.get()))
            self.settings_data.update(
                {
                    "position_base_pct": round(base_var.get(), 4),
                    "position_pct": round(base_var.get(), 4),
                    "leverage_min": max(1, lev_min.get()),
                    "leverage_max": max(lev_min.get(), lev_max.get()),
                    "top_n": max(1, int(top_n_var.get())),
                    "volatility_min": max(0.0, float(vol_var.get())),
                    "momentum_min_long": mom_long,
                    "momentum_min_short": mom_short,
                    "auto_tune_enabled": auto_tune_var.get(),
                    "auto_tune_mode": auto_tune_mode_var.get(),
                    "watch_limit": max(3, int(watch_limit_ui_var.get() or 20)),
                    "max_open_symbols": max(1, int(max_open_ui_var.get() or 5)),
                    "auto_tune_cooldown_min": max(1, int(auto_tune_cooldown_var.get() or 1)),
                    "max_loss_per_position": stop_loss_pct,
                    "min_margin_usdt": max(0.01, float(min_margin_var.get())),
                    "spike_guard_enabled": spike_guard_var.get(),
                    "auto_boost_position_pct": bool(self.settings_data.get("auto_boost_position_pct", False)),
                }
            )
            self.settings_data.pop("momentum_min", None)
            self.settings_data.pop("atr_min", None)
            self.settings_data.pop("show_command_buttons", None)

            try:
                partial_levels = self._parse_partial_levels_field(partial_levels_var.get())
            except ValueError as exc:
                self._show_warning("Partial Take Profit", str(exc))
                return
            self.settings_data["enable_profit_exit_layer"] = bool(pe_var.get())
            self.settings_data["enable_partial_take_profit"] = bool(partial_tp_var.get())
            self.settings_data["enable_atr_trailing_stop"] = bool(trail_tp_var.get())
            self.settings_data["enable_progress_stop"] = bool(progress_var.get())
            self.settings_data["partial_tp_levels"] = partial_levels
            self.settings_data["trail_atr_period"] = max(1, int(trail_period_var.get() or 22))
            self.settings_data["trail_atr_mult"] = max(0.1, float(trail_mult_var.get() or 3.0))
            self.settings_data["trail_activate_pnl_pct"] = self._normalize_ratio_input(trail_activate_var.get())
            self.settings_data["trail_recalc_interval_sec"] = max(1, int(trail_interval_var.get() or 5))
            self.settings_data["progress_stop_lookback_sec"] = max(60, int(prog_look_var.get() or 1800))
            self.settings_data["progress_stop_no_new_high_sec"] = max(60, int(prog_nohigh_var.get() or 1800))
            self.settings_data["progress_stop_drawdown_from_mfe"] = self._normalize_ratio_input(prog_drawdown_var.get(), clamp_zero_one=True)
            self.settings_data["progress_stop_min_pnl_pct"] = self._normalize_ratio_input(prog_minpnl_var.get())
            self.settings_data["progress_stop_action"] = prog_action_var.get().strip() or "partial_or_full"

            # ── 진입 필터 설정 저장 ───────────────────────────────────────────────
            self.settings_data["rsi_filter_enabled"] = bool(rsi_filter_var.get())
            self.settings_data["rsi_overbought"] = max(50.0, float(rsi_ob_var.get() or 75.0))
            self.settings_data["rsi_oversold"] = min(50.0, float(rsi_os_var.get() or 25.0))
            self.settings_data["composite_signal_enabled"] = bool(composite_var.get())
            self.settings_data["composite_min_score"] = max(0.1, float(composite_score_var.get() or 0.9))
            self.settings_data["breakeven_stop_enabled"] = bool(breakeven_var.get())
            self.settings_data["kelly_sizing_enabled"] = bool(kelly_var.get())
            self.settings_data["atr_risk_sizing_enabled"] = bool(atr_risk_var.get())
            self.settings_data["entry_risk_pct"] = max(0.001, min(0.1, float(entry_risk_var.get() or 0.01)))
            self.settings_data["funding_filter_enabled"] = bool(funding_filter_var.get())
            self.settings_data["maker_first_enabled"] = bool(maker_first_var.get())
            self.settings_data["diversify_watchlist"] = bool(diversify_var.get())
            # ── 방향 필터 (단기 EMA) ─────────────────────────────────────────────
            self.settings_data["enable_mtf_ema_confirm"] = bool(mtf_confirm_var.get())
            self.settings_data["short_ema_conflict_filter"] = bool(short_ema_conflict_var.get())
            self.settings_data["chop_use_short_ema_direction"] = bool(chop_ema_dir_var.get())
            # ── 시간 보호 ────────────────────────────────────────────────────────
            self.settings_data["min_hold_seconds"] = max(0, int(min_hold_var.get() or 45))
            self.settings_data["time_stop_seconds"] = max(0, int(time_stop_var.get() or 1800))

            self._save_json(CONFIG_PATH, self.settings_data)
            self._show_info(self._t("trade_settings_save_title","Saved"), self._t("trade_settings_saved","Trading settings saved."))
            self._update_positions_exit_summary()
            self._prompt_restart_after_setting_change(self._t("trade_settings_restart","Trading settings changed. Restart the engine now?"))


        def restore_trade_defaults():
            base_var.set(0.02)
            lev_min.set(1)
            lev_max.set(8)
            top_n_var.set(20)
            vol_var.set(0.002)
            mom_long_var.set(0.002)
            mom_short_var.set(-0.002)
            auto_tune_var.set(True)
            on_auto_tune_toggle()
            watch_limit_ui_var.set(20)
            max_open_ui_var.set(5)
            auto_tune_cooldown_var.set(10)
            max_loss_var.set(18.0)
            min_hold_var.set(300)
            time_stop_var.set(2700)
            change_mode("balanced", user=False)

        default_btn = ttk.Button(trade_btn_row, text="DEFAULT", command=restore_trade_defaults, style="EnvDefault.TButton")
        default_btn.pack(side=tk.LEFT, padx=(0, 12))
        save_btn = ttk.Button(trade_btn_row, text="SAVE", command=save_trade, style="EnvSave.TButton")
        save_btn.pack(side=tk.LEFT)

        set_trade_controls_locked(auto_tune_var.get())

    def _build_display_tab(self, frame):
        frame.columnconfigure(0, weight=1)
        # 타이틀 행 (오른쪽 끝에 세션 초기화 버튼)
        _title_row = tk.Frame(frame, bg="#181A20")
        _title_row.pack(fill="x", padx=40, pady=(30, 12))
        tk.Label(_title_row, text=self._t("display_title","화면 / 수동 매매 설정"),
                 bg="#181A20", fg="white", font=("Malgun Gothic", 13, "bold"),
                 anchor="w").pack(side=tk.LEFT)

        def _reset_session_state_inline():
            parent_w = getattr(self, "_active_modal", self.root)
            _en = self.language == "en"
            answer = tk.messagebox.askyesno(
                "Reset Session State" if _en else "세션 상태 초기화",
                ("Reset GUI session state.\n\n"
                 "  • Reset session start time\n"
                 "  • Reset win-rate / trade counters\n"
                 "  • Reset PnL card base time\n\n"
                 "Settings, API keys, and trade history are NOT affected.\nContinue?"
                 if _en else
                 "GUI 세션 상태를 초기화합니다.\n\n"
                 "  • 세션 시작 시각 초기화\n"
                 "  • 승률·거래수 카운터 초기화\n"
                 "  • 손익 카드 기준 시각 초기화\n\n"
                 "설정값·API 키·거래 기록은 영향 없습니다.\n계속하시겠습니까?"),
                parent=parent_w,
            )
            if not answer:
                return
            now = int(time.time() * 1000)
            _keep = {"card_modes"}
            preserved = {k: v for k, v in self.state_data.items() if k in _keep}
            self.state_data.clear()
            self.state_data.update(preserved)
            self._save_json(STATE_PATH, self.state_data)
            self.session_start_ms = None
            self.pnl_reset_ms     = now
            self.stat_resets["win_rate"]    = now
            self.stat_resets["trade_count"] = now
            self._set_stat_value("trade_count", "0")
            self._set_stat_value("win_rate",    "0%")
            pnl_keys = {"pnl_15", "pnl_60", "pnl_12h", "pnl_24h"}
            active = set(getattr(self, "card_modes", {}).values())
            for k in pnl_keys:
                if k in active:
                    self._set_stat_value(k, "+0 USDT")
            self._append_log("[DISPLAY] " + ("GUI session state reset" if _en else "GUI 세션 상태 초기화 완료"))
            tk.messagebox.showinfo(
                "Done" if _en else "완료",
                "Session state has been reset." if _en else "세션 상태가 초기화되었습니다.",
                parent=parent_w)

        tk.Button(_title_row,
                  text="🔄  Reset Session" if self.language == "en" else "🔄  세션 초기화",
                  command=_reset_session_state_inline,
                  bg="#1a2030", fg="#a0b4d0", relief=tk.FLAT,
                  font=("Segoe UI", 9), padx=10, pady=3,
                  cursor="hand2",
                  activebackground="#263040", activeforeground="#c0d4f0"
                  ).pack(side=tk.RIGHT)

        # ── 기본 카드 패널 설정 ──────────────────────────────────────────────
        _CARD_OPTIONS = [
            ("unrealized_total", self._t("card_opt_unrealized",  "미실현 손익 합계")),
            ("filter_pass_rate", self._t("card_opt_filter",      "필터 통과율")),
            ("top_symbol",       self._t("card_opt_top_symbol",  "최다 거래 심볼")),
            ("pnl_15",           self._t("card_opt_pnl15",       "15분 손익")),
            ("pnl_60",           self._t("card_opt_pnl60",       "60분 손익")),
            ("pnl_12h",          self._t("card_opt_pnl12h",      "12시간 손익")),
            ("pnl_24h",          self._t("card_opt_pnl24h",      "24시간 손익")),
            ("rr_ratio",         self._t("card_opt_rr",          "손익비 R:R")),
            ("expectancy",       self._t("card_opt_exp",         "거래당 기댓값")),
            ("max_consec_loss",  self._t("card_opt_consec",      "최대 연속 손실")),
        ]
        _card_section = tk.Frame(frame, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        _card_section.pack(fill="x", padx=40, pady=(0, 16))
        tk.Label(_card_section, text=self._t("display_card_defaults","기본 카드 패널 설정"),
                 bg="#181A20", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"),
                 anchor="w").pack(fill="x", padx=16, pady=(14, 6))
        tk.Label(_card_section, text=self._t("display_card_defaults_desc","메인 화면 상단 4개 카드의 기본 표시 항목을 설정합니다."),
                 bg="#181A20", fg="#9aa5c6", font=("Malgun Gothic", 9),
                 anchor="w").pack(fill="x", padx=16, pady=(0, 8))

        # saved default card modes (separate from runtime card_modes)
        _default_card_modes = self.settings_data.get("default_card_modes", {
            "card0": "unrealized_total", "card1": "filter_pass_rate",
            "card2": "top_symbol",       "card3": "pnl_24h",
        })
        # _card_vars: cid → StringVar(값=key)
        _card_vars = {}
        # key → label 매핑
        _key_to_label = {k: v for k, v in _CARD_OPTIONS}

        _grid_row = tk.Frame(_card_section, bg="#181A20")
        _grid_row.pack(fill="x", padx=16, pady=(0, 14))
        for _ci in range(4):
            _cid = f"card{_ci}"
            _col = tk.Frame(_grid_row, bg="#181A20")
            _col.pack(side=tk.LEFT, padx=(0, 16))
            tk.Label(_col, text=self._t(f"card_label_{_ci}", f"카드 {_ci+1}"),
                     bg="#181A20", fg="#9aa5c6",
                     font=("Malgun Gothic", 9, "bold")).pack(anchor="w", pady=(0, 4))
            _init_key = _default_card_modes.get(_cid, list(_key_to_label)[_ci])
            _var = tk.StringVar(value=_init_key)
            _card_vars[_cid] = _var

            # OptionMenu에 레이블 표시 — 선택 시 내부적으로 key로 역매핑
            _label_to_key = {v: k for k, v in _CARD_OPTIONS}
            _label_var = tk.StringVar(value=_key_to_label.get(_init_key, _init_key))

            _om = tk.OptionMenu(_col, _label_var,
                                *[v for _, v in _CARD_OPTIONS])
            _om.config(bg="#1f2534", fg="#d8def8", relief=tk.FLAT,
                       font=("Malgun Gothic", 9), width=14,
                       activebackground="#2a3248", activeforeground="#ffffff",
                       highlightthickness=0, bd=0)
            _om["menu"].config(bg="#1f2534", fg="#d8def8",
                               activebackground="#2a3248", activeforeground="#ffffff",
                               font=("Malgun Gothic", 9), bd=0)
            # label 선택 → key로 역매핑해서 _var에 저장
            def _sync_key(_lv=_label_var, _kv=_var, _l2k=_label_to_key, *_):
                _kv.set(_l2k.get(_lv.get(), _lv.get()))
            _label_var.trace_add("write", lambda *_, _lv=_label_var, _kv=_var, _l2k=_label_to_key:
                                 _kv.set(_l2k.get(_lv.get(), _lv.get())))
            _om.pack(anchor="w")

        manual_section = tk.Frame(frame, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        manual_section.pack(fill="x", padx=40, pady=(0, 20))
        tk.Label(manual_section, text=self._t("display_manual_symbols","수동 매매 심볼"), bg="#181A20", fg="#c0c6dc", font=("Malgun Gothic", 10, "bold"), anchor="w").pack(fill="x", padx=16, pady=(14, 4))

        symbol_choices = sorted({*(self.manual_symbol_options), *self.manual_symbols})
        if not symbol_choices:
            symbol_choices = ["BTCUSDT", "ETHUSDT"]
        manual_vars = []
        current = self.manual_symbols[:2]
        symbol_row = tk.Frame(manual_section, bg="#181A20")
        symbol_row.pack(fill="x", padx=16, pady=(0, 6))
        for idx in range(2):
            wrapper = tk.Frame(symbol_row, bg="#181A20")
            wrapper.pack(side=tk.LEFT, expand=True, fill="x", padx=(0 if idx == 0 else 8, 0))
            tk.Label(
                wrapper,
                text=f"{self._t('symbol_label','심볼')} {idx + 1}",
                bg="#181A20",
                fg="#c0c6dc",
                font=("Malgun Gothic", 9, "bold"),
                anchor="w",
            ).pack(fill="x", pady=(0, 2))
            var = tk.StringVar(value=current[idx] if idx < len(current) else "")
            manual_vars.append(var)
            combo = ttk.Combobox(
                wrapper,
                textvariable=var,
                values=[""] + symbol_choices,
                state="readonly",
                justify="center",
            )
            combo.pack(fill="x")
            combo.configure(font=("Malgun Gothic", 10))

        tk.Label(
            manual_section,
            text=self._t("display_max_symbols_hint","최대 2개까지 선택할 수 있습니다."),
            bg="#181A20",
            fg="#8e96b8",
            font=("Malgun Gothic", 9),
            anchor="w",
        ).pack(fill="x", padx=16, pady=(0, 10))

        show_manual_var = tk.BooleanVar(value=bool(self.settings_data.get("show_manual_panel", False)))

        def build_display_toggle(parent, text, variable):
            row = tk.Frame(parent, bg="#181A20")
            row.pack(fill="x", padx=16, pady=(0, 8))
            indicator = tk.Label(row, bg="#181A20")
            indicator.pack(side=tk.LEFT, padx=(0, 8))
            lbl = tk.Label(row, text=text, bg="#181A20", fg="#f5f7ff", font=("Malgun Gothic", 10, "bold"))
            lbl.pack(side=tk.LEFT)

            def refresh(*_):
                if variable.get() and self.checkbox_images.get("on"):
                    indicator.configure(image=self.checkbox_images["on"], text="")
                    indicator.image = self.checkbox_images["on"]
                elif not variable.get() and self.checkbox_images.get("off"):
                    indicator.configure(image=self.checkbox_images["off"], text="")
                    indicator.image = self.checkbox_images["off"]
                else:
                    indicator.configure(text="●" if variable.get() else "○", fg="#2EBD85" if variable.get() else "#f5f7ff")
                    indicator.image = None

            def toggle(_event=None):
                variable.set(not variable.get())
                refresh()

            for widget in (row, indicator, lbl):
                widget.bind("<Button-1>", toggle)
            refresh()
            return row

        build_display_toggle(manual_section, self._t("display_show_manual_panel","지정 코인 수동 매매 패널 표시"), show_manual_var)
        tk.Label(
            manual_section,
            text=self._t("display_hide_panel_hint","패널을 숨기면 메인 화면에서 포지션 창이 바로 위 카드 밑으로 이동합니다."),
            bg="#181A20",
            fg="#8e96b8",
            font=("Malgun Gothic", 9),
            anchor="w",
        ).pack(fill="x", padx=16, pady=(0, 14))

        def save_display():
            selected = []
            for var in manual_vars:
                val = var.get().strip().upper()
                if val:
                    if not val.endswith("USDT"):
                        val = f"{val.replace('USDT', '')}USDT"
                    selected.append(val)
            # Preserve previous symbols if combos empty
            filtered = []
            for sym in selected:
                if sym not in filtered:
                    filtered.append(sym)
            filtered = filtered[:2]
            if not filtered:
                self._show_warning(self._t("manual_symbol_title","Manual trade symbol"), self._t("manual_symbol_none_msg","Please select at least one symbol."))
                return
            self.manual_symbols = filtered
            self.settings_data["manual_symbols"] = filtered
            self.settings_data["show_manual_panel"] = bool(show_manual_var.get())
            # 카드 기본값 저장 + 런타임 즉시 반영
            _new_card_defaults = {_cid: _v.get() for _cid, _v in _card_vars.items()}
            self.settings_data["default_card_modes"] = _new_card_defaults
            for _cid, _mode in _new_card_defaults.items():
                if _cid in self.card_modes:
                    self.card_modes[_cid] = _mode
            self.state_data["card_modes"] = dict(self.card_modes)
            self._save_json(STATE_PATH, self.state_data)
            self._save_json(CONFIG_PATH, self.settings_data)
            self._render_manual_cards()
            self._apply_manual_panel_visibility()
            self._rebuild_stats_cards()
            self._show_info(self._t("trade_settings_save_title","Saved"), self._t("display_settings_saved","Display settings saved."))

        display_btn_container = tk.Frame(frame, bg="#181A20")
        display_btn_container.pack(side=tk.BOTTOM, fill="x", padx=40, pady=20)

        def restore_display_defaults():
            defaults = ["BTCUSDT", "ETHUSDT"]
            for idx, var in enumerate(manual_vars):
                var.set(defaults[idx] if idx < len(defaults) else "")
            show_manual_var.set(False)
            _default_card_reset = {
                "card0": "unrealized_total", "card1": "filter_pass_rate",
                "card2": "top_symbol",       "card3": "pnl_24h",
            }
            for _cid, _mode in _default_card_reset.items():
                if _cid in _card_vars:
                    _card_vars[_cid].set(_mode)

        # 세션 상태 초기화 버튼
        def _reset_session_state():
            parent_w = getattr(self, "_active_modal", self.root)
            _en = self.language == "en"
            answer = tk.messagebox.askyesno(
                "Reset Session State" if _en else "세션 상태 초기화",
                ("Reset GUI session state.\n\n"
                 "  • Reset session start time\n"
                 "  • Reset win-rate / trade counters\n"
                 "  • Reset PnL card base time\n\n"
                 "Settings, API keys, and trade history are NOT affected.\nContinue?"
                 if _en else
                 "GUI 세션 상태를 초기화합니다.\n\n"
                 "  • 세션 시작 시각 초기화\n"
                 "  • 승률·거래수 카운터 초기화\n"
                 "  • 손익 카드 기준 시각 초기화\n\n"
                 "설정값·API 키·거래 기록은 영향 없습니다.\n계속하시겠습니까?"),
                parent=parent_w,
            )
            if not answer:
                return
            now = int(time.time() * 1000)
            # state_data에서 세션 키 제거 (card_modes 보존)
            _keep = {"card_modes"}
            preserved = {k: v for k, v in self.state_data.items() if k in _keep}
            self.state_data.clear()
            self.state_data.update(preserved)
            self._save_json(STATE_PATH, self.state_data)
            # 인메모리 상태 초기화
            self.session_start_ms = None
            self.pnl_reset_ms     = now
            self.stat_resets["win_rate"]    = now
            self.stat_resets["trade_count"] = now
            self._set_stat_value("trade_count", "0")
            self._set_stat_value("win_rate",    "0%")
            pnl_keys = {"pnl_15", "pnl_60", "pnl_12h", "pnl_24h"}
            active = set(getattr(self, "card_modes", {}).values())
            for k in pnl_keys:
                if k in active:
                    self._set_stat_value(k, "+0 USDT")
            self._append_log("[DISPLAY] " + ("GUI session state reset" if _en else "GUI 세션 상태 초기화 완료"))
            tk.messagebox.showinfo(
                "Done" if _en else "완료",
                "Session state has been reset." if _en else "세션 상태가 초기화되었습니다.",
                parent=parent_w)

        display_btn_row = tk.Frame(display_btn_container, bg="#181A20")
        display_btn_row.pack(anchor="e", padx=20, pady=8)
        ttk.Button(display_btn_row, text="DEFAULT", command=restore_display_defaults, style="EnvDefault.TButton").pack(side=tk.LEFT, padx=(0, 12))
        ttk.Button(display_btn_row, text="SAVE", command=save_display, style="EnvSave.TButton").pack(side=tk.LEFT)

    # ------------------------------------------------------------------
    def _build_developer_tab(self, frame):
        frame.columnconfigure(0, weight=1)
        tk.Label(frame, text=self._t("dev_title","개발자 패널"), bg="#181A20", fg="white", font=("Malgun Gothic", 13, "bold"), anchor="w").pack(fill="x", padx=40, pady=(30, 12))
        if self.language == "ko":
            warning_text = (
                "⚠️ 고위험 경고 / 면책 / 사용자 책임 (필독)\n"
                "이 프로그램은 암호자산/선물 자동매매를 수행할 수 있는 소프트웨어입니다. 본 프로그램 및 개발자는 투자/재무/거래 조언을 제공하지 않으며, 어떤 수익도 보장하지 않습니다.\n"
                "암호자산·파생상품 거래는 매우 높은 위험을 수반하며 원금의 일부 또는 전부를 잃을 수 있습니다.\n\n"
                "사용자 책임\n"
                "프로그램 사용 여부, 전략/파라미터 설정, API 키 관리, 주문 실행 결과에 대한 책임은 전적으로 사용자에게 있습니다. 실거래 사용 전 테스트넷/소액으로 충분히 검증하세요.\n\n"
                "무보증 / 책임 제한\n"
                "본 소프트웨어는 \'있는 그대로(AS IS)\' 제공되며, 어떠한 보증도 제공하지 않습니다. 사용으로 인해 발생하는 거래 손실, 기회손실, 데이터 손상, 시스템 장애 등 모든 손해에 대해 개발자/기여자는 법이 허용하는 범위 내에서 책임을 지지 않습니다.\n"
                "거래소/서드파티 주문·체결·정산은 거래소/네트워크 상태 및 정책에 따라 달라질 수 있습니다. 사용자는 거래소 약관 및 API 규정을 준수해야 합니다."
            )
        else:
            warning_text = (
                "⚠️ High-Risk Warning / Disclaimer / User Responsibility (Please Read)\n"
                "This software can execute automated trading of crypto assets and futures. The program and its developers do not provide investment, financial, or trading advice and do not guarantee any returns.\n"
                "Crypto and derivatives trading involves very high risk. You may lose some or all of your principal.\n\n"
                "User Responsibility\n"
                "You are solely responsible for the decision to use this program, your strategy and parameter choices, API key management, and all order execution outcomes. Always test thoroughly on Testnet or with minimal funds before live trading.\n\n"
                "No Warranty / Limitation of Liability\n"
                "This software is provided AS IS without any warranty of any kind. To the maximum extent permitted by law, the developers and contributors are not liable for trading losses, opportunity costs, data corruption, system failures, or any other damages arising from its use.\n"
                "Exchange and third-party order execution, fills, and settlement may vary depending on exchange/network conditions and policies. Users must comply with exchange terms of service and API usage policies."
            )
        warning_label = tk.Label(
            frame,
            text=warning_text,
            bg="#181A20",
            fg="#cbd2f5",
            wraplength=760,
            justify="left",
            anchor="w",
        )
        warning_label.pack(fill="x", padx=40, pady=(0, 20))

        # A1: 2개 항목 분리 체크박스
        def _make_ack_row(parent, row_idx, ack_key, label_key, text_key, initial_val):
            ack_v = tk.BooleanVar(value=bool(self.settings_data.get(ack_key, False)))
            rf = tk.Frame(parent, bg="#181A20")
            rf.grid(row=row_idx, column=0, columnspan=2, sticky="ew", padx=8, pady=(8, 2))
            rf.columnconfigure(1, weight=1)
            ind = tk.Label(rf, bg="#181A20")
            ind.grid(row=0, column=0, padx=(4, 8), sticky="w")
            def _refresh_ind(ind=ind, ack_v=ack_v):
                if self.checkbox_images.get("on") and self.checkbox_images.get("off"):
                    ind.configure(image=self.checkbox_images["on"] if ack_v.get() else self.checkbox_images["off"], text="")
                    ind.image = self.checkbox_images["on" if ack_v.get() else "off"]
                else:
                    ind.configure(text="●" if ack_v.get() else "○", fg="#2EBD85" if ack_v.get() else "#f5f7ff")
            _refresh_ind()
            hdr = tk.Label(rf, text=self._t(label_key, label_key), bg="#181A20", fg="#F0B90B",
                           font=("Malgun Gothic", 9, "bold"), anchor="w")
            hdr.grid(row=0, column=1, sticky="w")
            txt = tk.Label(rf, text=self._t(text_key, text_key), bg="#181A20", fg="#f5f7ff",
                           font=("Malgun Gothic", 9), wraplength=620, justify="left", anchor="w")
            txt.grid(row=1, column=1, sticky="w", pady=(0, 4))
            def _toggle(ind=ind, ack_v=ack_v, ack_key=ack_key, _refresh=_refresh_ind):
                ack_v.set(not ack_v.get())
                _refresh()
                self._update_risk_ack(ack_v.get(), ack_key=ack_key)
            for w in (ind, hdr, txt):
                w.bind("<Button-1>", lambda _e, t=_toggle: t())
            return ack_v

        ack_frame = tk.Frame(frame, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        ack_frame.pack(fill="x", padx=40, pady=(0, 10))
        ack_frame.columnconfigure(1, weight=1)

        _make_ack_row(ack_frame, 0, "risk_ack1", "dev_ack1_label", "dev_ack1_text",
                      self.settings_data.get("risk_ack1", False))
        _make_ack_row(ack_frame, 1, "risk_ack2", "dev_ack2_label", "dev_ack2_text",
                      self.settings_data.get("risk_ack2", False))

        tk.Label(
            ack_frame,
            text=self._t("dev_ack_required_note", "※ 두 항목 모두 체크해야 자동/수동 매매가 활성화됩니다."),
            bg="#181A20", fg="#F0B90B", font=("Malgun Gothic", 9, "bold"), anchor="w",
        ).grid(row=2, column=0, columnspan=2, sticky="w", padx=12, pady=(4, 8))

        # D: Expert mode toggle — gates Aggressive preset
        expert_frame = tk.Frame(frame, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        expert_frame.pack(fill="x", padx=40, pady=(10, 20))
        expert_var = tk.BooleanVar(value=bool(self.settings_data.get("expert_mode_enabled", False)))
        expert_ind = tk.Label(expert_frame, bg="#181A20")
        expert_ind.pack(side=tk.LEFT, padx=(12, 8), pady=10)
        def _refresh_expert_ind():
            if self.checkbox_images.get("on") and self.checkbox_images.get("off"):
                expert_ind.configure(image=self.checkbox_images["on"] if expert_var.get() else self.checkbox_images["off"], text="")
                expert_ind.image = self.checkbox_images["on" if expert_var.get() else "off"]
            else:
                expert_ind.configure(text="●" if expert_var.get() else "○", fg="#F0B90B" if expert_var.get() else "#f5f7ff")
        _refresh_expert_ind()
        expert_lbl = tk.Label(
            expert_frame,
            text=self._t("expert_mode_label", "Expert Mode (Enable Aggressive Preset)"),
            bg="#181A20", fg="#f5f7ff", font=("Malgun Gothic", 10, "bold"), anchor="w"
        )
        expert_lbl.pack(side=tk.LEFT)
        def _toggle_expert():
            expert_var.set(not expert_var.get())
            _refresh_expert_ind()
            self.settings_data["expert_mode_enabled"] = expert_var.get()
            self._save_json(CONFIG_PATH, self.settings_data)
        for w in (expert_ind, expert_lbl):
            w.bind("<Button-1>", lambda _e: _toggle_expert())



    # ═══════════════════════════════════════════════════════════════════════════
    # 프리미엄 탭  — AI 학습 스코어러
    # ═══════════════════════════════════════════════════════════════════════════
    def _build_premium_tab(self, frame):
        """AI 학습 스코어러 UI — 프리미엄 독립 탭."""
        import tkinter as tk
        from tkinter import ttk, messagebox
        import webbrowser

        BG      = "#181A20"
        CARD    = "#10131d"
        BORDER  = "#343942"
        GOLD    = "#F0B90B"
        GREEN   = "#2ECC71"
        RED     = "#FF6B6B"
        BLUE    = "#4A90D9"
        MUTED   = "#8e96b8"
        WHITE   = "#f5f7ff"
        DARKBG  = "#0a0d14"
        LOCK_C  = "#FF6B35"   # 잠금 아이콘 색

        frame.configure(bg=BG)

        # ── 스크롤 컨테이너 ────────────────────────────────────────────────
        canvas  = tk.Canvas(frame, bg=BG, highlightthickness=0)
        vbar    = ttk.Scrollbar(frame, orient="vertical", command=canvas.yview)
        canvas.configure(yscrollcommand=vbar.set)
        vbar.pack(side=tk.RIGHT, fill=tk.Y)
        canvas.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        inner   = tk.Frame(canvas, bg=BG)
        win_id  = canvas.create_window((0, 0), window=inner, anchor="nw")
        inner.bind("<Configure>", lambda e: (
            canvas.configure(scrollregion=canvas.bbox("all")),
            canvas.itemconfig(win_id, width=canvas.winfo_width()),
        ))
        canvas.bind("<Configure>", lambda e: canvas.itemconfig(win_id, width=e.width))
        def _on_mousewheel(e):
            try:
                canvas.yview_scroll(int(-e.delta / 120), "units")
            except Exception:
                pass
        # Enter/Leave 방식: 마우스가 캔버스 위에 있을 때만 바인딩
        # bind_all 사용 시 창 닫혀도 이벤트 남아 TclError 발생하는 문제 해결
        inner.bind("<Enter>", lambda e: canvas.bind_all("<MouseWheel>", _on_mousewheel))
        inner.bind("<Leave>", lambda e: canvas.unbind_all("<MouseWheel>"))

        def make_card(parent, border=BORDER, pady=(4,4)):
            f = tk.Frame(parent, bg=CARD, highlightbackground=border, highlightthickness=1)
            f.pack(fill="x", padx=24, pady=pady)
            return f
        def sep(parent):
            tk.Frame(parent, bg=BORDER, height=1).pack(fill="x", padx=14, pady=(4,4))
        def card_title(parent, text, fg=WHITE, font_size=10):
            tk.Label(parent, text=text, bg=CARD, fg=fg,
                     font=("Malgun Gothic", font_size, "bold"), anchor="w"
                     ).pack(fill="x", padx=14, pady=(10,4))
        def muted_label(parent, text, wraplength=540):
            tk.Label(parent, text=text, bg=CARD, fg=MUTED,
                     font=("Malgun Gothic", 8), anchor="w",
                     justify="left", wraplength=wraplength
                     ).pack(fill="x", padx=16, pady=(0,4))

        import traceback as _tb
        import logging as _lg
        _prem_log = _lg.getLogger("premium_tab")

        def _safe_section(name, builder):
            """각 섹션을 try/except로 감싸 한 섹션 에러가 전체를 막지 않도록."""
            try:
                builder()
            except Exception as _e:
                _tb.print_exc()
                try:
                    err_f = tk.Frame(inner, bg="#2a0000", highlightbackground=RED, highlightthickness=1)
                    err_f.pack(fill="x", padx=24, pady=4)
                    tk.Label(err_f, text=f"⚠ {name} 렌더링 오류: {_e}",
                             bg="#2a0000", fg=RED, font=("Malgun Gothic", 8),
                             wraplength=550, anchor="w", justify="left"
                    ).pack(fill="x", padx=10, pady=6)
                except Exception:
                    pass

        _is_ko = (self.language == "ko")

        # ── 라이선스 키 검증 함수 (여러 섹션에서 공유) ──
        def _has_valid_key():
            k = self.settings_data.get("neural_license_key", "")
            if not k:
                return False
            try:
                from binance_futures_bot1_1.binance_futures_bot.license_gate import validate_key
                ok, _ = validate_key(k, "NEURAL", lang=self.language)
                return ok
            except Exception:
                return False

        # ══════════════════════════════════════════════════════════════════
        # HERO 배너
        # ══════════════════════════════════════════════════════════════════
        hero = tk.Frame(inner, bg=DARKBG, highlightbackground=GOLD, highlightthickness=1)
        hero.pack(fill="x", padx=24, pady=(20, 6))
        htop = tk.Frame(hero, bg=DARKBG)
        htop.pack(fill="x", padx=16, pady=(14,4))
        tk.Label(htop,
                 text="🧠  AI 학습 스코어러" if _is_ko else "🧠  AI Learning Scorer",
                 bg=DARKBG, fg=GOLD,
                 font=("Malgun Gothic", 12, "bold"), anchor="w").pack(side=tk.LEFT)
        tk.Label(htop, text=" PREMIUM ",
                 bg=GOLD, fg=DARKBG,
                 font=("Malgun Gothic", 8, "bold"), padx=4, pady=1).pack(side=tk.LEFT, padx=10)
        tk.Label(hero, text=(
            "v3 듀얼헤드 신경망: 승률 예측 + 기대 ROI를 동시에 분석합니다.\n"
            "20개 시장 피처 · 진입 strength 0.0x~2.0x 자동 조정 · 승률 25% 미만 진입 차단\n"
            "적응형 학습률로 시장 변화에 자동 적응합니다."
            if _is_ko else
            "v3 Dual-Head Neural Net: predicts win probability + expected ROI simultaneously.\n"
            "20 market features · Entry strength 0.0x–2.0x auto-adjust · Blocks entry below 25% win prob\n"
            "Adaptive learning rate automatically adapts to market changes."
        ), bg=DARKBG, fg=MUTED, font=("Malgun Gothic", 8),
           anchor="w", justify="left", wraplength=570
        ).pack(fill="x", padx=16, pady=(0,6))

        # ── AI 스코어러 ON/OFF 토글 (HERO 내부에 통합) ────────────────────
        tk.Frame(hero, bg=BORDER, height=1).pack(fill="x", padx=14, pady=(0,6))

        tk.Label(hero,
                 text=(
                     "ON  →  승률 + ROI 예측으로 진입 strength 0.0x~2.0x 보정. 승률 25% 미만 차단.\n"
                     "OFF →  거래 데이터 수집·학습만 진행. 나중에 ON 해도 학습 데이터 유지."
                     if _is_ko else
                     "ON  →  Win prob + ROI prediction adjusts entry strength 0.0x–2.0x. Blocks <25% win.\n"
                     "OFF →  Data collection & learning only. Accumulated data preserved when re-enabled."
                 ),
                 bg=DARKBG, fg=MUTED, font=("Malgun Gothic", 8),
                 anchor="w", justify="left", wraplength=560
        ).pack(fill="x", padx=16, pady=(0,6))

        # 토글 행: [상태 인디케이터(LEFT)] [설명(LEFT, expand)] [버튼(RIGHT)]
        scorer_toggle_row = tk.Frame(hero, bg=DARKBG)
        scorer_toggle_row.pack(fill="x", padx=16, pady=(0,14))

        toggle_btn = tk.Button(scorer_toggle_row,
                               relief="flat", cursor="hand2",
                               font=("Malgun Gothic", 9, "bold"), width=8)
        toggle_btn.pack(side=tk.RIGHT)

        scorer_toggle_left = tk.Frame(scorer_toggle_row, bg=DARKBG)
        scorer_toggle_left.pack(side=tk.LEFT, fill="x", expand=True)

        toggle_ind = tk.Label(scorer_toggle_left, bg=DARKBG,
                              font=("Malgun Gothic", 10, "bold"), anchor="w")
        toggle_ind.pack(side=tk.LEFT, padx=(0, 10))

        toggle_sub = tk.Label(scorer_toggle_left, bg=DARKBG,
                              font=("Malgun Gothic", 8), anchor="w")
        toggle_sub.pack(side=tk.LEFT)

        def _refresh_toggle():
            has_key = _has_valid_key()
            is_on   = bool(self.settings_data.get("neural_scorer_enabled", False)) and has_key
            if not has_key:
                toggle_ind.configure(text="🔒  LOCKED", fg=MUTED)
                toggle_btn.configure(
                    text="잠김" if _is_ko else "Locked",
                    bg="#1E2329", fg=MUTED,
                    activebackground="#1E2329", state="disabled")
                toggle_sub.configure(
                    text="라이선스 키를 먼저 입력하세요" if _is_ko else "Enter a license key first",
                    fg=MUTED)
            elif is_on:
                toggle_ind.configure(text="●  ON", fg=GREEN)
                toggle_btn.configure(
                    text="끄기" if _is_ko else "Turn OFF",
                    bg="#1E2329", fg=WHITE,
                    activebackground="#2a3040", state="normal")
                toggle_sub.configure(
                    text="예측 활성 — 진입 strength가 AI 확률로 보정됩니다" if _is_ko
                    else "Active — entry strength is being adjusted by AI probability",
                    fg=GREEN)
            else:
                toggle_ind.configure(text="○  OFF", fg=MUTED)
                toggle_btn.configure(
                    text="켜기" if _is_ko else "Turn ON",
                    bg=GOLD, fg=DARKBG,
                    activebackground="#d9a80a", state="normal")
                toggle_sub.configure(
                    text="예측 비활성 — 데이터 수집·학습만 진행 중" if _is_ko
                    else "Inactive — data collection & learning in progress",
                    fg=MUTED)

        def _do_toggle():
            if not _has_valid_key():
                return
            cur = bool(self.settings_data.get("neural_scorer_enabled", False))
            self.settings_data["neural_scorer_enabled"] = not cur
            self._save_json(CONFIG_PATH, self.settings_data)
            try:
                from binance_futures_bot1_1 import main as _eng_main
                _ce = getattr(_eng_main, "current_engine", None)
                if _ce:
                    _ce.config.neural_scorer_enabled = not cur
            except Exception:
                pass
            _refresh_toggle()
            _refresh_dashboard()

        toggle_btn.configure(command=_do_toggle)
        _refresh_toggle()

        # ══════════════════════════════════════════════════════════════════
        # AI 어시스턴트 (HERO 배너와 동일한 스타일)
        # ══════════════════════════════════════════════════════════════════
        def _open_ai_assistant():
            """AI 어시스턴트 창 열기/복원. 숨겨진 상태면 deiconify."""
            if (hasattr(self, "_ai_assistant_win") and self._ai_assistant_win
                    and self._ai_assistant_win.winfo_exists()):
                self._ai_assistant_win.deiconify()
                self._ai_assistant_win.lift()
                self._ai_assistant_win.focus_force()
                return
            self._create_ai_assistant_window()

        def _close_ai_assistant():
            """창만 숨김 (withdraw). 백그라운드 작업은 계속 유지."""
            if (hasattr(self, "_ai_assistant_win") and self._ai_assistant_win
                    and self._ai_assistant_win.winfo_exists()):
                self._ai_assistant_win.withdraw()

        ai_box = tk.Frame(inner, bg=DARKBG, highlightbackground=GOLD, highlightthickness=1)
        ai_box.pack(fill="x", padx=24, pady=(4, 6))

        # ── 타이틀 행 (HERO와 동일 폰트 12pt bold) ──
        ai_top = tk.Frame(ai_box, bg=DARKBG)
        ai_top.pack(fill="x", padx=16, pady=(14, 4))

        tk.Label(ai_top,
                 text="🤖  AI 트레이딩 어시스턴트" if _is_ko else "🤖  AI Trading Assistant",
                 bg=DARKBG, fg=GOLD,
                 font=("Malgun Gothic", 12, "bold"), anchor="w"
        ).pack(side=tk.LEFT)
        tk.Label(ai_top, text=" PREMIUM ",
                 bg=GOLD, fg=DARKBG,
                 font=("Malgun Gothic", 8, "bold"), padx=4, pady=1).pack(side=tk.LEFT, padx=10)

        # ── 설명 (HERO와 동일 폰트 8pt) ──
        tk.Label(ai_box, text=(
            "실시간 시장 분석 · 엔진 활동 모니터 · 패턴 분석 기반 자가 개선 제안"
            if _is_ko else
            "Real-time market analysis · Engine activity monitor · Pattern-based self-improvement"
        ), bg=DARKBG, fg=MUTED, font=("Malgun Gothic", 8),
           anchor="w", wraplength=570
        ).pack(fill="x", padx=16, pady=(0, 6))

        # ── 어시스턴트 ON/OFF 토글 (스코어러와 동일 스타일) ──
        tk.Frame(ai_box, bg=BORDER, height=1).pack(fill="x", padx=14, pady=(0,6))

        ai_toggle_row = tk.Frame(ai_box, bg=DARKBG)
        ai_toggle_row.pack(fill="x", padx=16, pady=(0,6))

        # 버튼 영역 (오른쪽): ON/OFF 버튼 + 숨기기/열기 버튼
        ai_btn_frame = tk.Frame(ai_toggle_row, bg=DARKBG)
        ai_btn_frame.pack(side=tk.RIGHT)

        ai_toggle_btn = tk.Button(ai_btn_frame,
                                  relief="flat", cursor="hand2",
                                  font=("Malgun Gothic", 9, "bold"), width=8)
        ai_toggle_btn.pack(side=tk.LEFT, padx=(0, 4))

        ai_visibility_btn = tk.Button(ai_btn_frame,
                                      relief="flat", cursor="hand2",
                                      font=("Malgun Gothic", 9, "bold"), width=8)
        ai_visibility_btn.pack(side=tk.LEFT)

        ai_toggle_left = tk.Frame(ai_toggle_row, bg=DARKBG)
        ai_toggle_left.pack(side=tk.LEFT, fill="x", expand=True)

        ai_status_lbl = tk.Label(ai_toggle_left, bg=DARKBG,
                                 font=("Malgun Gothic", 10, "bold"), anchor="w")
        ai_status_lbl.pack(side=tk.LEFT, padx=(0, 10))

        ai_sub_lbl = tk.Label(ai_toggle_left, bg=DARKBG,
                              font=("Malgun Gothic", 8), anchor="w")
        ai_sub_lbl.pack(side=tk.LEFT)

        # ── 실시간 미리보기 영역 (인라인 피드) ──
        _PREVIEW_BG = "#0d1017"
        ai_preview = tk.Frame(ai_box, bg=_PREVIEW_BG,
                              highlightbackground=BORDER, highlightthickness=1)
        ai_preview.pack(fill="x", padx=14, pady=(2, 6))

        # 시장 요약 라인
        ai_pv_summary = tk.Label(ai_preview,
                                 text="📊  " + ("시장 데이터 로딩 중..." if _is_ko else "Loading market data..."),
                                 bg=_PREVIEW_BG, fg=MUTED,
                                 font=("Malgun Gothic", 8), anchor="w",
                                 padx=8, pady=4)
        ai_pv_summary.pack(fill="x")

        # 최근 이벤트 2줄
        ai_pv_event1 = tk.Label(ai_preview, text="", bg=_PREVIEW_BG, fg=MUTED,
                                font=("Malgun Gothic", 8), anchor="w",
                                padx=8, pady=0)
        ai_pv_event1.pack(fill="x")

        ai_pv_event2 = tk.Label(ai_preview, text="", bg=_PREVIEW_BG, fg=MUTED,
                                font=("Malgun Gothic", 8), anchor="w",
                                padx=8, pady=0)
        ai_pv_event2.pack(fill="x", pady=(0, 2))

        # 구분선
        tk.Frame(ai_preview, bg=BORDER, height=1).pack(fill="x", padx=8, pady=2)

        # 개선 제안 라인
        ai_pv_suggest = tk.Label(ai_preview,
                                 text="💡  " + ("거래 데이터 수집 중..." if _is_ko else "Collecting trade data..."),
                                 bg=_PREVIEW_BG, fg=MUTED,
                                 font=("Malgun Gothic", 8), anchor="w",
                                 wraplength=540, justify="left",
                                 padx=8, pady=2)
        ai_pv_suggest.pack(fill="x", pady=(0, 4))

        # ── 토글 상태 관리 (라이선스 키 필요) ──
        def _is_ai_win_alive():
            """어시스턴트 창이 존재하는지 (숨김 상태 포함)."""
            return (hasattr(self, "_ai_assistant_win") and self._ai_assistant_win
                    and self._ai_assistant_win.winfo_exists())

        def _is_ai_win_visible():
            """어시스턴트 창이 보이는 상태인지."""
            if not _is_ai_win_alive():
                return False
            try:
                return self._ai_assistant_win.state() != "withdrawn"
            except Exception:
                return False

        def _turn_off_ai_assistant():
            """어시스턴트 기능 완전 종료 (창 파괴)."""
            if _is_ai_win_alive():
                try:
                    self._ai_assistant_win.destroy()
                except Exception:
                    pass
                self._ai_assistant_win = None

        def _refresh_ai_toggle():
            has_key = _has_valid_key()
            alive = _is_ai_win_alive()
            visible = _is_ai_win_visible()
            if not has_key:
                # 키 없으면 잠금 — 창이 살아있으면 완전 파괴
                if alive:
                    _turn_off_ai_assistant()
                ai_status_lbl.configure(text="🔒  LOCKED", fg=MUTED)
                ai_toggle_btn.configure(
                    text="잠김" if _is_ko else "Locked",
                    bg="#1E2329", fg=MUTED,
                    activebackground="#1E2329", state="disabled",
                )
                ai_visibility_btn.configure(
                    text="", bg=DARKBG, fg=DARKBG,
                    activebackground=DARKBG, state="disabled",
                    relief="flat", borderwidth=0, highlightthickness=0,
                    width=0, padx=0,
                )
                ai_visibility_btn.pack_forget()
                ai_sub_lbl.configure(
                    text="라이선스 키를 먼저 입력하세요" if _is_ko else "Enter a license key first",
                    fg=MUTED)
            elif alive:
                # 창 존재 (보이든 숨겨든) → ON 상태
                # ON/OFF 버튼: 항상 "끄기" 표시
                ai_status_lbl.configure(text="●  ON", fg=GREEN)
                ai_toggle_btn.configure(
                    text="끄기" if _is_ko else "Turn OFF",
                    bg="#8B0000", fg=WHITE,
                    activebackground="#a00000", state="normal",
                    command=lambda: (_turn_off_ai_assistant(), ai_box.after(200, _refresh_ai_toggle)),
                )
                # 숨기기/열기 버튼: 창 가시성 제어
                ai_visibility_btn.pack(side=tk.LEFT)
                if visible:
                    _vis_text = "숨기기" if _is_ko else "Hide"
                    _sub_text = "어시스턴트 실행 중 — 창 열림" if _is_ko else "Running — window open"
                else:
                    _vis_text = "열기" if _is_ko else "Show"
                    _sub_text = "백그라운드 실행 중 — 창 숨김" if _is_ko else "Running in background — hidden"
                ai_visibility_btn.configure(
                    text=_vis_text,
                    bg="#1E2329", fg=WHITE,
                    activebackground="#2a3040", state="normal",
                    relief="flat", borderwidth=1, highlightthickness=0,
                    width=8, padx=0,
                    command=lambda: (
                        (_close_ai_assistant() if _is_ai_win_visible() else _open_ai_assistant()),
                        ai_box.after(200, _refresh_ai_toggle),
                    ),
                )
                ai_sub_lbl.configure(text=_sub_text, fg=GREEN)
            else:
                # OFF 상태 — 켜기 버튼만 표시
                ai_status_lbl.configure(text="○  OFF", fg=MUTED)
                ai_toggle_btn.configure(
                    text="켜기" if _is_ko else "Turn ON",
                    bg=GOLD, fg=DARKBG,
                    activebackground="#d4a50a", state="normal",
                    command=lambda: (_open_ai_assistant(), ai_box.after(200, _refresh_ai_toggle)),
                )
                # 숨기기/열기 버튼 숨김 (OFF 상태이므로 불필요)
                ai_visibility_btn.pack_forget()
                ai_sub_lbl.configure(
                    text="어시스턴트 꺼짐" if _is_ko else "Assistant is off",
                    fg=MUTED)

        _refresh_ai_toggle()

        # ── 실시간 미리보기 갱신 (10초 간격) ──
        _ai_preview_advisor = [None]  # lazy init

        def _get_preview_advisor():
            if _ai_preview_advisor[0] is None:
                try:
                    from binance_futures_bot1_1.binance_futures_bot.ai_advisor import AIAdvisor
                    _ai_preview_advisor[0] = AIAdvisor(BASE_DIR, language=self.language)
                except Exception:
                    pass
            return _ai_preview_advisor[0]

        _preview_events_cache = []

        def _refresh_ai_preview():
            """인라인 미리보기 갱신."""
            try:
                if not inner.winfo_exists():
                    return

                # 토글 상태도 갱신
                _refresh_ai_toggle()

                adv = _get_preview_advisor()
                if not adv:
                    inner.after(10000, _refresh_ai_preview)
                    return

                # 언어 동기화
                adv.set_language(self.language)
                _ko = self.language == "ko"

                # 1) 시장 요약
                try:
                    stxt = adv.get_market_summary_text()
                    ai_pv_summary.configure(text="📊  " + stxt, fg=WHITE)
                except Exception:
                    pass

                # 2) 최근 이벤트
                try:
                    new_evs = adv.poll_new_events()
                    _preview_events_cache.extend(new_evs)
                    # 최근 20개만 유지
                    while len(_preview_events_cache) > 20:
                        _preview_events_cache.pop(0)

                    if len(_preview_events_cache) >= 1:
                        ev = _preview_events_cache[-1]
                        ts = ev.get("ts_str", "").split(" ")[-1][:5] if " " in ev.get("ts_str", "") else ""
                        ai_pv_event1.configure(
                            text=f"{ev.get('icon','')}  [{ts}] {ev.get('friendly_msg','')[:70]}",
                            fg=ev.get("color", MUTED))
                    if len(_preview_events_cache) >= 2:
                        ev = _preview_events_cache[-2]
                        ts = ev.get("ts_str", "").split(" ")[-1][:5] if " " in ev.get("ts_str", "") else ""
                        ai_pv_event2.configure(
                            text=f"{ev.get('icon','')}  [{ts}] {ev.get('friendly_msg','')[:70]}",
                            fg=ev.get("color", MUTED))
                except Exception:
                    pass

                # 3) 개선 제안
                try:
                    sug = adv.get_improvement_suggestions()
                    if sug:
                        top = sug[0]
                        msg = top.get("msg_ko" if _ko else "msg_en", "")
                        # 한줄로 정리 (첫 줄만)
                        first_line = msg.split("\n")[0]
                        _icon = top.get("icon", "💡")
                        _pri_color = {
                            "high": RED, "medium": GOLD, "low": GREEN
                        }.get(top.get("priority", ""), MUTED)
                        ai_pv_suggest.configure(
                            text=f"{_icon}  {first_line}",
                            fg=_pri_color)
                    else:
                        ai_pv_suggest.configure(
                            text="💡  " + ("현재 특이사항 없음 — 전략 안정" if _ko
                                          else "No issues — strategy is stable"),
                            fg=GREEN)
                except Exception:
                    pass

            except Exception:
                pass
            finally:
                try:
                    if inner.winfo_exists():
                        inner.after(10000, _refresh_ai_preview)
                except Exception:
                    pass

        # 첫 갱신 (2초 후)
        inner.after(2000, _refresh_ai_preview)
        _prem_log.info("[PREMIUM] AI assistant section OK, building toggle section...")

        def _refresh_lock_ui():
            _refresh_toggle()
            _refresh_ai_toggle()
            _refresh_pay_status()
            _refresh_dashboard()

        # ══════════════════════════════════════════════════════════════════
        # SECTION: 결제 / 잠금 해제  (어시스턴트 패널 아래)
        # ══════════════════════════════════════════════════════════════════
        pay_card = make_card(inner, border=LOCK_C, pady=(4,6))
        pay_top = tk.Frame(pay_card, bg=CARD)
        pay_top.pack(fill="x", padx=14, pady=(10,4))

        pay_title_lbl = tk.Label(pay_top,
                 text="🔐  기능 잠금 해제" if _is_ko else "🔐  Unlock Feature",
                 bg=CARD, fg=LOCK_C,
                 font=("Malgun Gothic", 10, "bold"), anchor="w")
        pay_title_lbl.pack(side=tk.LEFT)

        # 인증 완료 시 남은 일수 표시 라벨
        pay_verified_lbl = tk.Label(pay_top, text="", bg=CARD, fg=GREEN,
                                     font=("Malgun Gothic", 9, "bold"), anchor="w")
        pay_verified_lbl.pack(side=tk.LEFT, padx=(8,0))

        # 접기/펼치기 토글 버튼
        pay_collapse_btn = tk.Label(pay_top, text="", bg=CARD, fg=MUTED,
                                     font=("Malgun Gothic", 9), anchor="e", cursor="hand2")
        pay_collapse_btn.pack(side=tk.RIGHT)

        # ── 접기/펼치기 대상: pay_body ──
        pay_body = tk.Frame(pay_card, bg=CARD)
        pay_body.pack(fill="x")
        _pay_collapsed = [False]  # mutable for closure

        def _toggle_pay_body(event=None):
            if _pay_collapsed[0]:
                pay_body.pack(fill="x")
                _pay_collapsed[0] = False
                pay_collapse_btn.configure(text="▲ " + ("접기" if _is_ko else "Collapse"))
            else:
                pay_body.pack_forget()
                _pay_collapsed[0] = True
                pay_collapse_btn.configure(text="▼ " + ("펼치기" if _is_ko else "Expand"))
        pay_collapse_btn.bind("<Button-1>", _toggle_pay_body)

        sep(pay_body)

        tk.Label(pay_body,
                 text=(
                     "결제 완료 후 이메일로 라이선스 키를 발급해 드립니다.\n아래에서 구독 플랜을 선택하거나, 이미 키가 있으면 바로 입력하세요."
                     if _is_ko else
                     "A license key will be sent to your email after payment.\nChoose a subscription plan below, or enter your key directly if you already have one."
                 ),
                 bg=CARD, fg=MUTED, font=("Malgun Gothic", 8),
                 anchor="w", justify="left", wraplength=550
        ).pack(fill="x", padx=16, pady=(4,8))

        # ── [PATCH-12] Lemon Squeezy 구독 버튼 (월/연) ──
        pay_btns = tk.Frame(pay_body, bg=CARD)
        pay_btns.pack(fill="x", padx=14, pady=(0,4))

        def _open_payment(url, plan_name):
            if not url:
                messagebox.showinfo(
                    "준비 중" if _is_ko else "Coming Soon",
                    ("결제 시스템 준비 중입니다. 곧 오픈 예정입니다!"
                     if _is_ko else
                     "Payment system is being prepared. Coming soon!"),
                    parent=frame.winfo_toplevel())
                return
            try:
                webbrowser.open(url)
            except Exception:
                messagebox.showinfo(
                    "Purchase Link" if self.language == "en" else "구매 링크",
                    (f"Visit the following URL:\n{url}" if self.language == "en"
                     else f"아래 주소로 접속하세요:\n{url}"),
                    parent=frame.winfo_toplevel())

        tk.Button(pay_btns,
                  text=f"💳  월간 구독 ({PREMIUM_PRICE_MONTHLY}/월)" if _is_ko else f"💳  Monthly ({PREMIUM_PRICE_MONTHLY}/mo)",
                  command=lambda: _open_payment(LEMONSQUEEZY_MONTHLY_URL, "monthly"),
                  bg="#F0B90B", fg="#181A20",
                  activebackground="#d9a80a", activeforeground="#181A20",
                  relief="flat", cursor="hand2",
                  font=("Malgun Gothic", 9, "bold"), padx=16, pady=6
        ).pack(side=tk.LEFT, padx=(0,8))

        tk.Button(pay_btns,
                  text=f"💎  연간 구독 ({PREMIUM_PRICE_YEARLY} · 17% 할인)" if _is_ko else f"💎  Yearly ({PREMIUM_PRICE_YEARLY} · 17% off)",
                  command=lambda: _open_payment(LEMONSQUEEZY_YEARLY_URL, "yearly"),
                  bg=LOCK_C, fg="white",
                  activebackground="#e5541e", activeforeground="white",
                  relief="flat", cursor="hand2",
                  font=("Malgun Gothic", 9, "bold"), padx=16, pady=6
        ).pack(side=tk.LEFT, padx=(0,8))

        # 구독 안내 라벨
        pay_info = tk.Frame(pay_body, bg=CARD)
        pay_info.pack(fill="x", padx=14, pady=(2,8))
        tk.Label(pay_info,
                 text="💡 결제 후 이메일로 라이선스 키가 발송됩니다 · 언제든 해지 가능" if _is_ko else "💡 License key sent via email after payment · Cancel anytime",
                 bg=CARD, fg=MUTED, font=("Malgun Gothic", 8)
        ).pack(side=tk.LEFT)

        sep(pay_body)

        key_row = tk.Frame(pay_body, bg=CARD)
        key_row.pack(fill="x", padx=14, pady=(6,10))
        tk.Label(key_row,
                 text="라이선스 키" if _is_ko else "License Key",
                 bg=CARD, fg=WHITE, font=("Malgun Gothic", 9),
                 width=10, anchor="w").pack(side=tk.LEFT, padx=(0,8))

        key_var = tk.StringVar(value=self.settings_data.get("neural_license_key", ""))
        key_entry = tk.Entry(key_row, textvariable=key_var, width=34,
                             bg="#0D1117", fg=WHITE,
                             insertbackground=WHITE,
                             relief="flat", font=("Courier New", 9),
                             highlightbackground=BORDER, highlightthickness=1)
        key_entry.pack(side=tk.LEFT, padx=(0,8))

        key_msg = tk.Label(key_row, text="", bg=CARD,
                           font=("Malgun Gothic", 8))
        key_msg.pack(side=tk.LEFT)

        def _apply_key():
            from binance_futures_bot1_1.binance_futures_bot.license_gate import validate_key
            raw = key_var.get().strip()
            if not raw:
                key_msg.configure(text="키를 입력하세요" if _is_ko else "Please enter a key", fg=MUTED)
                return
            ok, msg = validate_key(raw, "NEURAL", lang=self.language)
            if ok:
                self.settings_data["neural_license_key"]     = raw
                self.settings_data["neural_scorer_enabled"]  = False
                self._save_json(CONFIG_PATH, self.settings_data)
                key_msg.configure(text=f"✅ {msg}", fg=GREEN)
                _refresh_lock_ui()
            else:
                self.settings_data["neural_scorer_enabled"] = False
                self._save_json(CONFIG_PATH, self.settings_data)
                key_msg.configure(text=f"❌ {msg}", fg=RED)
                _refresh_lock_ui()

        tk.Button(key_row,
                  text="확인" if _is_ko else "Apply",
                  command=_apply_key,
                  bg=GOLD, fg=DARKBG,
                  activebackground="#d9a80a", relief="flat",
                  font=("Malgun Gothic", 9, "bold"), cursor="hand2",
                  padx=10, pady=2
        ).pack(side=tk.LEFT)

        # 키 유효 시 잠금 해제 배너
        pay_status_lbl = tk.Label(pay_body, text="", bg=CARD,
                                  font=("Malgun Gothic", 8, "bold"), anchor="center")
        pay_status_lbl.pack(fill="x", padx=14, pady=(0, 6))

        def _refresh_pay_status():
            if _has_valid_key():
                # 남은 일수 계산
                _d_remain = -1
                try:
                    from binance_futures_bot1_1.binance_futures_bot.license_gate import days_remaining
                    _k = self.settings_data.get("neural_license_key", "")
                    if _k:
                        _d_remain = days_remaining(_k)
                except Exception:
                    pass

                if _d_remain >= 0:
                    _days_txt = f"D-{_d_remain}" + ("일" if _is_ko else " days")
                    pay_verified_lbl.configure(
                        text=("✅ 인증 완료 · " if _is_ko else "✅ Verified · ") + _days_txt,
                        fg=GREEN)
                else:
                    pay_verified_lbl.configure(
                        text="✅ 인증 완료" if _is_ko else "✅ Verified",
                        fg=GREEN)

                pay_title_lbl.configure(fg=GREEN)
                pay_status_lbl.configure(
                    text="✅  라이선스 인증 완료" if _is_ko else "✅  License Verified",
                    fg=GREEN)

                # 인증 완료 시 자동으로 접기 (이미 접힌 상태가 아니면)
                if not _pay_collapsed[0]:
                    _toggle_pay_body()
                pay_collapse_btn.configure(
                    text="▼ " + ("펼치기" if _is_ko else "Expand"))
                pay_card.configure(highlightbackground=GREEN, highlightthickness=1)
            else:
                pay_verified_lbl.configure(text="")
                pay_title_lbl.configure(fg=LOCK_C)
                pay_status_lbl.configure(
                    text="🔒  라이선스 키를 입력해 주세요" if _is_ko else "🔒  Please enter a license key",
                    fg=MUTED)
                # 미인증 시 펼치기
                if _pay_collapsed[0]:
                    _toggle_pay_body()
                pay_collapse_btn.configure(text="")
                pay_card.configure(highlightbackground=LOCK_C, highlightthickness=1)
        _refresh_pay_status()

        # ══════════════════════════════════════════════════════════════════
        # SECTION 3: 학습 현황 대시보드
        # ══════════════════════════════════════════════════════════════════
        dash_card = make_card(inner, pady=(4,4))
        card_title(dash_card, "📊  학습 현황 대시보드" if _is_ko else "📊  Learning Dashboard")
        sep(dash_card)

        dash_banner = tk.Label(dash_card, text="", bg=CARD,
                               font=("Malgun Gothic", 8, "bold"),
                               anchor="center", pady=5)
        dash_banner.pack(fill="x", padx=14, pady=(4,6))

        # 2행 × 4열 스탯 카드
        grid_frame = tk.Frame(dash_card, bg=CARD)
        grid_frame.pack(fill="x", padx=14, pady=(0,10))
        for c in range(4):
            grid_frame.columnconfigure(c, weight=1)

        STAT_DEFS = [
            ("trades", "학습 거래"  if _is_ko else "Trained"),
            ("wins",   "승리"       if _is_ko else "Wins"),
            ("losses", "패배"       if _is_ko else "Losses"),
            ("wr",     "승률"       if _is_ko else "Win Rate"),
            ("acc",    "예측 정확도" if _is_ko else "Accuracy"),
            ("roi",    "평균 ROI"   if _is_ko else "Avg ROI"),
            ("replay", "Replay 버퍼" if _is_ko else "Replay Buffer"),
            ("model",  "모델 상태"  if _is_ko else "Model Status"),
        ]
        stat_vals = {}
        for i, (key, label) in enumerate(STAT_DEFS):
            r, c = divmod(i, 4)
            cell = tk.Frame(grid_frame, bg="#0D1117",
                            highlightbackground=BORDER, highlightthickness=1)
            cell.grid(row=r, column=c, padx=2, pady=2, sticky="nsew")
            tk.Label(cell, text=label, bg="#0D1117", fg=MUTED,
                     font=("Malgun Gothic", 8), anchor="center"
            ).pack(pady=(6,2))
            vl = tk.Label(cell, text="—", bg="#0D1117", fg=WHITE,
                          font=("Malgun Gothic", 11, "bold"), anchor="center")
            vl.pack(pady=(0,6))
            stat_vals[key] = vl

        # 새로고침 버튼
        ref_row = tk.Frame(dash_card, bg=CARD)
        ref_row.pack(fill="x", padx=14, pady=(0,8))
        tk.Button(ref_row,
                  text="↻  새로고침" if _is_ko else "↻  Refresh",
                  command=lambda: _refresh_dashboard(),
                  bg="#1E2329", fg=MUTED, relief="flat",
                  font=("Malgun Gothic", 8), cursor="hand2",
                  padx=10, pady=3
        ).pack(side=tk.RIGHT)

        # ── 피처 중요도 바 ───────────────────────────────────────────────
        feat_card = make_card(inner, pady=(4,4))
        card_title(feat_card,
                   "📈  피처 가중치 — 모델이 중요하게 보는 시장 조건" if _is_ko
                   else "📈  Feature Weights — What the Model Focuses On")
        sep(feat_card)

        FEAT_NAMES = (
            ["상대 모멘텀", "변동성", "거래량 서지",
             "1m EMA 기울기", "5m EMA 기울기", "MTF 정렬도",
             "스프레드", "펀딩 방향", "레짐",
             "진입 시각(sin)", "진입 시각(cos)", "방향 일치",
             "24h 가격위치", "RSI-14", "ATR 비율",
             "기울기 괴리", "펀딩 크기", "변동성 레짐",
             "열린 포지션", "요일 주기"]
            if _is_ko else
            ["Rel. Momentum", "Volatility", "Vol. Surge",
             "1m EMA Slope", "5m EMA Slope", "MTF Align",
             "Spread", "Funding", "Regime",
             "Time(sin)", "Time(cos)", "Dir Match",
             "24h Price Pos", "RSI-14", "ATR Ratio",
             "Slope Diverg.", "Funding Mag.", "Vol Regime",
             "Open Positions", "Day of Week"]
        )
        BAR_MAX = 280
        feat_bars = []
        feat_frame = tk.Frame(feat_card, bg=CARD)
        feat_frame.pack(fill="x", padx=14, pady=(4,12))
        for name in FEAT_NAMES:
            row_f = tk.Frame(feat_frame, bg=CARD)
            row_f.pack(fill="x", pady=1)
            tk.Label(row_f, text=name, bg=CARD, fg=MUTED,
                     font=("Malgun Gothic", 8), width=14, anchor="e"
            ).pack(side=tk.LEFT, padx=(0,8))
            bg_bar = tk.Frame(row_f, bg="#1E2329", width=BAR_MAX, height=12)
            bg_bar.pack(side=tk.LEFT)
            bg_bar.pack_propagate(False)
            fill = tk.Frame(bg_bar, bg=BLUE, height=12)
            fill.place(x=0, y=0, height=12, width=0)
            pct_l = tk.Label(row_f, text="—", bg=CARD, fg=MUTED,
                             font=("Malgun Gothic", 8), width=5, anchor="w")
            pct_l.pack(side=tk.LEFT, padx=(6,0))
            feat_bars.append((fill, pct_l))

        def _refresh_feat_weights():
            try:
                import json as _j, os as _o
                import numpy as _np
                W1 = None
                # 1) 실행 중 엔진에서 직접 가져오기
                try:
                    from binance_futures_bot1_1 import main as _eng_main
                    _ce = getattr(_eng_main, "current_engine", None)
                    if _ce:
                        _ns = getattr(_ce, "neural_scorer", None)
                        if _ns: W1 = _ns.net.W1
                except Exception:
                    pass
                # 2) 파일 fallback
                if W1 is None:
                    for _p in [
                        _o.path.join(BASE_DIR, "binance_futures_bot1_1", "logs", "neural_scorer.json"),
                        "binance_futures_bot1_1/logs/neural_scorer.json",
                        "logs/neural_scorer.json",
                        "bot_data/neural_scorer.json",
                    ]:
                        if _o.path.exists(_p):
                            with open(_p) as f_: d_ = _j.load(f_)
                            if "net" in d_ and "W1" in d_["net"]:
                                W1 = _np.array(d_["net"]["W1"])
                            break
                if W1 is None:
                    return
                imp = _np.abs(W1).mean(axis=0)
                mx  = imp.max()
                if mx <= 0: return
                norm = imp / mx
                top3 = sorted(range(len(norm)), key=lambda x: -norm[x])[:3]
                for i, (fill, pct_l) in enumerate(feat_bars):
                    if i >= len(norm): break
                    p   = float(norm[i])
                    w   = max(2, int(BAR_MAX * p))
                    col = GOLD if i in top3 else BLUE
                    fill.place(width=w)
                    fill.configure(bg=col)
                    pct_l.configure(text=f"{p*100:.0f}%", fg=col)
            except Exception:
                pass

        # ── 대시보드 갱신 로직 ────────────────────────────────────────────
        def _get_scorer_status():
            try:
                # 1) 실행 중인 엔진에서 직접 조회
                try:
                    from binance_futures_bot1_1 import main as _eng_main
                    _ce = getattr(_eng_main, "current_engine", None)
                    if _ce:
                        ns = getattr(_ce, "neural_scorer", None)
                        if ns: return ns.status()
                except Exception:
                    pass
                # 2) 파일 fallback
                import json as _j, os as _o
                for _p in [
                    os.path.join(BASE_DIR, "binance_futures_bot1_1", "logs", "neural_scorer.json"),
                    "binance_futures_bot1_1/logs/neural_scorer.json",
                    "logs/neural_scorer.json",
                    "bot_data/neural_scorer.json",
                ]:
                    if _o.path.exists(_p):
                        d_ = _j.load(open(_p))
                        tr  = d_.get("tracker", {})
                        recs = tr.get("records", [])
                        rois = tr.get("roi_records", [])
                        rb  = d_.get("replay", {})
                        return {
                            "version":   d_.get("version", "2.0"),
                            "n_features": d_.get("n_features", 12),
                            "n_trained": d_.get("n_trained", 0),
                            "n_wins":    d_.get("n_wins",    0),
                            "n_losses":  d_.get("n_losses",  0),
                            "win_rate":  round(d_.get("n_wins",0)/max(d_.get("n_trained",1),1)*100,1),
                            "accuracy":  round(sum(recs)/len(recs)*100,1) if recs else 0.0,
                            "avg_roi":   round(sum(rois)/len(rois),2)     if rois else 0.0,
                            "replay_n":  len(rb.get("wins",[]))+len(rb.get("losses",[])),
                            "active":    d_.get("active", False),
                            "ready":     d_.get("n_trained",0) >= 50,
                            "lr":        d_.get("lr", 0.002),
                            "block_threshold": d_.get("block_threshold", 0.25),
                        }
            except Exception:
                pass
            return {"version":"3.0","n_features":20,"n_trained":0,"n_wins":0,"n_losses":0,"win_rate":0,
                    "accuracy":0,"avg_roi":0,"replay_n":0,"active":False,"ready":False,
                    "lr":0.002,"block_threshold":0.25}

        def _refresh_dashboard():
            st      = _get_scorer_status()
            enabled = bool(self.settings_data.get("neural_scorer_enabled", False))
            has_key = _has_valid_key()
            n       = st["n_trained"]

            if not has_key:
                dash_banner.configure(
                    text="🔒  라이선스 키 미등록 — 데이터 수집 중 (예측 미적용)" if _is_ko
                    else "🔒  No license key — collecting data (prediction inactive)",
                    bg=CARD, fg=MUTED)
            elif not enabled:
                dash_banner.configure(
                    text="○  예측 OFF — 데이터 수집·학습만 진행 중" if _is_ko
                    else "○  Prediction OFF — collecting & learning data only",
                    bg=CARD, fg=MUTED)
            elif not st["ready"]:
                dash_banner.configure(
                    text=f"⏳  냉각 기간 중 ({n}/50건) — 50건 이후 예측 시작" if _is_ko
                    else f"⏳  Warmup period ({n}/50 trades) — prediction starts after 50",
                    bg="#1a1500", fg=GOLD)
            elif not st["active"]:
                dash_banner.configure(
                    text="⚠️  예측 정확도 미달 — 자동 비활성화 (학습 계속)" if _is_ko
                    else "⚠️  Low accuracy — prediction auto-disabled (learning continues)",
                    bg="#1a0000", fg=RED)
            else:
                _ver = st.get("version", "2.0")
                _lr_s = f"{st.get('lr',0.002):.4f}"
                dash_banner.configure(
                    text=f"✅  v{_ver} 예측 활성 | 학습 {n}건 | 정확도 {st['accuracy']}% | LR {_lr_s}" if _is_ko
                    else f"✅  v{_ver} Active | {n} trades | Acc {st['accuracy']}% | LR {_lr_s}",
                    bg="#001a00", fg=GREEN)

            wr  = st["win_rate"]
            acc = st["accuracy"]
            roi = st["avg_roi"]
            stat_vals["trades"].configure(text=str(n))
            stat_vals["wins"].configure(  text=str(st["n_wins"]),   fg=GREEN if st["n_wins"]>0   else WHITE)
            stat_vals["losses"].configure(text=str(st["n_losses"]), fg=RED   if st["n_losses"]>0 else WHITE)
            stat_vals["wr"].configure(    text=f"{wr:.1f}%",  fg=GREEN if wr>=50  else RED)
            stat_vals["acc"].configure(   text=f"{acc:.1f}%", fg=GREEN if acc>=52 else (GOLD if acc>=48 else RED))
            stat_vals["roi"].configure(   text=f"{roi:+.2f}%",fg=GREEN if roi>0   else (RED if roi<0 else WHITE))
            stat_vals["replay"].configure(text=str(st["replay_n"]))
            if not has_key:
                stat_vals["model"].configure(text="🔒 미등록" if _is_ko else "🔒 Unregistered", fg=MUTED)
            elif not enabled:
                stat_vals["model"].configure(text="○ 수집 중" if _is_ko else "○ Collecting",   fg=MUTED)
            elif not st["ready"]:
                stat_vals["model"].configure(text=f"⏳ {n}/50",                                 fg=GOLD)
            elif not st["active"]:
                stat_vals["model"].configure(text="⚠️ 중단"   if _is_ko else "⚠️ Paused",       fg=RED)
            else:
                stat_vals["model"].configure(text="✅ 예측 중" if _is_ko else "✅ Predicting",   fg=GREEN)

            _refresh_feat_weights()

        # ══════════════════════════════════════════════════════════════════
        # SECTION 4: 작동 방식 설명
        # ══════════════════════════════════════════════════════════════════
        how_card = make_card(inner, pady=(4,4))
        card_title(how_card, "⚙️  작동 방식" if _is_ko else "⚙️  How It Works")
        sep(how_card)
        HOW_ITEMS = (
            [
                ("📥  진입 시 자동 수집 (항상 동작)",
                 "20개 시장 특성(모멘텀·변동성·거래량·EMA기울기·스프레드·펀딩·레짐·시각·\n"
                 "RSI·ATR·24h가격위치·변동성레짐·포지션수·요일주기 등)을\n"
                 "매 진입 시도 시 자동으로 기록합니다. ON/OFF와 무관하게 항상 실행됩니다."),
                ("🎓  청산 후 자동 학습 — 듀얼헤드 (항상 동작)",
                 "포지션 청산 시 승/패 + 실제 ROI%를 동시에 학습합니다.\n"
                 "분류 헤드: 승률 예측 (Binary CE loss)\n"
                 "회귀 헤드: 기대 ROI% 예측 (Huber loss)\n"
                 "Experience Replay + 적응형 학습률로 효율적 학습."),
                ("🎯  v3 신뢰도 기반 진입 게이팅 (ON일 때만)",
                 "50건 이상 학습 후, 승률 + 기대수익률을 동시 계산합니다.\n"
                 "승률 < 25% → ❌ 진입 완전 차단 (하드 블록)\n"
                 "승률 25~50% → strength ×0.0~1.0 (약한 진입)\n"
                 "승률 50%+ → strength ×1.0~2.0 (강한 진입)\n"
                 "기대 ROI가 음수면 승률 높아도 배율 감소."),
                ("🔒  ON/OFF 차이 요약",
                 "OFF: 데이터 수집 + 학습 O,  예측 반영 X  ← 데이터만 모이는 상태\n"
                 "ON:  데이터 수집 + 학습 O,  예측 반영 O  ← 실제로 진입에 영향을 주는 상태"),
            ] if _is_ko else [
                ("📥  Auto-Collect at Entry (always active)",
                 "20 market features (momentum, volatility, volume, EMA slope, spread, funding, regime,\n"
                 "time, RSI, ATR, 24h price position, vol regime, open positions, day-of-week, etc.)\n"
                 "are recorded at every entry attempt, regardless of ON/OFF state."),
                ("🎓  Dual-Head Auto-Learn after Close (always active)",
                 "Win/Loss + actual ROI% are learned simultaneously.\n"
                 "Classification head: win probability (Binary CE loss)\n"
                 "Regression head: expected ROI% (Huber loss)\n"
                 "Experience Replay + adaptive learning rate for efficient training."),
                ("🎯  v3 Confidence-Based Entry Gating (ON only)",
                 "After 50+ trades, win prob + expected ROI are calculated.\n"
                 "Win prob < 25% → entry BLOCKED (hard reject)\n"
                 "Win prob 25-50% → strength ×0.0–1.0 (weak entry)\n"
                 "Win prob 50%+   → strength ×1.0–2.0 (strong entry)\n"
                 "Negative expected ROI reduces multiplier even with high win prob."),
                ("🔒  ON vs OFF Summary",
                 "OFF: Data collection + learning  ON,  Prediction applied  OFF\n"
                 "ON:  Data collection + learning  ON,  Prediction applied  ON"),
            ]
        )
        for title, desc in HOW_ITEMS:
            tk.Label(how_card, text=title, bg=CARD, fg=GOLD,
                     font=("Malgun Gothic", 9, "bold"), anchor="w"
            ).pack(fill="x", padx=16, pady=(6,1))
            tk.Label(how_card, text=desc, bg=CARD, fg=MUTED,
                     font=("Malgun Gothic", 8), anchor="w",
                     justify="left", wraplength=555
            ).pack(fill="x", padx=30, pady=(0,2))
        tk.Frame(how_card, bg=CARD, height=6).pack()

        # ══════════════════════════════════════════════════════════════════
        # SECTION 5: 모델 초기화 + 서버 로드맵
        # ══════════════════════════════════════════════════════════════════
        ctrl_card = make_card(inner, pady=(4,4))
        crow = tk.Frame(ctrl_card, bg=CARD)
        crow.pack(fill="x", padx=14, pady=10)

        def _reset_model():
            if not self._show_yesno(
                    "모델 초기화" if _is_ko else "Reset Model",
                    "학습된 가중치와 Replay 버퍼를 모두 삭제합니다.\n계속하시겠습니까?"
                    if _is_ko else
                    "All learned weights and the Replay Buffer will be deleted.\nDo you want to continue?"):
                return
            try:
                from binance_futures_bot1_1 import main as _eng_main
                _ce = getattr(_eng_main, "current_engine", None)
                if _ce:
                    _ns = getattr(_ce, "neural_scorer", None)
                    if _ns: _ns.reset()
            except Exception: pass
            import os as _o
            for p in [
                os.path.join(BASE_DIR, "binance_futures_bot1_1", "logs", "neural_scorer.json"),
                "binance_futures_bot1_1/logs/neural_scorer.json",
                "logs/neural_scorer.json",
                "bot_data/neural_scorer.json",
            ]:
                try:
                    if _o.path.exists(p): _o.remove(p)
                except Exception: pass
            _refresh_dashboard()

        tk.Button(crow,
                  text="🗑  모델 초기화" if _is_ko else "🗑  Reset Model",
                  command=_reset_model,
                  bg="#1E2329", fg=RED, relief="flat",
                  font=("Malgun Gothic", 9), cursor="hand2",
                  padx=12, pady=4).pack(side=tk.LEFT)
        tk.Label(crow,
                 text="학습된 가중치와 Replay 버퍼 전체를 삭제합니다" if _is_ko
                 else "Deletes all learned weights and the Replay Buffer",
                 bg=CARD, fg=MUTED, font=("Malgun Gothic", 8)
        ).pack(side=tk.LEFT, padx=12)

        # 서버 로드맵 배너
        future = tk.Frame(inner, bg="#080c18",
                          highlightbackground=BLUE, highlightthickness=1)
        future.pack(fill="x", padx=24, pady=(4,28))
        ftop = tk.Frame(future, bg="#080c18")
        ftop.pack(fill="x", padx=16, pady=(12,4))
        tk.Label(ftop,
                 text="🌐  서버 연동 딥러닝  —  출시 예정" if _is_ko
                 else "🌐  Server-Side Deep Learning  —  Coming Soon",
                 bg="#080c18", fg=BLUE,
                 font=("Malgun Gothic", 9, "bold"), anchor="w").pack(side=tk.LEFT)
        tk.Label(ftop, text=" COMING SOON ",
                 bg="#0a1525", fg=BLUE,
                 font=("Malgun Gothic", 7, "bold"), padx=4).pack(side=tk.LEFT, padx=8)
        ROAD = (
            [
                "📡  클라이언트 학습 데이터 익명 수집 → 중앙 서버 집계",
                "🧬  연합 학습(Federated Learning)으로 개인 정보 노출 없이 통합 학습",
                "🔥  수만 건 기반 LSTM / Transformer 딥러닝 모델 구축 및 배포",
                "⚡  배포된 서버 모델 위에 개인 로컬 fine-tune 적용",
            ] if _is_ko else [
                "📡  Anonymized client learning data aggregated on a central server",
                "🧬  Federated Learning — collective training without exposing personal data",
                "🔥  LSTM / Transformer deep learning model trained on tens of thousands of trades",
                "⚡  Server model deployed and fine-tuned locally on each client",
            ]
        )
        for r in ROAD:
            tk.Label(future, text=r, bg="#080c18", fg=MUTED,
                     font=("Malgun Gothic", 8), anchor="w"
            ).pack(fill="x", padx=20, pady=1)
        tk.Frame(future, bg="#080c18", height=12).pack()

        # ── 초기 렌더링 ───────────────────────────────────────────────────
        _refresh_dashboard()
        def _auto():
            try:
                _refresh_dashboard()
                inner.after(30000, _auto)
            except Exception: pass
        inner.after(5000, _auto)

    # ══════════════════════════════════════════════════════════════════
    #  AI 어시스턴트 창 (프리미엄)
    # ══════════════════════════════════════════════════════════════════
    def _create_ai_assistant_window(self):
        """실시간 시장 분석 & 자가 개선 AI 어시스턴트 Toplevel 창 생성."""
        # ── 색상 & 상수 ──
        _BG      = "#0f1118"
        _CARD    = "#10131d"
        _BORDER  = "#343942"
        _GOLD    = "#F0B90B"
        _GREEN   = "#2ECC71"
        _RED     = "#FF6B6B"
        _BLUE    = "#4A90D9"
        _MUTED   = "#8e96b8"
        _WHITE   = "#f5f7ff"
        _DARKBG  = "#0a0d14"
        _is_ko   = self.language == "ko"
        _W, _H   = 440, 680

        win = tk.Toplevel(self.root)
        win.title("🧠 AI Trading Assistant")
        self._apply_icon(win)
        win.configure(bg=_BG)
        win.geometry(f"{_W}x{_H}")
        win.minsize(360, 500)
        # 메인 창 오른쪽에 배치
        try:
            rx = self.root.winfo_rootx() + self.root.winfo_width() + 8
            ry = self.root.winfo_rooty()
            win.geometry(f"{_W}x{_H}+{rx}+{ry}")
        except Exception:
            pass
        win.attributes("-topmost", True)
        win.after(100, lambda: win.attributes("-topmost", False))  # 초기 포커스 후 해제

        self._ai_assistant_win = win
        self._ai_assistant_refresh_job = None

        # ── AIAdvisor 초기화 ──
        try:
            from binance_futures_bot1_1.binance_futures_bot.ai_advisor import AIAdvisor
            advisor = AIAdvisor(BASE_DIR, language=self.language)
        except Exception as _e:
            logging.warning("AIAdvisor init error: %s", _e)
            advisor = None

        # ═════════════════════════════════════════════════════════════
        # 타이틀 바
        # ═════════════════════════════════════════════════════════════
        title_bar = tk.Frame(win, bg=_DARKBG)
        title_bar.pack(fill="x")
        tk.Label(title_bar,
                 text="🧠  AI Trading Assistant" if not _is_ko else "🧠  AI 트레이딩 어시스턴트",
                 bg=_DARKBG, fg=_GOLD,
                 font=("Malgun Gothic", 11, "bold"),
                 padx=12, pady=8
        ).pack(side=tk.LEFT)
        tk.Label(title_bar, text="PREMIUM", bg=_GOLD, fg=_DARKBG,
                 font=("Malgun Gothic", 7, "bold"), padx=4, pady=1
        ).pack(side=tk.LEFT, padx=4)

        # ═════════════════════════════════════════════════════════════
        # 시장 요약 카드
        # ═════════════════════════════════════════════════════════════
        summary_frame = tk.Frame(win, bg=_CARD, highlightbackground=_BORDER, highlightthickness=1)
        summary_frame.pack(fill="x", padx=10, pady=(6, 4))
        summary_label = tk.Label(summary_frame,
                                 text="📊 " + ("시장 분석 로딩 중..." if _is_ko else "Loading market analysis..."),
                                 bg=_CARD, fg=_MUTED,
                                 font=("Malgun Gothic", 9),
                                 anchor="w", padx=10, pady=8,
                                 wraplength=_W - 40, justify="left")
        summary_label.pack(fill="x")

        # ═════════════════════════════════════════════════════════════
        # 실시간 메시지 피드 (스크롤)
        # ═════════════════════════════════════════════════════════════
        feed_label = tk.Label(win,
                              text="  " + ("실시간 활동" if _is_ko else "Live Activity"),
                              bg=_BG, fg=_GOLD,
                              font=("Malgun Gothic", 9, "bold"), anchor="w")
        feed_label.pack(fill="x", padx=10, pady=(6, 2))

        feed_container = tk.Frame(win, bg=_DARKBG, highlightbackground=_BORDER, highlightthickness=1)
        feed_container.pack(fill="both", expand=True, padx=10, pady=(0, 4))

        feed_canvas = tk.Canvas(feed_container, bg=_DARKBG, highlightthickness=0, bd=0)
        feed_scrollbar = ttk.Scrollbar(feed_container, orient="vertical", command=feed_canvas.yview)
        feed_inner = tk.Frame(feed_canvas, bg=_DARKBG)

        feed_canvas.configure(yscrollcommand=feed_scrollbar.set)
        feed_scrollbar.pack(side=tk.RIGHT, fill="y")
        feed_canvas.pack(side=tk.LEFT, fill="both", expand=True)
        _feed_win_id = feed_canvas.create_window((0, 0), window=feed_inner, anchor="nw")

        def _on_feed_configure(_e=None):
            feed_canvas.configure(scrollregion=feed_canvas.bbox("all"))
        feed_inner.bind("<Configure>", _on_feed_configure)
        def _on_feed_canvas_configure(e):
            feed_canvas.itemconfig(_feed_win_id, width=e.width)
        feed_canvas.bind("<Configure>", _on_feed_canvas_configure)

        # 마우스 휠 스크롤
        def _on_mousewheel(e):
            feed_canvas.yview_scroll(int(-1 * (e.delta / 120)), "units")
        feed_canvas.bind_all("<MouseWheel>", _on_mousewheel)

        _auto_scroll = [True]
        _msg_widgets = []
        _MAX_MSGS = 200

        def _append_feed_message(icon, color, ts_str, text):
            """메시지 피드에 새 항목 추가."""
            if len(_msg_widgets) >= _MAX_MSGS:
                oldest = _msg_widgets.pop(0)
                try:
                    oldest.destroy()
                except Exception:
                    pass

            row = tk.Frame(feed_inner, bg=_DARKBG)
            row.pack(fill="x", padx=4, pady=1)
            # 시간
            time_part = ts_str.split(" ")[-1] if " " in ts_str else ts_str
            if len(time_part) > 5:
                time_part = time_part[:5]
            tk.Label(row, text=time_part, bg=_DARKBG, fg=_MUTED,
                     font=("Malgun Gothic", 7), width=5, anchor="e"
            ).pack(side=tk.LEFT, padx=(2, 4))
            # 아이콘
            tk.Label(row, text=icon, bg=_DARKBG, fg=color,
                     font=("Malgun Gothic", 9)
            ).pack(side=tk.LEFT, padx=(0, 4))
            # 메시지 (여러 줄 가능)
            tk.Label(row, text=text, bg=_DARKBG, fg=_WHITE,
                     font=("Malgun Gothic", 8), anchor="w",
                     justify="left", wraplength=_W - 120
            ).pack(side=tk.LEFT, fill="x", expand=True)

            _msg_widgets.append(row)

            if _auto_scroll[0]:
                feed_canvas.update_idletasks()
                feed_canvas.yview_moveto(1.0)

        # ═════════════════════════════════════════════════════════════
        # AI 분석 패널 (하단)
        # ═════════════════════════════════════════════════════════════
        insight_label = tk.Label(win,
                                 text="  💡 " + ("AI 분석 & 개선 제안" if _is_ko else "AI Analysis & Improvement"),
                                 bg=_BG, fg=_GOLD,
                                 font=("Malgun Gothic", 9, "bold"), anchor="w")
        insight_label.pack(fill="x", padx=10, pady=(6, 2))

        insight_frame = tk.Frame(win, bg=_CARD, highlightbackground=_GOLD, highlightthickness=1)
        insight_frame.pack(fill="x", padx=10, pady=(0, 10))

        insight_text = tk.Label(insight_frame,
                                text=("아직 충분한 거래 데이터가 없습니다.\n거래가 쌓이면 AI가 패턴을 분석하고 개선 방안을 제안합니다."
                                      if _is_ko else
                                      "Not enough trade data yet.\nOnce trades accumulate, AI will analyze patterns and suggest improvements."),
                                bg=_CARD, fg=_MUTED,
                                font=("Malgun Gothic", 8),
                                anchor="w", justify="left",
                                wraplength=_W - 50,
                                padx=12, pady=10)
        insight_text.pack(fill="x")

        # ═════════════════════════════════════════════════════════════
        # 주기적 갱신 루프
        # ═════════════════════════════════════════════════════════════
        _last_summary_ts = [0.0]
        _last_insight_ts = [0.0]

        def _refresh_ai():
            """AI 어시스턴트 갱신 (3초 간격)."""
            if not win.winfo_exists():
                return
            try:
                if not advisor:
                    return

                # 1) 새 이벤트 폴링 & 메시지 피드
                new_events = advisor.poll_new_events()
                for ev in new_events:
                    _append_feed_message(
                        ev.get("icon", "ℹ"),
                        ev.get("color", _MUTED),
                        ev.get("ts_str", ""),
                        ev.get("friendly_msg", ev.get("raw_msg", "")),
                    )

                now = time.time()

                # 2) 시장 요약 (60초 간격)
                if now - _last_summary_ts[0] > 60:
                    try:
                        stxt = advisor.get_market_summary_text()
                        summary_label.configure(text="📊 " + stxt, fg=_WHITE)
                    except Exception:
                        pass
                    _last_summary_ts[0] = now

                # 3) AI 분석 & 개선 제안 (30초 간격)
                if now - _last_insight_ts[0] > 30:
                    try:
                        suggestions = advisor.get_improvement_suggestions()
                        if suggestions:
                            lines = []
                            for s in suggestions[:3]:
                                _msg = s.get("msg_ko" if _is_ko else "msg_en", "")
                                _icon = s.get("icon", "💡")
                                _pri = s.get("priority", "low")
                                lines.append(f"{_icon}  {_msg}")
                            insight_text.configure(
                                text="\n\n".join(lines),
                                fg=_WHITE if any(s["priority"] == "high" for s in suggestions) else _MUTED,
                            )
                        else:
                            patterns = advisor.get_trade_patterns()
                            if patterns["total"] > 0:
                                _summary = (
                                    f"최근 {patterns['total']}건 분석 완료.\n평균 ROI: {patterns['avg_roi']:+.2f}%\n"
                                    f"현재 특이사항 없음 — 전략이 안정적입니다."
                                    if _is_ko else
                                    f"Analyzed {patterns['total']} recent trades.\nAvg ROI: {patterns['avg_roi']:+.2f}%\n"
                                    f"No issues detected — strategy is stable."
                                )
                                insight_text.configure(text=f"📈  {_summary}", fg=_GREEN)
                    except Exception:
                        pass
                    _last_insight_ts[0] = now

            except Exception as _e:
                logger.debug("AI assistant refresh error: %s", _e)
            finally:
                if win.winfo_exists():
                    self._ai_assistant_refresh_job = win.after(3000, _refresh_ai)

        # 시작 메시지
        _now_str = time.strftime("%H:%M")
        _append_feed_message("🧠", _GOLD, _now_str,
                             "AI 어시스턴트가 시작되었습니다. 시장을 모니터링합니다." if _is_ko
                             else "AI Assistant started. Monitoring market activity.")

        # 창 닫기 처리 — 숨기기(withdraw)로 백그라운드 유지
        def _on_close():
            try:
                feed_canvas.unbind_all("<MouseWheel>")
            except Exception:
                pass
            # destroy 대신 withdraw → 백그라운드에서 계속 학습/분석
            win.withdraw()

        def _on_destroy():
            """실제 파괴 시에만 (앱 종료 등) 리소스 정리."""
            if self._ai_assistant_refresh_job:
                try:
                    win.after_cancel(self._ai_assistant_refresh_job)
                except Exception:
                    pass
            self._ai_assistant_win = None

        win.protocol("WM_DELETE_WINDOW", _on_close)
        win.bind("<Destroy>", lambda _e: _on_destroy() if _e.widget is win else None)

        # 첫 갱신 시작
        _refresh_ai()

    # ------------------------------------------------------------------
    def _build_engine_config(self):
        try:
            from binance_futures_bot1_1.binance_futures_bot.config import EngineConfig
        except Exception:
            return None
        spike_guard_enabled = bool(self.settings_data.get("spike_guard_enabled", True))
        spike_guard_pct = max(0.0, float(self.settings_data.get("spike_guard_pct", 5.0))) / 100.0  # fallback 5%
        spike_guard_window = max(1, int(self.settings_data.get("spike_guard_window", 8)))
        spike_guard_interval = max(1, int(self.settings_data.get("spike_guard_interval", 2)))
        if not spike_guard_enabled:
            spike_guard_pct = 0.0

        # [PATCH-4] 포지션 크기 AutoTuner 범위 유효성 검증
        _raw_pos_pct = float(self.settings_data.get("position_pct", 0.05))
        _tune_mode = str(self.settings_data.get("auto_tune_mode", "balanced"))
        _pos_bounds = {"aggressive": (0.03, 0.18), "balanced": (0.02, 0.12), "conservative": (0.02, 0.08)}
        _lo, _hi = _pos_bounds.get(_tune_mode, (0.02, 0.12))
        if self.settings_data.get("auto_tune_enabled", False) and _raw_pos_pct < _lo:
            logging.warning(f"[PATCH-4] position_pct {_raw_pos_pct:.4f}이 {_tune_mode} 모드 최소값 {_lo:.3f}보다 낮아 {_lo:.3f}로 조정합니다.")
            _raw_pos_pct = _lo
        elif self.settings_data.get("auto_tune_enabled", False) and _raw_pos_pct > _hi:
            logging.warning(f"[PATCH-4] position_pct {_raw_pos_pct:.4f}이 {_tune_mode} 모드 최대값 {_hi:.3f}을 초과하여 {_hi:.3f}로 조정합니다.")
            _raw_pos_pct = _hi

        return EngineConfig(
            top_n=int(self.settings_data.get("top_n", 20)),
            position_pct=_raw_pos_pct,
            leverage_min=int(self.settings_data.get("leverage_min", 5)),
            leverage_max=int(self.settings_data.get("leverage_max", 25)),
            volatility_min=float(self.settings_data.get("volatility_min", 0.002)),
            momentum_min_long=float(self.settings_data.get("momentum_min_long", 0.002)),
            momentum_min_short=float(self.settings_data.get("momentum_min_short", -0.002)),
            momentum_min=float(self.settings_data.get("momentum_min_long", 0.002)),
            auto_tune_enabled=bool(self.settings_data.get("auto_tune_enabled", False)),
            auto_tune_mode=str(self.settings_data.get("auto_tune_mode", "balanced")),
            total_risk_budget=float(self.settings_data.get("position_pct", 0.05)),
            watch_limit=int(self.settings_data.get("watch_limit", 10)),
            max_open_symbols=int(self.settings_data.get("max_open_symbols", 10)),
            # [수정] 기본값 55.0 → 18.0: config.py·defaults와 통일. 55%는 레버리지 환경에서 치명적 손실 허용.
            max_loss_per_position=float(self.settings_data.get("max_loss_per_position", 18.0)),
            spike_guard_enabled=bool(self.settings_data.get("spike_guard_enabled", True)),
            spike_guard_return_pct=spike_guard_pct,
            spike_guard_window=spike_guard_window,
            spike_guard_check_interval_s=spike_guard_interval,
            global_spike_cooldown_min=int(self.settings_data.get("global_spike_cooldown_min", 5)),
            spark_reentry_candles=int(self.settings_data.get("spark_reentry_candles", 3)),
            session_loss_limit_pct=float(self.settings_data.get("session_loss_limit_pct", 3.0)),
            session_loss_window_minutes=int(self.settings_data.get("session_loss_window_minutes", 1440)),
            kill_switch_cooldown_min=int(self.settings_data.get("kill_switch_cooldown_min", 30)),
            auto_tune_cooldown_min=int(self.settings_data.get("auto_tune_cooldown_min", 10)),
            auto_boost_position_pct=bool(self.settings_data.get("auto_boost_position_pct", False)),
            maker_fee_pct=float(self.settings_data.get("maker_fee_pct", 0.0002)),
            taker_fee_pct=float(self.settings_data.get("taker_fee_pct", 0.0005)),
            # ── 이하: 지난 세션에서 tick_engine/config에 추가된 신규 파라미터 ──────
            # [추가] GUI 설정이 엔진에 전달되도록 연결. 미포함 시 config.py 기본값만 사용됨.
            enable_profit_exit_layer=bool(self.settings_data.get("enable_profit_exit_layer", True)),
            enable_partial_take_profit=bool(self.settings_data.get("enable_partial_take_profit", True)),
            enable_atr_trailing_stop=bool(self.settings_data.get("enable_atr_trailing_stop", True)),
            enable_progress_stop=bool(self.settings_data.get("enable_progress_stop", True)),
            trail_atr_period=int(self.settings_data.get("trail_atr_period", 22)),
            trail_atr_mult=float(self.settings_data.get("trail_atr_mult", 3.0)),
            trail_activate_pnl_pct=float(self.settings_data.get("trail_activate_pnl_pct", 0.03)),
            trail_use_highest_since_entry=bool(self.settings_data.get("trail_use_highest_since_entry", True)),
            trail_recalc_interval_sec=int(self.settings_data.get("trail_recalc_interval_sec", 5)),
            progress_stop_lookback_sec=int(self.settings_data.get("progress_stop_lookback_sec", 600)),
            progress_stop_no_new_high_sec=int(self.settings_data.get("progress_stop_no_new_high_sec", 300)),
            progress_stop_drawdown_from_mfe=float(self.settings_data.get("progress_stop_drawdown_from_mfe", 0.07)),
            progress_stop_min_pnl_pct=float(self.settings_data.get("progress_stop_min_pnl_pct", 0.05)),
            progress_stop_action=str(self.settings_data.get("progress_stop_action", "partial_or_full")),
            # [추가] ATR 리스크 사이징: 포지션당 계좌의 entry_risk_pct% 이상 손실 방지
            atr_risk_sizing_enabled=bool(self.settings_data.get("atr_risk_sizing_enabled", True)),
            entry_risk_pct=float(self.settings_data.get("entry_risk_pct", 0.01)),
            min_margin_usdt=max(1.0, float(self.settings_data.get("min_margin_usdt", 1.0))),  # [PATCH-6f] 최소 1.0 USDT 보장
            # [추가] RSI 과열 필터 (틱 기반 모멘텀 상한): 추세 끝자락 진입 차단
            rsi_filter_enabled=bool(self.settings_data.get("rsi_filter_enabled", False)),
            rsi_overbought=float(self.settings_data.get("rsi_overbought", 75.0)),
            rsi_oversold=float(self.settings_data.get("rsi_oversold", 25.0)),
            # [추가] 복합 신호 스코어링: 모멘텀·거래량서지·MTF 가중 합산으로 신호 품질 향상
            composite_signal_enabled=bool(self.settings_data.get("composite_signal_enabled", True)),
            composite_min_score=float(self.settings_data.get("composite_min_score", 1.2)),
            # [추가] Breakeven stop: 첫 TP 발동 후 손익분기점으로 스탑 자동 이동
            breakeven_stop_enabled=bool(self.settings_data.get("breakeven_stop_enabled", True)),
            breakeven_buffer_pct=float(self.settings_data.get("breakeven_buffer_pct", 0.001)),
            # [추가] Kelly 사이징: 승률 기반 동적 포지션 비율 (Quarter-Kelly)
            kelly_sizing_enabled=bool(self.settings_data.get("kelly_sizing_enabled", True)),
            kelly_fraction=float(self.settings_data.get("kelly_fraction", 0.25)),
            kelly_min_samples=int(self.settings_data.get("kelly_min_samples", 10)),
            # [추가] 펀딩레이트 편향 필터: 과도한 포지션 쏠림 시 신호 강도 감소
            funding_filter_enabled=bool(self.settings_data.get("funding_filter_enabled", True)),
            funding_bias_threshold=float(self.settings_data.get("funding_bias_threshold", 0.001)),
            funding_bias_penalty=float(self.settings_data.get("funding_bias_penalty", 0.30)),
            # ── 진입 필터 (방향 검증 / 단기 EMA) ─────────────────────────────────
            enable_mtf_ema_confirm=bool(self.settings_data.get("enable_mtf_ema_confirm", True)),
            short_ema_conflict_filter=bool(self.settings_data.get("short_ema_conflict_filter", True)),
            chop_use_short_ema_direction=bool(self.settings_data.get("chop_use_short_ema_direction", True)),
            # ── 진입 시간 보호 / Signal Decay ────────────────────────────────────
            min_hold_seconds=int(self.settings_data.get("min_hold_seconds", 180)),
            time_stop_seconds=int(self.settings_data.get("time_stop_seconds", 1800)),
            signal_decay_threshold=float(self.settings_data.get("signal_decay_threshold", 0.25)),
            signal_decay_min_profit=float(self.settings_data.get("signal_decay_min_profit", 2.0)),
            sl_atr_mult=float(self.settings_data.get("sl_atr_mult", 0.7)),
            partial_tp_levels=self.settings_data.get("partial_tp_levels", None) or [
                {"r": 0.7, "close_frac": 0.35},
                {"r": 1.2, "close_frac": 0.35},
                {"r": 2.0, "close_frac": 1.00},
            ],
            # ── Maker-first / 워치리스트 다양성 ──────────────────────────────────
            maker_first_enabled=bool(self.settings_data.get("maker_first_enabled", True)),
            diversify_watchlist=bool(self.settings_data.get("diversify_watchlist", False)),
            ui_language=self.language,   # "ko" or "en" — engine log language
            expert_mode_enabled=bool(self.settings_data.get("expert_mode_enabled", False)),  # D: non-expert leverage cap
            neural_scorer_enabled=bool(self.settings_data.get("neural_scorer_enabled", False)),
            neural_license_key=str(self.settings_data.get("neural_license_key", "")),
            # A3: Fail-Closed — engine blocks all entries until consent is verified
            consent_verified=bool(self.risk_acknowledged),
            # C2/C3/E: commercial safety config
            startup_grace_sec=int(self.settings_data.get("startup_grace_sec", 60)),
            entry_slippage_cap_bps=float(self.settings_data.get("entry_slippage_cap_bps", 20.0)),
            max_consecutive_rollbacks=int(self.settings_data.get("max_consecutive_rollbacks", 5)),
        )

    def _get_api_credentials(self):
        self._refresh_system_env()
        env_specific = {
            "TESTNET": ("TESTNET_API_KEY", "TESTNET_API_SECRET"),
            "LIVE": ("BINANCE_API_KEY", "BINANCE_API_SECRET"),
        }
        default_pair = ("BINANCE_API_KEY", "BINANCE_API_SECRET")
        mode = self.env_mode
        env_pair = env_specific.get(mode, (None, None))
        env_key_name, env_secret_name = env_pair
        meta = {"api_key_source": None, "api_secret_source": None}

        def pick_env_value(candidates, meta_key):
            for name in candidates:
                if name and os.environ.get(name):
                    meta[meta_key] = f"환경변수({name})"
                    return os.environ.get(name)
            return ""

        api_key = pick_env_value(list(filter(None, [env_key_name, default_pair[0]])), "api_key_source")
        api_secret = pick_env_value(list(filter(None, [env_secret_name, default_pair[1]])), "api_secret_source")

        def record_settings_source(field):
            # [수정] "환경설정 입력값" → "환경변수 미등록" 으로 변경.
            # 사유: API Key 입력창 제거 후 settings_data 폴백이 없으므로
            #       환경변수에 없으면 미등록 상태임을 명확히 표시.
            if field == "api_key_source":
                meta[field] = meta[field] or ("환경변수 미등록(TESTNET_API_KEY)" if mode == "TESTNET" else "환경변수 미등록(BINANCE_API_KEY)")
            else:
                meta[field] = meta[field] or ("환경변수 미등록(TESTNET_API_SECRET)" if mode == "TESTNET" else "환경변수 미등록(BINANCE_API_SECRET)")

        # [제거] API 입력창 폐지로 settings_data 내 api_key/api_secret 폴백 불필요.
        # 환경변수에서 키를 못 찾으면 빈 문자열 → _require_credentials에서 오류 안내.
        # 아래 코드는 보안상 평문 저장 위험이 있어 의도적으로 비활성화.
        # if not api_key or not api_secret:
        #     if mode == "TESTNET":
        #         api_key = api_key or self.settings_data.get("api_key", "").strip()
        #         api_secret = api_secret or self.settings_data.get("api_secret", "").strip()
        #     else:
        #         api_key = api_key or self.settings_data.get("fallback_api_key", "").strip()
        #         api_secret = api_secret or self.settings_data.get("fallback_api_secret", "").strip()

        record_settings_source("api_key_source")
        record_settings_source("api_secret_source")
        return {"api_key": api_key, "api_secret": api_secret, "testnet": mode == "TESTNET", "meta": meta}

    def _require_credentials(self, *, silent=False, action=""):
        creds = self._get_api_credentials()
        issues_missing = []
        issues_invalid = []

        for field in ("api_key", "api_secret"):
            value = (creds.get(field) or "").strip()
            if not value:
                issues_missing.append(field)
            elif len(value) < 16 or not value.isascii():
                issues_invalid.append(field)

        if issues_missing:
            if not silent:
                self._show_api_error(creds, "missing", issues_missing, action=action)
                # [제거] self.open_settings_modal() 호출 삭제.
                # 사유: API Key 입력창이 GUI에서 제거됨. 설정 모달을 열어도
                #       입력할 곳이 없으므로 오류 안내만 표시하고 종료.
            self._append_log(
                "[WARN] " + ("API credentials missing. Check environment variables." if self.language == "en"
                             else "API 자격 증명 부족으로 작업이 중단되었습니다. 환경변수를 확인하세요."))
            return None
        if issues_invalid:
            if not silent:
                self._show_api_error(creds, "invalid", issues_invalid, action=action)
            self._append_log(
                "[WARN] " + ("API credential format error. Check environment variable values." if self.language == "en"
                             else "API 자격 증명 형식 오류로 작업이 중단되었습니다. 환경변수 값을 확인하세요."))
            return None
        return creds

    def _show_api_error(self, creds, issue, fields, action=""):
        meta = creds.get("meta", {})
        mode_label = ("Testnet" if self.env_mode == "TESTNET" else "Live") if self.language == "en" else ("테스트넷" if self.env_mode == "TESTNET" else "라이브")
        if issue == "missing":
            body = ("Could not load the following API fields." if self.language == "en" else "다음 API 항목 값을 불러오지 못했습니다.")
            status = ("Value is empty." if self.language == "en" else "값이 비어 있습니다.")
            box = self._show_error
        else:
            body = ("The following API field values have an invalid format." if self.language == "en" else "다음 API 항목 값의 형식이 올바르지 않습니다.")
            status = ("Invalid format." if self.language == "en" else "형식이 잘못되었습니다.")
            box = self._show_warning

        lines = []
        label_map = {"api_key": "API Key", "api_secret": "API Secret"}
        for field in fields:
            source = meta.get(f"{field}_source", ("Settings input" if self.language == "en" else "환경설정 입력값"))
            lines.append(f"- {label_map.get(field, field)}: {source} → {status}")

        action_line = (f"Action: {action}\n" if self.language == "en" else f"작업: {action}\n") if action else ""
        # [수정] "환경설정 탭" 안내 제거 — API Key 입력창 폐지로 설정 탭에서 할 일 없음.
        message = (
            f"[{mode_label}] {body}\n{action_line}\n"
            + "\n".join(lines)
            + ("\n\nCheck your system environment variables.\n"
            + "(Control Panel → System → Advanced → Environment Variables)\n"
            + "Restart the GUI after making changes."
            if self.language == "en" else
            "\n\n시스템 환경변수를 확인해 주세요.\n"
            + "(제어판 → 시스템 → 고급 시스템 설정 → 환경 변수)\n"
            + "변경 후 GUI를 재시작해야 적용됩니다.")
        )
        box(("API Credentials Error" if self.language == "en" else "API 자격 증명 오류"), message)
        # API 키 오류 후 시작하기 다이얼로그 띄움
        has_key = bool(os.environ.get("BINANCE_API_KEY") or os.environ.get("TESTNET_API_KEY"))
        if not has_key:
            self.root.after(300, self._show_referral_onboarding)

    def _startup_script_path(self):
        startup_dir = os.path.join(os.environ.get("APPDATA", ""), "Microsoft", "Windows", "Start Menu", "Programs", "Startup")
        return os.path.join(startup_dir, "bot_gui_autostart.bat")

    def _configure_auto_start(self, enabled: bool):
        script_path = self._startup_script_path()
        try:
            if enabled:
                os.makedirs(os.path.dirname(script_path), exist_ok=True)
                with open(script_path, "w", encoding="utf-8") as fh:
                    fh.write(
                        f"@echo off\n"
                        f"cd /d \"{BASE_DIR}\"\n"
                        f"\"{sys.executable}\" \"{os.path.abspath(__file__)}\"\n"
                    )
            else:
                if os.path.exists(script_path):
                    os.remove(script_path)
        except Exception as exc:
            self._append_log(f"[WARN] 자동 실행 설정 실패: {exc}")

    def _load_checkbox_images(self):
        images = {}
        try:
            images["on"] = tk.PhotoImage(file=os.path.join(BASE_DIR, "assets", "check_on.png"))
            images["off"] = tk.PhotoImage(file=os.path.join(BASE_DIR, "assets", "check_off.png"))
        except Exception as exc:
            self._append_log(f"[WARN] 체크박스 이미지를 불러오지 못했습니다: {exc}")
            images["on"] = images["off"] = None
        return images

    def _run_async(self, coro, success_cb=None, fail_msg=None):
        future = asyncio.run_coroutine_threadsafe(coro, self.loop)

        def callback(fut):
            try:
                result = fut.result()
                if success_cb:
                    self.root.after(0, lambda: success_cb(result))
            except Exception as exc:
                if fail_msg:
                    err_text = f"{fail_msg}\n{exc}"
                    self.root.after(0, lambda e=err_text: self._show_error(self._t("error_title","Error"), e))

        future.add_done_callback(callback)

    # ------------------------------------------------------------------
    def _reset_runtime_stats(self, timestamp=None):
        ts = int(timestamp or time.time() * 1000)
        self.stat_resets["trade_count"] = ts
        self.stat_resets["win_rate"] = ts
        self._set_stat_value("trade_count", "0")
        self._set_stat_value("win_rate", "0%")

    def start_engine(self):
        if self.engine_running:
            return
        # 방어 코드: 잔여 grab이 있으면 해제하여 버튼 클릭 불가 현상 방지
        try:
            self.root.grab_release()
        except Exception:
            pass
        if not self.risk_acknowledged:
            self._show_warning(self._t("engine_start_risk_title","Risk acknowledgement required"), self._t("engine_start_risk_msg","To start the engine, acknowledge the risk notice in the Agreement tab."))
            self.open_settings_modal(initial_tab="dev")
            return
        config = self._build_engine_config()
        if not config:
            self._show_error(self._t("error_title","Error"), self._t("engine_config_error","Failed to load engine configuration."))
            return
        creds = self._require_credentials(action="봇 시작")
        if not creds:
            # [제거] self.open_settings_modal() 삭제.
            # 사유: API Key 입력창 제거됨. 오류 팝업(_show_api_error)만으로 안내 충분.
            return
        self.session_start_ms = int(time.time() * 1000)
        self.state_data["session_start_ms"] = self.session_start_ms
        self._reset_runtime_stats(self.session_start_ms)
        self._save_json(STATE_PATH, self.state_data)
        self.engine_running = True
        self.start_btn.configure(state=tk.DISABLED, bg="#2f6f40")
        self.stop_btn.configure(state=tk.NORMAL, bg="#c62828")
        self.status_label.configure(text=f'{self._t("bot_status","봇 상태")}: RUNNING', fg="#1f9d55")
        self._append_log("[INFO] Tick engine starting")

        try:
            open(self.notification_path, "w", encoding="utf-8").close()
            self.notification_pointer = 0
        except Exception as exc:
            self._append_log(f"[WARN] 알림 파일 초기화 실패: {exc}")

        from binance_futures_bot1_1.main import start_engine as async_start

        def _on_engine_started(_result):
            if getattr(self, "_restart_notice_pending", False):
                self._restart_notice_pending = False
                self._show_info(self._t("engine_restart_done_title","Restart complete"), self._t("engine_restart_done_msg","The trading engine has been restarted."))

        def _on_engine_failed(err_text):
            # [추가] 엔진 시작 실패 시 engine_running 상태 롤백.
            # 사유: _run_async 예외 발생 시 GUI가 "실행 중" 상태로 굳어버리는 버그 수정.
            self.engine_running = False
            self.start_btn.configure(state=tk.NORMAL, bg="#1f9d55")
            self.stop_btn.configure(state=tk.DISABLED, bg="#55565e")
            self.status_label.configure(text=f'● {self._t("bot_status","봇 상태")}: STOPPED', fg="#c62828")
            self._append_log(f"[ERROR] 엔진 시작 실패: {err_text}")
            self._show_error(self._t("engine_start_fail_title","Engine start failed"), err_text)

        def _run_async_with_rollback(coro):
            import asyncio
            future = asyncio.run_coroutine_threadsafe(coro, self.loop)
            def callback(fut):
                try:
                    result = fut.result()
                    if _on_engine_started:
                        self.root.after(0, lambda: _on_engine_started(result))
                except Exception as exc:
                    self.root.after(0, lambda e=str(exc): _on_engine_failed(e))
            future.add_done_callback(callback)

        _run_async_with_rollback(async_start(config, creds, self.notification_path))
        if (not getattr(self, "_shutdown_alert_suppressed", False)) and (not getattr(self, "_restart_notice_pending", False)):
            self._show_info(self._t("bot_start_title","Bot starting"), self._t("bot_start_msg","Starting the trading engine."))

    def stop_engine(self, on_stopped=None, *, suppress_alert=False):
        if not self.engine_running:
            if (not suppress_alert) and (not getattr(self, "_restart_notice_pending", False)):
                self._show_info(self._t("bot_stop_title","Bot stopped"), self._t("bot_stop_already","The engine is already stopped."))
            if on_stopped:
                self.root.after(0, on_stopped)
            return
        self.engine_running = False
        if self.position_refresh_job:
            self.root.after_cancel(self.position_refresh_job)
            self.position_refresh_job = None
        self.start_btn.configure(state=tk.NORMAL, bg="#1f9d55")
        self.stop_btn.configure(state=tk.DISABLED, bg="#55565e")
        self.status_label.configure(text=f'● {self._t("bot_status","봇 상태")}: STOPPED', fg="#c62828")
        self._append_log("[INFO] Tick engine stop requested")

        from binance_futures_bot1_1.main import stop_engine as async_stop

        def handle_stop(_result):
            if on_stopped:
                self.root.after(0, on_stopped)
            self.session_start_ms = None
            if "session_start_ms" in self.state_data:
                self.state_data.pop("session_start_ms", None)
                self._save_json(STATE_PATH, self.state_data)
            show_alert = (not suppress_alert) and (not getattr(self, "_shutdown_alert_suppressed", False)) and (not getattr(self, "_restart_notice_pending", False))
            if show_alert:
                self._show_info(self._t("bot_stop_title","Bot stopped"), self._t("bot_stop_msg","The trading engine has been stopped."))
            self._shutdown_alert_suppressed = False

        self._run_async(async_stop(), success_cb=handle_stop, fail_msg=self._t("engine_stop_fail","Engine stop failed"))

    def run_test_trade(self):
        creds = self._require_credentials(action="테스트 주문")
        if not creds:
            # [제거] self.open_settings_modal() 삭제.
            # 사유: API Key 입력창 제거됨. 오류 팝업만으로 안내 충분.
            return
        from binance_futures_bot1_1.main import place_test_order

        self._append_log("[INFO] 테스트 주문 전송")

        def on_success(result):
            msg = (
                f"Test order filled\nSymbol: {result['symbol']}\nQty: {result['quantity']}\nOrderID: {result['orderId']}"
                if self.language == "en" else
                f"테스트 주문 체결\n심볼: {result['symbol']}\n수량: {result['quantity']}\n"
                f"주문ID: {result['orderId']}"
            )
            self._append_log(f"[FILL] {msg}")
            if getattr(self, "alert_enabled", True):
                self._show_info(self._t("test_order_title","Test order"), msg)

        self._run_async(place_test_order(creds), success_cb=on_success, fail_msg="Test order failed")

    def show_log_analysis(self):
        analysis = self._analyze_filter_logs()
        dialog = tk.Toplevel(self.root)
        dialog.title("Log Analysis" if self.language == "en" else "실시간 로그 분석")
        self._apply_icon(dialog)
        width, height = 520, 360
        dialog.geometry(f"{width}x{height}")
        dialog.configure(bg="#181A20")
        dialog.transient(self.root)
        dialog.focus_force()
        self._center_modal(dialog, width, height)

        header = tk.Label(dialog, text=self._t("filter_reject_title","최근 필터/체결 거부 사유"), bg="#181A20", fg="white", font=("Malgun Gothic", 13, "bold"))
        header.pack(fill="x", pady=(18, 6))

        text = tk.Text(dialog, bg="#121521", fg="#c0c6dc", insertbackground="white", font=("Malgun Gothic", 10), wrap="word")
        text.pack(fill="both", expand=True, padx=14, pady=10)
        text.insert(tk.END, analysis)
        text.configure(state=tk.DISABLED)

        ttk.Button(dialog, text=self._t("close","닫기"), command=dialog.destroy).pack(pady=(0, 12))

    def show_symbol_watchlist(self):
        tracked = self._extract_symbol_watchlist()
        dialog = tk.Toplevel(self.root)
        dialog.title(self._t("watchlist_title","거래 대상 심볼 상태"))
        self._apply_icon(dialog)
        width, height = 520, 380
        dialog.geometry(f"{width}x{height}")
        dialog.configure(bg="#181A20")
        dialog.transient(self.root)
        dialog.focus_force()
        self._center_modal(dialog, width, height)

        tk.Label(dialog, text=self._t("watchlist_header","최근 모니터링 심볼"), bg="#181A20", fg="white", font=("Malgun Gothic", 13, "bold")).pack(fill="x", pady=(18, 6))
        text = tk.Text(dialog, bg="#121521", fg="#c0c6dc", insertbackground="white", font=("Malgun Gothic", 10), wrap="word")
        text.pack(fill="both", expand=True, padx=14, pady=10)
        if tracked:
            for entry in tracked:
                text.insert(tk.END, f"• {entry}\n")
        else:
            text.insert(tk.END, self._t("watchlist_none","로그에서 추출한 심볼이 없습니다. 엔진이 실행 중인지 확인하세요."))
        text.configure(state=tk.DISABLED)
        ttk.Button(dialog, text=self._t("close","닫기"), command=dialog.destroy).pack(pady=(0, 12))

    def _build_manual_panel(self):
        panel = tk.Frame(
            self.main_area,
            bg="#181A20",
            highlightbackground="#343942",
            highlightthickness=1,
        )
        panel.columnconfigure(0, weight=1)
        self.manual_panel = panel
        # ── 헤더 (제목 + 토글 우측) ─────────────────────────────────────
        header = tk.Frame(panel, bg="#181A20")
        header.pack(fill="x", padx=12, pady=(8, 4))

        self.manual_toggle_btn = tk.Button(
            header, text="▲",
            command=self._toggle_manual_collapse,
            bg="#181A20", fg="#6a7490",
            relief=tk.FLAT, font=("Segoe UI Symbol", 9),
            padx=4, pady=0, bd=0, cursor="hand2",
            activebackground="#181A20", activeforeground="#c0c6dc",
        )
        self.manual_toggle_btn.pack(side=tk.RIGHT)

        tk.Label(header, text=self._t("manual_panel_title", "지정 코인 수동 매매"),
                 bg="#181A20", fg="white",
                 font=("Malgun Gothic", 11, "bold")).pack(side=tk.LEFT)

        # ── 슬라이더 영역 — 기본 펼침, 토글로 접기 가능 ─────────────────
        self._manual_collapsed = False  # 기본 펼침
        self._manual_slider_frame = tk.Frame(panel, bg="#181A20")
        self._manual_slider_frame.pack(fill="x", padx=12, pady=(4, 2))  # 초기부터 표시

        # DoubleVar / IntVar — 설정 탭과 별개로 수동 패널 전용
        _init_pct = round(float(self.settings_data.get("position_pct", 0.05)), 4)
        _init_pct = max(0.01, min(0.99, _init_pct))
        self.manual_pct_var = tk.DoubleVar(value=_init_pct)

        _lev_min_cfg = 1     # 수동 매매 패널은 항상 1x 시작
        _lev_max_cfg = 100   # 신규 계정 20x 제한 등 고려, 100x로 통일
        _init_lev = min(10, _lev_max_cfg)   # 기본값 10x
        self.manual_lev_var2 = tk.IntVar(value=_init_lev)

        # 진입 크기 레이블
        _pct_text = tk.StringVar()
        _hint_text = tk.StringVar()

        def _update_pct_lbl(*_):
            _lbl  = "진입 크기" if self.language == "ko" else "Entry size"
            _acct = "계좌의"   if self.language == "ko" else "account"
            _pct_text.set(f"{_lbl} ({_acct} {self.manual_pct_var.get() * 100:.1f}%)")

        self.manual_pct_var.trace_add("write", _update_pct_lbl)
        _update_pct_lbl()

        _hint_lbl = None  # 슬라이더 초기 upd() 호출 시 참조 오류 방지

        def _guidance(val):
            if val < 0.03:
                msg = ("3% 미만: 수익 실현이 어려울 수 있습니다." if self.language == "ko"
                       else "Below 3%: position size may be too small to realize gains meaningfully.")
                color = "#F7C948"
            elif val > 0.06:
                msg = ("6% 초과: 손실 위험이 큽니다. 레버리지와 손절 설정을 재확인하세요." if self.language == "ko"
                       else "Above 6%: risk of amplified losses. Re-check leverage and stop-loss settings.")
                color = "#FF5F5F"
            else:
                msg = ("권장 3–6%: 변동성이 클 때는 3% 근처, 확신이 강할 때 5% 수준을 유지하세요." if self.language == "ko"
                       else "Recommended 3–6%: stay near 3% in choppy markets; use ~5% only with strong conviction.")
                color = "#2EBD85"
            _hint_text.set(msg)
            if _hint_lbl is not None and _hint_lbl.winfo_exists():
                _hint_lbl.configure(fg=color)
            return color

        _sf = self._manual_slider_frame

        # ── 한 줄 슬라이더 빌더 (레이블+값 | 슬라이더 | 우측 라벨) ────
        def _build_canvas_slider(parent, variable, min_val, max_val,
                                 step=0.01, length=200, state_cb=None, int_mode=False):
            _safe_max = max_val if max_val > min_val else min_val + (1 if int_mode else 0.01)
            c = tk.Canvas(parent, width=length, height=26,
                          bg="#181A20", highlightthickness=0, bd=0)
            ty, mg = 13, 10
            sw = max(1, length - mg * 2)
            c.create_line(mg, ty, length - mg, ty, fill="#3f4659", width=3, capstyle=tk.ROUND)
            kr = 7
            knob = c.create_oval(0, 0, kr*2, kr*2, fill="#2EBD85", outline="#2EBD85", width=2)
            _th = {"id": None}

            def v2x(v):
                _range = _safe_max - min_val
                if _range == 0:
                    return mg + sw / 2
                return mg + (v - min_val) / _range * sw

            def upd(*_):
                if not c.winfo_exists():
                    if _th["id"]:
                        try: variable.trace_remove("write", _th["id"])
                        except tk.TclError: pass
                    return
                v = max(min_val, min(_safe_max, variable.get()))
                x = v2x(v)
                c.coords(knob, x - kr, ty - kr, x + kr, ty + kr)
                col = "#2EBD85"
                if state_cb:
                    col = state_cb(v) or col
                c.itemconfigure(knob, fill=col, outline=col)

            def on_drag(ev):
                _range = _safe_max - min_val
                if _range == 0:
                    return
                ratio = max(0.0, min(1.0, (ev.x - mg) / sw))
                v = min_val + ratio * _range
                v = round(round(v / step) * step, 6)
                if int_mode:
                    v = int(round(v))
                v = max(min_val, min(_safe_max, v))
                variable.set(v)

            c.bind("<Button-1>", on_drag)
            c.bind("<B1-Motion>", on_drag)
            c.bind("<Destroy>", lambda _: (_th["id"] and
                    [variable.trace_remove("write", _th["id"]) for _ in [None]]))
            _th["id"] = variable.trace_add("write", upd)
            upd()
            return c

        # ── 명목금/증거금 갱신 함수 ─────────────────────────────────────
        def _refresh_notional(*_):
            bal = float(getattr(self, "last_account_balance", 0.0))
            lbl = getattr(self, "_manual_notional_lbl", None)
            if not lbl or not lbl.winfo_exists():
                return
            if bal <= 0:
                lbl.config(text="Loading balance…" if self.language == "en" else "잔고 조회 중…")
                return
            try:
                pct      = float(self.manual_pct_var.get())
                lev      = max(1, int(self.manual_lev_var2.get()))
                notional = bal * pct * lev
                margin   = bal * pct
                lbl.config(text=f"명목 {notional:,.1f} U  /  증거금 {margin:,.1f} U")
            except Exception:
                lbl.config(text="")

        # ── 경고 문구 함수 ───────────────────────────────────────────────
        _hint_lbl = None

        def _guidance(val):
            if val < 0.03:
                msg   = ("3% 미만: 수익 실현이 어려울 수 있습니다." if self.language == "ko"
                         else "Below 3%: position size may be too small.")
                color = "#F7C948"
            elif val > 0.06:
                msg   = ("6% 초과: 손실 위험이 큽니다. 레버리지·손절 설정을 확인하세요." if self.language == "ko"
                         else "Above 6%: high loss risk. Re-check leverage and stop-loss.")
                color = "#FF5F5F"
            else:
                msg   = ("권장 3–6%: 변동성이 클 땐 3%, 확신이 강할 때 5% 수준을 유지하세요." if self.language == "ko"
                         else "Recommended 3–6%: near 3% in choppy markets, ~5% with strong conviction.")
                color = "#2EBD85"
            _hint_text.set(msg)
            if _hint_lbl is not None and _hint_lbl.winfo_exists():
                _hint_lbl.configure(fg=color)
            return color

        # ── 한 줄: [크기 X.X%][슬라이더] | [배율 Xx][슬라이더] | [명목/증거금] ──
        _lev_min_eff = _lev_min_cfg
        _lev_max_eff = _lev_max_cfg if _lev_max_cfg > _lev_min_cfg else _lev_min_cfg + 20

        _combined_row = tk.Frame(_sf, bg="#181A20")
        _combined_row.pack(fill="x", pady=(2, 0))

        # 크기 레이블
        _pct_lbl_text = tk.StringVar()
        def _upd_pct_lbl(*_):
            _pct_lbl_text.set(
                f"{'크기' if self.language=='ko' else 'Size'} {self.manual_pct_var.get()*100:.1f}%")
        self.manual_pct_var.trace_add("write", lambda *_: (_upd_pct_lbl(), _refresh_notional()))
        _upd_pct_lbl()

        tk.Label(_combined_row, textvariable=_pct_lbl_text, bg="#181A20", fg="#c0c6dc",
                 font=("Malgun Gothic", 9, "bold"), width=9, anchor="w").pack(side=tk.LEFT)
        _build_canvas_slider(_combined_row, self.manual_pct_var, 0.01, 0.99,
                             step=0.001, length=160, state_cb=_guidance).pack(side=tk.LEFT, padx=(2, 10))

        # 배율 레이블
        _lev_lbl_text = tk.StringVar()
        def _upd_lev_lbl(*_):
            _unit = "배" if self.language == "ko" else "x"
            _lev_lbl_text.set(
                f"{'배율' if self.language=='ko' else 'Lev'} {self.manual_lev_var2.get()}{_unit}")
        self.manual_lev_var2.trace_add("write", lambda *_: (_upd_lev_lbl(), _refresh_notional()))
        _upd_lev_lbl()

        tk.Label(_combined_row, textvariable=_lev_lbl_text, bg="#181A20", fg="#c0c6dc",
                 font=("Malgun Gothic", 9, "bold"), width=8, anchor="w").pack(side=tk.LEFT)
        _build_canvas_slider(_combined_row, self.manual_lev_var2,
                             _lev_min_eff, _lev_max_eff,
                             step=1, length=130, int_mode=True).pack(side=tk.LEFT, padx=(2, 8))

        # 명목금/증거금 라벨
        self._manual_notional_lbl = tk.Label(
            _combined_row, text="", bg="#181A20", fg="#6a7a9a",
            font=("Consolas", 8), anchor="e")
        self._manual_notional_lbl.pack(side=tk.LEFT, fill="x", expand=True)

        # 경고 문구
        _hint_text = tk.StringVar()
        _hint_lbl = tk.Label(_sf, textvariable=_hint_text, bg="#181A20",
                             fg="#c0c6dc", font=("Malgun Gothic", 8), anchor="w")
        _hint_lbl.pack(fill="x", padx=2, pady=(0, 4))
        _guidance(_init_pct)

        # 초기 명목금 표시
        _refresh_notional()
        # ── 카드 컨테이너 ────────────────────────────────────────────────
        self.manual_cards_container = tk.Frame(panel, bg="#181A20")
        self.manual_cards_container.pack(fill="x", padx=10, pady=(0, 10))
        self._render_manual_cards()
        self.refresh_top_symbols()

        self._apply_manual_panel_visibility()
    def _toggle_manual_collapse(self):
        """슬라이더 + 카드 영역 펼침/접힘 토글."""
        self._manual_collapsed = not getattr(self, "_manual_collapsed", True)
        container = getattr(self, "manual_cards_container", None)
        slider_fr = getattr(self, "_manual_slider_frame", None)
        btn       = getattr(self, "manual_toggle_btn", None)
        if self._manual_collapsed:
            if slider_fr and slider_fr.winfo_ismapped():
                slider_fr.pack_forget()
            if container and container.winfo_ismapped():
                container.pack_forget()
        else:
            # 슬라이더를 카드보다 먼저 표시 (pack 순서 보장)
            if container and container.winfo_ismapped():
                container.pack_forget()
            if slider_fr:
                slider_fr.pack(fill="x", padx=12, pady=(4, 2))
            if container:
                container.pack(fill="x", padx=10, pady=(0, 10))
        if btn and btn.winfo_exists():
            btn.config(text="▼" if self._manual_collapsed else "▲")
    def _render_manual_cards(self):
        container = getattr(self, "manual_cards_container", None)
        if not container:
            return
        for child in container.winfo_children():
            child.destroy()
        if not self.manual_symbols:
            tk.Label(container, text=self._t("manual_no_symbols","수동 매매 심볼이 없습니다"), bg="#181A20", fg="#c0c6dc").pack(fill="x", pady=8)
            return
        for idx, symbol in enumerate(self.manual_symbols):
            container.columnconfigure(idx, weight=1)
            card = tk.Frame(container, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
            card.grid(row=0, column=idx, sticky="nsew", padx=6, pady=4)
            label = symbol.replace("USDT", "") or symbol
            tk.Label(card, text=f"{label} {self._t('manual_card_suffix','수동 매매')}", bg="#181A20", fg="white", font=("Malgun Gothic", 11, "bold")).grid(row=0, column=0, columnspan=2, sticky="w", padx=10, pady=(10, 6))
            button_defs = [
                (1, 0, f"{label} {self._t('manual_long_limit','롱 지정가 진입')}", "long_limit", "#2EBD85"),
                (1, 1, f"{label} {self._t('manual_long_market','롱 시장가 진입')}", "long_market", "#2EBD85"),
                (2, 0, f"{label} {self._t('manual_short_limit','숏 지정가 진입')}", "short_limit", "#F6465D"),
                (2, 1, f"{label} {self._t('manual_short_market','숏 시장가 진입')}", "short_market", "#F6465D"),
                (3, 0, f"{label} {self._t('manual_close_limit','지정가 청산')}", "close_limit", "#4a4c7d"),
                (3, 1, f"{label} {self._t('manual_close_market','시장가 청산')}", "close_market", "#4a4c7d"),
            ]
            _btn_hover = {
                "#2EBD85": "#1da070",
                "#F6465D": "#d03040",
                "#4a4c7d": "#5c5e9a",
            }
            for row, col, text_label, action, color in button_defs:
                _hov = _btn_hover.get(color, color)
                tk.Button(
                    card,
                    text=text_label,
                    command=lambda s=symbol, a=action, t=text_label: self._manual_trade(s, a, label=t),
                    bg=color,
                    fg="white",
                    relief=tk.FLAT,
                    font=("Malgun Gothic", 9, "bold"),
                    cursor="hand2",
                    activebackground=_hov,
                    activeforeground="white",
                ).grid(row=row, column=col, sticky="nsew", padx=6, pady=3)
            for r in range(1, 4):
                card.rowconfigure(r, weight=1)
            card.columnconfigure(0, weight=1)
            card.columnconfigure(1, weight=1)

    def _toggle_manual_panel_section(self):
        pass  # 패널 표시는 화면설정 탭에서 제어 — 직접 토글 불필요

    def _apply_manual_panel_visibility(self):
        show = bool(self.settings_data.get("show_manual_panel", False))
        panel = getattr(self, "manual_panel", None)
        positions = getattr(self, "positions_panel", None)
        if panel and panel.winfo_exists():
            if show:
                panel.grid(row=1, column=0, sticky="ew", pady=(10, 4))
            else:
                panel.grid_remove()
        if positions and positions.winfo_exists():
            pad = (0, 0) if show else (10, 0)
            positions.grid(row=2, column=0, sticky="nsew", pady=pad)
        if hasattr(self, "main_area"):
            self.main_area.rowconfigure(1, weight=0)
            self.main_area.rowconfigure(2, weight=1)

    def _manual_trade(self, symbol, action, *, silent=False, label=None):
        if not self.risk_acknowledged:
            if not silent:
                self._show_warning(self._t("risk_ack_title","위험 고지 동의 필요"), self._t("risk_ack_msg","지정 코인 수동 매매를 사용하려면 '필수 동의' 탭에서 동의 체크박스를 활성화하세요."))
                self.open_settings_modal(initial_tab="dev")
            return
        creds = self._require_credentials(silent=silent, action=f"{symbol} manual order")
        if not creds:
            # [제거] self.open_settings_modal() 삭제.
            # 사유: API Key 입력창 제거됨. 오류 팝업만으로 안내 충분.
            return
        # 수동 매매 패널 캔버스 슬라이더 값 읽기 (범위 이미 보장)
        _pct_var2 = getattr(self, "manual_pct_var", None)
        _lev_var2 = getattr(self, "manual_lev_var2", None)
        try:
            _pct_raw = float(_pct_var2.get()) if _pct_var2 else float(self.settings_data.get("position_pct", 0.05))
            _pct_raw = max(0.01, min(0.99, _pct_raw))
        except Exception:
            _pct_raw = float(self.settings_data.get("position_pct", 0.05))
        percent = round(_pct_raw * 100, 4)  # _manual_trade 내부는 % 단위
        try:
            _manual_leverage = max(1, int(_lev_var2.get())) if _lev_var2 else None
        except Exception:
            _manual_leverage = None
        order_map = {
            "long_limit": ("BUY", "LIMIT", False),
            "long_market": ("BUY", "MARKET", False),
            "short_limit": ("SELL", "LIMIT", False),
            "short_market": ("SELL", "MARKET", False),
            "close_limit": (None, "LIMIT", True),
            "close_market": (None, "MARKET", True),
            "close_limit_then_market": (None, "LIMIT_THEN_MARKET", True),
        }
        if action not in order_map:
            return
        side, order_type, reduce_only = order_map[action]
        position = None
        if reduce_only:
            position = self.current_positions.get(symbol)
            if not position:
                if not silent:
                    self._show_warning(self._t("alert","알림"), (f"No open position for {symbol}" if self.language=="en" else f"{symbol} 포지션이 없습니다"))
                return
            side = "SELL" if position["side"] == "LONG" else "BUY"
            quantity = abs(position["amount"])
            notional_percent = None
        else:
            quantity = None
            notional_percent = percent
        price = None
        price_source = None
        # [제거] price = None 중복 선언 삭제. 사유: 바로 위에서 이미 None 초기화됨.
        if order_type in ("LIMIT", "LIMIT_THEN_MARKET"):
            price = self.symbol_price_map.get(symbol)
            price_source = "symbol_price_map"
            if (not price) and position:
                price = position.get("markPrice") or position.get("entryPrice") or position.get("breakEvenPrice")
                price_source = "position"
            if not price:
                warn_msg = (f"{symbol}: Could not fetch current price. Try 'Refresh Symbol Prices'." if self.language == "en"
                    else f"{symbol} 현재가를 불러오지 못했습니다. '심볼 시세 새로고침'을 눌러주세요.")
                if not silent:
                    self._show_warning(self._t("price_missing_title","가격 정보 없음"), warn_msg)
                else:
                    self._append_log(f"[WARN] {warn_msg}")
                return
            self._mark_manual_highlight(symbol)
        elif order_type == "MARKET":
            price = self.symbol_price_map.get(symbol)
            price_source = "symbol_price_map"
        if price is not None:
            try:
                price = float(price)
            except (TypeError, ValueError):
                warn_msg = (f"{symbol}: Cannot convert price to number ({price_source}): {price}" if self.language == "en"
                    else f"{symbol} 가격 값을 숫자로 변환할 수 없습니다 ({price_source}): {price}")
                if not silent:
                    self._show_warning(self._t("price_convert_error_title","Price conversion error"), warn_msg)
                else:
                    self._append_log(f"[WARN] {warn_msg}")
                return
        if not reduce_only and bool(self.settings_data.get("auto_boost_position_pct", False)):
            percent = self._ensure_min_notional_percent(symbol, percent, price, creds)
        if reduce_only:
            quantity, price = self._apply_order_filters(symbol, quantity, price, reduce_only=True, creds=creds)
            if quantity is None or quantity <= 0:
                warn_msg = (f"{symbol}: Failed to adjust close quantity to filter constraints." if self.language == "en"
                    else f"{symbol} 청산 수량을 필터에 맞게 조정하지 못했습니다.")
                self._append_log(f"[WARN] {warn_msg}")
                if not silent:
                    self._show_warning(self._t("qty_error_title","Quantity error"), warn_msg)
                return
        elif order_type == "MARKET":
            price = self.symbol_price_map.get(symbol)
        try:
            if price is not None:
                price = float(price)
        except (TypeError, ValueError):
            price = None
        action_label_map = {
            "long_limit":   self._t("action_long_limit",  "Long limit entry"),
            "long_market":  self._t("action_long_market", "Long market entry"),
            "short_limit":  self._t("action_short_limit", "Short limit entry"),
            "short_market": self._t("action_short_market","Short market entry"),
            "close_limit":  self._t("action_close_limit", "Close (limit)"),
            "close_market": self._t("action_close_market","Close (market)"),
            "close_limit_then_market": self._t("action_close_limit_then_market","Close (limit→market)"),
        }
        action_label = label or action_label_map.get(action, action)
        if not silent:
            if not self._confirm_manual_trade(
                symbol,
                action_label,
                reduce_only=reduce_only,
                order_type=order_type,
                percent=None if reduce_only else percent,
                quantity=quantity,
                price=price,
                position=position,
            ):
                return
        from binance_futures_bot1_1.main import place_manual_order

        def on_success(resp):
            if not silent and getattr(self, "alert_enabled", True):
                if reduce_only:
                    _msg = self._t("order_sent_close", "{sym} close order sent").format(sym=symbol)
                else:
                    _msg = self._t("order_sent_entry", "{sym} entry order sent").format(sym=symbol)
                self._show_info(self._t("order_sent_title", "Order"), f"{_msg}\nID: {resp.get('orderId')}")
            try:
                order_id = resp.get("orderId") if isinstance(resp, dict) else None
                _verb = "청산" if reduce_only else "진입"
                self._write_manual_notification(
                    f"{symbol} {_verb} {order_type.upper()} 주문 전송 (reduceOnly={reduce_only}, orderId={order_id})"
                )
            except Exception as exc:
                self._append_log(f"[WARN] 수동 알림 작성 실패: {exc}")
            if reduce_only and order_type == "MARKET":
                _pos_snap = position or self.current_positions.get(symbol)
                self._log_manual_trade_event(symbol, _pos_snap, resp, label or action_label)
                # trade_log 기록 직후 즉시 패널 갱신 (API 불필요)
                self._refresh_panels_from_trade_log()
            self._burst_refresh_positions()
            self._trigger_stat_refresh(2000)   # API 갱신은 2초 후

        # 진입 주문 시 레버리지 먼저 적용
        if not reduce_only and _manual_leverage:
            async def _set_lev_and_place():
                from binance_futures_bot1_1.main import place_manual_order as _pmo
                import re as _re
                _actual_lev = _manual_leverage
                try:
                    from binance_futures_bot1_1.main import _ensure_client
                    _c = await _ensure_client(
                        creds.get("api_key",""), creds.get("api_secret",""), creds.get("testnet", True))
                    try:
                        await _c.futures_change_leverage(symbol=symbol, leverage=_actual_lev)
                        self._append_log(f"[INFO] {symbol} 레버리지 {_actual_lev}x 설정")
                    except Exception as _le:
                        _err = str(_le)
                        # 바이낸스 신규 계정 제한: 최대 허용 배율 파싱 후 재시도
                        _m = _re.search(r"maximum.*?(\d+)x|(\d+)x.*?maximum|leverage.*?(\d+)", _err, _re.I)
                        if _m:
                            _allowed = int(next(g for g in _m.groups() if g))
                            _allowed = max(1, min(_allowed, _actual_lev))
                            self._append_log(
                                f"[WARN] {symbol} 레버리지 {_actual_lev}x 불가 → {_allowed}x로 재시도: {_err}")
                            try:
                                await _c.futures_change_leverage(symbol=symbol, leverage=_allowed)
                                _actual_lev = _allowed
                                self._append_log(f"[INFO] {symbol} 레버리지 {_actual_lev}x 설정 완료")
                            except Exception as _le2:
                                self._append_log(f"[WARN] {symbol} 레버리지 재시도 실패, 현재 배율 유지: {_le2}")
                        else:
                            self._append_log(f"[WARN] {symbol} 레버리지 변경 실패, 현재 배율 유지: {_err}")
                except Exception as _outer:
                    self._append_log(f"[WARN] {symbol} 레버리지 클라이언트 오류: {_outer}")
                return await _pmo(creds, symbol, side or "BUY",
                                  order_type if order_type != "LIMIT_THEN_MARKET" else "LIMIT",
                                  quantity=quantity, price=price,
                                  reduce_only=reduce_only, notional_percent=notional_percent)
            self._run_async(
                _set_lev_and_place(),
                success_cb=lambda resp: self._handle_manual_trade_success(resp, on_success),
                fail_msg=self._t("order_fail_prefix", "{sym} manual order failed").format(sym=symbol),
            )
            return

        # LIMIT_THEN_MARKET: 지정가 GTC 제출 → 3초 대기 → 미체결 시 취소 후 시장가 재시도
        _actual_order_type = "LIMIT" if order_type == "LIMIT_THEN_MARKET" else order_type
        if order_type == "LIMIT_THEN_MARKET":
            async def _limit_then_market_coro():
                import asyncio as _aio
                from binance_futures_bot1_1.main import place_manual_order as _pmo
                from binance_futures_bot1_1.binance_futures_bot.tick_engine import TickEngine
                # 지정가 주문 제출
                resp = await _pmo(creds, symbol, side or "SELL", "LIMIT",
                                  quantity=quantity, price=price,
                                  reduce_only=reduce_only, notional_percent=notional_percent)
                oid = (resp or {}).get("orderId")
                if not oid:
                    return resp
                # 3초 폴링
                _deadline = _aio.get_event_loop().time() + 3.0
                while _aio.get_event_loop().time() < _deadline:
                    await _aio.sleep(0.3)
                    try:
                        from binance_futures_bot1_1.main import get_order_status
                        chk = await get_order_status(creds, symbol, oid)
                        if (chk or {}).get("status") == "FILLED":
                            return chk  # 체결 성공
                    except Exception:
                        break
                # 취소 후 시장가 fallback
                try:
                    from binance_futures_bot1_1.main import cancel_order
                    await cancel_order(creds, symbol, oid)
                except Exception:
                    pass
                self._append_log(f"[INFO] {symbol} 지정가 미체결 → 시장가 fallback")
                return await _pmo(creds, symbol, side or "SELL", "MARKET",
                                  quantity=quantity, price=None,
                                  reduce_only=reduce_only, notional_percent=notional_percent)
            self._run_async(
                _limit_then_market_coro(),
                success_cb=lambda resp: self._handle_manual_trade_success(resp, on_success),
                fail_msg=self._t("order_fail_prefix", "{sym} manual order failed").format(sym=symbol),
            )
        else:
            self._run_async(
                place_manual_order(
                    creds,
                    symbol,
                    side or "SELL",
                    _actual_order_type,
                    quantity=quantity,
                    price=price,
                    reduce_only=reduce_only,
                    notional_percent=notional_percent,
                ),
                success_cb=lambda resp: self._handle_manual_trade_success(resp, on_success),
                fail_msg=self._t("order_fail_prefix", "{sym} manual order failed").format(sym=symbol),
            )

    def _handle_manual_trade_success(self, resp, callback):
        if callback:
            try:
                callback(resp)
            except Exception:
                self._append_log('[WARN] manual trade callback failed')
        self._trigger_stat_refresh()

    def _close_all_positions(self, order_type: str):
        if not self.risk_acknowledged:
            self._show_warning(self._t("risk_ack_title","Risk acknowledgement required"), self._t("close_risk_msg","Enable the risk acknowledgement checkbox in the Agreement tab to use close functions."))
            self.open_settings_modal(initial_tab="dev")
            return
        if not self.current_positions:
            self._show_info(self._t("close_all_title","Close all"), self._t("close_positions_none","No open positions"))
            return
        symbols = list(self.current_positions.keys())
        for sym in symbols:
            self._manual_trade(sym, f"close_{order_type}", silent=True)
        self._burst_refresh_positions()
        self._show_info(self._t("close_all_title","Close all"), self._t("close_all_sent","Close order sent for all positions"))

    def _close_positions_by_pnl(self, order_type: str, *, positive: bool):
        if not self.risk_acknowledged:
            self._show_warning(self._t("risk_ack_title","Risk acknowledgement required"), self._t("close_risk_msg","Enable the risk acknowledgement checkbox in the Agreement tab to use close functions."))
            self.open_settings_modal(initial_tab="dev")
            return
        if not self.current_positions:
            self._show_info(self._t("close_positions_title","Close positions"), self._t("close_positions_none","No open positions"))
            return
        comparator = (lambda pnl: pnl >= 0.0) if positive else (lambda pnl: pnl < 0.0)
        filtered = [sym for sym, pos in self.current_positions.items() if comparator(float(pos.get("unRealizedProfit", 0.0)))]
        if not filtered:
            if positive:
                self._show_info(self._t("close_positions_title","Close positions"), self._t("close_positions_no_profit","No profitable positions are open"))
            else:
                self._show_info(self._t("close_positions_title","Close positions"), self._t("close_positions_no_loss","No losing positions are open"))
            return
        for sym in filtered:
            self._manual_trade(sym, f"close_{order_type}", silent=True)
        self._burst_refresh_positions()
        self._trigger_stat_refresh(2000)
        _n = len(filtered)
        if positive:
            self._show_info(self._t("close_positions_title","Close positions"), self._t("close_positions_sent_profit","{n}개 수익 포지션에 청산 명령 전송").format(n=_n))
        else:
            self._show_info(self._t("close_positions_title","Close positions"), self._t("close_positions_sent_loss","{n}개 손실 포지션에 청산 명령 전송").format(n=_n))

    def _confirm_manual_trade(self, symbol, label, *, reduce_only, order_type, percent=None, quantity=None, price=None, position=None):
        ref_price = None
        try:
            if price is not None:
                ref_price = float(price)
        except (TypeError, ValueError):
            ref_price = None
        if ref_price is None:
            try:
                ref_price = float(self.symbol_price_map.get(symbol))
            except (TypeError, ValueError, AttributeError):
                ref_price = None
        try:
            qty_display = float(quantity) if quantity is not None else None
        except (TypeError, ValueError):
            qty_display = None
        entry_amount = None
        margin_amount = None
        if reduce_only:
            pos = position or self.current_positions.get(symbol)
            if pos:
                qty_display = abs(float(pos.get("amount", qty_display or 0.0)))
                try:
                    entry_price = float(pos.get("entryPrice", ref_price or 0.0))
                except (TypeError, ValueError):
                    entry_price = ref_price or 0.0
                entry_amount = qty_display * entry_price if entry_price else None
                try:
                    margin_amount = float(pos.get("marginValue", 0.0))
                except (TypeError, ValueError):
                    margin_amount = None
        else:
            notional_amount = None
            if percent is not None and self.last_account_balance:
                notional_amount = (self.last_account_balance * percent) / 100.0
            leverage_hint = None
            try:
                leverage_hint = float(self.settings_data.get("leverage_min", 5))
            except (TypeError, ValueError):
                leverage_hint = None
            if notional_amount:
                if leverage_hint and leverage_hint > 0:
                    margin_amount = notional_amount / leverage_hint
                else:
                    margin_amount = notional_amount
                if ref_price and ref_price > 0:
                    qty_display = notional_amount / ref_price
            entry_amount = None
        _type_label = self._t("trade_confirm_close_label","Close") if reduce_only else self._t("trade_confirm_entry_label","Entry")
        lines = [
            f"{self._t('trade_confirm_symbol','Symbol')}: {symbol}",
            f"{self._t('trade_confirm_order','Order')}: {label}",
            f"{self._t('trade_confirm_type','Type')}: {_type_label} / {order_type}",
        ]
        if percent is not None and not reduce_only:
            lines.append(f"Position %: {percent:.2f}%")
        if ref_price:
            lines.append(f"{self._t('trade_confirm_ref_price','Ref. price')}: {ref_price:.4f} USDT")
        if qty_display:
            lines.append(f"{self._t('trade_confirm_qty','Est. qty')}: {qty_display:.4f}")
        if entry_amount and reduce_only:
            lines.append(f"{self._t('trade_confirm_pos_amount','Est. position')}: {entry_amount:.2f} USDT")
        if not reduce_only:
            leverage_hint = None
            try:
                leverage_hint = float(self.settings_data.get("leverage_min", 5))
            except (TypeError, ValueError):
                leverage_hint = None
            if leverage_hint and leverage_hint > 0 and margin_amount:
                leveraged_amount = margin_amount * leverage_hint
                lines.append(f"{self._t('trade_confirm_notional','Est. notional')}: {leveraged_amount:.2f} USDT ({self._t('field_leverage','Leverage')} {leverage_hint:.0f}x)")
        if margin_amount:
            lines.append(f"{self._t('trade_confirm_margin','Est. margin')}: {margin_amount:.2f} USDT")
        lines.append("")
        lines.append(self._t("trade_confirm_question","Confirm this order?"))
        message = "\n".join(lines)

        dialog = tk.Toplevel(self.root)
        dialog.title(self._t("trade_confirm_title","Order confirmation"))
        self._apply_icon(dialog)
        dialog.configure(bg="#0C1017", highlightthickness=1, highlightbackground="#394058")
        dialog.resizable(False, False)
        dialog.transient(self.root)
        dialog.grab_set()
        width, height = 420, 360
        dialog.geometry(f"{width}x{height}")
        self._center_modal(dialog, width, height)

        container = tk.Frame(dialog, bg="#0C1017")
        container.pack(fill="both", expand=True, padx=18, pady=18)

        icon_label = tk.Label(
            container,
            text="⚠",
            bg="#0C1017",
            fg="#F7C948",
            font=("Segoe UI Symbol", 28, "bold"),
        )
        icon_label.pack(anchor="center", pady=(0, 6))

        text_frame = tk.Frame(container, bg="#0C1017")
        text_frame.pack(fill="both", expand=True)
        label_widget = tk.Label(
            text_frame,
            text=message,
            bg="#0C1017",
            fg="#e0e6ff",
            font=("Malgun Gothic", 11),
            justify="left",
            anchor="nw",
        )
        label_widget.pack(fill="both", expand=True)

        result = tk.BooleanVar(value=False)

        def choose(value):
            result.set(value)
            dialog.destroy()

        btn_row = tk.Frame(container, bg="#0C1017")
        btn_row.pack(fill="x", pady=(12, 0))
        btn_style = {
            "font": ("Malgun Gothic", 10, "bold"),
            "bg": "#0C1017",
            "fg": "#f7f9ff",
            "relief": tk.FLAT,
            "bd": 1,
            "highlightbackground": "#394058",
            "highlightcolor": "#394058",
            "activebackground": "#141b2a",
            "activeforeground": "#ffffff",
            "padx": 14,
            "pady": 6,
        }
        yes_btn = tk.Button(
            btn_row,
            text=self._t("trade_confirm_yes","Execute (Y)"),
            command=lambda: choose(True),
            **btn_style,
        )
        yes_btn.pack(side=tk.LEFT, expand=True, padx=(0, 6))
        no_btn = tk.Button(
            btn_row,
            text=self._t("trade_confirm_no","Cancel (N)"),
            command=lambda: choose(False),
            **btn_style,
        )
        no_btn.pack(side=tk.LEFT, expand=True, padx=(6, 0))

        dialog.bind("<Return>", lambda _e: choose(True))
        dialog.bind("<Escape>", lambda _e: choose(False))
        dialog.protocol("WM_DELETE_WINDOW", lambda: choose(False))
        yes_btn.focus_set()
        self.root.wait_window(dialog)
        return result.get()

    def _center_modal(self, window, width, height):
        try:
            self.root.update_idletasks()
            root_x = self.root.winfo_rootx()
            root_y = self.root.winfo_rooty()
            root_w = self.root.winfo_width()
            root_h = self.root.winfo_height()
            x = root_x + max((root_w - width) // 2, 0)
            y = root_y + max((root_h - height) // 2, 0)
            window.geometry(f"{width}x{height}+{x}+{y}")
        except Exception:
            pass

    def _apply_icon(self, toplevel):
        """Toplevel 창에 앱 아이콘을 적용한다."""
        try:
            _ico = os.path.join(BASE_DIR, "assets", "bot_converted.ico")
            if not os.path.isfile(_ico):
                _ico = os.path.join(BASE_DIR, "assets", "bot.ico")
            if os.path.isfile(_ico):
                toplevel.iconbitmap(_ico)
        except Exception:
            try:
                _png = os.path.join(BASE_DIR, "assets", "bot.ico")
                _img = tk.PhotoImage(file=_png)
                toplevel.iconphoto(True, _img)
            except Exception:
                pass

    def _trim_file(self, path, *, max_bytes=MAX_LOG_BYTES, keep_bytes=LOG_KEEP_BYTES):
        if keep_bytes >= max_bytes:
            keep_bytes = max_bytes // 2
        try:
            size = os.path.getsize(path)
        except OSError:
            return
        if size <= max_bytes:
            return
        try:
            with open(path, "rb") as fh:
                fh.seek(-keep_bytes, os.SEEK_END)
                data = fh.read()
            with open(path, "wb") as fh:
                fh.write(data)
        except OSError as exc:
            self._append_log(f"[WARN] 로그 정리 실패({path}): {exc}")

    def _build_sidebar_close_buttons(self):
        pass

    def _apply_sidebar_manual_state(self):
        pass

    def _toggle_sidebar_manual_section(self):
        pass

    def _add_tooltip(self, widget, text):
        """위젯에 마우스 오버 툴팁 추가."""
        tip = {"win": None}

        def _enter(_e):
            x = widget.winfo_rootx() + widget.winfo_width() // 2
            y = widget.winfo_rooty() + widget.winfo_height() + 4
            tip["win"] = tw = tk.Toplevel(self.root)
            tw.wm_overrideredirect(True)
            tw.wm_geometry(f"+{x}+{y}")
            tk.Label(tw, text=text, bg="#23293a", fg="#e0e6ff",
                     font=("Malgun Gothic", 8), relief="flat",
                     padx=8, pady=4).pack()

        def _leave(_e):
            if tip["win"]:
                try: tip["win"].destroy()
                except: pass
                tip["win"] = None

        widget.bind("<Enter>", _enter)
        widget.bind("<Leave>", _leave)


    def _add_settings_button(self):
        button = tk.Button(
            self.main_area,
            text="⚙",
            command=self.open_settings_modal,
            font=("Segoe UI Symbol", 14, "bold"),
            bg="#181A20",
            fg="#f5f7ff",
            relief=tk.FLAT,
            bd=0,
            activebackground="#1f2128",
            activeforeground="#ffffff",
            cursor="hand2",
            width=3,
        )
        button.place(relx=0.98, rely=0.02, anchor="ne")


    def _build_positions_panel(self):
        container = tk.Frame(self.main_area, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        container.grid(row=2, column=0, sticky="nsew", pady=(0, 0))
        container.columnconfigure(0, weight=1)
        container.rowconfigure(2, weight=1)
        self.positions_panel = container

        header = tk.Frame(container, bg="#181A20", highlightbackground="#343942", highlightthickness=1)
        header.grid(row=0, column=0, sticky="ew", pady=(0, 6))
        title = tk.Label(header, text=self._t("monitoring_panel", "모니터링"), fg="#f7f9ff", bg="#181A20", font=("Segoe UI", 13, "bold"))
        title.pack(side=tk.LEFT, padx=(10, 0))

        # ── RIGHT 쪽 배치 순서 (pack RIGHT는 역순으로 화면에 표시됨) ──
        # 최우측: 선택 청산
        self.position_quick_close_btn = tk.Button(
            header,
            text=self._t("close_selected","선택 청산"),
            command=self._close_selected_position,
            state=tk.DISABLED,
            bg="#1e2438", fg="#5a6280",
            relief=tk.FLAT, font=("Malgun Gothic", 8),
            padx=5, pady=2,
            activebackground="#2a3555", activeforeground="#ffffff",
            disabledforeground="#5a6280", cursor="hand2",
        )
        self.position_quick_close_btn.pack(side=tk.RIGHT, padx=(2, 6))
        self._add_tooltip(self.position_quick_close_btn,
                          "포지션 목록에서 선택 후 클릭하면 지정가 우선으로 청산합니다")

        # 구분선
        tk.Label(header, text="│", bg="#181A20", fg="#2a3040",
                 font=("Segoe UI", 11)).pack(side=tk.RIGHT, padx=(2, 2))

        # 새로고침 버튼 + 카운트다운 (선택 청산 왼쪽)
        refresh_btn = tk.Button(
            header, text="⟳",
            command=lambda: self.refresh_positions(auto=False),
            bg="#181A20", fg="#f7f9ff",
            relief=tk.FLAT, font=("Segoe UI Symbol", 11, "bold"),
            width=2, activebackground="#181b2a", activeforeground="#ffffff",
            cursor="hand2",
        )
        refresh_btn.pack(side=tk.RIGHT, padx=(2, 0))
        self._add_tooltip(refresh_btn, "포지션 목록 즉시 새로고침")
        self.position_refresh_label = tk.Label(
            header, text="--s",
            bg="#181A20", fg="#9aa5c6",
            font=("Segoe UI", 9, "bold"), width=5, anchor="e",
        )
        self.position_refresh_label.pack(side=tk.RIGHT, padx=(4, 0))

        # 구분선
        tk.Label(header, text="│", bg="#181A20", fg="#2a3040",
                 font=("Segoe UI", 11)).pack(side=tk.RIGHT, padx=(2, 2))

        # 일괄청산 토글 텍스트 (클릭으로 펼치기/접기)
        self._close_btns_hidden = False
        self._close_btn_toggle = None  # 버튼 없음
        self._close_btn_toggle_lbl = tk.Label(
            header,
            text=self._t("bulk_close_collapse", "일괄 청산 접기"),
            bg="#181A20", fg="#4a5270",
            font=("Malgun Gothic", 8), cursor="hand2",
        )
        self._close_btn_toggle_lbl.pack(side=tk.RIGHT, padx=(0, 6))
        self._close_btn_toggle_lbl.bind("<Button-1>", lambda _: self._toggle_close_btn_group())
        self._close_btn_toggle_lbl.bind("<Enter>", lambda _: self._close_btn_toggle_lbl.config(fg="#9aa5c6"))
        self._close_btn_toggle_lbl.bind("<Leave>", lambda _: self._close_btn_toggle_lbl.config(fg="#4a5270"))

        # 구분선 (일괄청산 버튼 그룹 왼쪽)
        self._close_btn_sep = tk.Label(header, text="│", bg="#181A20", fg="#2a3040",
                                       font=("Segoe UI", 11))
        self._close_btn_sep.pack(side=tk.RIGHT, padx=(2, 2))

        # 일괄 청산 버튼 그룹
        _profit_cfg = dict(bg="#163326", fg="#2EBD85", relief=tk.FLAT,
                           font=("Malgun Gothic", 8), padx=5, pady=2,
                           activebackground="#1e4a38", activeforeground="#5eddaa",
                           cursor="hand2")
        _loss_cfg   = dict(bg="#2e1a1a", fg="#F6465D", relief=tk.FLAT,
                           font=("Malgun Gothic", 8), padx=5, pady=2,
                           activebackground="#4a2222", activeforeground="#ff7a8a",
                           cursor="hand2")

        self._close_btn_group = tk.Frame(header, bg="#181A20")
        self._close_btn_group.pack(side=tk.RIGHT)
        tk.Button(self._close_btn_group, text=self._t("sidebar_close_loss_market","손실 · 시장가"),
                  command=lambda: self._close_positions_by_pnl("market", positive=False),
                  **_loss_cfg).pack(side=tk.LEFT, padx=(2, 0))
        tk.Button(self._close_btn_group, text=self._t("sidebar_close_loss_limit","손실 · 지정가"),
                  command=lambda: self._close_positions_by_pnl("limit", positive=False),
                  **_loss_cfg).pack(side=tk.LEFT, padx=(2, 0))
        tk.Button(self._close_btn_group, text=self._t("sidebar_close_profit_market","수익 · 시장가"),
                  command=lambda: self._close_positions_by_pnl("market", positive=True),
                  **_profit_cfg).pack(side=tk.LEFT, padx=(2, 0))
        tk.Button(self._close_btn_group, text=self._t("sidebar_close_profit_limit","수익 · 지정가"),
                  command=lambda: self._close_positions_by_pnl("limit", positive=True),
                  **_profit_cfg).pack(side=tk.LEFT, padx=(2, 0))

        self.positions_exit_label = tk.Label(container, text="", bg="#181A20", fg="#8e96b8", font=("Malgun Gothic", 9), anchor="w")
        self.positions_exit_label.grid(row=1, column=0, sticky="ew", padx=4, pady=(0, 2))
        self._update_positions_exit_summary()

        style = ttk.Style()
        try:
            style.theme_use("clam")
        except Exception:
            pass
        style.layout("Positions.Treeview", [("Treeview.treearea", {"sticky": "nswe"})])
        style.configure(
            "Positions.Treeview",
            background="#181A20",
            foreground="#f2f6ff",
            fieldbackground="#181A20",
            rowheight=26,
            font=("Segoe UI", 10),
            borderwidth=0,
            relief=tk.FLAT,
        )
        style.map("Positions.Treeview", background=[("selected", "#1f2a42")], foreground=[("selected", "#ffffff")])
        style.configure(
            "Positions.Treeview.Heading",
            background="#0c0e13",
            foreground="#f7f9ff",
            font=("Segoe UI", 11, "bold"),
            anchor="center",
            padding=(12, 6),
            borderwidth=0,
        )
        style.configure("PositionsNotebook.TNotebook", background="#181A20", borderwidth=0, tabmargins=(8, 8, 8, 0))
        style.configure("PositionsNotebook.TNotebook.Tab", background="#0e111a", foreground="#8b93b7", padding=(12, 4, 12, 2), font=("Malgun Gothic", 10, "bold"), borderwidth=0)
        style.map(
            "PositionsNotebook.TNotebook.Tab",
            background=[("selected", "#25304a"), ("!selected", "#0e111a")],
            foreground=[("selected", "#ffffff"), ("!selected", "#8b93b7")],
            padding=[("selected", (18, 8, 18, 6)), ("!selected", (12, 4, 12, 2))],
        )

        notebook = ttk.Notebook(container, style="PositionsNotebook.TNotebook")
        notebook.grid(row=2, column=0, sticky="nsew", padx=2, pady=(0, 4))

        positions_tab = tk.Frame(notebook, bg="#181A20")
        exit_tab = tk.Frame(notebook, bg="#181A20")
        monitor_tab = tk.Frame(notebook, bg="#0f1118")
        log_tab = tk.Frame(notebook, bg="#0a0d14")
        notebook.add(positions_tab, text="  " + self._t("positions_tab","포지션 목록") + "  ")
        notebook.add(exit_tab,     text="  " + self._t("exits_tab","익절 · Trailing") + "  ")
        notebook.add(monitor_tab,  text="  " + self._t("monitor_tab","모니터링") + "  ")
        notebook.add(log_tab,      text="  " + self._t("log_tab","로그") + "  ")
        self._build_monitor_tab(monitor_tab)
        self._build_log_tab(log_tab)

        table_frame = tk.Frame(positions_tab, bg="#181A20", highlightthickness=0)
        table_frame.pack(fill="both", expand=True)
        table_frame.columnconfigure(0, weight=1)
        table_frame.rowconfigure(0, weight=1)
        cols = ("symbol", "side", "leverage", "entry", "margin", "pnl")
        self.position_tree = ttk.Treeview(table_frame, columns=cols, show="headings", height=6, style="Positions.Treeview")
        headings = ["Symbol", "Long / Short", "Leverage (Entry x)", "Entry Price", "Margin (USDT)", "PNL / Fee"]
        widths = [120, 120, 150, 140, 130, 310]
        for col, header_text, width in zip(cols, headings, widths):
            self.position_tree.heading(col, text=header_text)
            self.position_tree.column(col, width=width, anchor="center")
        self.position_tree.grid(row=0, column=0, sticky="nsew")
        self.position_tree.bind("<<TreeviewSelect>>", self._update_position_action_state)
        # 컬럼 헤더 클릭으로 정렬
        for _col in cols:
            self.position_tree.heading(_col, command=lambda c=_col: self._sort_position_tree(c))
        self._pos_sort_col = None
        self._pos_sort_rev = False
        scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=self.position_tree.yview)
        scrollbar.grid(row=0, column=1, sticky="ns")
        self.position_tree.configure(yscrollcommand=scrollbar.set)
        self.position_tree.tag_configure("profit", foreground="#2EBD85")
        self.position_tree.tag_configure("loss", foreground="#F5475D")
        self.position_tree.tag_configure("odd", background="#181A20")
        self.position_tree.tag_configure("even", background="#1f2128")
        self.position_tree.tag_configure("long", foreground="#2EBD85")
        self.position_tree.tag_configure("short", foreground="#F6465D")
        self.position_tree.tag_configure("text-white", foreground="#f2f6ff")
        self.position_tree.tag_configure("manual", background="#332e12", foreground="#f7d66d")

        exit_frame = tk.Frame(exit_tab, bg="#181A20")
        exit_frame.pack(fill="both", expand=True)
        exit_frame.columnconfigure(0, weight=1)
        exit_frame.rowconfigure(0, weight=1)
        exit_cols = ("symbol", "roi", "partial", "trail", "progress")
        self.position_exit_tree = ttk.Treeview(exit_frame, columns=exit_cols, show="headings", height=6, style="Positions.Treeview")
        exit_headings = ["Symbol", "ROI %", "Partial", "Trailing", "Progress"]
        exit_widths = [140, 120, 220, 220, 220]
        for col, header_text, width in zip(exit_cols, exit_headings, exit_widths):
            self.position_exit_tree.heading(col, text=header_text)
            self.position_exit_tree.column(col, width=width, anchor="center")
        self.position_exit_tree.grid(row=0, column=0, sticky="nsew")
        exit_scroll = ttk.Scrollbar(exit_frame, orient="vertical", command=self.position_exit_tree.yview)
        exit_scroll.grid(row=0, column=1, sticky="ns")
        self.position_exit_tree.configure(yscrollcommand=exit_scroll.set)
        self.position_exit_tree.tag_configure("profit", foreground="#2EBD85")
        self.position_exit_tree.tag_configure("loss", foreground="#F5475D")
        self.position_exit_tree.tag_configure("odd", background="#181A20")
        self.position_exit_tree.tag_configure("even", background="#1f2128")

        self.refresh_positions()
        self._apply_manual_panel_visibility()

    def refresh_top_symbols(self):
        creds = self._require_credentials(silent=True, action="심볼 데이터 갱신")
        if not creds:
            return
        from binance_futures_bot1_1.main import get_top_symbols

        def on_success(result):
            self.top_symbols = result
            self.symbol_price_map = {item["symbol"]: item.get("price", 0) for item in result}

        limit = max(10, len(self.manual_symbols) * 5)
        self._run_async(get_top_symbols(creds, limit=limit), success_cb=on_success, fail_msg="심볼 시세 갱신 실패")

    def refresh_positions(self, auto=True):
        current_token = self.env_request_token
        creds = self._require_credentials(silent=auto, action="포지션 정보 갱신")
        if not creds:
            self._schedule_positions_refresh()
            return
        from binance_futures_bot1_1.main import fetch_open_positions

        if self.position_refresh_label and self.position_refresh_label.winfo_exists():
            self.position_refresh_label.config(text="…")

        def on_success(result, request_token=current_token):
            if request_token != self.env_request_token:
                return
            self.current_positions = {pos["symbol"]: pos for pos in result}
            for row in self.position_tree.get_children():
                self.position_tree.delete(row)
            exit_tree = getattr(self, "position_exit_tree", None)
            if exit_tree:
                for row in exit_tree.get_children():
                    exit_tree.delete(row)
            total_margin = 0.0
            total_nominal = 0.0
            now = time.time()
            active_symbols = set()
            for idx, pos in enumerate(result):
                symbol = pos["symbol"]
                active_symbols.add(symbol)
                direction = pos["side"]
                direction_display = "LONG" if direction == "LONG" else ("SHORT" if direction == "SHORT" else direction)
                leverage_raw = pos.get("leverage")
                try:
                    leverage_val = float(leverage_raw)
                except (TypeError, ValueError):
                    leverage_val = 0.0
                if leverage_val <= 0:
                    try:
                        amt_abs = abs(float(pos.get("amount", 0.0)))
                        mark_price = float(pos.get("markPrice", 0.0))
                        margin_val = float(pos.get("marginValue", 0.0))
                        if amt_abs > 0 and margin_val > 0:
                            leverage_val = (amt_abs * mark_price) / margin_val
                    except (TypeError, ValueError):
                        leverage_val = 0.0
                lev_text = f"{leverage_val:.0f}x" if leverage_val > 0 else "-"
                try:
                    qty_abs = abs(float(pos.get("amount", 0.0)))
                    entry_px = float(pos.get("entryPrice", 0.0))
                    mark_price = float(pos.get("markPrice", entry_px))
                    # ═══════════════════════════════════════════════════════════
                    # 🔧 수정: 진입 + 청산 모두 테이커 수수료 (0.05%)
                    # ═══════════════════════════════════════════════════════════
                    taker_fee = float(self.settings_data.get("taker_fee_pct", 0.0005) or 0.0005)
                    # 진입 수수료: 테이커
                    entry_fee = entry_px * qty_abs * taker_fee
                    # 청산 수수료: 테이커
                    exit_fee = mark_price * qty_abs * taker_fee
                    est_fee = max(entry_fee + exit_fee, 0.0)
                except (TypeError, ValueError):
                    est_fee = 0.0
                _pnl_net = float(pos["unRealizedProfit"]) - est_fee
                _net_sign = "+" if _pnl_net > 0 else ""
                _fee_part = f" | 수수료-{est_fee:.3f}" if est_fee >= 0.001 else ""
                pnl_text = f"{pos['unRealizedProfit']:.2f} ({pos['roiPercent']:.2f}%){_fee_part} → 순{_net_sign}{_pnl_net:.2f} USDT"
                margin_val = float(pos.get("marginValue", 0.0))
                tags = ["profit" if pos["unRealizedProfit"] >= 0 else "loss", "even" if idx % 2 == 0 else "odd"]
                highlight = self.manual_highlight.get(symbol)
                if highlight and highlight.get("expires", 0) > now:
                    highlight["activated"] = True
                    tags.append("manual")
                self.position_tree.insert(
                    "",
                    tk.END,
                    values=(
                        symbol,
                        direction_display,
                        lev_text,
                        f"{pos['entryPrice']:.4f}",
                        f"{margin_val:.2f}",
                        pnl_text,
                    ),
                    tags=tags,
                )
                if exit_tree:
                    roi_val = float(pos.get("roiPercent", 0.0) or 0.0)
                    partial_text = self._format_partial_status(roi_val)
                    trail_text = self._format_trail_status(roi_val)
                    progress_text = self._format_progress_status(roi_val)
                    exit_tags = ["profit" if roi_val >= 0 else "loss", "even" if idx % 2 == 0 else "odd"]
                    exit_tree.insert(
                        "",
                        tk.END,
                        values=(
                            symbol,
                            f"{roi_val:.2f}%",
                            partial_text,
                            trail_text,
                            progress_text,
                        ),
                        tags=exit_tags,
                    )
                total_margin += margin_val
                total_nominal += abs(pos.get("amount", 0.0)) * float(pos.get("markPrice", 0.0))
            # 미실현 합계 카드 즉시 반영 (활성 모드일 때만)
            try:
                if "unrealized_total" in getattr(self, "card_modes", {}).values():
                    _total_net = self._calc_unrealized_net()
                    _total_gross = self._calc_unrealized_gross()
                    _total_fee = _total_gross - _total_net
                    _sign = "+" if _total_net > 0 else ""
                    _col = "#2EBD85" if _total_net > 0 else ("#F6465D" if _total_net < 0 else "#ffffff")
                    _lbl = self.stats_labels.get("unrealized_total")
                    if _lbl:
                        # 메인: 순손익만 표시 (수수료 제거 → 짤림 방지)
                        _lbl.config(text=self._trim_usdt(f"{_sign}{_total_net:.2f} USDT"), fg=_col)
                    # 서브라인: 수수료 (카드 아래 작은 글씨)
                    _sub = getattr(self, "_card_sub_labels", {}).get("unrealized_total")
                    if _sub:
                        try:
                            if abs(_total_fee) >= 0.001:
                                _sub.config(text=f"수수료 -{_total_fee:.3f} USDT 차감 후")
                            else:
                                _sub.config(text="")
                        except Exception:
                            pass
            except Exception:
                pass

            self._cleanup_manual_highlight(active_symbols)
            self.last_total_margin = total_margin
            label = self.stats_labels.get("notional")
            if label:
                label.config(text=f"{total_margin:.2f} / {total_nominal:.2f} USDT")
            self._update_positions_exit_summary()

        self._run_async(fetch_open_positions(creds), success_cb=on_success, fail_msg="포지션 정보 실패")
        self._schedule_positions_refresh()
        self._update_position_action_state()

    def _sort_position_tree(self, col):
        """포지션 트리뷰 헤더 클릭 정렬."""
        tree = getattr(self, "position_tree", None)
        if not tree:
            return
        reverse = (self._pos_sort_col == col and not self._pos_sort_rev)
        self._pos_sort_col = col
        self._pos_sort_rev = reverse
        data = [(tree.set(k, col), k) for k in tree.get_children("")]
        try:
            data.sort(key=lambda t: float(t[0].replace("%","").replace(" USDT","").replace(",","").split()[0]), reverse=reverse)
        except (ValueError, IndexError):
            data.sort(key=lambda t: t[0].lower(), reverse=reverse)
        for i, (_, k) in enumerate(data):
            tree.move(k, "", i)
            tree.item(k, tags=("even" if i % 2 == 0 else "odd",))
        # 정렬 방향 표시
        for c in ("symbol","side","leverage","entry","margin","pnl"):
            txt = tree.heading(c)["text"].rstrip(" ▲▼")
            tree.heading(c, text=txt + (" ▲" if c==col and not reverse else " ▼" if c==col else ""))


    def _update_position_action_state(self, *_):
        btn = getattr(self, "position_quick_close_btn", None)
        tree = getattr(self, "position_tree", None)
        if not btn or not tree:
            return
        try:
            selection = tree.selection()
        except Exception:
            selection = ()
        if selection:
            btn.configure(state=tk.NORMAL, fg="#d8def8", bg="#253055")
        else:
            btn.configure(state=tk.DISABLED, fg="#5a6280", bg="#1e2438")

    def _close_selected_position(self):
        tree = getattr(self, "position_tree", None)
        if not tree:
            return
        try:
            selection = tree.selection()
        except Exception:
            selection = ()
        if not selection:
            self._show_info(self._t("close_selected","Close selected"), self._t("close_sel_select_first","Select a symbol from the positions list first."))
            return
        try:
            values = tree.item(selection[0], "values")
        except Exception:
            values = None
        symbol = values[0] if values else None
        if not symbol:
            self._show_warning(self._t("close_selected","Close selected"), self._t("close_sel_no_sym","Could not find a symbol in the selected row."))
            return
        self._manual_trade(symbol, "close_limit_then_market")

    def _schedule_notification_poll(self):
        self.root.after(4000, self._poll_notifications)

    def _burst_refresh_positions(self, delay_ms=1500):
        self.refresh_positions(auto=False)
        if delay_ms and delay_ms > 0:
            self.root.after(delay_ms, lambda: self.refresh_positions(auto=False))

    def _clear_env_specific_views(self):
        self.current_positions = {}
        self.manual_highlight = {}
        try:
            _lbl = self.stats_labels.get("unrealized_total")
            if _lbl:
                _lbl.config(text="0.00 USDT", fg="#ffffff")
            _sub = getattr(self, "_card_sub_labels", {}).get("unrealized_total")
            if _sub:
                _sub.config(text="")
        except Exception:
            pass
        self.symbol_filters_cache = {}
        if hasattr(self, "position_tree") and self.position_tree:
            for row in self.position_tree.get_children():
                self.position_tree.delete(row)
        if hasattr(self, "position_exit_tree") and self.position_exit_tree:
            for row in self.position_exit_tree.get_children():
                self.position_exit_tree.delete(row)
        for _tree_attr in ("monitor_watch_tree", "monitor_reject_tree", "monitor_autotune_tree"):
            _t = getattr(self, _tree_attr, None)
            if _t:
                try:
                    for row in _t.get_children(): _t.delete(row)
                except Exception: pass
        self.monitor_filter_stat_labels = {}
        self.top_symbols = []
        self.symbol_price_map = {}
        self.last_total_margin = 0.0
        self.next_position_refresh_ts = None
        if self.position_refresh_label and self.position_refresh_label.winfo_exists():
            self.position_refresh_label.config(text="--s")
        for key, label in self.stats_labels.items():
            default_value = self.stat_defaults.get(key)
            if default_value is not None:
                try:
                    label.config(text=default_value)
                except Exception:
                    continue
        self.last_account_balance = 0.0

    def _mark_manual_highlight(self, symbol, lifespan=900):
        if not symbol:
            return
        try:
            lifespan = max(120, int(lifespan))
        except (TypeError, ValueError):
            lifespan = 900
        self.manual_highlight[symbol] = {"expires": time.time() + lifespan, "activated": False}

    def _cleanup_manual_highlight(self, active_symbols=None):
        if not getattr(self, "manual_highlight", None):
            return
        now = time.time()
        active = set(active_symbols or [])
        for symbol, data in list(self.manual_highlight.items()):
            expires = data.get("expires", now)
            activated = data.get("activated", False)
            if expires <= now:
                self.manual_highlight.pop(symbol, None)
            elif activated and symbol not in active:
                self.manual_highlight.pop(symbol, None)

    def _update_risk_ack(self, value, *, ack_key="risk_acknowledged"):
        """체크박스 상태 변경 → 동의 완료 여부 재계산 + 증적 로그 기록."""
        acknowledged = bool(value)
        self.settings_data[ack_key] = acknowledged
        # 완전 동의 = ack1 AND ack2 AND 버전 일치
        ack1 = bool(self.settings_data.get("risk_ack1", False))
        ack2 = bool(self.settings_data.get("risk_ack2", False))
        fully_acked = ack1 and ack2
        self.risk_acknowledged = fully_acked
        self.settings_data["risk_acknowledged"] = fully_acked
        if fully_acked:
            self.settings_data["consent_version"] = CONSENT_VERSION
        self._save_json(CONFIG_PATH, self.settings_data)
        # A2: 동의 증적 로그 (로컬 파일)
        self._write_consent_audit(ack1=ack1, ack2=ack2, fully_acked=fully_acked)
        if fully_acked:
            self._append_log("[INFO] " + self._t("risk_ack_complete", "위험 고지 동의 완료"))
            # 동의 완료 후 API 키 미설정 시 → 시작하기(레퍼럴 온보딩) 다이얼로그
            has_key = bool(os.environ.get("BINANCE_API_KEY") or os.environ.get("TESTNET_API_KEY"))
            if not has_key:
                self.root.after(300, self._show_referral_onboarding)
        elif not acknowledged:
            self._append_log("[WARN] " + self._t("risk_ack_revoked", "위험 고지 동의가 해제되었습니다"))

    def _write_consent_audit(self, *, ack1: bool, ack2: bool, fully_acked: bool):
        """A2: 동의 시각·버전·항목 상태를 로컬 파일에 기록."""
        import datetime
        try:
            log_dir = os.path.join(BASE_DIR, "binance_futures_bot1_1", "logs")
            os.makedirs(log_dir, exist_ok=True)
            audit_path = os.path.join(log_dir, "consent_audit.log")
            ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            entry = (
                f"{ts} | app={TITLE} | consent_version={CONSENT_VERSION} | ack1={ack1} | ack2={ack2} "
                f"| fully_acked={fully_acked}\n"
            )
            with open(audit_path, "a", encoding="utf-8") as f:
                f.write(entry)
        except Exception as exc:
            self._append_log(f"[WARN] consent audit write failed: {exc}")

    def _check_consent_version_early(self):
        """앱 시작 시 동의 버전 확인 (로그 없이 조용히 처리)."""
        saved = self.settings_data.get("consent_version", "")
        if saved and saved != CONSENT_VERSION:
            self.risk_acknowledged = False
            self.settings_data["risk_acknowledged"] = False
            self.settings_data["risk_ack1"] = False
            self.settings_data["risk_ack2"] = False
            self.settings_data["consent_version"] = ""
            self._save_json(CONFIG_PATH, self.settings_data)

    def _check_consent_version(self):
        """A1: 저장된 동의 버전이 현재 버전과 다르면 재동의 필요."""
        saved_version = self.settings_data.get("consent_version", "")
        if saved_version != CONSENT_VERSION and self.risk_acknowledged:
            self.risk_acknowledged = False
            self.settings_data["risk_acknowledged"] = False
            self.settings_data["risk_ack1"] = False
            self.settings_data["risk_ack2"] = False
            self.settings_data["consent_version"] = ""
            self._save_json(CONFIG_PATH, self.settings_data)
            self._append_log("[WARN] " + self._t(
                "consent_version_changed",
                "고지문이 업데이트되었습니다. 필수 동의 탭에서 다시 동의해주세요."
            ))

    def _toggle_close_btn_group(self):
        """수익/손실 일괄 청산 버튼 그룹 숨기기/보이기."""
        self._close_btns_hidden = not getattr(self, "_close_btns_hidden", False)
        grp  = getattr(self, "_close_btn_group", None)
        sep  = getattr(self, "_close_btn_sep", None)
        lbl  = getattr(self, "_close_btn_toggle_lbl", None)
        if self._close_btns_hidden:
            if grp and grp.winfo_ismapped():
                grp.pack_forget()
            if sep and sep.winfo_ismapped():
                sep.pack_forget()
            if lbl and lbl.winfo_exists():
                lbl.config(text=self._t("bulk_close_expand","일괄 청산 펼치기"))
        else:
            if sep and not sep.winfo_ismapped():
                sep.pack(side=tk.RIGHT, padx=(2, 2))
            if grp and not grp.winfo_ismapped():
                grp.pack(side=tk.RIGHT)
            if lbl and lbl.winfo_exists():
                lbl.config(text=self._t("bulk_close_collapse","일괄 청산 접기"))


    def _update_positions_exit_summary(self):
        label = getattr(self, "positions_exit_label", None)
        if not label or not label.winfo_exists():
            return
        stop_pct = self._format_percent_display(self.settings_data.get("max_loss_per_position", 0.0))
        partial_text = "Partial: OFF"
        if self.settings_data.get("enable_partial_take_profit", True):
            partial_levels = self.settings_data.get("partial_tp_levels", [])
            if partial_levels:
                partial_text = f"Partial: {self._format_partial_levels_for_display(partial_levels)}"
        trail_text = "Trail: OFF"
        if self.settings_data.get("enable_atr_trailing_stop", True):
            trail_text = f"Trail: >= {self._format_percent_display(self.settings_data.get('trail_activate_pnl_pct', 0.0))} ATR x {self.settings_data.get('trail_atr_mult', 3.0):.1f}"
        progress_text = "Progress: OFF"
        if self.settings_data.get("enable_progress_stop", True):
            drawdown = self._format_percent_display(self.settings_data.get('progress_stop_drawdown_from_mfe', 0.15))
            stale = int(self.settings_data.get('progress_stop_no_new_high_sec', 1800))
            progress_text = f"Progress: stale {stale}s / DD {drawdown}"
        _pe_label = self._t("profit_exit_based", "(based on Profit Exit settings)")
        summary = f"{partial_text} | {trail_text} | {progress_text} · {_pe_label}"
        label.configure(text=summary)

    def _set_auto_boost_position_pct(self, enabled: bool):
        self.settings_data["auto_boost_position_pct"] = bool(enabled)
        state = "ON" if enabled else "OFF"
        self._append_log(
            f"[INFO] " + (f"Min-margin auto-boost: {state}" if self.language == "en"
                          else f"최소 증거금 자동 보정: {state}"))
        self._sync_position_pct_status()

    def _refresh_system_env(self):
        if winreg is None:
            return
        env_vars = [
            "TESTNET_API_KEY",
            "TESTNET_API_SECRET",
            "BINANCE_API_KEY",
            "BINANCE_API_SECRET",
        ]
        def query_env(name):
            paths = [
                (winreg.HKEY_CURRENT_USER, r"Environment"),
                (winreg.HKEY_LOCAL_MACHINE, r"SYSTEM\\CurrentControlSet\\Control\\Session Manager\\Environment"),
            ]
            for hive, subkey in paths:
                try:
                    with winreg.OpenKey(hive, subkey) as key:
                        value, _ = winreg.QueryValueEx(key, name)
                        if value:
                            return value
                except FileNotFoundError:
                    continue
                except OSError:
                    continue
            return None
        for name in env_vars:
            value = query_env(name)
            if value:
                os.environ[name] = value

    def _get_symbol_filters(self, symbol, creds):
        if not symbol:
            return None
        now = time.time()
        cache = self.symbol_filters_cache.get(symbol)
        if cache and cache.get("expires", 0) > now:
            return cache.get("filters")
        try:
            from binance_futures_bot1_1.main import get_symbol_filters
            future = asyncio.run_coroutine_threadsafe(get_symbol_filters(creds, symbol), self.loop)
            filters = future.result(timeout=5)
            if filters:
                self.symbol_filters_cache[symbol] = {"filters": filters, "expires": now + 600}
            return filters
        except Exception as exc:
            self._append_log(f"[WARN] 심볼 필터 조회 실패({symbol}): {exc}")
            return None

    def _min_notional_from_filters(self, filters):
        if not filters:
            return 0.0
        for f in filters:
            if f.get("filterType") in {"MIN_NOTIONAL", "NOTIONAL"}:
                try:
                    value = f.get("notional") or f.get("minNotional")
                    return float(value)
                except (TypeError, ValueError):
                    continue
        return 0.0

    def _blocking_fetch_available_balance(self, creds):
        try:
            from binance_futures_bot1_1.main import fetch_account_balance
            future = asyncio.run_coroutine_threadsafe(fetch_account_balance(creds), self.loop)
            data = future.result(timeout=5)
            available = data.get("availableBalance")
            if available is None:
                available = data.get("totalWalletBalance", 0.0)
            return float(available or 0.0)
        except Exception as exc:
            self._append_log(f"[WARN] 계좌 잔고 확인 실패: {exc}")
            return float(self.last_account_balance or 0.0)

    def _ensure_min_notional_percent(self, symbol, percent, ref_price, creds):
        if percent is None or percent <= 0 or not ref_price or ref_price <= 0:
            return percent
        if not bool(self.settings_data.get("auto_boost_position_pct", False)):
            return percent
        filters = self._get_symbol_filters(symbol, creds)
        min_notional = self._min_notional_from_filters(filters)
        if min_notional <= 0:
            return percent
        available = float(self.last_account_balance or 0.0)
        if available <= 0:
            available = self._blocking_fetch_available_balance(creds)
        if available <= 0:
            return percent
        needed_percent = (min_notional / available) * 100.0
        if needed_percent <= percent:
            return percent
        boosted = min(100.0, needed_percent * 1.05)
        self._append_log(
            f"[INFO] " + (f"Boosting {symbol} position ratio {percent:.2f}% → {boosted:.2f}% to meet min margin" if self.language == "en"
                          else f"최소 증거금 충족을 위해 {symbol} 포지션 비율을 {percent:.2f}% → {boosted:.2f}% 로 조정합니다.")
        )
        return boosted

    def _apply_order_filters(self, symbol, quantity, price, *, reduce_only, creds):
        filters = self._get_symbol_filters(symbol, creds)
        if not filters:
            return quantity, price
        min_qty = 0.0
        step = 0.0
        tick = 0.0
        for f in filters:
            ftype = f.get("filterType")
            if ftype == "LOT_SIZE":
                try:
                    min_qty = float(f.get("minQty", min_qty))
                    step = float(f.get("stepSize", step))
                except (TypeError, ValueError):
                    pass
            elif ftype == "PRICE_FILTER":
                try:
                    tick = float(f.get("tickSize", tick))
                except (TypeError, ValueError):
                    pass
        def quantize(value, step_size, *, round_up=False):
            if value is None or step_size <= 0:
                return value
            if round_up:
                return math.ceil(value / step_size) * step_size
            return math.floor(value / step_size) * step_size
        if quantity is not None:
            round_up = not reduce_only
            if step > 0:
                quantity = quantize(quantity, step, round_up=round_up)
            if min_qty > 0 and quantity < min_qty:
                quantity = min_qty
        if price is not None and tick > 0:
            price = quantize(price, tick, round_up=not reduce_only)
        return quantity, price

    def _apply_env_switch(self, target_env):
        if target_env not in {"TESTNET", "LIVE"}:
            return
        # ── 현재 환경 패널 값 저장 (복원용) ─────────────────────────────────
        self._save_env_panel_snapshot(self.env_mode)
        self.env_request_token += 1
        self.env_mode = target_env
        self.state_data["env_label"] = self.env_mode
        self.settings_data["default_env_testnet"] = self.env_mode == "TESTNET"
        self._save_json(STATE_PATH, self.state_data)
        self._save_json(CONFIG_PATH, self.settings_data)
        self._render_env_toggle()
        self._clear_env_specific_views()
        self._append_log(f"[INFO] 환경 전환: {target_env}")
        # ── 새 환경의 이전 캐시 복원 (있으면) ────────────────────────────────
        self._restore_env_panel_snapshot(target_env)
        self.refresh_positions(auto=False)
        self._refresh_stats_after_trade()

    def _save_env_panel_snapshot(self, env_key: str):
        """현재 환경의 상단 패널 값을 캐시에 저장."""
        if not env_key or not hasattr(self, "stats_labels"):
            return
        if not hasattr(self, "_env_panel_cache"):
            self._env_panel_cache = {}
        snapshot = {}
        for key, lbl in self.stats_labels.items():
            try:
                snapshot[key] = {"text": lbl.cget("text"), "fg": lbl.cget("fg")}
            except Exception:
                pass
        # last_account_balance도 저장
        snapshot["_balance"] = getattr(self, "last_account_balance", 0.0)
        snapshot["_positions"] = dict(getattr(self, "current_positions", {}))
        self._env_panel_cache[env_key] = snapshot

    def _restore_env_panel_snapshot(self, env_key: str):
        """환경의 이전 캐시가 있으면 상단 패널에 복원."""
        cache = getattr(self, "_env_panel_cache", {})
        snapshot = cache.get(env_key)
        if not snapshot:
            return
        # 잔고/포지션 복원
        bal = snapshot.get("_balance", 0.0)
        self.last_account_balance = float(bal)
        self.current_positions = dict(snapshot.get("_positions", {}))
        # 라벨 복원
        for key, state in snapshot.items():
            if key.startswith("_"):
                continue
            lbl = self.stats_labels.get(key)
            if lbl:
                try:
                    lbl.config(text=state.get("text", "—"), fg=state.get("fg", "#ffffff"))
                except Exception:
                    pass

    def _start_position_countdown(self):
        if self.position_countdown_job:
            try:
                self.root.after_cancel(self.position_countdown_job)
            except Exception:
                pass
            self.position_countdown_job = None
        self._update_position_countdown()

    def _update_position_countdown(self):
        if not self.position_refresh_label or not self.position_refresh_label.winfo_exists():
            self.position_countdown_job = None
            return
        if self.next_position_refresh_ts is None:
            self.position_refresh_label.config(text="--s")
            self.position_countdown_job = None
            return
        remaining = max(0.0, self.next_position_refresh_ts - time.time())
        self.position_refresh_label.config(text=f"{remaining:0.1f}s")
        if remaining <= 0:
            self.position_countdown_job = None
            return
        self.position_countdown_job = self.root.after(200, self._update_position_countdown)

    def _schedule_positions_refresh(self):
        if self.position_refresh_job:
            self.root.after_cancel(self.position_refresh_job)
        interval_ms = 10000
        self.next_position_refresh_ts = time.time() + (interval_ms / 1000.0)
        self.position_refresh_job = self.root.after(interval_ms, lambda: self.refresh_positions(auto=True))
        self._start_position_countdown()

    def _schedule_stats_refresh(self):
        if self.stats_refresh_job:
            self.root.after_cancel(self.stats_refresh_job)
        self._update_account_balance()
        self._update_income_stats()
        self.stats_refresh_job = self.root.after(20000, self._schedule_stats_refresh)

    def _refresh_panels_from_trade_log(self):
        """API 없이 trade_history.jsonl만으로 패널 즉시 갱신."""
        try:
            now_ms = int(time.time() * 1000)
            tl = self._reconstruct_income_from_trade_log(now_ms)
            self._apply_income_history([], now_ms)
        except Exception as exc:
            self._append_log(f"[WARN] trade_log 패널 갱신 실패: {exc}")

    def _refresh_stats_after_trade(self):
        try:
            self._refresh_panels_from_trade_log()   # trade_log로 즉시 갱신
            self._update_account_balance()
            self._update_income_stats()             # API로 추가 갱신
        except Exception as exc:
            self._append_log(f"[WARN] 통계 갱신 실패: {exc}")

    def _trigger_stat_refresh(self, delay_ms=0):
        if delay_ms <= 0:
            self._refresh_stats_after_trade()
        else:
            self.root.after(delay_ms, self._refresh_stats_after_trade)

    def _set_auto_tune_state_value(self, key, text, color=None):
        label = self.auto_tune_status_labels.get(key)
        if not label:
            return
        try:
            if not label.winfo_exists():
                return
        except Exception:
            return
        label.configure(text=text if text is not None else "--", fg=color or "#f5f7ff")

    def _sync_position_pct_status(self, value=None):
        if value is None:
            field = getattr(self, "trade_field_vars", {}).get("position_pct")
            if field is not None:
                try:
                    value = float(field.get())
                except (TypeError, ValueError, tk.TclError):
                    value = None
        if value is None:
            try:
                value = float(self.settings_data.get("position_pct", 0.05))
            except (TypeError, ValueError):
                value = 0.0
        pct_text = f"{max(0.0, value) * 100:.2f}%"
        if bool(self.settings_data.get("auto_boost_position_pct", False)):
            pct_text += " (AUTO)"
        self._set_auto_tune_state_value("position_pct", pct_text)

    def _reset_kill_switch(self):
        """킬스위치 수동 초기화: 엔진의 kill_switch_triggered를 해제하고 즉시 진입 재개."""
        _msg = (
            "킬스위치를 수동으로 초기화합니다.\n진입 차단이 해제되고 즉시 거래가 재개됩니다.\n\n계속하시겠습니까?"
            if self.language == "ko" else
            "Manually reset Kill Switch.\nEntry blocking will be lifted and trading resumes immediately.\n\nContinue?"
        )
        if not self._ask_yes_no("Kill Switch Reset", _msg):
            return
        # [PATCH-13] self.engine는 존재하지 않음 → main.current_engine 참조
        try:
            from binance_futures_bot1_1 import main as _eng_main
            engine = getattr(_eng_main, "current_engine", None)
        except Exception:
            engine = None
        if engine is not None:
            engine.kill_switch_triggered = False
            engine.kill_switch_release_ts = 0.0
            engine.kill_switch_reason = ""
            self._append_log("[INFO] 킬스위치 수동 초기화 완료 — 진입 차단 해제")
            _done = (
                "킬스위치가 초기화되었습니다.\n진입 차단이 해제되어 즉시 거래가 가능합니다."
                if self.language == "ko" else
                "Kill Switch has been reset.\nEntry blocking is lifted — trading can resume immediately."
            )
            self._show_info("Kill Switch Reset", _done)
        else:
            self._append_log("[INFO] 킬스위치 초기화: 엔진이 실행 중이 아닙니다.")
            _no_engine = (
                "엔진이 실행 중이 아닙니다.\n엔진 시작 시 킬스위치는 자동으로 초기 상태입니다."
                if self.language == "ko" else
                "Engine is not running.\nKill switch is automatically inactive on engine start."
            )
            self._show_info("Kill Switch Reset", _no_engine)

    def _reset_auto_tune_state(self):
        state_path = os.path.join(os.path.dirname(LOG_PATH), "auto_tuner_state.json")
        _msg = (
            "Auto-Tune 상태 파일과 오염된 파라미터(모멘텀/레버리지 등)를 초기값으로 복원합니다. 계속하시겠습니까?"
            if self.language == "ko" else
            "Reset Auto-Tune state and restore polluted params (momentum/leverage) to defaults. Continue?"
        )
        if not self._ask_yes_no("Reset Auto-Tune", _msg):
            return
        try:
            if os.path.exists(state_path):
                os.remove(state_path)
        except OSError as exc:
            self._append_log(f"[WARN] Reset Auto-Tune 실패: {exc}")
            self._show_error("Reset Auto-Tune", self._t("auto_tune_reset_fail","Failed to delete the state file.") + f"\n{exc}")
            return

        # ── 오염된 파라미터 기본값 복원 ────────────────────────────────────
        # auto-tune이 config에 극단값(momentum=0.007, leverage=40x 등)을 쓴 경우
        # settings.json에서도 기본값으로 덮어씀
        _defaults_to_restore = {
            "momentum_min_long":  0.003,
            "momentum_min_short": -0.003,
            "volatility_min":     0.001,
            "leverage_min":       5,
            "leverage_max":       20,
            "max_loss_per_position": 18.0,
        }
        _restored = []
        for _key, _default in _defaults_to_restore.items():
            _cur = self.settings_data.get(_key)
            if _cur is not None:
                # 기존값이 기본값에서 크게 벗어난 경우만 복원
                try:
                    _cur_f = float(_cur)
                    _def_f = float(_default)
                    if _key == "momentum_min_long" and _cur_f > 0.005:
                        self.settings_data[_key] = _default
                        _restored.append(f"{_key}: {_cur_f:.4f}→{_default}")
                    elif _key == "leverage_min" and _cur_f > 20:
                        self.settings_data[_key] = _default
                        _restored.append(f"{_key}: {_cur_f:.0f}→{_default}")
                    elif _key == "leverage_max" and _cur_f > 30:
                        self.settings_data[_key] = _default
                        _restored.append(f"{_key}: {_cur_f:.0f}→{_default}")
                    elif _key == "max_loss_per_position" and _cur_f < 5.0:
                        self.settings_data[_key] = _default
                        _restored.append(f"{_key}: {_cur_f:.1f}→{_default}")
                except (TypeError, ValueError):
                    pass
        if _restored:
            self._save_json(CONFIG_PATH, self.settings_data)
            self._append_log("[INFO] 오염 파라미터 복원: " + ", ".join(_restored))

        self._append_log("[INFO] Auto-Tune 상태 파일 삭제 완료. 엔진 재시작 후 기본값 적용됩니다.")
        _done_msg = (
            f"학습 상태가 초기화되었습니다.{chr(10)}복원된 파라미터: {', '.join(_restored) if _restored else '없음 (정상 범위)'}{chr(10)}엔진을 재시작하면 새 기본값이 적용됩니다."
            if self.language == "ko" else
            f"State reset complete.{chr(10)}Restored: {', '.join(_restored) if _restored else 'none (within normal range)'}{chr(10)}Restart engine to apply defaults."
        )
        self._show_info("Reset Auto-Tune", _done_msg)
        self._prompt_restart_after_setting_change(self._t("auto_tune_reset_restart","Auto-Tune state has been reset. Restart the engine now?"))

    def _prompt_restart_after_setting_change(self, reason: str = ""):
        if not reason:
            reason = self._t("trade_settings_restart", "Settings changed. Restart the engine now?")
        if not self.engine_running or self._restart_in_progress:
            return
        if getattr(self, "_restart_prompt_open", False):
            return
        self._restart_prompt_open = True
        try:
            if self._ask_yes_no(self._t("engine_restart_title","Engine restart required"), reason):
                self._append_log(self._t("engine_restart_log","[INFO] Restarting engine after settings change"))
                self._restart_in_progress = True
                self._restart_notice_pending = True

                def _restart_then_clear():
                    try:
                        self.start_engine()
                    finally:
                        self._restart_in_progress = False

                self.stop_engine(on_stopped=_restart_then_clear)
            else:
                self._restart_notice_pending = False
                self._append_log(self._t("engine_restart_pending","[INFO] Engine restart deferred"))
        finally:
            self._restart_prompt_open = False

    def _maybe_auto_launch_engine(self):
        if getattr(self, "_auto_launch_checked", False):
            return
        self._auto_launch_checked = True
        if not bool(self.settings_data.get("auto_start", False)):
            return
        if self.engine_running:
            return
        if not self.risk_acknowledged:
            self._append_log("[INFO] 자동 실행이 설정되어 있지만 위험 동의가 필요합니다.")
            return
        self._append_log("[INFO] 자동 실행 설정에 따라 엔진을 시작합니다.")
        self.start_engine()

    def _refresh_auto_tune_state(self):
        # ── stale 위젯 정리 ──────────────────────────────────────────────
        if self.auto_tune_status_labels:
            stale = []
            for key, widget in self.auto_tune_status_labels.items():
                try:
                    if not widget.winfo_exists():
                        stale.append(key)
                except Exception:
                    stale.append(key)
            for key in stale:
                self.auto_tune_status_labels.pop(key, None)
        if self.auto_tune_last_update_label is not None:
            try:
                if not self.auto_tune_last_update_label.winfo_exists():
                    self.auto_tune_last_update_label = None
            except Exception:
                self.auto_tune_last_update_label = None
        if not self.auto_tune_status_labels:
            return
        data = self._load_json(AUTO_TUNE_STATE_PATH, default={})
        manual_override = not bool(self.settings_data.get("auto_tune_enabled", True))
        if not data:
            for key in self.auto_tune_status_labels:
                self._set_auto_tune_state_value(key, "--", "#62697f")
            if self.auto_tune_last_update_label:
                label_text = ("Updated: No data" if self.language == "en" else "업데이트: 데이터 없음")
                if manual_override:
                    label_text += (" · Manual mode" if self.language == "en" else " · 수동 모드")
                self.auto_tune_last_update_label.configure(
                    text=label_text,
                    fg="#F0B90B" if manual_override else "#F6465D",
                )
            # 데이터 없음 → 배너 숨김 (리셋 후 잔류 방지)
            _banner = getattr(self, "_trade_block_banner", None)
            if _banner and getattr(self, "_trade_block_banner_visible", False):
                try:
                    if _banner.winfo_exists():
                        _banner.pack_forget()
                        self._trade_block_banner_visible = False
                except Exception:
                    pass
            return

        current = data.get("current") or {}
        meta = data.get("meta") or {}
        lifecycle = data.get("lifecycle") or {}
        active_snapshot = lifecycle.get("active") or {}
        staged_snapshot = lifecycle.get("staged") or {}
        proposed_snapshot = lifecycle.get("proposed") or {}
        active_metrics = active_snapshot.get("metrics") or {}
        hyst = data.get("hysteresis") or {}

        def fmt_float(value):
            try:
                return f"{float(value):.4f}"
            except (TypeError, ValueError):
                return "--"

        def fmt_int(value):
            try:
                return f"{int(value)}"
            except (TypeError, ValueError):
                return "--"

        def fmt_percent(value, digits=1):
            try:
                return f"{float(value) * 100:.{digits}f}%"
            except (TypeError, ValueError):
                return "--"

        def fmt_small(value, digits=4):
            try:
                return f"{float(value):.{digits}f}"
            except (TypeError, ValueError):
                return "--"

        self._set_auto_tune_state_value("momentum_long", fmt_float(current.get("momentum_min_long")))
        self._set_auto_tune_state_value("momentum_short", fmt_float(current.get("momentum_min_short")))
        self._set_auto_tune_state_value("volatility_min", fmt_float(current.get("volatility_min")))
        # watch_limit/max_open_symbols: 거래설정 탭 필드에서 직접 관리 → 여기선 표시 생략

        entry_pct = current.get("position_pct")
        if entry_pct is not None:
            try:
                pct_text = f"{float(entry_pct) * 100:.1f}%"
                if bool(self.settings_data.get("auto_boost_position_pct", False)):
                    pct_text += " (AUTO)"
            except (TypeError, ValueError):
                pct_text = "--"
            self._set_auto_tune_state_value("position_pct", pct_text)
        else:
            self._set_auto_tune_state_value("position_pct", "--")

        lev_min = current.get("leverage_min")
        lev_max = current.get("leverage_max")
        if lev_min is not None and lev_max is not None:
            try:
                lev_text = f"{int(lev_min)}x ~ {int(lev_max)}x"
            except (TypeError, ValueError):
                lev_text = "--"
            self._set_auto_tune_state_value("leverage_range", lev_text)
        else:
            self._set_auto_tune_state_value("leverage_range", "--")

        stop_loss = current.get("max_loss_per_position")
        if stop_loss is not None:
            try:
                stop_text = f"{float(stop_loss):.1f}%"
            except (TypeError, ValueError):
                stop_text = "--"
            self._set_auto_tune_state_value("stop_loss", stop_text)
        else:
            self._set_auto_tune_state_value("stop_loss", "--")

        # confidence + noise 통합 (1셀)
        _conf = active_metrics.get("confidence")
        _noise = active_metrics.get("noise_index")
        if _conf is not None:
            _conf_text = fmt_percent(_conf, 1)
            if _noise is not None:
                _conf_text += f"  |  noise {fmt_small(_noise)}"
        else:
            _conf_text = "--"
        self._set_auto_tune_state_value("confidence", _conf_text)
        # pass_rate + entry_rate 통합 (1셀) — fill_rate 제거(중복)
        pass_rate = active_metrics.get("pass_rate")
        entry_rate = active_metrics.get("entry_rate")
        if pass_rate is not None or entry_rate is not None:
            pass_entry_text = f"Pass {fmt_percent(pass_rate, 1)} / Entry {fmt_percent(entry_rate, 2)}"
        else:
            pass_entry_text = "--"
        self._set_auto_tune_state_value("pass_entry", pass_entry_text)
        pnl_30m = active_metrics.get("pnl_30m")
        pnl_fee = active_metrics.get("pnl_slow_fee", 0.0)
        if pnl_30m is not None:
            try:
                # REALIZED_PNL + COMMISSION = 순손익 (수수료는 음수로 들어옴)
                net = float(pnl_30m) + float(pnl_fee or 0.0)
                sign = "+" if net >= 0 else ""
                pnl_text = f"{sign}{net:.2f} USDT"
            except (TypeError, ValueError):
                pnl_text = "--"
        else:
            pnl_text = "--"
        self._set_auto_tune_state_value("pnl_30m", pnl_text)

        regime = (hyst.get("current_regime") or "-").lower()
        regime_text = regime.upper()
        # trend_up / trend_down → 초록, chop → 노랑, neutral/unknown → 흰색
        regime_color = (
            "#2EBD85" if regime in ("trend", "trend_up", "trend_down")
            else "#F0B90B" if regime == "chop"
            else "#f5f7ff"
        )
        self._set_auto_tune_state_value("regime", regime_text, regime_color)

        # hits (U/D/C) — 내부 디버그용, UI에서 제거

        configured_cd = int(self.settings_data.get("auto_tune_cooldown_min", 10) or 10)
        cooldown_until = data.get("cooldown_until")
        cooldown_color = "#2EBD85"
        _wait = "대기" if self.language == "ko" else "Idle"
        _cfg = "설정" if self.language == "ko" else "cfg"
        _min_unit = "분" if self.language == "ko" else "min"
        cooldown_text = f"{_wait} ({_cfg} {configured_cd}{_min_unit})"
        if cooldown_until:
            try:
                remaining = float(cooldown_until) - time.time()
                if remaining > 1:
                    mins = int(max(0, remaining) // 60)
                    secs = int(max(0, remaining) % 60)
                    if mins > 0:
                        _m_unit = "분" if self.language == "ko" else "m"
                        _s_unit = "초" if self.language == "ko" else "s"
                        _remain = "후 재시작" if self.language == "ko" else "remaining"
                        cooldown_text = f"{mins}{_m_unit} {secs}{_s_unit} {_remain} / {_cfg} {configured_cd}{_min_unit}"
                    else:
                        _s_unit = "초" if self.language == "ko" else "s"
                        _remain = "후 재시작" if self.language == "ko" else "remaining"
                        cooldown_text = f"{secs}{_s_unit} {_remain} / {_cfg} {configured_cd}{_min_unit}"
                    cooldown_color = "#F7C948" if remaining <= 300 else "#F6465D"
                else:
                    cooldown_text = "재가동 중" if self.language == "ko" else "Restarting"
            except (TypeError, ValueError):
                cooldown_text = "정보 없음" if self.language == "ko" else "Unknown"
                cooldown_color = "#F6465D"
        self._set_auto_tune_state_value("cooldown", cooldown_text, cooldown_color)

        # shadow_mode — 내부 상태, UI에서 제거
        shadow_active = bool(data.get("shadow_active"))

        def fmt_snapshot(snapshot):
            if not snapshot:
                return "없음" if self.language == "ko" else "None"
            regime = str(snapshot.get("regime") or "-").upper()
            rationale = snapshot.get("rationale") or "-"
            updated_at = snapshot.get("updated_at")
            if updated_at:
                try:
                    time_text = time.strftime("%H:%M:%S", time.localtime(float(updated_at)))
                except (OSError, ValueError, TypeError):
                    time_text = "--:--"
            else:
                time_text = "--:--"
            return f"{regime} · {rationale} · {time_text}"

        # ── 활성 스냅샷 + 적용 상태 시각화 ────────────────────────────────────
        _shadow_on = bool(data.get("shadow_active", False))
        _applied_at = active_snapshot.get("updated_at", 0)
        _last_reason = meta.get("last_reason") or ""

        if _last_reason == "applied" or "applied" in (active_snapshot.get("rationale") or ""):
            _status = ("✅ 적용됨" if self.language == "ko" else "✅ Applied")
            _s_color = "#2EBD85"
        elif _shadow_on:
            _status = ("🔍 Shadow 검증 중" if self.language == "ko" else "🔍 Shadow Testing")
            _s_color = "#F0B90B"
        elif "low_confidence" in _last_reason:
            _status = ("⏸ 신뢰도 부족" if self.language == "ko" else "⏸ Low Confidence")
            _s_color = "#8892a4"
        elif "cooldown" in _last_reason or time.time() < float(data.get("cooldown_until", 0) or 0):
            _status = ("⏳ 쿨다운" if self.language == "ko" else "⏳ Cooldown")
            _s_color = "#F6465D"
        elif active_snapshot:
            _status = ("✅ 적용됨" if self.language == "ko" else "✅ Applied")
            _s_color = "#2EBD85"
        else:
            _status = ("— 초기화됨" if self.language == "ko" else "— Reset")
            _s_color = "#8892a4"

        self._set_auto_tune_state_value("active_snapshot",
            f"{_status}  |  {fmt_snapshot(active_snapshot)}",
            _s_color)

        def fmt_ts(ts):
            if not ts:
                return "--"
            try:
                return time.strftime("%Y-%m-%d %H:%M:%S", time.localtime(float(ts)))
            except (OSError, ValueError, TypeError):
                return "--"

        # last_apply → last_update_label에 통합 표시 (별도 행 불필요)

        # ── 진입 차단 경고 배너 업데이트 ──────────────────────────────────────
        _banner = getattr(self, "_trade_block_banner", None)
        if _banner:
            try:
                if _banner.winfo_exists():
                    _mom_val = current.get("momentum_min_long", 0.0) or 0.0
                    _baseline_mom = 0.005  # 기본 baseline
                    _regime_cur = (hyst.get("current_regime") or "").lower()
                    _warn_msg = ""
                    if float(_mom_val) >= 0.007:
                        if self.language == "ko":
                            _warn_msg = (
                                f"⚠️  진입 차단 중 — 모멘텀 임계값이 {float(_mom_val):.4f}로 높습니다 "
                                f"(레짐: {_regime_cur.upper()}). 대부분 심볼이 이 조건을 충족하지 못해 진입이 막히고 있습니다. "
                                f"Auto-Tune Reset 후 재시작을 권장합니다."
                            )
                        else:
                            _warn_msg = (
                                f"⚠️  TRADES BLOCKED — momentum threshold {float(_mom_val):.4f} is too high "
                                f"(regime: {_regime_cur.upper()}). Most symbols fail this filter. "
                                f"Consider Auto-Tune Reset and restart."
                            )
                    elif float(_mom_val) >= 0.006:
                        if self.language == "ko":
                            _warn_msg = f"⚡ 모멘텀 임계값 {float(_mom_val):.4f} — 진입 기회가 제한됩니다."
                        else:
                            _warn_msg = f"⚡ Momentum threshold {float(_mom_val):.4f} — entry opportunities are limited."
                    if _warn_msg:
                        _banner.configure(
                            text=_warn_msg,
                            bg="#3a1a00" if float(_mom_val) >= 0.007 else "#1a2a0a",
                            fg="#ffaa44" if float(_mom_val) >= 0.007 else "#aaff88",
                        )
                        if not getattr(self, "_trade_block_banner_visible", False):
                            _banner.pack(fill="x", padx=18, pady=(0, 8), before=_banner.master.winfo_children()[1] if len(_banner.master.winfo_children()) > 1 else None)
                            self._trade_block_banner_visible = True
                    else:
                        if getattr(self, "_trade_block_banner_visible", False):
                            _banner.pack_forget()
                            self._trade_block_banner_visible = False
            except Exception:
                pass

        if not manual_override:
            field_vars = getattr(self, "trade_field_vars", {})
            auto_values = {
                "position_pct": current.get("position_pct"),
                "leverage_min": current.get("leverage_min"),
                "leverage_max": current.get("leverage_max"),
                "volatility_min": current.get("volatility_min"),
                "momentum_min_long": current.get("momentum_min_long"),
                "momentum_min_short": current.get("momentum_min_short"),
                # watch_limit/max_open_symbols는 사용자 설정값 유지 (auto-tune으로 덮어쓰지 않음)
                # "watch_limit": current.get("watch_limit"),
                # "max_open_symbols": current.get("max_open_symbols"),
                "max_loss_per_position": current.get("max_loss_per_position"),
            }
            for key, value in auto_values.items():
                if value is None:
                    continue
                var = field_vars.get(key)
                if not var:
                    continue
                try:
                    if key in ("watch_limit", "max_open_symbols", "top_n"):
                        var.set(int(value))
                    elif key in ("leverage_min", "leverage_max"):
                        var.set(int(round(value)))
                    else:
                        var.set(float(value))
                except (TypeError, ValueError, tk.TclError):
                    continue

        if self.auto_tune_last_update_label:
            try:
                timestamp = os.path.getmtime(AUTO_TUNE_STATE_PATH)
                time_text = time.strftime("%H:%M:%S", time.localtime(timestamp))
                label_text = (f"Updated: {time_text}" if self.language == "en" else f"업데이트: {time_text}")
            except OSError:
                label_text = ("Updated: (time unknown)" if self.language == "en" else "업데이트: 시간 정보 없음")
            if manual_override:
                label_text += (" · Manual mode" if self.language == "en" else " · 수동 모드")
                color = "#F0B90B"
            else:
                color = "#6c738a"
            try:
                if self.auto_tune_last_update_label.winfo_exists():
                    self.auto_tune_last_update_label.configure(text=label_text, fg=color)
            except Exception:
                pass

    def _schedule_auto_tune_state_refresh(self):
        self._refresh_auto_tune_state()
        self.auto_tune_refresh_job = self.root.after(5000, self._schedule_auto_tune_state_refresh)

    def _update_account_balance(self):
        creds = self._require_credentials(silent=True, action="계좌 잔고 갱신")
        if not creds:
            return
        from binance_futures_bot1_1.main import fetch_account_balance

        def on_success(data):
            available = data.get("availableBalance")
            if available is None:
                available = data.get("totalWalletBalance", 0.0) - getattr(self, "last_total_margin", 0.0)
            try:
                balance_value = float(available)
            except (TypeError, ValueError):
                balance_value = 0.0
            self.last_account_balance = balance_value
            self._set_stat_value("account_balance", f"{balance_value:.2f} USDT")
            # 수동 매매 패널 명목금 라벨 갱신
            _nlbl = getattr(self, "_manual_notional_lbl", None)
            if _nlbl and _nlbl.winfo_exists():
                try:
                    pct = float(self.manual_pct_var.get())
                    lev = max(1, int(self.manual_lev_var2.get()))
                    notional = balance_value * pct * lev
                    margin   = balance_value * pct
                    _nlbl.config(text=f"명목 {notional:,.1f} U  /  증거금 {margin:,.1f} U")
                except Exception:
                    pass
            self._record_balance_snapshot(int(time.time() * 1000), balance_value)

        self._run_async(fetch_account_balance(creds), success_cb=on_success, fail_msg="잔고 조회 실패")

    def _update_income_stats(self):
        creds = self._require_credentials(silent=True, action="손익 데이터 갱신")
        if not creds:
            return
        from binance_futures_bot1_1.main import fetch_income_history
        now_ms = int(time.time() * 1000)
        start_ms = now_ms - (24 * 60 * 60 * 1000)

        def on_success(rows, current_time=now_ms):
            self._apply_income_history(rows, current_time)

        self._run_async(
            fetch_income_history(creds, start_ms),
            success_cb=on_success,
            fail_msg="손익 데이터 갱신 실패",
        )

    def _apply_income_history(self, rows, now_ms):
        rows = list(rows or [])
        source = "api" if rows else "none"
        # 현재 카드에 활성화된 PnL 윈도우만 계산
        _pnl_window_map = {
            "pnl_15":  15 * 60 * 1000,
            "pnl_60":  60 * 60 * 1000,
            "pnl_12h": 12 * 60 * 60 * 1000,
            "pnl_24h": 24 * 60 * 60 * 1000,
        }
        _active_modes = set(self.card_modes.values())
        windows = [
            (k, v) for k, v in _pnl_window_map.items()
            if k in _active_modes or k == "pnl_24h"
        ]
        if not rows:
            synthetic_rows = self._reconstruct_income_from_log(now_ms)
            if synthetic_rows:
                rows = synthetic_rows
                source = "log"
            else:
                trade_rows = self._reconstruct_income_from_trade_log(now_ms)
                if trade_rows:
                    rows = trade_rows
                    source = "trade_log"
                    self._append_log("[INFO] 손익 패널: trade history fallback 사용")
        # 손익 패널: trade_log(수동+자동) + API rows 병합 → 명목금(pnl/leverage) 표시
        # trade_log는 명목금으로 저장됨, API rows는 leverage 나눠서 정규화
        tl_rows = self._reconstruct_income_from_trade_log(now_ms)


        if tl_rows:
            # trade_log가 있으면 단독 사용 (수동+자동 모두 포함)
            pnl_source = tl_rows
        else:
            # API rows 폴백: REALIZED_PNL+COMMISSION 합산 (Binance 실제 손익)
            pnl_source = [
                r for r in rows
                if str(r.get("incomeType") or r.get("type") or "REALIZED_PNL").upper()
                   in {"REALIZED_PNL", "COMMISSION"}
            ] or rows

        for key, delta in windows:
            start_time = now_ms - delta
            if self.pnl_reset_ms:
                start_time = max(start_time, self.pnl_reset_ms)
            total = sum(
                float(item.get("income", 0.0))
                for item in pnl_source
                if int(item.get("time", 0)) >= start_time
            )
            sign = "+" if total > 0 else ""
            self._set_stat_value(key, f"{sign}{total:.2f} USDT")

        def filter_rows(reset_key=None, dataset=None):
            data = rows if dataset is None else dataset
            cutoff = 0
            if self.session_start_ms:
                cutoff = max(cutoff, int(self.session_start_ms))
            if reset_key:
                reset_ts = self.stat_resets.get(reset_key)
                if reset_ts:
                    cutoff = max(cutoff, int(reset_ts))
            if cutoff:
                return [item for item in data if int(item.get("time", 0)) >= cutoff]
            return data

        trade_history_rows = self._reconstruct_income_from_trade_log(now_ms)

        def _is_partial(item):
            """부분청산(PARTIAL_TP_X)은 거래 건수/승률 집계에서 제외."""
            trigger = str(item.get("trigger", "")).upper()
            return trigger.startswith("PARTIAL")

        def _is_closed_trade(item):
            """완전 청산 거래만 허용 (income != 0, partial 아닌 것)."""
            return abs(float(item.get("income", 0.0))) > 1e-6 and not _is_partial(item)

        # trade_log = 수동+자동 청산 포함 → 우선 사용
        # API rows = 자동 거래만 → trade_log 없을 때 폴백
        # 두 소스를 합산: 중복 가능하지만 통계에는 큰 영향 없음
        if trade_history_rows:
            # trade_log 있으면 단독 사용 (mode=manual/auto 모두 포함)
            stats_dataset = trade_history_rows
        else:
            # API rows 폴백 (REALIZED_PNL만)
            stats_dataset = [
                r for r in rows
                if abs(float(r.get("income", 0.0))) > 1e-6
                and str(r.get("incomeType", "REALIZED_PNL")).upper() == "REALIZED_PNL"
            ]

        win_rows   = filter_rows("win_rate",    dataset=stats_dataset)
        trade_rows = filter_rows("trade_count", dataset=stats_dataset)

        # 완전 청산 건만 집계
        win_rows   = [item for item in win_rows   if _is_closed_trade(item)]
        trade_rows = [item for item in trade_rows if _is_closed_trade(item)]
        trade_count = len(trade_rows)
        wins = sum(1 for item in win_rows if float(item.get("income", 0.0)) > 0) if win_rows else 0
        win_rate = (wins / len(win_rows) * 100) if win_rows else 0.0
        self._set_stat_value("trade_count", str(trade_count))
        self._set_stat_value("win_rate", f"{win_rate:.1f}%")
        if trade_history_rows:
            self._append_log(
                f"[STATS] trade_log={len(trade_history_rows)}건 "
                f"→ filtered={len(trade_rows)}건 wins={wins} "
                f"session_cutoff={self.session_start_ms}"
            )

        # ── 전환 가능 카드: 현재 활성화된 모드별 갱신 ──────────────
        _active = set(self.card_modes.values()) if hasattr(self, "card_modes") else set()
        if "unrealized_total" in _active:
            try:
                total_unrealized = self._calc_unrealized_net()
                sign  = "+" if total_unrealized > 0 else ""
                color = "#2EBD85" if total_unrealized > 0 else ("#F6465D" if total_unrealized < 0 else "#ffffff")
                lbl   = self.stats_labels.get("unrealized_total")
                if lbl:
                    lbl.config(text=self._trim_usdt(f"{sign}{total_unrealized:.2f} USDT"), fg=color)
            except Exception:
                pass
        if "filter_pass_rate" in _active:
            try:
                self._update_filter_pass_rate_card()
            except Exception:
                pass
        if "top_symbol" in _active:
            try:
                self._update_top_symbol_card(trade_rows)
            except Exception:
                pass
        if "rr_ratio" in _active:
            try:
                self._update_rr_ratio_card()
            except Exception:
                pass
        if "expectancy" in _active:
            try:
                self._update_expectancy_card()
            except Exception:
                pass
        if "max_consec_loss" in _active:
            try:
                self._update_max_consec_loss_card()
            except Exception:
                pass

    def _calc_unrealized_net(self) -> float:
        """현재 포지션 미실현 손익에서 진입+청산 수수료를 차감한 순 손익.

        ═══════════════════════════════════════════════════════════
        🔧 수정: 진입 + 청산 수수료 모두 테이커 기준 (0.05%)
        ═══════════════════════════════════════════════════════════
        - 진입 수수료: 테이커 0.05%
        - 청산 수수료: 테이커 0.05%
        - 보수적 계산: 실제보다 약간 적게 보여주어 안전 마진 확보
        """
        # maker_fee = float(self.settings_data.get("maker_fee_pct", 0.0002) or 0.0002)
        taker_fee = float(self.settings_data.get("taker_fee_pct", 0.0005) or 0.0005)
        # maker_first = bool(self.settings_data.get("maker_first_enabled", True))
        total_net = 0.0
        for pos in self.current_positions.values():
            try:
                unrealized = float(pos.get("unRealizedProfit", 0.0))
                qty        = abs(float(pos.get("amount", 0.0)))
                entry_px   = float(pos.get("entryPrice", 0.0) or 0.0)
                mark_price = float(pos.get("markPrice", 0.0) or entry_px)
                
                # ═══════════════════════════════════════════════════════════
                # 🔧 수정: 진입 + 청산 모두 테이커 수수료 적용
                # ═══════════════════════════════════════════════════════════
                # 진입 수수료: 테이커 0.05%
                entry_fee  = qty * entry_px  * taker_fee
                
                # 청산 수수료: 테이커 0.05%
                exit_fee   = qty * mark_price * taker_fee
                
                total_net += unrealized - entry_fee - exit_fee
            except (TypeError, ValueError):
                continue
        return total_net

    def _calc_unrealized_gross(self) -> float:
        """수수료 차감 전 미실현 손익 합계."""
        return sum(
            float(p.get("unRealizedProfit", 0.0))
            for p in self.current_positions.values()
            if p.get("unRealizedProfit") is not None
        )

    def _update_rr_ratio_card(self):
        """손익비 R:R 카드 업데이트."""
        lbl = self.stats_labels.get("rr_ratio")
        if not lbl:
            return
        try:
            all_rows = self._reconstruct_income_from_trade_log(int(time.time() * 1000))
            wins   = [float(r.get("income", 0.0)) for r in all_rows if float(r.get("income", 0.0)) > 0]
            losses = [float(r.get("income", 0.0)) for r in all_rows if float(r.get("income", 0.0)) < 0]
            if wins and losses:
                avg_win  = sum(wins)  / len(wins)
                avg_loss = abs(sum(losses) / len(losses))
                rr = avg_win / avg_loss if avg_loss > 0 else 0.0
                _en = self.language == "en"
                sub = f"  +{avg_win:.2f} / -{avg_loss:.2f} USDT" if not _en else f"  +{avg_win:.2f} / -{avg_loss:.2f}"
                lbl.config(text=f"{rr:.2f}x{sub}", fg="#F7C948" if rr >= 1 else "#F6465D")
            else:
                lbl.config(text="—", fg="#8892a4")
        except Exception:
            lbl.config(text="—", fg="#8892a4")

    def _update_expectancy_card(self):
        """거래당 기댓값 카드 업데이트."""
        lbl = self.stats_labels.get("expectancy")
        if not lbl:
            return
        try:
            all_rows = self._reconstruct_income_from_trade_log(int(time.time() * 1000))
            vals = [float(r.get("income", 0.0)) for r in all_rows]
            if vals:
                exp = sum(vals) / len(vals)
                sign  = "+" if exp >= 0 else ""
                color = "#2EBD85" if exp >= 0 else "#F6465D"
                lbl.config(text=self._trim_usdt(f"{sign}{exp:.4f} USDT") + f"  ({len(vals)}건)", fg=color)
            else:
                lbl.config(text="—", fg="#8892a4")
        except Exception:
            lbl.config(text="—", fg="#8892a4")

    def _update_max_consec_loss_card(self):
        """최대 연속 손실 카드 업데이트."""
        lbl = self.stats_labels.get("max_consec_loss")
        if not lbl:
            return
        try:
            all_rows = self._reconstruct_income_from_trade_log(int(time.time() * 1000))
            sorted_rows = sorted(all_rows, key=lambda r: int(r.get("time", 0)))
            max_cl = cur_cl = 0
            for r in sorted_rows:
                if float(r.get("income", 0.0)) < 0:
                    cur_cl += 1
                    max_cl = max(max_cl, cur_cl)
                else:
                    cur_cl = 0
            _en = self.language == "en"
            unit = " times" if _en else "회"
            color = "#F6465D" if max_cl >= 3 else ("#F7C948" if max_cl >= 2 else "#2EBD85")
            lbl.config(text=f"{max_cl}{unit}" if max_cl > 0 else "—", fg=color)
        except Exception:
            lbl.config(text="—", fg="#8892a4")

    def _update_filter_pass_rate_card(self):
        """필터 통과율 카드 업데이트: notifications.log의 Filter summary 파싱."""
        import re as _re
        _en = self.language == "en"
        pattern = _re.compile(
            r"Filter summary: input=(\d+) topN=(\d+) passed=(\d+)"
        )
        latest = None
        if os.path.exists(self.notification_path):
            try:
                with open(self.notification_path, "r", encoding="utf-8", errors="ignore") as fh:
                    lines = fh.readlines()
                for line in reversed(lines):
                    m = pattern.search(line)
                    if m:
                        inp = int(m.group(1))
                        passed = int(m.group(3))
                        latest = (inp, passed)
                        break
            except Exception:
                pass
        # bot.log fallback
        if latest is None:
            try:
                log_lines = self._read_recent_log_lines(limit=500)
                for line in reversed(log_lines):
                    m = pattern.search(line)
                    if m:
                        latest = (int(m.group(1)), int(m.group(3)))
                        break
            except Exception:
                pass

        lbl = self.stats_labels.get("filter_pass_rate")
        if not lbl:
            return
        if latest:
            inp, passed = latest
            pct = (passed / inp * 100) if inp > 0 else 0.0
            bar = "▓" * int(pct / 10) + "░" * (10 - int(pct / 10))
            color = "#2EBD85" if pct >= 30 else ("#F7C948" if pct >= 10 else "#F6465D")
            lbl.config(text=f"{pct:.1f}%  {passed}/{inp}", fg=color)
        else:
            lbl.config(text="— %", fg="#8892a4")

    def _update_top_symbol_card(self, trade_rows):
        """최다 거래 심볼 카드 업데이트: trade_history.jsonl 직접 읽기."""
        lbl = self.stats_labels.get("top_symbol")
        if not lbl:
            return
        # _reconstruct_income_from_trade_log 은 symbol 필드를 포함하지 않으므로
        # 항상 TRADE_LOG_PATH 를 직접 읽어 symbol + pnl 집계
        counts = {}
        pnl_by_sym = {}
        try:
            if os.path.exists(TRADE_LOG_PATH):
                with open(TRADE_LOG_PATH, "r", encoding="utf-8") as fh:
                    for line in fh:
                        line = line.strip()
                        if not line:
                            continue
                        try:
                            data = json.loads(line)
                        except Exception:
                            continue
                        sym = data.get("symbol", "")
                        if not sym:
                            continue
                        pnl = float(data.get("pnl", 0.0))
                        counts[sym] = counts.get(sym, 0) + 1
                        pnl_by_sym[sym] = pnl_by_sym.get(sym, 0.0) + pnl
        except Exception:
            pass
        if not counts:
            lbl.config(text="—", fg="#8892a4")
            return
        top = max(counts, key=counts.get)
        n = counts[top]
        sym_pnl = pnl_by_sym.get(top, 0.0)
        sign = "+" if sym_pnl >= 0 else ""
        color = "#2EBD85" if sym_pnl >= 0 else "#F6465D"
        clean_sym = top.replace("USDT", "").replace("usdt", "")
        _unit = "건" if self.language == "ko" else "trades"
        lbl.config(text=f"{clean_sym}  {n}{_unit}", fg=color)

    def _reconstruct_income_from_log(self, now_ms):
        entries = []
        if not os.path.exists(self.notification_path):
            return entries
        try:
            with open(self.notification_path, "r", encoding="utf-8") as fh:
                lines = fh.readlines()
        except OSError:
            return entries
        pnl_by_symbol = {}
        for line in lines[-1000:]:
            if "주문 체결" not in line:
                continue
            try:
                parts = line.strip().split("|")
                if len(parts) < 3:
                    continue
                _, _, body = parts
                if "주문 체결" not in body:
                    continue
                symbol = None
                direction = None
                if " LONG " in body:
                    direction = "LONG"
                elif " SHORT " in body:
                    direction = "SHORT"
                try:
                    symbol = body.split(" ")[2]
                except Exception:
                    continue
                if not symbol:
                    continue
                pnl_by_symbol.setdefault(symbol, {"entry_ts": now_ms, "last_price": 0.0})
            except Exception:
                continue
        reconstructed = []
        for symbol, info in pnl_by_symbol.items():
            reconstructed.append({"time": info.get("entry_ts", now_ms), "income": info.get("pnl", 0.0)})
        return reconstructed

    def _reconstruct_income_from_trade_log(self, now_ms):
        path = TRADE_LOG_PATH
        if not os.path.exists(path):
            return []
        rows = []
        try:
            with open(path, "r", encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if not line:
                        continue
                    try:
                        data = json.loads(line)
                    except json.JSONDecodeError:
                        continue
                    ts = data.get("ts")
                    try:
                        ts_ms = int(float(ts) * 1000)
                    except (TypeError, ValueError):
                        ts_ms = now_ms
                    pnl_value = data.get("pnl", 0.0)
                    try:
                        pnl_value = float(pnl_value)
                    except (TypeError, ValueError):
                        pnl_value = 0.0
                    roi_pct = data.get("roi_pct", None)
                    try:
                        roi_pct = float(roi_pct) if roi_pct is not None else None
                    except (TypeError, ValueError):
                        roi_pct = None
                    rows.append({
                        "time": ts_ms,
                        "income": pnl_value,   # 실제 손익 USDT (수수료 포함)
                        "roi_pct": roi_pct,
                        "trigger": data.get("trigger", ""),
                    })
        except OSError:
            return []
        return rows

    def _log_manual_trade_event(self, symbol, position, resp, trigger):
        if not position:
            self._append_log(f"[WARN] _log_manual_trade_event: position is None for {symbol}")
            return
        try:
            qty = abs(float(position.get("amount", 0.0)))
            entry_price = float(position.get("entryPrice", 0.0))
        except (TypeError, ValueError):
            return
        if qty <= 0 or entry_price <= 0:
            self._append_log(f"[WARN] _log_manual_trade_event: invalid qty={qty} entry={entry_price}")
            return

        side = position.get("side", "LONG")
        try:
            lev = max(1.0, float(position.get("leverage") or position.get("lev") or 1.0))
        except (TypeError, ValueError):
            lev = 1.0

        # ── PnL 계산: unRealizedProfit(Binance 계산값) + 수수료 차감 ──
        margin_value = float(position.get("marginValue", 0.0) or 0.0)
        if margin_value <= 0:
            margin_value = (entry_price * qty) / lev

        # 체결가: 주문 응답 avgPrice → markPrice → symbol_price_map 순서
        exit_price = self._extract_order_fill_price(resp)
        if not exit_price or exit_price <= 0:
            try:
                exit_price = float(position.get("markPrice") or position.get("breakEvenPrice") or 0.0)
            except (TypeError, ValueError):
                exit_price = 0.0
        if not exit_price or exit_price <= 0:
            try:
                exit_price = float(self.symbol_price_map.get(symbol, 0.0))
            except (TypeError, ValueError, AttributeError):
                exit_price = 0.0
        if not exit_price or exit_price <= 0:
            exit_price = entry_price

        # unRealizedProfit = Binance가 계산한 수수료 미포함 손익
        unrealized = float(position.get("unRealizedProfit", 0.0) or 0.0)
        try:
            fee_rate = float(self.settings_data.get("taker_fee_pct", 0.0005) or 0.0)
        except (TypeError, ValueError):
            fee_rate = 0.0005
        fees = (entry_price + exit_price) * qty * fee_rate
        # 실제 PnL = 미실현손익 - 수수료 (Binance unrealizedProfit 우선)
        if unrealized != 0.0:
            pnl = unrealized - fees
        else:
            # unrealizedProfit 없으면 직접 계산
            direction = 1.0 if side == "LONG" else -1.0
            pnl = direction * (exit_price - entry_price) * qty - fees

        roi_pct = (pnl / margin_value * 100) if margin_value > 0 else 0.0

        event = {
            "ts": time.time(),
            "symbol": symbol,
            "side": side,
            "quantity": qty,
            "entry_price": entry_price,
            "exit_price": exit_price,
            "pnl": pnl,
            "roi_pct": roi_pct,
            "leverage": lev,
            "trigger": trigger or "manual",
            "order_id": resp.get("orderId") if isinstance(resp, dict) else None,
            "mode": "manual",
            "env": "testnet" if getattr(self, "env_mode", "TESTNET") == "TESTNET" else "live",
        }
        self._append_log(
            f"[INFO] 수동 청산 기록: {symbol} {side} pnl={pnl:.4f} USDT "
            f"entry={entry_price} exit={exit_price} qty={qty} lev={lev}x"
        )
        try:
            os.makedirs(os.path.dirname(TRADE_LOG_PATH), exist_ok=True)
            with open(TRADE_LOG_PATH, "a", encoding="utf-8") as fh:
                json.dump(event, fh, ensure_ascii=False)
                fh.write("\n")
        except OSError as exc:
            self._append_log(f"[WARN] 수동 거래 로그 실패: {exc}")

    def _extract_order_fill_price(self, resp):
        if not isinstance(resp, dict):
            return None
        price_fields = ["avgPrice", "avg_price", "executedPrice", "price"]
        for field in price_fields:
            value = resp.get(field)
            if value:
                try:
                    return float(value)
                except (TypeError, ValueError):
                    continue
        fills = resp.get("fills")
        if isinstance(fills, list) and fills:
            try:
                totals = sum(float(item.get("price", 0.0)) * float(item.get("qty", item.get("quantity", 0.0))) for item in fills)
                qty_sum = sum(float(item.get("qty", item.get("quantity", 0.0))) for item in fills)
                if qty_sum > 0:
                    return totals / qty_sum
            except (TypeError, ValueError):
                pass
        return None

    def _compute_manual_pnl(self, entry_price, exit_price, qty, side, leverage: float = 1.0):
        try:
            entry_price = float(entry_price)
            exit_price = float(exit_price)
            qty = float(qty)
        except (TypeError, ValueError):
            return 0.0
        direction = 1.0 if str(side).upper() == "LONG" else -1.0
        pnl = direction * (exit_price - entry_price) * qty
        fee_rate = 0.0
        try:
            fee_rate = float(self.settings_data.get("taker_fee_pct", 0.0005) or 0.0)
        except (TypeError, ValueError):
            fee_rate = 0.0
        if fee_rate > 0:
            fees = (abs(entry_price) + abs(exit_price)) * qty * fee_rate
            pnl -= fees
        return pnl  # 실제 계좌 손익 USDT (leverage 반영, 수수료 포함)

    def _record_balance_snapshot(self, timestamp_ms, balance):
        history = getattr(self, "balance_history", None)
        if history is None:
            return
        try:
            balance_value = float(balance)
        except (TypeError, ValueError):
            return
        timestamp_ms = int(timestamp_ms)
        cutoff = timestamp_ms - (24 * 60 * 60 * 1000)
        while history and history[0]["time"] < cutoff:
            history.popleft()
        if history and abs(history[-1]["balance"] - balance_value) < 1e-6:
            history[-1]["time"] = timestamp_ms
            return
        history.append({"time": timestamp_ms, "balance": balance_value})

    def _fallback_income_from_balance(self):
        history = getattr(self, "balance_history", None)
        if not history or len(history) < 2:
            return []
        rows = []
        items = list(history)
        prev = items[0]
        for curr in items[1:]:
            delta = float(curr["balance"] - prev["balance"])
            if abs(delta) > 1e-4:
                rows.append({"time": curr["time"], "income": delta})
            prev = curr
        return rows

    def _release_client_and_exit(self, callback):
        from binance_futures_bot1_1.main import close_client_session

        pending = [task for task in asyncio.all_tasks(self.loop) if not task.done()]
        for task in pending:
            task.cancel()

        future = asyncio.run_coroutine_threadsafe(close_client_session(), self.loop)

        def finish():
            try:
                future.result()
            except Exception:
                pass
            callback()

        self.root.after(0, finish)

    def _on_close(self):
        def finalize():
            try:
                self.loop.call_soon_threadsafe(self.loop.stop)
            except Exception:
                pass
            try:
                self.loop_thread.join(timeout=1)
            except Exception:
                pass
            self.root.destroy()

        if self.position_refresh_job:
            self.root.after_cancel(self.position_refresh_job)
            self.position_refresh_job = None
        if self.position_countdown_job:
            try:
                self.root.after_cancel(self.position_countdown_job)
            except Exception:
                pass
            self.position_countdown_job = None
        if self.stats_refresh_job:
            self.root.after_cancel(self.stats_refresh_job)
            self.stats_refresh_job = None
        if self.auto_tune_refresh_job:
            self.root.after_cancel(self.auto_tune_refresh_job)
            self.auto_tune_refresh_job = None

        if self.engine_running:
            def after_stop():
                self._release_client_and_exit(finalize)

            self._shutdown_alert_suppressed = True
            self.stop_engine(on_stopped=after_stop)
        else:
            self._release_client_and_exit(finalize)

    def _poll_notifications(self):
        self._trim_file(self.notification_path)
        try:
            size = os.path.getsize(self.notification_path)
        except FileNotFoundError:
            return self._schedule_notification_poll()
        if size < self.notification_pointer:
            self.notification_pointer = 0
        try:
            with open(self.notification_path, "r", encoding="utf-8") as fh:
                fh.seek(self.notification_pointer)
                new_data = fh.read()
                self.notification_pointer = fh.tell()
        except FileNotFoundError:
            new_data = ""
        stats_refresh_needed = False
        positions_refresh_needed = False
        if new_data:
            for line in new_data.strip().splitlines():
                parts = line.split("|", 2)
                if len(parts) != 3:
                    continue
                _, level, message = parts
                level = level.strip()
                message = message.strip()
                self._append_log(f"[{level}] {message}")
                normalized = message.lower()
                trade_tokens = ["fill", "체결", "청산", "order", "주문", "거래"]
                if any(token in normalized for token in trade_tokens):
                    stats_refresh_needed = True
                    positions_refresh_needed = True
                alert_tokens = ["fill", "체결", "청산", "order", "주문", "거래", "long", "short", "buy", "sell", "매수", "매도"]
                if level == "ALERT" and any(token in normalized for token in alert_tokens):
                    self._show_auto_closing_alert(level, message)
        if stats_refresh_needed:
            self._trigger_stat_refresh()
            self._trigger_stat_refresh(2000)
            self._trigger_stat_refresh(5000)
        if positions_refresh_needed:
            self._burst_refresh_positions(delay_ms=1500)
        self._schedule_notification_poll()

    def _show_auto_closing_alert(self, level, message):
        self.active_alerts = [dlg for dlg in self.active_alerts if dlg.winfo_exists()]
        while self.active_alerts:
            dlg = self.active_alerts.pop()
            try:
                dlg.destroy()
            except Exception:
                pass
        if not getattr(self, "alert_enabled", True):
            return
        dialog = tk.Toplevel(self.root)
        _title = self._t("alert", "Notice")
        dialog.title(_title)
        self._apply_icon(dialog)
        width, height = 380, 200
        dialog.configure(bg="#0C1017", highlightthickness=1, highlightbackground="#394058")
        dialog.resizable(False, False)
        dialog.geometry(f"{width}x{height}")
        self._center_modal(dialog, width, height)
        # 레벨별 아이콘/색상
        _icon_map = {"ALERT": ("🔔", "#F7C948"), "WARN": ("⚠", "#F6465D"), "INFO": ("ℹ", "#5EB0FF")}
        _icon, _color = _icon_map.get(level.upper(), ("🔔", "#F7C948"))
        container = tk.Frame(dialog, bg="#0C1017")
        container.pack(fill="both", expand=True, padx=18, pady=12)
        tk.Label(container, text=_icon, bg="#0C1017", fg=_color, font=("Segoe UI Symbol", 20)).pack(anchor="center", pady=(0, 4))
        tk.Label(container, text=message, bg="#0C1017", fg="#e0e6ff", wraplength=340,
                 font=("Malgun Gothic", 10), justify="center").pack(fill="both", expand=True)
        tk.Button(container, text=self._t("dlg_ok", "OK"), command=dialog.destroy,
                  bg="#1f5c3a", fg="#ffffff", relief=tk.FLAT,
                  activebackground="#2EBD85", activeforeground="#ffffff",
                  font=("Malgun Gothic", 10, "bold"), padx=20, pady=5, cursor="hand2").pack(pady=(8, 0))
        dialog.after(10000, lambda: dialog.winfo_exists() and dialog.destroy())
        self.active_alerts.append(dialog)

    # ─────────────────────────────────────────────────────────────
    # 모니터링 탭
    # ─────────────────────────────────────────────────────────────
    def _build_monitor_tab(self, parent):
        """모니터링 탭 UI — 내부 서브 노트북 3탭 구조."""
        BG   = "#0f1118"
        BG2  = "#141822"
        FG   = "#e0e6ff"
        FG2  = "#8b93b7"
        ACC  = "#2EBD85"
        WARN = "#F7C948"
        ERR  = "#F6465D"
        FONT_BOLD  = ("Segoe UI", 10, "bold")
        FONT_SMALL = ("Segoe UI", 9)

        parent.columnconfigure(0, weight=1)
        parent.rowconfigure(0, weight=0)   # 공통 필터 바
        parent.rowconfigure(1, weight=1)   # 서브 노트북

        # ── 공통 필터 현황 바 (항상 표시) ────────────────────────
        stat_bar = tk.Frame(parent, bg=BG, highlightbackground="#2a3145",
                            highlightthickness=1)
        stat_bar.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 3))

        tk.Label(stat_bar, text=f"  {self._t('monitor_filter_title','필터 현황')}",
                 bg=BG, fg=FG, font=FONT_BOLD).pack(side=tk.LEFT, padx=(4, 10), pady=4)

        stat_chips = [
            ("monitor_input",    "—", "#394058"),
            ("monitor_topn",     "—", "#394058"),
            ("monitor_passed",   "—", ACC),
            ("monitor_vol_fail", "—", WARN),
            ("monitor_mom_fail", "—", ERR),
        ]
        self.monitor_filter_stat_labels = {}
        for key, default, color in stat_chips:
            chip = tk.Frame(stat_bar, bg=color)
            chip.pack(side=tk.LEFT, padx=3, pady=4)
            tk.Label(chip, text=self._t(key, key), bg=color, fg="#fff",
                     font=FONT_SMALL, padx=5, pady=2).pack(side=tk.LEFT)
            val_lbl = tk.Label(chip, text=default, bg=color, fg="#fff",
                               font=("Segoe UI", 10, "bold"), padx=5, pady=2)
            val_lbl.pack(side=tk.LEFT)
            self.monitor_filter_stat_labels[key] = (chip, val_lbl)

        self.monitor_last_update_label = tk.Label(
            stat_bar, text="", bg=BG, fg=FG2, font=FONT_SMALL)
        self.monitor_last_update_label.pack(side=tk.RIGHT, padx=10)

        tk.Button(stat_bar, text=f"⟳  {self._t('monitor_refresh','새로고침')}",
                  command=self._refresh_monitor_tab, bg="#1f2538", fg=FG,
                  relief=tk.FLAT, font=FONT_SMALL, padx=8, pady=2,
                  activebackground="#2a3352", activeforeground="#fff",
                  cursor="hand2").pack(side=tk.RIGHT, padx=4)

        # ── 서브 노트북 스타일 ────────────────────────────────
        style = ttk.Style()
        style.configure("MonitorNB.TNotebook",
                        background=BG, borderwidth=0, tabmargins=(6, 6, 6, 0))
        style.configure("MonitorNB.TNotebook.Tab",
                        background="#161c2a", foreground="#7a84a8",
                        padding=(14, 4, 14, 2),
                        font=("Malgun Gothic", 10, "bold"), borderwidth=0)
        style.map("MonitorNB.TNotebook.Tab",
                  background=[("selected", "#253050"), ("!selected", "#161c2a")],
                  foreground=[("selected", "#ffffff"), ("!selected", "#7a84a8")],
                  padding=[("selected", (18, 6, 18, 4)), ("!selected", (14, 4, 14, 2))])

        sub_nb = ttk.Notebook(parent, style="MonitorNB.TNotebook")
        sub_nb.grid(row=1, column=0, sticky="nsew", padx=8, pady=(0, 4))

        # ── 탭 1: 모니터링 심볼 ──────────────────────────────
        watch_tab = tk.Frame(sub_nb, bg=BG)
        sub_nb.add(watch_tab, text=self._t("monitor_watch_title", "모니터링 심볼"))
        watch_tab.columnconfigure(0, weight=1)
        watch_tab.rowconfigure(0, weight=1)

        wf = tk.Frame(watch_tab, bg=BG2, highlightbackground="#2a3145",
                      highlightthickness=1)
        wf.grid(row=0, column=0, sticky="nsew", padx=6, pady=6)
        wf.columnconfigure(0, weight=1)
        wf.rowconfigure(0, weight=1)

        self.monitor_watch_tree = ttk.Treeview(
            wf, columns=("symbol", "last_msg"), show="headings",
            style="Positions.Treeview")
        self.monitor_watch_tree.heading("symbol",   text=self._t("monitor_col_symbol", "심볼"))
        self.monitor_watch_tree.heading("last_msg", text=self._t("monitor_col_status", "상태"))
        self.monitor_watch_tree.column("symbol",   width=120, anchor="center")
        self.monitor_watch_tree.column("last_msg", width=500, anchor="w")
        self.monitor_watch_tree.grid(row=0, column=0, sticky="nsew")
        _ws = ttk.Scrollbar(wf, orient="vertical", command=self.monitor_watch_tree.yview)
        _ws.grid(row=0, column=1, sticky="ns")
        self.monitor_watch_tree.configure(yscrollcommand=_ws.set)
        self.monitor_watch_tree.tag_configure("watch",        foreground="#5EB0FF")
        self.monitor_watch_tree.tag_configure("system_watch", foreground="#9aa5c6", font=("Segoe UI", 9))
        self.monitor_watch_tree.tag_configure("odd",           background=BG2)
        self.monitor_watch_tree.tag_configure("even",          background="#181e2c")

        # ── 탭 2: 필터 탈락 / 스킵 ───────────────────────────
        reject_tab = tk.Frame(sub_nb, bg=BG)
        sub_nb.add(reject_tab, text=self._t("monitor_reject_title", "필터 탈락 / 스킵"))
        reject_tab.columnconfigure(0, weight=1)
        reject_tab.rowconfigure(1, weight=1)

        # 탭 2 헤더 — 탈락/스킵 범례
        legend = tk.Frame(reject_tab, bg=BG)
        legend.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 2))
        tk.Label(legend, text="●", bg=BG, fg="#F6465D",
                 font=("Segoe UI", 11)).pack(side=tk.LEFT)
        tk.Label(legend, text=("Rejected" if self.language == "en" else "탈락"),
                 bg=BG, fg=FG2, font=FONT_SMALL).pack(side=tk.LEFT, padx=(2, 14))
        tk.Label(legend, text="●", bg=BG, fg="#F7C948",
                 font=("Segoe UI", 11)).pack(side=tk.LEFT)
        tk.Label(legend, text=("Skipped" if self.language == "en" else "스킵"),
                 bg=BG, fg=FG2, font=FONT_SMALL).pack(side=tk.LEFT, padx=(2, 0))

        rf = tk.Frame(reject_tab, bg=BG2, highlightbackground="#2a3145",
                      highlightthickness=1)
        rf.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 6))
        rf.columnconfigure(0, weight=1)
        rf.rowconfigure(0, weight=1)

        self.monitor_reject_tree = ttk.Treeview(
            rf, columns=("symbol", "reason", "time"), show="headings",
            style="Positions.Treeview")
        self.monitor_reject_tree.heading("symbol", text=self._t("monitor_col_symbol", "심볼"))
        self.monitor_reject_tree.heading("reason", text=self._t("monitor_col_reason", "사유"))
        self.monitor_reject_tree.heading("time",   text=self._t("monitor_col_time",   "시각"))
        self.monitor_reject_tree.column("symbol", width=120, anchor="center")
        self.monitor_reject_tree.column("reason", width=460, anchor="w")
        self.monitor_reject_tree.column("time",   width=80,  anchor="center")
        self.monitor_reject_tree.grid(row=0, column=0, sticky="nsew")
        _rs = ttk.Scrollbar(rf, orient="vertical", command=self.monitor_reject_tree.yview)
        _rs.grid(row=0, column=1, sticky="ns")
        self.monitor_reject_tree.configure(yscrollcommand=_rs.set)
        self.monitor_reject_tree.tag_configure("rejected", foreground="#F6465D")
        self.monitor_reject_tree.tag_configure("skipped",  foreground="#F7C948")
        self.monitor_reject_tree.tag_configure("odd",  background=BG2)
        self.monitor_reject_tree.tag_configure("even", background="#181e2c")

        # ── 탭 3: Auto-tune 상태 ─────────────────────────────
        atune_tab = tk.Frame(sub_nb, bg=BG)
        sub_nb.add(atune_tab, text=self._t("monitor_autotune_title", "Auto-tune 상태"))
        atune_tab.columnconfigure(0, weight=1)
        atune_tab.rowconfigure(1, weight=1)

        # 탭 3 헤더 — 색 범례
        at_legend = tk.Frame(atune_tab, bg=BG)
        at_legend.grid(row=0, column=0, sticky="ew", padx=8, pady=(6, 2))
        for dot_color, dot_label in [
            ("#F7C948", "Tightened" if self.language == "en" else "조임"),
            ("#2EBD85", "Relaxed"   if self.language == "en" else "완화"),
            ("#e0e6ff", "Neutral"   if self.language == "en" else "기본"),
        ]:
            tk.Label(at_legend, text="●", bg=BG, fg=dot_color,
                     font=("Segoe UI", 11)).pack(side=tk.LEFT)
            tk.Label(at_legend, text=dot_label, bg=BG, fg=FG2,
                     font=FONT_SMALL).pack(side=tk.LEFT, padx=(2, 12))

        atf = tk.Frame(atune_tab, bg=BG2, highlightbackground="#2a3145",
                       highlightthickness=1)
        atf.grid(row=1, column=0, sticky="nsew", padx=6, pady=(0, 6))
        atf.columnconfigure(0, weight=1)
        atf.rowconfigure(0, weight=1)

        self.monitor_autotune_tree = ttk.Treeview(
            atf, columns=("param", "current", "base"), show="headings",
            style="Positions.Treeview")
        self.monitor_autotune_tree.heading("param",   text=self._t("monitor_col_param",  "파라미터"))
        self.monitor_autotune_tree.heading("current", text=self._t("monitor_col_value",  "현재값"))
        self.monitor_autotune_tree.heading("base",    text=self._t("monitor_col_base",   "기본값"))
        self.monitor_autotune_tree.column("param",   width=240, anchor="w")
        self.monitor_autotune_tree.column("current", width=160, anchor="center")
        self.monitor_autotune_tree.column("base",    width=160, anchor="center")
        self.monitor_autotune_tree.grid(row=0, column=0, sticky="nsew")
        _ats = ttk.Scrollbar(atf, orient="vertical",
                             command=self.monitor_autotune_tree.yview)
        _ats.grid(row=0, column=1, sticky="ns")
        self.monitor_autotune_tree.configure(yscrollcommand=_ats.set)
        self.monitor_autotune_tree.tag_configure("tightened", foreground="#F7C948")
        self.monitor_autotune_tree.tag_configure("relaxed",   foreground="#2EBD85")
        self.monitor_autotune_tree.tag_configure("neutral",   foreground="#e0e6ff")
        self.monitor_autotune_tree.tag_configure("odd",  background=BG2)
        self.monitor_autotune_tree.tag_configure("even", background="#181e2c")

        self.root.after(500, self._refresh_monitor_tab)

    def _refresh_monitor_tab(self):
        """모니터링 탭 데이터를 로그에서 파싱해 갱신."""
        if not (hasattr(self, "monitor_watch_tree") and self.monitor_watch_tree):
            return

        import re as _re
        _en = self.language == "en"

        # ── 1. 로그 읽기 ──────────────────────────────────────
        log_lines = self._read_recent_log_lines(limit=800)
        notif_entries = []
        if os.path.exists(self.notification_path):
            try:
                with open(self.notification_path, "r", encoding="utf-8", errors="ignore") as fh:
                    notif_entries = fh.readlines()[-400:]
            except Exception:
                pass

        # ── 2. Filter summary 파싱 ────────────────────────────
        # 실제 엔진 포맷:
        # Filter summary: input=211 topN=20 passed=3 status=0 blocked=0 spike=0 vol_fail=5 quality=0
        filter_pattern = _re.compile(
            r"Filter summary: input=(\d+) topN=(\d+) passed=(\d+)"
            r".*?vol_fail=(\d+).*?quality=(\d+)"
        )
        ts_pattern = _re.compile(r"^(\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2})")
        latest_filter = None
        latest_filter_ts = ""
        # Filter summary는 notifications.log(WATCH) 우선, fallback bot.log
        _filter_sources = list(reversed(notif_entries)) + list(reversed(log_lines))
        for line in _filter_sources:
            m = filter_pattern.search(line)
            if m:
                latest_filter = {
                    "input": m.group(1), "top": m.group(2),
                    "passed": m.group(3), "vol": m.group(4), "quality": m.group(5)
                }
                # 타임스탬프: 앞 19자에서 HH:MM:SS 추출
                stripped_l = line.strip()
                ts_m = ts_pattern.match(stripped_l)
                if ts_m:
                    latest_filter_ts = ts_m.group(1)[11:]
                break

        if latest_filter:
            chip_data = {
                "monitor_input":    (latest_filter["input"],   "#394058"),
                "monitor_topn":     (latest_filter["top"],     "#394058"),
                "monitor_passed":   (latest_filter["passed"],
                                     "#2EBD85" if int(latest_filter["passed"]) > 0 else "#F6465D"),
                "monitor_vol_fail": (latest_filter["vol"],
                                     "#F7C948" if int(latest_filter["vol"]) > 0 else "#394058"),
                "monitor_mom_fail": (latest_filter["quality"],
                                     "#F7C948" if int(latest_filter["quality"]) > 0 else "#394058"),
            }
            for key, (val, color) in chip_data.items():
                entry = self.monitor_filter_stat_labels.get(key)
                if entry:
                    chip_frame, val_lbl = entry
                    try:
                        val_lbl.config(text=val)
                        chip_frame.config(bg=color)
                        for child in chip_frame.winfo_children():
                            child.config(bg=color)
                    except Exception:
                        pass
            _ts_txt = (f"{self._t('monitor_last_filter','최근 필터 실행')}: {latest_filter_ts}"
                       if latest_filter_ts else "")
            try:
                self.monitor_last_update_label.config(text=_ts_txt)
            except Exception:
                pass

        # ── 3. Watchlist (notifications.log |WATCH|) ─────────
        watch_map = {}    # symbol → msg  (심볼 이벤트)
        system_watch = []  # 심볼 없는 시스템 WATCH 항목
        for line in notif_entries:
            if "|WATCH|" in line:
                parts = line.strip().split("|", 2)
                if len(parts) == 3:
                    ts_raw, _, msg = parts
                    msg = msg.strip()
                    # 메시지 첫 단어가 USDT 심볼인지 먼저 확인 (가장 일반적인 포맷)
                    first_word = msg.split()[0] if msg else ""
                    if first_word.upper().endswith("USDT") and len(first_word) >= 5:
                        sym = first_word.upper()
                        watch_map[sym] = msg
                    else:
                        # 심볼 없는 WATCH — 시스템 메시지(AutoTune 등)
                        ts_str = ts_raw.strip()[11:] if len(ts_raw.strip()) >= 19 else ts_raw.strip()
                        system_watch.append((ts_str, msg))

        try:
            for row in self.monitor_watch_tree.get_children():
                self.monitor_watch_tree.delete(row)
            rows_added = 0
            # 심볼 이벤트 (최신순)
            for idx, (sym, msg) in enumerate(reversed(list(watch_map.items()))):
                status = msg.replace(sym, "").strip(" :-|")[:80]
                tag = ("watch", "even" if idx % 2 == 0 else "odd")
                self.monitor_watch_tree.insert("", tk.END, values=(sym, status), tags=tag)
                rows_added += 1
            # 시스템 WATCH (AutoTune 등) — 별도 섹션으로 최신 5개
            for idx, (ts, msg) in enumerate(reversed(system_watch[-5:])):
                label = f"⚙ {ts}" if ts else "⚙ System"
                tag = ("system_watch", "even" if (rows_added + idx) % 2 == 0 else "odd")
                self.monitor_watch_tree.insert("", tk.END, values=(label, msg[:80]), tags=tag)
                rows_added += 1
            if rows_added == 0:
                self.monitor_watch_tree.insert("", tk.END,
                    values=("—", self._t("monitor_no_data","데이터 없음 — 엔진 실행 후 표시됩니다")),
                    tags=("watch",))
        except Exception:
            pass

        # ── 4. Rejected / Skipped (notifications.log |WATCH| 라인) ──────
        # 엔진이 notifications.log에 쓰는 포맷:
        #   YYYY-MM-DD HH:MM:SS|WATCH|SIGNAL_REJECT BTCUSDT 사유...
        #   YYYY-MM-DD HH:MM:SS|WATCH|FILTER_REJECT BTCUSDT FILTER_REJECT_VOL vol=...
        #   YYYY-MM-DD HH:MM:SS|WATCH|ENTRY_BLOCKED_GLOBAL BTCUSDT: MAX_OPEN
        #   YYYY-MM-DD HH:MM:SS|WATCH|ENTRY_BLOCKED_SPIKE BTCUSDT 30.0s
        #   YYYY-MM-DD HH:MM:SS|WATCH|심볼 제외: BTCUSDT 반복 실패
        #   YYYY-MM-DD HH:MM:SS|WATCH|Symbol excluded: BTCUSDT repeated failures
        reject_pat = _re.compile(
            r"(?:SIGNAL_REJECT|FILTER_REJECT|ENTRY_BLOCKED[_A-Z]*)\s+([A-Z0-9]{3,20})\s*(.*)",
            _re.IGNORECASE
        )
        skip_pat = _re.compile(
            r"(?:심볼 제외|Symbol excluded)[:\s]+([A-Z0-9]{3,20})\s*(.*)",
            _re.IGNORECASE
        )
        # 사유 라벨 한/영 변환
        reason_labels = {
            "FILTER_REJECT_VOL":      ("변동성 부족",        "Low volatility"),
            "FILTER_REJECT_BLOCKED":  ("심볼 차단",          "Symbol blocked"),
            "FILTER_REJECT_SPIKE":    ("스파이크 쿨다운",    "Spike cooldown"),
            "FILTER_REJECT_SPREAD":   ("스프레드 초과",      "Spread too wide"),
            "FILTER_REJECT_MARK_GAP": ("마크갭 초과",        "Mark gap too wide"),
            "FILTER_REJECT_QUALITY":  ("퀄리티 스코어 부족", "Low quality score"),
            "FILTER_REJECT_STATUS":   ("거래 불가 상태",     "Non-tradable status"),
            "MAX_OPEN_SYMBOLS":       ("최대 포지션 도달",   "Max positions reached"),
            "MAX_OPEN":               ("최대 포지션 도달",   "Max positions reached"),
            "ALREADY_OPEN":           ("이미 포지션 존재",   "Position already open"),
        }

        reject_entries = []
        skip_entries   = []

        for line in reversed(notif_entries):
            stripped = line.strip()
            if "|WATCH|" not in stripped:
                continue
            # 타임스탬프 추출 (앞 19자)
            ts_str = stripped[:19] if len(stripped) >= 19 else ""
            ts_str = ts_str[11:]  # HH:MM:SS만
            # |WATCH| 이후 메시지
            parts = stripped.split("|WATCH|", 1)
            msg = parts[1].strip() if len(parts) == 2 else ""

            if len(reject_entries) < 50:
                rm = reject_pat.match(msg)
                if rm:
                    sym    = rm.group(1).upper()
                    detail = rm.group(2).strip()
                    # FILTER_REJECT_XXX 라벨 변환
                    label = detail
                    for code, (ko, en) in reason_labels.items():
                        if code in detail.upper():
                            label = en if _en else ko
                            break
                    # ENTRY_BLOCKED 사유 정리
                    if "ENTRY_BLOCKED" in msg.upper():
                        block_type = msg.split()[0].replace("ENTRY_BLOCKED_", "")
                        bl_labels = {
                            "GLOBAL": ("글로벌 차단", "Global block"),
                            "SPIKE":  ("스파이크 쿨다운", "Spike cooldown"),
                            "MARK_GAP": ("마크갭", "Mark gap"),
                        }
                        bl = bl_labels.get(block_type, (block_type, block_type))
                        label = (bl[1] if _en else bl[0]) + (f": {detail}" if detail else "")
                    reject_entries.append((sym, label[:90], ts_str, "rejected"))

            if len(skip_entries) < 20:
                sm = skip_pat.match(msg)
                if sm:
                    sym    = sm.group(1).upper()
                    reason = sm.group(2).strip()
                    label  = f"{'Repeated fail' if _en else '반복 실패'}" + (f" — {reason}" if reason else "")
                    skip_entries.append((sym, label[:90], ts_str, "skipped"))

            if len(reject_entries) >= 50 and len(skip_entries) >= 20:
                break

        try:
            for row in self.monitor_reject_tree.get_children():
                self.monitor_reject_tree.delete(row)
            combined = reject_entries + skip_entries
            combined.sort(key=lambda x: x[2], reverse=True)
            if combined:
                for idx, (sym, reason, ts, tag) in enumerate(combined):
                    row_tags = (tag, "even" if idx % 2 == 0 else "odd")
                    self.monitor_reject_tree.insert("", tk.END,
                        values=(sym, reason, ts), tags=row_tags)
            else:
                self.monitor_reject_tree.insert("", tk.END,
                    values=("—", self._t("monitor_no_data","데이터 없음 — 엔진 실행 후 표시됩니다"), ""),
                    tags=("skipped",))
        except Exception:
            pass

        # ── 5. Auto-tune 파라미터 ─────────────────────────────
        # gui_config.json을 직접 재읽어 가장 최신 설정값을 반영
        fresh_settings = self._load_json(CONFIG_PATH, default=self.settings_data)
        at_data = self._load_json(AUTO_TUNE_STATE_PATH, default={})

        # 모드 = gui_config 기준이 유일한 진실 소스
        # (auto_tuner_state의 "mode"는 Auto-tune 사이클이 실제로 돌아야만 갱신되므로
        #  재시작 직후에는 이전 세션 값이 남아 있어 신뢰할 수 없음)
        cfg_mode = fresh_settings.get("auto_tune_mode",
                   self.settings_data.get("auto_tune_mode", "balanced"))

        base_defaults = {
            "volatility_min":     fresh_settings.get("volatility_min", 0.001),
            "momentum_min_long":  fresh_settings.get("momentum_min_long", 0.001),
            "momentum_min_short": fresh_settings.get("momentum_min_short", -0.001),
            "position_pct":       fresh_settings.get("position_pct", 0.05),
            "leverage_min":       fresh_settings.get("leverage_min", 5),
            "leverage_max":       fresh_settings.get("leverage_max", 25),
        }
        label_map = {
            "volatility_min":     self._t("monitor_autotune_vol", "변동성 최소"),
            "momentum_min_long":  self._t("monitor_autotune_mom", "모멘텀 long"),
            "momentum_min_short": ("Momentum short" if _en else "모멘텀 short"),
            "position_pct":       self._t("monitor_autotune_pos", "포지션 크기"),
            "leverage_min":       ("Leverage min" if _en else "레버리지 min"),
            "leverage_max":       ("Leverage max" if _en else "레버리지 max"),
        }
        effective = at_data.get("effective_params", at_data.get("params",
                    at_data.get("current", {})))
        score = at_data.get("score", "—")
        # 상태 파일의 updated_at 이용: 최근 Auto-tune 사이클 시각 표시용
        updated_at = at_data.get("updated_at") or at_data.get("meta", {}).get("updated_at")

        try:
            for row in self.monitor_autotune_tree.get_children():
                self.monitor_autotune_tree.delete(row)

            # 모드 행 — gui_config 기준 (항상 최신)
            mode_label = f"{'Mode (config)' if _en else '모드 (설정값)'}"
            self.monitor_autotune_tree.insert("", tk.END,
                values=(mode_label, cfg_mode, "—"),
                tags=("neutral",))

            # 스코어 행
            if score != "—":
                s_str = f"{score:.3f}" if isinstance(score, float) else str(score)
                self.monitor_autotune_tree.insert("", tk.END,
                    values=(f"{'Score' if _en else '스코어'}", s_str, "—"),
                    tags=("neutral",))

            # 마지막 Auto-tune 사이클 시각
            if updated_at:
                import datetime
                try:
                    ts_str = datetime.datetime.fromtimestamp(float(updated_at)).strftime("%m-%d %H:%M")
                    cycle_label = f"{'Last cycle' if _en else '마지막 튜닝'}"
                    self.monitor_autotune_tree.insert("", tk.END,
                        values=(cycle_label, ts_str, "—"),
                        tags=("neutral",))
                except Exception:
                    pass
            for key, label in label_map.items():
                base = base_defaults.get(key, "—")
                current = effective.get(key, base)
                try:
                    c_f, b_f = float(current), float(base)
                    if abs(c_f) < abs(b_f) - 1e-9:
                        tag = "tightened"
                    elif abs(c_f) > abs(b_f) + 1e-9:
                        tag = "relaxed"
                    else:
                        tag = "neutral"
                    c_str = f"{c_f:.4f}" if abs(c_f) < 1 else f"{c_f:.2f}"
                    b_str = f"{b_f:.4f}" if abs(b_f) < 1 else f"{b_f:.2f}"
                except (TypeError, ValueError):
                    tag, c_str, b_str = "neutral", str(current), str(base)
                self.monitor_autotune_tree.insert("", tk.END,
                    values=(label, c_str, b_str), tags=(tag,))
        except Exception:
            pass

        # ── 6. 10초 자동 갱신 ─────────────────────────────────
        if getattr(self, "_monitor_refresh_job", None):
            try:
                self.root.after_cancel(self._monitor_refresh_job)
            except Exception:
                pass
        self._monitor_refresh_job = self.root.after(10000, self._refresh_monitor_tab)

    def _read_recent_log_lines(self, limit=400):
        self._trim_file(LOG_PATH)
        if not os.path.exists(LOG_PATH):
            return []
        with open(LOG_PATH, "r", encoding="utf-8", errors="ignore") as fh:
            lines = fh.readlines()
        return lines[-limit:]

    def _extract_symbol_watchlist(self):
        if not os.path.exists(self.notification_path):
            return []
        with open(self.notification_path, "r", encoding="utf-8", errors="ignore") as fh:
            lines = fh.readlines()
        watch_entries = []
        for line in reversed(lines):
            if "|WATCH|" in line:
                parts = line.strip().split("|", 2)
                if len(parts) == 3:
                    _, _, message = parts
                    watch_entries.append(message)
            if len(watch_entries) >= 40:
                break
        return list(reversed(watch_entries))

    def _analyze_filter_logs(self):
        lines = self._read_recent_log_lines()
        _en = self.language == "en"
        if not lines:
            return ("Log file not found. The engine may not have run yet." if _en
                    else "로그 파일을 찾을 수 없습니다. 아직 엔진이 실행되지 않았을 수 있습니다.")

        summary_line = None
        warning_line = None
        rejection_samples = []
        skip_samples = []
        for line in reversed(lines):
            stripped = line.strip()
            if not summary_line and "Filter summary" in stripped:
                summary_line = stripped
            if not warning_line and "No tradable symbols" in stripped:
                warning_line = stripped
            if "FILTER_REJECT_" in stripped and len(rejection_samples) < 5:
                rejection_samples.append(stripped)
            if "Skip count" in stripped and len(skip_samples) < 5:
                skip_samples.append(stripped)
            if summary_line and len(rejection_samples) >= 5 and len(skip_samples) >= 5:
                break

        parts = ["Recent log analysis" if _en else "최근 로그 분석 결과", "" ]
        pattern = re.compile(r"Filter summary: input=(?P<input>\d+) topN=(?P<top>\d+) passed=(?P<passed>\d+).*?vol_fail=(?P<vol>\d+).*?quality=(?P<quality>\d+)")
        if summary_line:
            match = pattern.search(summary_line)
            if match:
                data = match.groupdict()
                if _en:
                    parts.append("- Filter stats: checked top {top} of {input} symbols, {passed} passed".format(**data))
                    parts.append(f"- Volatility fail: {data['vol']}, Quality fail: {data['quality']}")
                else:
                    parts.append("- 최근 필터 통계: 전체 {input}개 중 상위 {top}개를 검사, {passed}개 통과".format(**data))
                    parts.append(f"- 변동성 조건 미달: {data['vol']}개, 품질 조건 미달: {data['quality']}개")
                if data["passed"] == "0":
                    parts.append("")
                    if _en:
                        parts.append("→ All symbols rejected — filters may be too strict. Try lowering volatility/momentum thresholds or check market conditions.")
                    else:
                        parts.append("→ 현재 조건이 엄격해 모든 심볼이 탈락했습니다. 변동성/모멘텀 임계값을 조정하거나 시장 상황을 확인하세요.")
            else:
                parts.append(summary_line)
        else:
            parts.append("No recent filter summary log. Check if the engine has been running long enough." if _en else "최근 필터 요약 로그가 없습니다. 엔진이 충분히 실행되었는지 확인하세요.")

        if warning_line:
            parts.append("")
            parts.append(f"{'Warning' if _en else '경고'}: {warning_line}")

        if rejection_samples:
            parts.append("")
            parts.append("Rejected samples (latest 5):" if _en else "필터 탈락 샘플 (최신 5개):")
            for entry in rejection_samples[:5]:
                parts.append(f"  • {entry}")

        if skip_samples:
            parts.append("")
            parts.append("Skip counts (latest 5):" if _en else "스킵 카운트 (최신 5개):")
            for entry in skip_samples[:5]:
                parts.append(f"  • {entry}")

        if len(parts) == 2:
            parts.append("Insufficient log data. Try again after the engine has run for a while." if _en else "분석 가능한 로그가 부족합니다. 엔진 실행 후 다시 시도하세요.")
        return "\n".join(parts)

    # ------------------------------------------------------------------
    def run(self):
        try:
            asyncio.set_event_loop_policy(asyncio.WindowsSelectorEventLoopPolicy())
        except Exception:
            pass
        self.root.mainloop()


if __name__ == "__main__":
    try:
        BotGUI().run()
    finally:
        try:
            from binance_futures_bot1_1.main import _close_client_session_sync
            _close_client_session_sync()
        except Exception:
            pass